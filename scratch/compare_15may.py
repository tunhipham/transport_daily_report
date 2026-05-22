# -*- coding: utf-8 -*-
"""Compare DB vs local file for 15/05/2026"""
import sys, os, json, re
from collections import defaultdict
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, r'g:\My Drive\DOCS\transport_daily_report\script')
from data_pipeline.config import load_clickhouse_config
from openpyxl import load_workbook
import requests

cfg = load_clickhouse_config()
p = {'user': cfg['user'], 'password': cfg['password'], 'database': cfg['database']}
def q(sql):
    r = requests.get(cfg['base_url'], params={**p, 'query': sql}, timeout=120)
    r.raise_for_status()
    return r.text.strip()

KRC = '5fdc170ebd89c10006f15b7c'

# Load master data
from lib.sources import MASTER_SHEET_URL
from io import BytesIO
r_m = requests.get(MASTER_SHEET_URL, allow_redirects=True, timeout=120)
r_m.raise_for_status()
wb = load_workbook(BytesIO(r_m.content), read_only=True, data_only=True)
ws = wb.worksheets[0]
master = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    bc = str(row[0].value or "").strip()
    if not bc: continue
    tl = row[25].value
    if tl is not None:
        try:
            w = float(tl)
            if w > 0: master[bc] = w
        except: pass
wb.close()
print(f"Master: {len(master)} barcodes\n")

def extract_wt_kg(name):
    if not name: return 0
    text = name.upper()
    for pat, mult in [(r'(\d+(?:[.,]\d+)?)\s*KG\b',1.0),(r'(\d+(?:[.,]\d+)?)\s*G\b',0.001)]:
        m = re.findall(pat, text)
        if m:
            try: return float(m[-1].replace(",",".")) * mult
            except: continue
    return 0

# ═══════════════════════════════════════════════════════════
# 1) LOCAL FILE — 15/05/2026
# ═══════════════════════════════════════════════════════════
print("=" * 70)
print("  LOCAL FILE: po_krc_15052026.xlsx")
print("=" * 70)

local_file = r'G:\My Drive\DOCS\DAILY\po_krc\po_krc_15052026.xlsx'
if os.path.exists(local_file):
    wb = load_workbook(local_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    local_tons = 0
    local_rows = 0
    local_po_set = set()
    local_with_kg = 0
    local_with_master = 0
    local_with_name = 0
    local_no_wt = 0
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        date_val = row[15].value  # Col P
        if date_val is None: continue
        if isinstance(date_val, datetime):
            ds = date_val.strftime('%d/%m/%Y')
        else:
            ds = str(date_val).strip()
        if ds != '15/05/2026': continue
        
        local_rows += 1
        po = str(row[3].value or '').strip()
        local_po_set.add(po)
        bc = str(row[18].value or '').strip()  # Col S
        pname = str(row[19].value or '').strip()  # Col T
        
        try: qty = float(row[21].value or 0)  # Col V
        except: qty = 0
        try: kg = float(row[22].value or 0)   # Col W
        except: kg = 0
        
        if kg > 0:
            local_tons += kg / 1000
            local_with_kg += 1
        elif bc and bc in master:
            wt_g = master[bc]
            local_tons += qty * wt_g / 1_000_000
            local_with_master += 1
        else:
            wkg = extract_wt_kg(pname)
            if wkg > 0:
                local_tons += qty * wkg / 1000
                local_with_name += 1
            else:
                local_no_wt += 1
    
    wb.close()
    print(f"  Rows (15/05): {local_rows}")
    print(f"  POs: {len(local_po_set)}")
    print(f"  TOTAL: {local_tons:.2f} tấn")
    print(f"  Sources: col_W(kg)={local_with_kg}, master={local_with_master}, name={local_with_name}, skip={local_no_wt}")
else:
    print(f"  ⚠ File not found: {local_file}")

# ═══════════════════════════════════════════════════════════
# 2) DB — 15/05/2026
# ═══════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("  DB: ClickHouse — 15/05/2026")
print("=" * 70)

r1 = q(f"""
SELECT
    ri.product_barcode AS barcode,
    ri.product_name AS pname,
    ri.qty AS qty,
    ri.net_weight AS net_weight,
    ri.purchase_code AS po_code
FROM kf_receipt_items ri
INNER JOIN kf_purchase_order po
    ON ri.purchase_code = po.code
    AND po.branch_id = '{KRC}'
    AND po.deleted = 0
WHERE ri.branch_id = '{KRC}'
  AND toDate(fromUnixTimestamp(toUInt32(po.delivery_date))) = '2026-05-15'
FORMAT JSONEachRow
""")

db_tons = 0
db_rows = 0
db_po_set = set()
db_with_nw = 0
db_with_master = 0
db_with_name = 0
db_no_wt = 0
db_tons_nw_only = 0

for line in r1.split('\n'):
    if not line.strip(): continue
    obj = json.loads(line)
    db_rows += 1
    
    bc = str(obj.get('barcode','')).strip()
    pname = str(obj.get('pname','')).strip()
    qty = float(obj.get('qty', 0))
    nw = float(obj.get('net_weight', 0))
    po = str(obj.get('po_code','')).strip()
    db_po_set.add(po)
    
    if qty <= 0: continue
    
    if nw > 0:
        tons = qty * nw / 1_000_000
        db_tons += tons
        db_tons_nw_only += tons
        db_with_nw += 1
    elif bc and bc in master:
        db_tons += qty * master[bc] / 1_000_000
        db_with_master += 1
    else:
        wkg = extract_wt_kg(pname)
        if wkg > 0:
            db_tons += qty * wkg / 1000
            db_with_name += 1
        else:
            db_no_wt += 1

print(f"  Rows: {db_rows}")
print(f"  POs: {len(db_po_set)}")
print(f"  TOTAL: {db_tons:.2f} tấn")
print(f"  (net_weight only: {db_tons_nw_only:.2f} tấn)")
print(f"  Sources: DB_nw={db_with_nw}, master={db_with_master}, name={db_with_name}, skip={db_no_wt}")

# ═══════════════════════════════════════════════════════════
# 3) Compare PO codes
# ═══════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("  PO CODE COMPARISON")
print("=" * 70)
if os.path.exists(local_file):
    both = local_po_set & db_po_set
    only_local = local_po_set - db_po_set
    only_db = db_po_set - local_po_set
    print(f"  Local POs: {len(local_po_set)}")
    print(f"  DB POs:    {len(db_po_set)}")
    print(f"  Both:      {len(both)}")
    print(f"  Only local: {len(only_local)}")
    print(f"  Only DB:    {len(only_db)}")
    
    if only_db:
        print(f"\n  Sample POs only in DB (first 10):")
        for po in list(only_db)[:10]:
            print(f"    {po}")

# ═══════════════════════════════════════════════════════════
# 4) DIFF — What's causing the gap?
# ═══════════════════════════════════════════════════════════
if os.path.exists(local_file):
    print(f"\n{'='*70}")
    print(f"  DIFF: DB ({db_tons:.2f}T) vs Local ({local_tons:.2f}T) = Δ{db_tons-local_tons:+.2f}T")
    print("=" * 70)
    
    # Tons from DB-only POs
    db_only_tons = 0
    for line in r1.split('\n'):
        if not line.strip(): continue
        obj = json.loads(line)
        po = str(obj.get('po_code','')).strip()
        if po in only_db:
            qty = float(obj.get('qty',0))
            nw = float(obj.get('net_weight',0))
            bc = str(obj.get('barcode','')).strip()
            pname = str(obj.get('pname','')).strip()
            if nw > 0: db_only_tons += qty * nw / 1_000_000
            elif bc in master: db_only_tons += qty * master[bc] / 1_000_000
    
    print(f"  Tonnage from DB-only POs: {db_only_tons:.2f} tấn")
    print(f"  → DB có nhiều hơn {len(only_db)} POs = +{db_only_tons:.2f}T")

print("\nDone.")
