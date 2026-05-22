# -*- coding: utf-8 -*-
"""Deep dive 8/5 and 15/5 — after dedup, why still high?"""
import sys, os, json, re
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, r'g:\My Drive\DOCS\transport_daily_report\script')
from data_pipeline.config import load_clickhouse_config
import requests

cfg = load_clickhouse_config()
p = {'user': cfg['user'], 'password': cfg['password'], 'database': cfg['database']}
def q(sql):
    r = requests.get(cfg['base_url'], params={**p, 'query': sql}, timeout=120)
    r.raise_for_status()
    return r.text.strip()

KRC = '5fdc170ebd89c10006f15b7c'

# Load master
from lib.sources import MASTER_SHEET_URL
from io import BytesIO
from openpyxl import load_workbook as lw
r_m = requests.get(MASTER_SHEET_URL, allow_redirects=True, timeout=120)
r_m.raise_for_status()
wb = lw(BytesIO(r_m.content), read_only=True, data_only=True)
ws = wb.worksheets[0]
master = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    bc = str(row[0].value or "").strip()
    if not bc: continue
    tl = row[25].value
    if tl is not None:
        try:
            w = float(tl)
            if w > 0: master[bc] = w
        except: pass
wb.close()

def resolve_wt(nw, bc, pname):
    if nw > 0: return nw, 'db'
    if bc and bc in master: return master[bc], 'master'
    text = (pname or '').upper()
    for pat, mult in [(r'(\d+(?:[.,]\d+)?)\s*KG\b',1000),(r'(\d+(?:[.,]\d+)?)\s*G\b',1)]:
        m = re.findall(pat, text)
        if m:
            try: return float(m[-1].replace(",","."))*mult, 'name'
            except: pass
    return 0, 'none'

for target_date in ['2026-05-08', '2026-05-15']:
    d_label = target_date[8:10] + '/05'
    
    # Deduped query (same as capacity_forecast.py)
    r1 = q(f"""
    SELECT
        ri.purchase_code AS po_code,
        ri.product_barcode AS barcode,
        any(ri.product_name) AS pname,
        any(ri.po_qty) AS qty,
        any(ri.net_weight) AS net_weight,
        any(ri.qty) AS actual_qty,
        any(ri.qty_receipt) AS qty_receipt
    FROM kf_receipt_items ri
    INNER JOIN kf_purchase_order po
        ON ri.purchase_code = po.code
        AND po.branch_id = '{KRC}' AND po.deleted = 0
    WHERE ri.branch_id = '{KRC}'
      AND toDate(fromUnixTimestamp(toUInt32(po.delivery_date_vendor_confirm))) = '{target_date}'
    GROUP BY ri.purchase_code, ri.product_barcode
    FORMAT JSONEachRow
    """)
    
    items = []
    total_tons = 0
    for line in r1.split('\n'):
        if not line.strip(): continue
        obj = json.loads(line)
        qty = float(obj.get('qty', 0))
        nw = float(obj.get('net_weight', 0))
        bc = str(obj.get('barcode','')).strip()
        pname = str(obj.get('pname','')).strip()
        po = obj.get('po_code','')
        aq = float(obj.get('actual_qty', 0))
        qr = float(obj.get('qty_receipt', 0))
        
        wg, src = resolve_wt(nw, bc, pname)
        if qty <= 0 or wg <= 0: continue
        tons = qty * wg / 1_000_000
        total_tons += tons
        items.append({'po': po, 'bc': bc, 'pname': pname, 'qty': qty, 'aq': aq, 'qr': qr, 'wg': wg, 'tons': tons, 'src': src})
    
    print(f"\n{'='*70}")
    print(f"  {d_label}: {total_tons:.2f} tấn ({len(items)} items after dedup)")
    print(f"{'='*70}")
    
    # Compare po_qty vs qty vs qty_receipt
    total_po_qty_tons = sum(i['tons'] for i in items)
    total_actual_tons = sum(i['aq'] * i['wg'] / 1_000_000 for i in items if i['wg'] > 0)
    total_receipt_tons = sum(i['qr'] * i['wg'] / 1_000_000 for i in items if i['wg'] > 0)
    
    print(f"\n  Tonnage by qty field:")
    print(f"    po_qty (đặt):     {total_po_qty_tons:.2f}T")
    print(f"    qty (actual):     {total_actual_tons:.2f}T")
    print(f"    qty_receipt (nhận): {total_receipt_tons:.2f}T")
    
    # Count POs
    po_set = set(i['po'] for i in items)
    print(f"\n  POs: {len(po_set)}")
    
    # Top items
    items.sort(key=lambda x: x['tons'], reverse=True)
    print(f"\n  Top 10 items:")
    print(f"  {'Barcode':<15} {'po_qty':>7} {'qty':>7} {'rcpt':>7} {'Wt(g)':>7} {'Tons':>7} {'Src':>6}  Product")
    for i in items[:10]:
        print(f"  {i['bc']:<15} {i['qty']:>7.0f} {i['aq']:>7.0f} {i['qr']:>7.0f} {i['wg']:>7.0f} {i['tons']:>7.3f} {i['src']:>6}  {i['pname'][:35]}")
    
    # Check: items where po_qty >> qty (over-ordered?)
    over = [i for i in items if i['qty'] > i['aq'] * 1.5 and i['aq'] > 0]
    if over:
        over_tons = sum(i['tons'] - i['aq']*i['wg']/1_000_000 for i in over)
        print(f"\n  ⚠ po_qty >> qty items: {len(over)} (extra {over_tons:.2f}T)")
        for i in over[:5]:
            diff = i['qty'] - i['aq']
            print(f"    {i['bc']} po_qty={i['qty']:.0f} vs qty={i['aq']:.0f} Δ={diff:.0f} (+{diff*i['wg']/1_000_000:.3f}T)")

    # Also: compare with local file if available
    local_file = rf'G:\My Drive\DOCS\DAILY\po_krc\po_krc_{target_date[8:10]}052026.xlsx'
    if os.path.exists(local_file):
        wb2 = lw(local_file, read_only=True, data_only=True)
        ws2 = wb2.worksheets[0]
        local_rows = sum(1 for _ in ws2.iter_rows(min_row=2)) 
        wb2.close()
        print(f"\n  📎 Local file exists: {os.path.basename(local_file)} ({local_rows} rows)")
    else:
        print(f"\n  📎 No local file for {d_label}")

print("\nDone.")
