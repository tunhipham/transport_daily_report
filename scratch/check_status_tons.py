# -*- coding: utf-8 -*-
"""Check tonnage with status filter + master data weight mapping"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
BASE = r'g:\My Drive\DOCS\transport_daily_report'
sys.path.insert(0, os.path.join(BASE, 'script'))
from data_pipeline.config import load_clickhouse_config
import requests

cfg = load_clickhouse_config()
params = {'user': cfg['user'], 'password': cfg['password'], 'database': cfg['database']}

def q(sql):
    r = requests.get(cfg['base_url'], params={**params, 'query': sql}, timeout=120)
    r.raise_for_status()
    return r.text.strip()

KRC = '5fdc170ebd89c10006f15b7c'

# Load master data (same source as transfer PT)
print("Loading master data...")
from lib.sources import MASTER_SHEET_URL
from io import BytesIO
from openpyxl import load_workbook

r_master = requests.get(MASTER_SHEET_URL, allow_redirects=True, timeout=120)
r_master.raise_for_status()
wb = load_workbook(BytesIO(r_master.content), read_only=True, data_only=True)
ws = wb.worksheets[0]
master = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    bc = str(row[0].value or "").strip()
    if not bc: continue
    tl = row[25].value
    if tl is not None:
        try:
            w = float(tl)
            if w > 0: master[bc] = w
        except: pass
wb.close()
print(f"  Master: {len(master)} barcodes")

# Fetch row-level data for T5/2026
print("\nQuerying row-level PO KRC T5/2026...")
r1 = q(f"""
SELECT
    formatDateTime(fromUnixTimestamp(toUInt32(po.delivery_date)), '%d/%m/%Y') AS del_date,
    ri.product_barcode AS barcode,
    ri.product_name AS pname,
    ri.qty AS qty,
    ri.net_weight AS net_weight,
    po.status AS po_status
FROM kf_receipt_items ri
INNER JOIN kf_purchase_order po
    ON ri.purchase_code = po.code
    AND po.branch_id = '{KRC}'
    AND po.deleted = 0
WHERE ri.branch_id = '{KRC}'
  AND fromUnixTimestamp(toUInt32(po.delivery_date)) >= '2026-05-01'
  AND fromUnixTimestamp(toUInt32(po.delivery_date)) < '2026-06-01'
FORMAT JSONEachRow
""")

import re
def extract_weight_kg(name):
    if not name: return 0
    text = name.upper()
    for pat, mult in [(r'(\d+(?:[.,]\d+)?)\s*KG\b', 1.0), (r'(\d+(?:[.,]\d+)?)\s*G\b', 0.001),
                       (r'(\d+(?:[.,]\d+)?)\s*(?:LÍT|LIT)\b', 1.0), (r'(\d+(?:[.,]\d+)?)\s*L\b', 1.0),
                       (r'(\d+(?:[.,]\d+)?)\s*ML\b', 0.001)]:
        m = re.findall(pat, text)
        if m:
            try: return float(m[-1].replace(",", ".")) * mult
            except: continue
    return 0

from collections import defaultdict

# Calculate tonnage per day per status filter
# all = status 3+5+7, confirmed = status 5+7, done = status 7 only
tons_all = defaultdict(float)
tons_confirmed = defaultdict(float)  # status 5 + 7
tons_done = defaultdict(float)       # status 7 only

for line in r1.split('\n'):
    if not line.strip(): continue
    obj = json.loads(line)
    
    d = obj.get('del_date', '')
    qty = float(obj.get('qty', 0))
    nw = float(obj.get('net_weight', 0))
    bc = str(obj.get('barcode', '')).strip()
    pname = str(obj.get('pname', '')).strip()
    status = int(obj.get('po_status', 0))
    
    if not d or qty <= 0: continue
    
    # Resolve weight (same logic as transfer PT)
    if nw > 0:
        wg = nw
    elif bc and bc in master:
        wg = master[bc]
    else:
        wkg = extract_weight_kg(pname)
        if wkg > 0:
            wg = wkg * 1000
        else:
            continue
    
    tons = qty * wg / 1_000_000
    
    tons_all[d] += tons
    if status in (5, 7):
        tons_confirmed[d] += tons
    if status == 7:
        tons_done[d] += tons

# Print comparison
print(f"\n{'='*70}")
print(f"  T5/2026: Tonnage with master data + status filter")
print(f"{'='*70}")
print(f"\n  {'Date':<12} {'ALL(3+5+7)':>11} {'Confirmed':>11} {'Done(7)':>11}")
print(f"  {'':12} {'':>11} {'(5+7)':>11} {'':>11}")
print(f"  {'─'*12} {'─'*11} {'─'*11} {'─'*11}")

all_dates = sorted(set(list(tons_all.keys()) + list(tons_confirmed.keys())),
                   key=lambda d: d.split('/')[::-1])

for d in all_dates:
    a = tons_all.get(d, 0)
    c = tons_confirmed.get(d, 0)
    dn = tons_done.get(d, 0)
    marker = ' ←' if c > 0 else ''
    print(f"  {d:<12} {a:>11.2f} {c:>11.2f} {dn:>11.2f}{marker}")

# Averages
days_with_confirmed = [d for d in all_dates if tons_confirmed.get(d, 0) > 0]
if days_with_confirmed:
    avg_all = sum(tons_all[d] for d in days_with_confirmed) / len(days_with_confirmed)
    avg_conf = sum(tons_confirmed[d] for d in days_with_confirmed) / len(days_with_confirmed)
    avg_done = sum(tons_done[d] for d in days_with_confirmed) / len(days_with_confirmed)
    print(f"\n  {'AVG':<12} {avg_all:>11.2f} {avg_conf:>11.2f} {avg_done:>11.2f}")
    print(f"  (trên {len(days_with_confirmed)} ngày có status 5+7)")

print("\nDone.")
