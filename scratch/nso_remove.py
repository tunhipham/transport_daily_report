# -*- coding: utf-8 -*-
"""Remove cancelled stores from NSO master, then run inject pipeline."""
import json, sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(REPO, "data", "nso", "nso_stores.json")
XLSX_PATH = os.path.join(REPO, "data", "nso", "nso_master.xlsx")

# Load current master
stores = json.load(open(JSON_PATH, 'r', encoding='utf-8'))
print(f"Master before: {len(stores)} stores")

# Stores to remove
REMOVE = [
    "6/1 Đường 13",       # Thạnh Xuân 13 - bị hủy
    "VHGP Q9 - BS1501",   # không có trong mail mới
]

def normalize(s):
    return re.sub(r'[\s\-–]+', ' ', s.lower().strip())

removed = []
kept = []
for s in stores:
    nm = normalize(s.get('name_mail') or s.get('name_full') or '')
    should_remove = False
    for r in REMOVE:
        if normalize(r) in nm or nm in normalize(r):
            should_remove = True
            removed.append(s)
            break
    if not should_remove:
        kept.append(s)

print(f"\nRemoved {len(removed)} stores:")
for s in removed:
    nm = s.get('name_mail') or s.get('name_full') or '?'
    print(f"  ❌ {nm} | {s.get('opening_date','')} | {s.get('code','--')}")

# Save JSON
with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(kept, f, ensure_ascii=False, indent=2)
print(f"\nJSON saved: {len(kept)} stores")

# Now rebuild xlsx from the NsoMaster class
sys.path.insert(0, os.path.join(REPO, "script"))
from domains.nso.nso_master import NsoMaster
master = NsoMaster()
master.stores = kept
# Load existing history from xlsx if available
import openpyxl
if os.path.exists(XLSX_PATH):
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        if "History" in wb.sheetnames:
            ws_h = wb["History"]
            H_MAP = {
                "Thời gian": "timestamp", "Code": "code", "Tên": "name",
                "Thay đổi": "action", "Giá trị cũ": "old_value",
                "Giá trị mới": "new_value", "Nguồn": "source",
            }
            raw_h = [c.value for c in ws_h[1]]
            h_headers = [H_MAP.get(h, h) for h in raw_h]
            for row in ws_h.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                d = {}
                for i, h in enumerate(h_headers):
                    if i < len(row):
                        d[h] = row[i] or ""
                master.history.append(d)
        wb.close()
    except Exception as e:
        print(f"  Warning loading history: {e}")

# Log removals
from datetime import datetime
for s in removed:
    master.history.append({
        "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "code": s.get("code") or "",
        "name": s.get("name_mail") or s.get("name_full") or "",
        "action": "Xóa (hủy KT)",
        "old_value": s.get("opening_date", ""),
        "new_value": "",
        "source": "Manual",
    })

master.save()
print(f"Master after removal: {len(master.stores)} stores")
