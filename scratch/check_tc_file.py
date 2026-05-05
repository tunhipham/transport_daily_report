import os, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

base = r'G:\My Drive\DOCS\DAILY'
target_dir = None
for d in os.listdir(base):
    if 'GIAO' in d and 'HANG' in d:
        target_dir = os.path.join(base, d)
        print(f'Dir: {d}')
        break

if not target_dir:
    print('Directory not found!')
    sys.exit(1)

files = [f for f in os.listdir(target_dir) if f.endswith('.xlsx') and not f.startswith('~')]
for f in sorted(files):
    fpath = os.path.join(target_dir, f)
    size = os.path.getsize(fpath)
    print(f'  {f} ({size:,} bytes)')

# Find T04 file
t04_file = None
for f in files:
    if '04' in f and '2026' in f:
        t04_file = os.path.join(target_dir, f)
        break

if not t04_file:
    print('T04 file not found!')
    sys.exit(1)

print(f'\nReading T04: {os.path.basename(t04_file)}')
wb = load_workbook(t04_file, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
print(f'Sheet: {ws.title}')

# Print header
print('\nHeader:')
for hdr in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, c in enumerate(hdr):
        print(f'  Col {i}: {c}')

# Count by date
from collections import Counter
date_counts = Counter()
total = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    date_val = row[2] if len(row) > 2 else None
    actual_val = row[10] if len(row) > 10 else None
    if date_val and hasattr(date_val, 'strftime'):
        total += 1
        date_counts[date_val.strftime('%d/%m/%Y')] += 1

print(f'\nTotal rows with date: {total}')
print(f'Unique dates: {len(date_counts)}')
print('\nDate distribution:')
for d in sorted(date_counts.keys(), key=lambda x: (int(x.split('/')[1]), int(x.split('/')[0]))):
    print(f'  {d}: {date_counts[d]} rows')

wb.close()
