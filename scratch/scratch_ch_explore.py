"""Export active stores from ClickHouse to Excel."""
import json, requests, sys, io
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open(r"g:\My Drive\DOCS\transport_daily_report\config\mcp_clickhouse.json") as f:
    cfg = json.load(f)

BASE = cfg["base_url"]
PARAMS = cfg["params"]

def query(sql):
    r = requests.get(BASE, params={**PARAMS, "query": sql})
    r.raise_for_status()
    return r.text

sql = """
    SELECT 
        branch_code,
        branch_name,
        branch_name_abbreviate,
        CASE branch_store_type 
            WHEN 1 THEN 'MART'
            WHEN 3 THEN 'MINI'
            ELSE toString(branch_store_type)
        END AS store_type,
        CASE branch_type
            WHEN 1 THEN 'Store'
            WHEN 3 THEN 'DC/Kho'
            ELSE toString(branch_type)
        END AS branch_type,
        branch_status,
        toString(toDate(branch_opening_date)) AS opening_date,
        coalesce(location_name, '') AS location,
        coalesce(locality_name, '') AS area,
        coalesce(locality_code, '') AS area_code
    FROM kdb.kf_branch_location 
    WHERE branch_status = 1
    ORDER BY branch_type, store_type, branch_code
    FORMAT TabSeparatedWithNames
"""

result = query(sql)
lines = result.strip().split("\n")
header = lines[0].split("\t")
rows = [line.split("\t") for line in lines[1:] if line.strip()]

# ── Build Excel ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Siêu thị hoạt động"

# Styles
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
data_font = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center")

# Vietnamese headers
vn_headers = ["STT", "Mã chi nhánh", "Tên siêu thị", "Viết tắt", "Loại ST", "Loại CN", "Trạng thái", "Ngày mở", "Địa điểm", "Khu vực", "Mã KV"]

# Write header
for col, h in enumerate(vn_headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write data
for i, row in enumerate(rows, 1):
    ws.cell(row=i+1, column=1, value=i).font = data_font
    ws.cell(row=i+1, column=1).alignment = center_align
    ws.cell(row=i+1, column=1).border = thin_border
    
    for j, val in enumerate(row):
        cell = ws.cell(row=i+1, column=j+2, value=val)
        cell.font = data_font
        cell.border = thin_border
        if j in (3, 4, 5, 6, 9):  # center-align short columns
            cell.alignment = center_align

# Auto-fit column widths
col_widths = [5, 14, 50, 10, 8, 10, 10, 12, 55, 12, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:K{len(rows)+1}"

# ── Summary sheet ──
ws2 = wb.create_sheet("Thống kê")
ws2.cell(row=1, column=1, value="Thống kê siêu thị đang hoạt động").font = Font(name="Arial", bold=True, size=14, color="2E86C1")

# By type
ws2.cell(row=3, column=1, value="Loại").font = Font(name="Arial", bold=True, size=11)
ws2.cell(row=3, column=2, value="Số lượng").font = Font(name="Arial", bold=True, size=11)

from collections import Counter
type_counts = Counter(r[3] for r in rows)
r = 4
for t, c in type_counts.most_common():
    ws2.cell(row=r, column=1, value=t).font = data_font
    ws2.cell(row=r, column=2, value=c).font = data_font
    r += 1
ws2.cell(row=r, column=1, value="TỔNG").font = Font(name="Arial", bold=True, size=11)
ws2.cell(row=r, column=2, value=len(rows)).font = Font(name="Arial", bold=True, size=11)

# By area
r += 2
ws2.cell(row=r, column=1, value="Khu vực").font = Font(name="Arial", bold=True, size=11)
ws2.cell(row=r, column=2, value="Số lượng").font = Font(name="Arial", bold=True, size=11)
r += 1
area_counts = Counter(r_[8] for r_ in rows)
for area, cnt in sorted(area_counts.items()):
    ws2.cell(row=r, column=1, value=area).font = data_font
    ws2.cell(row=r, column=2, value=cnt).font = data_font
    r += 1

ws2.column_dimensions["A"].width = 55
ws2.column_dimensions["B"].width = 12

# Save
out_path = r"g:\My Drive\DOCS\transport_daily_report\output\danh_sach_sieu_thi_hoat_dong.xlsx"
wb.save(out_path)
print(f"✅ Exported {len(rows)} records to: {out_path}")
