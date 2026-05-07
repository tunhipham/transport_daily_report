"""Export Excel: 2 tabs — 83 stores dư + 62 stores giao 2 chuyến on 22/04."""
import sys, io, json, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from collections import Counter
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

cache_path = r'G:\My Drive\DOCS\transport_daily_report\output\state\trip_cache_T04.json'
with open(cache_path, 'r', encoding='utf-8') as f:
    cache = json.load(f)
rows = cache['rows']

plan_path = r'G:\My Drive\DOCS\transport_daily_report\output\state\monthly_plan_T04.json'
with open(plan_path, 'r', encoding='utf-8') as f:
    plan_data = json.load(f)
dm_plan = plan_data['plan'].get('ĐÔNG MÁT', [])

# ── Common styles ──
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
title_font = Font(bold=True, size=14)
thin_bd = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
red_fill = PatternFill("solid", fgColor="FFC7CE")

# ── Data prep ──
# KH ĐÔNG 22/04
plan_22_dong = set(r['store'] for r in dm_plan if '22/04/2026' in str(r.get('date','')) and str(r.get('tuyen','')).startswith('HĐ'))

# Actual ĐÔNG 22/04
dong_actual_rows = [r for r in rows if r.get('date') == '2026-04-22' and r.get('kho') == 'ĐÔNG MÁT' and r.get('sub_kho') == 'ĐÔNG']
dong_actual_stores = set(r['dest'] for r in dong_actual_rows)
extra_stores = sorted(dong_actual_stores - plan_22_dong)

# KH lookup per day
kh_by_day = {}
for offset in range(-7, 8):
    d = date(2026, 4, 22) + timedelta(days=offset)
    d_str = d.strftime('%d/%m/%Y')
    d_stores = set(r['store'] for r in dm_plan if d_str in str(r.get('date','')) and str(r.get('tuyen','')).startswith('HĐ'))
    if d_stores:
        kh_by_day[d] = d_stores

# Actual per day
actual_by_day = {}
for offset in range(-3, 4):
    d = date(2026, 4, 22) + timedelta(days=offset)
    d_actual = set(r['dest'] for r in rows if r.get('date') == d.isoformat() and r.get('kho') == 'ĐÔNG MÁT' and r.get('sub_kho') == 'ĐÔNG')
    if d_actual:
        actual_by_day[d] = d_actual

# ── Create workbook ──
wb = Workbook()

# ════════════════════════════════════════════
# TAB 1: 83 stores dư
# ════════════════════════════════════════════
ws1 = wb.active
ws1.title = "83 Stores Dư"

# Title
ws1.merge_cells('A1:F1')
ws1.cell(1, 1, "83 Stores ĐÔNG giao ngày 22/04 nhưng KHÔNG có trong KH hàng đông").font = title_font

# Headers
headers1 = ["STT", "Store", "KH giao ngày nào", "Actual giao 21/04?", "Actual giao 22/04?", "Actual giao 23/04?"]
for i, h in enumerate(headers1, 1):
    c = ws1.cell(3, i, h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = thin_bd
    c.alignment = center

dong_21 = actual_by_day.get(date(2026, 4, 21), set())
dong_23 = actual_by_day.get(date(2026, 4, 23), set())

for idx, store in enumerate(extra_stores, 1):
    row = idx + 3
    # KH days
    kh_days = []
    for d in sorted(kh_by_day.keys()):
        if store in kh_by_day[d]:
            day_name = ['T2','T3','T4','T5','T6','T7','CN'][d.weekday()]
            kh_days.append(f'{d.strftime("%d/%m")}({day_name})')
    kh_str = ', '.join(kh_days) if kh_days else 'Không có KH'

    on_21 = "Có" if store in dong_21 else "Không"
    on_22 = "Có"
    on_23 = "Có" if store in dong_23 else "Không"

    ws1.cell(row, 1, idx).border = thin_bd
    ws1.cell(row, 1).alignment = center
    ws1.cell(row, 2, store).border = thin_bd
    ws1.cell(row, 2).alignment = center
    ws1.cell(row, 3, kh_str).border = thin_bd
    ws1.cell(row, 3).alignment = left_align
    
    c21 = ws1.cell(row, 4, on_21)
    c21.border = thin_bd
    c21.alignment = center
    
    c22 = ws1.cell(row, 5, on_22)
    c22.border = thin_bd
    c22.alignment = center
    c22.fill = yellow_fill
    
    c23 = ws1.cell(row, 6, on_23)
    c23.border = thin_bd
    c23.alignment = center

# Widths
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 55
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 16
ws1.column_dimensions['F'].width = 16
ws1.freeze_panes = "A4"

# ════════════════════════════════════════════
# TAB 2: 62 stores giao 2 chuyến
# ════════════════════════════════════════════
ws2 = wb.create_sheet("62 Stores 2 Chuyến")

ws2.merge_cells('A1:G1')
ws2.cell(1, 1, "62 Stores ĐÔNG giao 2 chuyến xe khác nhau ngày 22/04/2026").font = title_font

headers2 = ["STT", "Store", "Chuyến 1", "Giờ đến 1", "Chuyến 2", "Giờ đến 2", "Ngoài KH?"]
for i, h in enumerate(headers2, 1):
    c = ws2.cell(3, i, h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = thin_bd
    c.alignment = center

dest_counter = Counter(r['dest'] for r in dong_actual_rows)
dups = sorted([k for k, v in dest_counter.items() if v > 1])

for idx, dest in enumerate(dups, 1):
    row = idx + 3
    details = sorted([r for r in dong_actual_rows if r['dest'] == dest], key=lambda r: r.get('arrival_raw', ''))
    
    t1 = details[0]['trip_id']
    a1 = details[0].get('arrival_raw', '').replace('22/04/2026 ', '')
    t2 = details[1]['trip_id']
    a2 = details[1].get('arrival_raw', '').replace('22/04/2026 ', '').replace('23/04/2026 ', '23/04 ')
    is_extra = "Có" if dest in extra_stores else "Không"
    
    ws2.cell(row, 1, idx).border = thin_bd
    ws2.cell(row, 1).alignment = center
    ws2.cell(row, 2, dest).border = thin_bd
    ws2.cell(row, 2).alignment = center
    ws2.cell(row, 3, t1).border = thin_bd
    ws2.cell(row, 3).alignment = center
    ws2.cell(row, 4, a1).border = thin_bd
    ws2.cell(row, 4).alignment = center
    ws2.cell(row, 5, t2).border = thin_bd
    ws2.cell(row, 5).alignment = center
    ws2.cell(row, 6, a2).border = thin_bd
    ws2.cell(row, 6).alignment = center
    
    c_extra = ws2.cell(row, 7, is_extra)
    c_extra.border = thin_bd
    c_extra.alignment = center
    if is_extra == "Có":
        c_extra.fill = red_fill

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 20
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 12
ws2.freeze_panes = "A4"

# Save
out_path = r'G:\My Drive\DOCS\transport_daily_report\output\artifacts\performance\DONG_22_04_CHECK.xlsx'
wb.save(out_path)
print(f'✅ Saved: {out_path}')
print(f'  Tab 1: {len(extra_stores)} stores dư')
print(f'  Tab 2: {len(dups)} stores giao 2 chuyến')
