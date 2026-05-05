import os, sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from collections import Counter

path = r'G:\My Drive\DOCS\DAILY\BÁO CÁO GIAO HÀNG MIỀN ĐÔNG\04.2026 BÁO CÁO GIAO HÀNG KINGFOOD.xlsx'
print(f'File: {os.path.basename(path)}')
print(f'Size: {os.path.getsize(path):,} bytes')

wb = load_workbook(path, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
print(f'Sheet: {ws.title}')

# Print header
print('\nHeader:')
for hdr in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, c in enumerate(hdr):
        if c:
            print(f'  Col {i}: {c}')

# Detect columns
date_col, store_col, tuyen_col, plan_col, actual_col = 0, 1, 4, 7, 8
for hdr in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, c in enumerate(hdr):
        h = str(c or '').strip().upper()
        if 'NGÀY' in h or 'NGAY' in h:
            date_col = i
        elif 'MÃ' in h and ('CH' in h or 'ĐIỂM' in h or 'DIEM' in h):
            store_col = i
        elif 'TUYẾN' in h or 'TUYEN' in h:
            tuyen_col = i
        elif 'DỰ KIẾN' in h or 'DU KIEN' in h:
            plan_col = i
        elif 'TG ĐẾN' in h or 'TG DEN' in h or 'ĐẾN CỬA' in h:
            actual_col = i
print(f'\nDetected: date={date_col}, store={store_col}, tuyen={tuyen_col}, plan={plan_col}, actual={actual_col}')

# Count by date
date_counts = Counter()
total = 0
no_actual = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    date_val = row[date_col] if len(row) > date_col else None
    actual_val = row[actual_col] if len(row) > actual_col else None
    store_val = row[store_col] if len(row) > store_col else None
    if date_val and hasattr(date_val, 'strftime') and store_val:
        total += 1
        date_counts[date_val.strftime('%d/%m/%Y')] += 1
        if not actual_val:
            no_actual += 1

print(f'\nTotal rows with date+store: {total}')
print(f'Rows without actual_time: {no_actual}')
print(f'Unique dates: {len(date_counts)}')
print('\nDate distribution:')
for d in sorted(date_counts.keys(), key=lambda x: (int(x.split('/')[2]), int(x.split('/')[1]), int(x.split('/')[0]))):
    print(f'  {d}: {date_counts[d]} rows')

wb.close()
