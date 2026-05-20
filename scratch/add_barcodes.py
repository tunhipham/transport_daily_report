# -*- coding: utf-8 -*-
"""Add 26 unclassified barcodes to ABA Master Data."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

ABA_MASTER_PATH = r'G:\My Drive\DOCS\DAILY\ton_aba\data\master_data\Master Data.xlsx'

# 26 barcodes to add — 2 ĐÔNG, rest MÁT
NEW_BARCODES = [
    # (barcode, name, classification)
    ("11570", "QUÝT NỘI ĐỊA TRUNG TÚI 1.5KG", "MÁT"),
    ("8850393991995", "BETAGEN - SỮA CHUA UỐNG HƯƠNG CAM 700ML", "MÁT"),
    ("8850393800402", "BETAGEN - SỮA CHUA UỐNG ÍT ĐƯỜNG 85ML", "MÁT"),
    ("8934717402432", "BIBIGO - BÁNH XẾP KIỂU HÀN QUỐC 365 HẢI SẢN", "ĐÔNG"),  # ĐÔNG
    ("8850393919968", "BETAGEN - SỮA CHUA UỐNG HƯƠNG TỰ NHIÊN 700ML", "MÁT"),
    ("8850393919975", "BETAGEN - SỮA CHUA UỐNG HƯƠNG TỰ NHIÊN 400ML", "MÁT"),
    ("8934717266713", "CJ CẦU TRE - CHẢ GIÒ DA GẠO HẢI SẢN 400G", "ĐÔNG"),  # ĐÔNG
    ("8850393991865", "BETAGEN - SỮA CHUA UỐNG HƯƠNG CAM 85ML", "MÁT"),
    ("8850393991926", "BETAGEN - SỮA CHUA UỐNG HƯƠNG CAM 300ML", "MÁT"),
    ("8936183060444", "GFU - CƠM RƯỢU NẾP CÁI HOA VÀNG 170G", "MÁT"),
    ("8936183060451", "GFU - CƠM RƯỢU NẾP CẨM 170G", "MÁT"),
    ("11532", "TÁO XANH MỸ TÚI 900G UP", "MÁT"),
    ("11711", "TÁO QUEEN NEW ZEALAND HỘP 6 QUẢ", "MÁT"),
    ("8850393801379", "BETAGEN - SỮA CHUA UỐNG HƯƠNG DÂU 300ML", "MÁT"),
    ("8935101607211", "O'FOOD - SỦI CẢO NHÂN THỊT HEO RAU CỦ 450G", "MÁT"),
    ("11697", "VỈ CANH KHOAI MỠ 300G", "MÁT"),
    ("11681", "HỖN HỢP MĂNG TÂY, BẮP NON, CÀ RỐT 300G", "MÁT"),
    ("8850393800679", "BETAGEN - SCU HƯƠNG DỨA CHAI 300ML", "MÁT"),
    ("8935117700104", "ÁNH HỒNG - BÁNH FLAN CARAMEL 6*100G", "MÁT"),
    ("11682", "BÔNG CẢI TRẮNG SƠ CHẾ 200G", "MÁT"),
]

# Also add the remaining 6 smaller ones from the check
# Need to get them from DB first
sys.path.insert(0, os.path.join(r'g:\My Drive\DOCS\transport_daily_report', 'script'))
from data_pipeline.config import load_clickhouse_config
import requests, json

cfg = load_clickhouse_config()
params = {'user': cfg['user'], 'password': cfg['password'], 'database': cfg['database']}

def q(sql):
    r = requests.get(cfg['base_url'], params={**params, 'query': sql}, timeout=30)
    r.raise_for_status()
    return r.text.strip()

# Get all unclassified barcodes
print("  → Getting remaining unclassified barcodes from DB...")
existing_bcs = set(bc for bc, _, _ in NEW_BARCODES)

# Load current master
wb_check = load_workbook(ABA_MASTER_PATH, read_only=True, data_only=True)
ws_check = wb_check.worksheets[0]
existing_in_master = set()
for row in ws_check.iter_rows(min_row=2, values_only=False):
    bc = str(row[1].value or "").strip()
    if bc:
        existing_in_master.add(bc)
wb_check.close()

# Query remaining unclassified from DB
DONG_MAT_BRANCH = '61d4ffa72997ae0007f5ad19'
r1 = q(f"""
SELECT
    p.base_barcode,
    p.name,
    SUM(t.transfer_quantity) as qty,
    round(SUM(t.transfer_quantity * p.base_net_weight / 1000000), 3) as tons
FROM kf_transfer_mart t
LEFT JOIN kf_product_static p ON t.product_id = p.id
WHERE toDate(t.transfer_date) = '2026-05-20'
  AND t.deleted = 0 AND t.status != 6
  AND t.from_branch_id = '{DONG_MAT_BRANCH}'
GROUP BY p.base_barcode, p.name
HAVING base_barcode NOT IN ({','.join(f"'{bc}'" for bc in existing_in_master)})
ORDER BY tons DESC
""")

for line in r1.split('\n'):
    if not line.strip():
        continue
    parts = line.split('\t')
    bc = parts[0].strip()
    name = parts[1].strip() if len(parts) > 1 else ''
    if bc and bc not in existing_bcs and bc not in existing_in_master:
        # Default to MÁT for remaining small items
        NEW_BARCODES.append((bc, name, "MÁT"))
        existing_bcs.add(bc)

print(f"  Total barcodes to add: {len(NEW_BARCODES)}")
print(f"  ĐÔNG: {sum(1 for _,_,c in NEW_BARCODES if c=='ĐÔNG')}")
print(f"  MÁT: {sum(1 for _,_,c in NEW_BARCODES if c=='MÁT')}")

# Filter out any that already exist in master
to_add = [(bc, name, cl) for bc, name, cl in NEW_BARCODES if bc not in existing_in_master]
already = [(bc, name, cl) for bc, name, cl in NEW_BARCODES if bc in existing_in_master]

if already:
    print(f"\n  ⚠ {len(already)} already in master (skipping):")
    for bc, name, cl in already:
        print(f"    {bc}: {name[:40]}")

print(f"\n  → Adding {len(to_add)} new barcodes...")

# Open for editing (not read_only)
wb = load_workbook(ABA_MASTER_PATH)
ws = wb.worksheets[0]

start_row = ws.max_row + 1
added = 0
for bc, name, cl in to_add:
    row = start_row + added
    ws.cell(row=row, column=2, value=bc)       # Col B = Mã hàng
    ws.cell(row=row, column=3, value=name)      # Col C = Tên
    ws.cell(row=row, column=5, value=cl)        # Col E = Phân Loại
    added += 1
    tag = "🧊" if cl == "ĐÔNG" else "🌡️"
    print(f"    {tag} Row {row}: {bc} = {cl} — {name[:45]}")

wb.save(ABA_MASTER_PATH)
wb.close()

print(f"\n  ✅ Added {added} barcodes to ABA Master Data")
print(f"     File: {ABA_MASTER_PATH}")

# Verify
wb_v = load_workbook(ABA_MASTER_PATH, read_only=True, data_only=True)
ws_v = wb_v.worksheets[0]
total = 0
dong = 0
mat = 0
for row in ws_v.iter_rows(min_row=2, values_only=False):
    bc = str(row[1].value or "").strip()
    pl = str(row[4].value or "").strip().upper()
    if bc and pl in ('MÁT', 'ĐÔNG'):
        total += 1
        if pl == 'ĐÔNG':
            dong += 1
        else:
            mat += 1
wb_v.close()
print(f"  Verified: {total} barcodes total (ĐÔNG: {dong}, MÁT: {mat})")
