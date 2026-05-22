# -*- coding: utf-8 -*-
"""Dump full PO KRC data for 22/05/2026 — no filters, all columns"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, r'g:\My Drive\DOCS\transport_daily_report\script')
from data_pipeline.config import load_clickhouse_config
import requests

cfg = load_clickhouse_config()
p = {'user': cfg['user'], 'password': cfg['password'], 'database': cfg['database']}
def q(sql):
    r = requests.get(cfg['base_url'], params={**p, 'query': sql}, timeout=120)
    r.raise_for_status()
    return r.text.strip()

KRC = '5fdc170ebd89c10006f15b7c'

# ═══════════════════════════════════════════════════════════
# 1) kf_purchase_order columns
# ═══════════════════════════════════════════════════════════
print("=" * 70)
print("  TABLE: kf_purchase_order — ALL COLUMNS")
print("=" * 70)
r1 = q("DESCRIBE kf_purchase_order")
po_cols = []
for line in r1.split('\n'):
    parts = line.split('\t')
    col = parts[0]
    typ = parts[1] if len(parts) > 1 else ''
    po_cols.append(col)
    print(f"  {col:<40} {typ}")

# ═══════════════════════════════════════════════════════════
# 2) kf_receipt_items columns
# ═══════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("  TABLE: kf_receipt_items — ALL COLUMNS")
print("=" * 70)
r2 = q("DESCRIBE kf_receipt_items")
ri_cols = []
for line in r2.split('\n'):
    parts = line.split('\t')
    col = parts[0]
    typ = parts[1] if len(parts) > 1 else ''
    ri_cols.append(col)
    print(f"  {col:<40} {typ}")

# ═══════════════════════════════════════════════════════════
# 3) Export raw data 22/05 to Excel for user review
# ═══════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("  EXPORT: Full data 22/05/2026 → Excel")
print("=" * 70)

# Get PO headers for 22/05
r_po = q(f"""
SELECT *
FROM kf_purchase_order
WHERE branch_id = '{KRC}'
  AND toDate(fromUnixTimestamp(toUInt32(delivery_date))) = '2026-05-22'
FORMAT JSONEachRow
""")

po_rows = []
for line in r_po.split('\n'):
    if not line.strip(): continue
    po_rows.append(json.loads(line))
print(f"  PO headers (22/05): {len(po_rows)}")

# Get receipt items for those POs
r_ri = q(f"""
SELECT *
FROM kf_receipt_items ri
WHERE ri.branch_id = '{KRC}'
  AND ri.purchase_code IN (
    SELECT code FROM kf_purchase_order
    WHERE branch_id = '{KRC}'
      AND toDate(fromUnixTimestamp(toUInt32(delivery_date))) = '2026-05-22'
  )
FORMAT JSONEachRow
""")

ri_rows = []
for line in r_ri.split('\n'):
    if not line.strip(): continue
    ri_rows.append(json.loads(line))
print(f"  Receipt items: {len(ri_rows)}")

# Write to Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

out_path = r'g:\My Drive\DOCS\transport_daily_report\scratch\po_krc_db_raw_22052026.xlsx'
wb = Workbook()

# Sheet 1: PO Headers
ws1 = wb.active
ws1.title = "PO_Headers"
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

if po_rows:
    cols = list(po_rows[0].keys())
    for ci, col in enumerate(cols, 1):
        cell = ws1.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
    for ri, row in enumerate(po_rows, 2):
        for ci, col in enumerate(cols, 1):
            val = row.get(col, '')
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            ws1.cell(row=ri, column=ci, value=val)

# Sheet 2: Receipt Items
ws2 = wb.create_sheet("Receipt_Items")
if ri_rows:
    cols = list(ri_rows[0].keys())
    for ci, col in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
    for ri_idx, row in enumerate(ri_rows, 2):
        for ci, col in enumerate(cols, 1):
            val = row.get(col, '')
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            ws2.cell(row=ri_idx, column=ci, value=val)

# Sheet 3: Column mapping vs local file
ws3 = wb.create_sheet("Column_Mapping")
ws3.cell(row=1, column=1, value="Local File Column").font = Font(bold=True)
ws3.cell(row=1, column=2, value="Local Col Index").font = Font(bold=True)
ws3.cell(row=1, column=3, value="DB Table").font = Font(bold=True)
ws3.cell(row=1, column=4, value="DB Column").font = Font(bold=True)
ws3.cell(row=1, column=5, value="Notes").font = Font(bold=True)

mapping = [
    ("Mã PO (D)", 3, "kf_purchase_order", "code", ""),
    ("Trạng Thái (E)", 4, "kf_purchase_order", "status", "3=duyệt, 5=xác nhận, 7=nhận hàng"),
    ("Mã NCC (L)", 11, "kf_purchase_order", "vendor_id", ""),
    ("Ngày giao hàng dự kiến (P)", 15, "kf_purchase_order", "delivery_date", "epoch seconds → toDate()"),
    ("Ngày NCC xác nhận giao hàng (Q)", 16, "kf_purchase_order", "delivery_date_vendor_confirm", "epoch seconds"),
    ("Barcode (S)", 18, "kf_receipt_items", "product_barcode", ""),
    ("Tên sản phẩm (T)", 19, "kf_receipt_items", "product_name", ""),
    ("Số lượng (V)", 21, "kf_receipt_items", "qty", "= số lượng đặt"),
    ("Khối lượng kg (W)", 22, "kf_receipt_items", "net_weight", "grams, nhiều item = 0"),
    ("Số lượng thực nhận (X)", 23, "kf_receipt_items", "qty_receipt", ""),
    ("Số lượng đặt PO", -1, "kf_receipt_items", "po_qty", "qty đặt trên PO"),
    ("Ngày nhận hàng", -1, "kf_receipt_items", "receipted_at", "datetime hoặc null"),
    ("Ngày tạo PO", -1, "kf_receipt_items", "po_created_at", ""),
    ("Ngày NCC confirm", -1, "kf_receipt_items", "delivery_date_vendor_confirm", "datetime"),
    ("Ngày giao (PO header epoch)", -1, "kf_receipt_items", "po_delivery_date", "datetime from epoch"),
]

for ri, (local, idx, table, dbcol, note) in enumerate(mapping, 2):
    ws3.cell(row=ri, column=1, value=local)
    ws3.cell(row=ri, column=2, value=idx if idx >= 0 else "N/A")
    ws3.cell(row=ri, column=3, value=table)
    ws3.cell(row=ri, column=4, value=dbcol)
    ws3.cell(row=ri, column=5, value=note)

wb.save(out_path)
print(f"\n  ✅ Saved: {out_path}")
print(f"     Sheet 1: PO_Headers ({len(po_rows)} rows, {len(po_rows[0]) if po_rows else 0} cols)")
print(f"     Sheet 2: Receipt_Items ({len(ri_rows)} rows, {len(ri_rows[0]) if ri_rows else 0} cols)")
print(f"     Sheet 3: Column_Mapping (local file ↔ DB)")

# Quick summary of key date fields
print(f"\n{'='*70}")
print("  KEY DATE FIELDS — sample from 22/05")
print("=" * 70)
if po_rows:
    sample = po_rows[0]
    print(f"  delivery_date:                {sample.get('delivery_date','?')}")
    print(f"  delivery_date_vendor_confirm: {sample.get('delivery_date_vendor_confirm','?')}")
    print(f"  created_at:                   {sample.get('created_at','?')}")
    
    from datetime import datetime
    try:
        epoch = int(sample['delivery_date'])
        dt = datetime.fromtimestamp(epoch)
        print(f"  → delivery_date parsed:       {dt.strftime('%d/%m/%Y %H:%M')}")
    except: pass
    try:
        epoch2 = int(sample['delivery_date_vendor_confirm'])
        dt2 = datetime.fromtimestamp(epoch2)
        print(f"  → vendor_confirm parsed:      {dt2.strftime('%d/%m/%Y %H:%M')}")
    except: pass

if ri_rows:
    sample = ri_rows[0]
    print(f"\n  Receipt item sample:")
    print(f"  qty:              {sample.get('qty','?')}")
    print(f"  po_qty:           {sample.get('po_qty','?')}")
    print(f"  qty_receipt:      {sample.get('qty_receipt','?')}")
    print(f"  net_weight:       {sample.get('net_weight','?')}")
    print(f"  po_delivery_date: {sample.get('po_delivery_date','?')}")
    print(f"  delivery_date_vendor_confirm: {sample.get('delivery_date_vendor_confirm','?')}")
    print(f"  receipted_at:     {sample.get('receipted_at','?')}")

print("\nDone.")
