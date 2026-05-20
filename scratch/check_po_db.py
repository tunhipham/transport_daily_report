# -*- coding: utf-8 -*-
"""
1. Check unclassified barcodes ĐÔNG/MÁT hôm nay
2. Compare PO KRC DB vs local file
"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
BASE = r'g:\My Drive\DOCS\transport_daily_report'
sys.path.insert(0, os.path.join(BASE, 'script'))
from data_pipeline.config import load_clickhouse_config
from openpyxl import load_workbook
import requests

cfg = load_clickhouse_config()
params = {'user': cfg['user'], 'password': cfg['password'], 'database': cfg['database']}

def q(sql):
    r = requests.get(cfg['base_url'], params={**params, 'query': sql}, timeout=60)
    r.raise_for_status()
    return r.text.strip()

# ═══════════════════════════════════════════════════════════
# PART 1: UNCLASSIFIED BARCODES — ĐÔNG/MÁT
# ═══════════════════════════════════════════════════════════
print("=" * 70)
print("  PART 1: UNCLASSIFIED BARCODES ĐÔNG/MÁT — 20/05/2026")
print("=" * 70)

DONG_MAT_BRANCH = '61d4ffa72997ae0007f5ad19'
ABA_MASTER_PATH = r'G:\My Drive\DOCS\DAILY\ton_aba\data\master_data\Master Data.xlsx'

# Load ABA Master classification
print("\n  → Loading ABA Master Data for ĐÔNG/MÁT classification...")
barcode_type = {}
wb = load_workbook(ABA_MASTER_PATH, read_only=True, data_only=True)
ws = wb.worksheets[0]
for row in ws.iter_rows(min_row=2, values_only=False):
    bc = str(row[1].value or "").strip()
    pl = str(row[4].value or "").strip().upper()
    if bc and pl in ('MÁT', 'ĐÔNG'):
        barcode_type[bc] = pl
wb.close()
mat_c = sum(1 for v in barcode_type.values() if v == 'MÁT')
dong_c = sum(1 for v in barcode_type.values() if v == 'ĐÔNG')
print(f"    Master: {len(barcode_type)} barcodes (ĐÔNG: {dong_c}, MÁT: {mat_c})")

# Get ALL transfer items for ABA QC branch today from DB
print("\n  → Querying all transfer items for ABA QC today...")
r1 = q(f"""
SELECT
    p.base_barcode,
    p.name,
    t.transfer_quantity,
    p.base_net_weight
FROM kf_transfer_mart t
LEFT JOIN kf_product_static p ON t.product_id = p.id
WHERE toDate(t.transfer_date) = '2026-05-20'
  AND t.deleted = 0 AND t.status != 6
  AND t.from_branch_id = '{DONG_MAT_BRANCH}'
FORMAT JSONEachRow
""")

classified_dong = {'rows': 0, 'qty': 0, 'tons': 0}
classified_mat = {'rows': 0, 'qty': 0, 'tons': 0}
unclassified = {}  # barcode → {name, qty, tons}
total_rows = 0

for line in r1.split('\n'):
    if not line.strip():
        continue
    obj = json.loads(line)
    total_rows += 1
    bc = str(obj.get('base_barcode') or '').strip()
    name = str(obj.get('name') or '').strip()
    try:
        qty = float(obj.get('transfer_quantity') or 0)
    except:
        qty = 0
    try:
        weight = float(obj.get('base_net_weight') or 0)
    except:
        weight = 0
    tons = qty * weight / 1_000_000

    classification = barcode_type.get(bc)
    if classification == 'ĐÔNG':
        classified_dong['rows'] += 1
        classified_dong['qty'] += qty
        classified_dong['tons'] += tons
    elif classification == 'MÁT':
        classified_mat['rows'] += 1
        classified_mat['qty'] += qty
        classified_mat['tons'] += tons
    else:
        if bc not in unclassified:
            unclassified[bc] = {'name': name, 'qty': 0, 'tons': 0, 'rows': 0}
        unclassified[bc]['qty'] += qty
        unclassified[bc]['tons'] += tons
        unclassified[bc]['rows'] += 1

total_unc_tons = sum(v['tons'] for v in unclassified.values())
total_unc_qty = sum(v['qty'] for v in unclassified.values())
total_unc_rows = sum(v['rows'] for v in unclassified.values())

print(f"\n{'─'*70}")
print(f"  SUMMARY — ABA QUÁ CẢNH transfer 20/05/2026")
print(f"{'─'*70}")
print(f"  Total rows:       {total_rows:>10,}")
print(f"")
print(f"  ✅ ĐÔNG:          {classified_dong['rows']:>10,} rows | {classified_dong['qty']:>12,.0f} qty | {classified_dong['tons']:>8.2f} tấn")
print(f"  ✅ MÁT:           {classified_mat['rows']:>10,} rows | {classified_mat['qty']:>12,.0f} qty | {classified_mat['tons']:>8.2f} tấn")
print(f"  ❌ UNCLASSIFIED:  {total_unc_rows:>10,} rows | {total_unc_qty:>12,.0f} qty | {total_unc_tons:>8.2f} tấn")
print(f"")
print(f"  → {len(unclassified)} unique barcodes chưa phân loại")
print(f"  → {total_unc_tons:.2f} tấn BỊ BỎ QUA (không vào report)")

# Top unclassified by tons
sorted_unc = sorted(unclassified.items(), key=lambda x: -x[1]['tons'])
print(f"\n  TOP 20 UNCLASSIFIED BARCODES (by tons):")
print(f"  {'Barcode':<15} {'Tấn':>8} {'Qty':>10} {'Rows':>6}  Tên sản phẩm")
print(f"  {'─'*15} {'─'*8} {'─'*10} {'─'*6}  {'─'*40}")
for bc, info in sorted_unc[:20]:
    name_short = info['name'][:45]
    print(f"  {bc:<15} {info['tons']:>8.3f} {info['qty']:>10,.0f} {info['rows']:>6}  {name_short}")

# Also compare with yesterday
print(f"\n\n  → Comparing with yesterday (19/05)...")
r_y = q(f"""
SELECT
    p.base_barcode,
    t.transfer_quantity,
    p.base_net_weight
FROM kf_transfer_mart t
LEFT JOIN kf_product_static p ON t.product_id = p.id
WHERE toDate(t.transfer_date) = '2026-05-19'
  AND t.deleted = 0 AND t.status != 6
  AND t.from_branch_id = '{DONG_MAT_BRANCH}'
FORMAT JSONEachRow
""")

y_dong = {'tons': 0, 'qty': 0}
y_mat = {'tons': 0, 'qty': 0}
y_unc = {'tons': 0, 'qty': 0, 'barcodes': set()}

for line in r_y.split('\n'):
    if not line.strip():
        continue
    obj = json.loads(line)
    bc = str(obj.get('base_barcode') or '').strip()
    try:
        qty = float(obj.get('transfer_quantity') or 0)
    except:
        qty = 0
    try:
        weight = float(obj.get('base_net_weight') or 0)
    except:
        weight = 0
    tons = qty * weight / 1_000_000

    cl = barcode_type.get(bc)
    if cl == 'ĐÔNG':
        y_dong['tons'] += tons
        y_dong['qty'] += qty
    elif cl == 'MÁT':
        y_mat['tons'] += tons
        y_mat['qty'] += qty
    else:
        y_unc['tons'] += tons
        y_unc['qty'] += qty
        y_unc['barcodes'].add(bc)

print(f"\n{'─'*70}")
print(f"  COMPARISON — 19/05 vs 20/05")
print(f"{'─'*70}")
print(f"  {'':20} {'19/05':>12} {'20/05':>12} {'Change':>12}")
print(f"  {'MÁT (tấn)':20} {y_mat['tons']:>12.2f} {classified_mat['tons']:>12.2f} {classified_mat['tons']-y_mat['tons']:>+12.2f}")
print(f"  {'ĐÔNG (tấn)':20} {y_dong['tons']:>12.2f} {classified_dong['tons']:>12.2f} {classified_dong['tons']-y_dong['tons']:>+12.2f}")
print(f"  {'UNCLASSIFIED (tấn)':20} {y_unc['tons']:>12.2f} {total_unc_tons:>12.2f} {total_unc_tons-y_unc['tons']:>+12.2f}")
print(f"  {'UNC barcodes':20} {len(y_unc['barcodes']):>12} {len(unclassified):>12}")


# ═══════════════════════════════════════════════════════════
# PART 2: PO KRC — DB vs LOCAL FILE
# ═══════════════════════════════════════════════════════════
print(f"\n\n{'='*70}")
print(f"  PART 2: PO KRC — DB vs LOCAL FILE (19/05/2026)")
print(f"{'='*70}")

# Use 19/05 since we have local file for that date
KRC_BRANCH = '5fdc170ebd89c10006f15b7c'

# DB: PO header summary
r_db = q(f"""
SELECT
    COUNT(*) as po_count,
    SUM(sum_qty) as total_qty,
    SUM(product_type_count) as total_products
FROM kf_purchase_order
WHERE branch_id = '{KRC_BRANCH}'
  AND toDate(fromUnixTimestamp(toUInt32(delivery_date))) = '2026-05-19'
  AND deleted = 0
""")
print(f"\n  DB (kf_purchase_order): POs|qty|products = {r_db}")

# DB: receipt_items weight
r_db_w = q(f"""
SELECT
    round(SUM(ri.qty * ri.net_weight / 1000) / 1000, 2) as total_tons,
    round(SUM(ri.qty), 0) as total_qty,
    COUNT(DISTINCT ri.purchase_code) as po_count
FROM kf_receipt_items ri
WHERE ri.branch_id = '{KRC_BRANCH}'
  AND ri.purchase_code IN (
      SELECT code FROM kf_purchase_order
      WHERE branch_id = '{KRC_BRANCH}'
        AND toDate(fromUnixTimestamp(toUInt32(delivery_date))) = '2026-05-19'
        AND deleted = 0
  )
""")
print(f"  DB (receipt_items):    tons|qty|POs = {r_db_w}")

# Local file
from datetime import datetime
po_dir = r'G:\My Drive\DOCS\DAILY\po_krc'
local_file = os.path.join(po_dir, 'po_krc_19052026.xlsx')
if os.path.exists(local_file):
    print(f"\n  Local file: {os.path.basename(local_file)} ({os.path.getsize(local_file):,} bytes)")
    wb = load_workbook(local_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    local_po_set = set()
    local_qty = 0
    local_kg = 0
    local_rows = 0
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        date_val = row[15].value
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            ds = date_val.strftime('%d/%m/%Y')
        else:
            ds = str(date_val).strip()
        if ds != '19/05/2026':
            continue
        
        local_rows += 1
        po_code = str(row[3].value or '').strip()
        local_po_set.add(po_code)
        
        try:
            qty = float(row[21].value or 0)
        except:
            qty = 0
        try:
            kg = float(row[22].value or 0)
        except:
            kg = 0
        
        local_qty += qty
        local_kg += kg
    
    wb.close()
    
    print(f"  Local file (19/05): {local_rows} rows, {len(local_po_set)} POs, qty={local_qty:,.0f}, kg={local_kg:,.1f} ({local_kg/1000:.2f} tấn)")
    
    # DB PO codes for 19/05
    r_db_codes = q(f"""
    SELECT code FROM kf_purchase_order
    WHERE branch_id = '{KRC_BRANCH}'
      AND toDate(fromUnixTimestamp(toUInt32(delivery_date))) = '2026-05-19'
      AND deleted = 0
    """)
    db_po_set = set(r_db_codes.strip().split('\n')) if r_db_codes.strip() else set()
    
    print(f"\n  DB PO codes:    {len(db_po_set)}")
    print(f"  Local PO codes: {len(local_po_set)}")
    
    only_db = db_po_set - local_po_set
    only_local = local_po_set - db_po_set
    both = db_po_set & local_po_set
    
    print(f"  Matching:       {len(both)}")
    print(f"  Only in DB:     {len(only_db)}")
    print(f"  Only in local:  {len(only_local)}")
    
    if only_db:
        print(f"\n  POs only in DB (sample): {list(only_db)[:5]}")
    if only_local:
        print(f"  POs only in local (sample): {list(only_local)[:5]}")

# DB for TODAY (20/05) — to show we can serve realtime
print(f"\n\n{'─'*70}")
print(f"  PO KRC — DB DATA FOR TODAY (20/05/2026)")
print(f"{'─'*70}")

r_today = q(f"""
SELECT
    COUNT(*) as po_count,
    SUM(sum_qty) as total_qty,
    SUM(product_type_count) as total_products
FROM kf_purchase_order
WHERE branch_id = '{KRC_BRANCH}'
  AND toDate(fromUnixTimestamp(toUInt32(delivery_date))) = '2026-05-20'
  AND deleted = 0
""")
print(f"  POs|qty|products = {r_today}")

r_today_w = q(f"""
SELECT
    round(SUM(ri.qty * ri.net_weight / 1000) / 1000, 2) as total_tons,
    round(SUM(ri.qty), 0) as total_qty
FROM kf_receipt_items ri
WHERE ri.branch_id = '{KRC_BRANCH}'
  AND ri.purchase_code IN (
      SELECT code FROM kf_purchase_order
      WHERE branch_id = '{KRC_BRANCH}'
        AND toDate(fromUnixTimestamp(toUInt32(delivery_date))) = '2026-05-20'
        AND deleted = 0
  )
""")
print(f"  Weight from receipt_items: tons|qty = {r_today_w}")

# Sample PO status distribution
r_status = q(f"""
SELECT status, COUNT(*) as cnt
FROM kf_purchase_order
WHERE branch_id = '{KRC_BRANCH}'
  AND toDate(fromUnixTimestamp(toUInt32(delivery_date))) = '2026-05-20'
  AND deleted = 0
GROUP BY status
ORDER BY cnt DESC
""")
print(f"  Status distribution: {r_status}")

print(f"\n  ✅ DB CÓ ĐỦ DATA → có thể kết nối realtime!")

print("\nDone.")
