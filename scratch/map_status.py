# -*- coding: utf-8 -*-
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
sys.path.insert(0, r'g:\My Drive\DOCS\transport_daily_report\script')
from data_pipeline.config import load_clickhouse_config
import requests

cfg = load_clickhouse_config()
p = {'user': cfg['user'], 'password': cfg['password'], 'database': cfg['database']}

def q(sql):
    r = requests.get(cfg['base_url'], params={**p, 'query': sql}, timeout=60)
    r.raise_for_status()
    return r.text.strip()

# Get PO codes from local file (19/05) — all "Hoàn tất"
wb = load_workbook(r'G:\My Drive\DOCS\DAILY\po_krc\po_krc_19052026.xlsx', read_only=True, data_only=True)
ws = wb.worksheets[0]
local_pos = set()
for row in ws.iter_rows(min_row=2, values_only=False):
    po = str(row[3].value or '').strip()
    if po: local_pos.add(po)
wb.close()
print(f'Local file POs (19/05, all Hoàn tất): {len(local_pos)} unique POs')

# Check these PO codes' status in DB
sample = list(local_pos)
bc_list = "','".join(sample)
r = q(f"SELECT code, status, sub_status FROM kf_purchase_order WHERE code IN ('{bc_list}') FORMAT JSONEachRow")

status_count = {}
for line in r.split('\n'):
    if not line.strip(): continue
    obj = json.loads(line)
    s = obj['status']
    ss = obj['sub_status']
    key = f"status={s}, sub={ss}"
    status_count[key] = status_count.get(key, 0) + 1

print(f'\nDB status of "Hoàn tất" POs:')
for k, v in sorted(status_count.items()):
    print(f'  {k}: {v} POs')

print(f'\n→ "Hoàn tất" trên app = status {set(int(k.split("=")[1].split(",")[0]) for k in status_count)} trên DB')
