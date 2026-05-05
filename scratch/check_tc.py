import json, sys
from datetime import date, timedelta
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Check thitca_actual in monthly_plan_T04
with open('output/state/monthly_plan_T04.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

tc = data.get('thitca_actual', [])
print(f'Total thitca_actual rows in T04 plan: {len(tc)}')

# Check date range
dates = sorted(set(r['date'] for r in tc))
print(f'Date range: {dates[0] if dates else "N/A"} to {dates[-1] if dates else "N/A"}')
print(f'Unique dates: {len(dates)}')

# Check by week
for w in [14,15,16,17,18]:
    monday = date.fromisocalendar(2026, w, 1)
    days = set()
    for i in range(7):
        d = monday + timedelta(days=i)
        days.add(d.strftime('%d/%m/%Y'))
    count = sum(1 for r in tc if r['date'] in days)
    print(f'  W{w} ({monday}): {count} rows')

# Check for actual_time presence
with_actual = [r for r in tc if r.get('actual_time')]
print(f'\nWith actual_time: {len(with_actual)} / {len(tc)}')

# Sample first 3 rows
print('\nSample rows:')
for r in tc[:3]:
    print(f"  {r['date']} | {r['store']} | plan={r.get('planned_time','')} | actual={r.get('actual_time','')} | tuyen={r.get('tuyen','')}")

# Now check load_thitca_data output
sys.path.insert(0, '.')
from script.domains.performance.generate import load_thitca_data, parse_arrival_time, parse_date_str

print('\n--- load_thitca_data([4]) ---')
tc_rows = load_thitca_data([4])
print(f'Loaded rows: {len(tc_rows)}')

# Check by week  
for w in [14,15,16,17,18]:
    monday = date.fromisocalendar(2026, w, 1)
    week_dates = set(monday + timedelta(days=i) for i in range(7))
    count = sum(1 for r in tc_rows if r.get('date') in week_dates)
    with_arr = sum(1 for r in tc_rows if r.get('date') in week_dates and r.get('arrival_time'))
    print(f'  W{w}: {count} rows, {with_arr} with arrival')

# Now also check the raw external file
print('\n--- Raw external file check ---')
from openpyxl import load_workbook
import os
path = r'G:\My Drive\DOCS\DAILY\BÁO CÁO GIAO HÀNG MIỀN ĐÔNG\04.2026 BAO CAO GIAO HANG KINGFOOD.xlsx'
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
print(f'Sheet: {ws.title}')

# Read header
for hdr in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, c in enumerate(hdr):
        print(f'  Col {i}: {c}')

# Count rows with data
total = 0
no_actual = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    date_val = row[2] if len(row) > 2 else None
    actual_val = row[10] if len(row) > 10 else None
    if date_val:
        total += 1
        if not actual_val:
            no_actual += 1

print(f'\nTotal data rows: {total}')
print(f'Rows without actual_time (col 10): {no_actual}')
wb.close()
