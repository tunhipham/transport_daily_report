"""
fetch_weekly_plan.py - Fetch delivery schedule data for a whole week
Usage: python script/fetch_weekly_plan.py [--week W14] [--start DD/MM/YYYY]

Fetches data from KRC, KFM (KSL/DRY), KH MEAT/ĐÔNG/MÁT sources
and outputs JSON grouped by warehouse for email composition.

Output: output/weekly_plan_W{week}.json
"""
import os, sys, re, json
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "data")
OUTPUT = os.path.join(BASE, "output")

# ── Online data source URLs ──
KRC_SHEET_URL = "https://docs.google.com/spreadsheets/d/1tWamqjpOI2j2MrYW3Ah6ptmT524CAlQvEP8fCkxfuII/export?format=xlsx"
KFM_SHEET_URL = "https://docs.google.com/spreadsheets/d/1LkJFJhOQ8F2WEB3uCk7kA2Phvu8IskVi3YBfVr7pBx0/export?format=xlsx"

# Google Drive folder URLs for KH data:
KH_DRIVE_FOLDERS = [
    ("KH MEAT", "THỊT CÁ", 11, 22, -1, "https://drive.google.com/drive/folders/1GIzH8nmCbLhWfpdmxFIn9cHTvQNbnwWr"),
    ("KH HÀNG ĐÔNG", "ĐÔNG MÁT", 9, 17, 18, "https://drive.google.com/drive/folders/1pQ8coYeV-K0dcHlkvXcJ8KngmH22xp1Z"),
    ("KH HÀNG MÁT", "ĐÔNG MÁT", 9, 19, 20, "https://drive.google.com/drive/folders/1c2zfgcXM8O9ezkOZYj0p4t_ihaJmb98f"),
]


def read_xlsx_from_url(url):
    import requests
    r = requests.get(url, allow_redirects=True, timeout=120)
    r.raise_for_status()
    return load_workbook(BytesIO(r.content), read_only=True, data_only=True)


def _list_drive_folder(folder_url):
    import requests as _req
    import html as _html
    folder_id = re.search(r'folders/([a-zA-Z0-9_-]+)', folder_url).group(1)
    url = f"https://drive.google.com/drive/folders/{folder_id}"
    r = _req.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    text = _html.unescape(r.text)
    files = []
    seen = set()
    for m in re.finditer(r'"(1[a-zA-Z0-9_-]{25,50})"', text):
        fid = m.group(1)
        if fid in seen:
            continue
        after = text[m.end():m.end() + 500]
        name_match = re.search(r'"([^"]+\.xlsx)"', after)
        if name_match:
            fname = name_match.group(1)
            if "Microsoft" not in fname and "Shared" not in fname:
                seen.add(fid)
                files.append((fid, fname))
    return files


def read_kh_from_drive(folder_url, folder_name, date_for_file):
    import requests as _req
    files = _list_drive_folder(folder_url)
    if not files:
        print(f"    ⚠ Không list được folder")
        return None
    target = None
    for fid, fname in files:
        if date_for_file in fname:
            target = (fid, fname)
            break
    if not target:
        return None
    print(f"    ↳ {target[1]}")
    dl_url = f"https://drive.google.com/uc?export=download&confirm=t&id={target[0]}"
    r = _req.get(dl_url, allow_redirects=True, timeout=60)
    if r.status_code != 200 or len(r.content) < 500:
        print(f"    ⚠ Download failed (status={r.status_code})")
        return None
    return load_workbook(BytesIO(r.content), read_only=True, data_only=True)


def safe_val(row, idx):
    if idx < len(row):
        return str(row[idx].value or "").strip()
    return ""


def format_time(time_val):
    if not time_val:
        return ""
    s = str(time_val).strip()
    m = re.match(r'(\d{1,2}):(\d{2})', s)
    if m:
        return f"{int(m.group(1))}:{m.group(2)}"
    if hasattr(time_val, 'hour'):
        return f"{time_val.hour}:{time_val.minute:02d}"
    try:
        fval = float(time_val)
        if 0 <= fval < 1:
            total_minutes = round(fval * 24 * 60)
            hours = total_minutes // 60
            minutes = total_minutes % 60
            return f"{hours}:{minutes:02d}"
    except (ValueError, TypeError):
        pass
    return s


def format_date_vn(dt):
    """Format date as DD/MM/YYYY"""
    return dt.strftime("%d/%m/%Y")


def format_date_us(dt):
    """Format date as M/D/YYYY"""
    return f"{dt.month}/{dt.day}/{dt.year}"


def format_date_file(dt):
    """Format date for file matching: DD.MM.YYYY"""
    return dt.strftime("%d.%m.%Y")


def fetch_week_data(start_date, end_date):
    """Fetch delivery data for all days in the week, grouped by warehouse."""
    all_rows = {
        "KRC": [],
        "DRY": [],   # KSL/KFM = DRY
        "ĐÔNG MÁT": [],
        "THỊT CÁ": [],
    }
    warnings = []

    # Generate list of dates in the week
    dates = []
    current = start_date
    while current <= end_date:
        dates.append(current)
        current += timedelta(days=1)

    date_strs = [format_date_vn(d) for d in dates]
    print(f"  Dates: {', '.join(date_strs)}")

    # ── 1. KRC (all dates at once from single sheet) ──
    print("\n  → KRC (Google Sheets)...")
    try:
        wb = read_xlsx_from_url(KRC_SHEET_URL)
        ws = wb["KRC"]
        krc_count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            scv = str(row[0].value or "").strip()
            if scv in date_strs:
                diem_den = str(row[6].value or "").strip()  # Col G
                gio_den = row[7].value                       # Col H
                if diem_den:
                    gio_str = format_time(gio_den)
                    all_rows["KRC"].append({
                        "date": scv,
                        "diem_den": diem_den,
                        "gio_den": gio_str,
                    })
                    krc_count += 1
        wb.close()
        print(f"    {krc_count} rows")
    except Exception as e:
        warnings.append(f"KRC: lỗi — {e}")
        print(f"    ⚠ {e}")

    # ── 2. KFM/KSL = DRY (all dates at once from single sheet) ──
    print("\n  → KFM/DRY (Google Sheets)...")
    try:
        wb = read_xlsx_from_url(KFM_SHEET_URL)
        ws = None
        for name in wb.sheetnames:
            if "DRY" in name:
                ws = wb[name]
                break
        if not ws:
            ws = wb.worksheets[1]

        kfm_count = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            scv = str(row[0].value or "").strip()
            if scv in date_strs:
                diem_den = str(row[6].value or "").strip()   # Col G
                gio_load = row[3].value                       # Col D = giờ load hàng
                gio_den = row[7].value                        # Col H = giờ đến
                if diem_den:
                    gio_str = format_time(gio_den)
                    gio_load_str = format_time(gio_load)
                    all_rows["DRY"].append({
                        "date": scv,
                        "diem_den": diem_den,
                        "gio_load": gio_load_str,
                        "gio_den": gio_str,
                    })
                    kfm_count += 1
        wb.close()
        print(f"    {kfm_count} rows")
    except Exception as e:
        warnings.append(f"KFM: lỗi — {e}")
        print(f"    ⚠ {e}")

    # ── 3. KH MEAT / ĐÔNG / MÁT (per date) ──
    for dt in dates:
        date_str = format_date_vn(dt)
        date_for_file = format_date_file(dt)

        for folder_name, kho_name, tuyen_col, time_col, loai_hang_col, folder_url in KH_DRIVE_FOLDERS:
            print(f"\n  → {folder_name} — {date_str} (Google Drive)...")
            try:
                wb = read_kh_from_drive(folder_url, folder_name, date_for_file)
                if not wb:
                    warnings.append(f"{folder_name} ({date_str}): file {date_for_file} not found")
                    print(f"    ⚠ file not found")
                    continue
                ws = wb.worksheets[0]

                actual_time_col = time_col
                actual_loai_col = loai_hang_col
                for hdr_row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                    for i, cell in enumerate(hdr_row):
                        h = str(cell.value or "").strip().lower()
                        if "kien" in h and "giao" in h:
                            actual_time_col = i
                        elif "loại" in h or "loai" in h:
                            actual_loai_col = i

                count = 0
                for row in ws.iter_rows(min_row=2, values_only=False):
                    diem_den = safe_val(row, 2)
                    gio_den = row[actual_time_col].value if actual_time_col < len(row) else None
                    loai_hang_val = ""
                    if actual_loai_col >= 0 and actual_loai_col < len(row):
                        loai_hang_val = str(row[actual_loai_col].value or "").strip()
                    if diem_den:
                        gio_str = format_time(gio_den) if gio_den else ""
                        entry = {
                            "date": date_str,
                            "diem_den": diem_den,
                            "gio_den": gio_str,
                        }
                        if kho_name == "ĐÔNG MÁT":
                            entry["loai_hang"] = loai_hang_val
                        all_rows[kho_name].append(entry)
                        count += 1
                wb.close()
                print(f"    {count} rows")
            except Exception as e:
                warnings.append(f"{folder_name} ({date_str}): lỗi — {e}")
                print(f"    ⚠ {e}")

    return all_rows, warnings


def calc_current_week():
    """Calculate current week number and start date based on delivery date (D+1).
    
    Anchor: W14 starts Monday 30/03/2026. Each week = +7 days, +1 week number.
    """
    ANCHOR_WEEK = 14
    ANCHOR_START = datetime(2026, 3, 30)  # Monday
    
    tomorrow = datetime.now() + timedelta(days=1)
    days_diff = (tomorrow - ANCHOR_START).days
    weeks_diff = days_diff // 7
    
    week_num = ANCHOR_WEEK + weeks_diff
    week_start = ANCHOR_START + timedelta(weeks=weeks_diff)
    return f"W{week_num}", format_date_vn(week_start)


def main():
    # Parse arguments — auto-calculate if not provided
    default_week, default_start = calc_current_week()
    week_num = default_week
    start_str = default_start

    if "--week" in sys.argv:
        idx = sys.argv.index("--week")
        week_num = sys.argv[idx + 1]
    if "--start" in sys.argv:
        idx = sys.argv.index("--start")
        start_str = sys.argv[idx + 1]

    parts = start_str.split("/")
    start_date = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
    end_date = start_date + timedelta(days=6)  # Mon-Sun

    print("=" * 60)
    print(f"  KẾ HOẠCH GIAO HÀNG - {week_num}")
    print(f"  {format_date_vn(start_date)} → {format_date_vn(end_date)}")
    print("=" * 60)

    data, warnings = fetch_week_data(start_date, end_date)

    for w in warnings:
        print(f"  ⚠️ {w}")

    # Save JSON
    os.makedirs(OUTPUT, exist_ok=True)
    out_path = os.path.join(OUTPUT, f"weekly_plan_{week_num}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({
            "week": week_num,
            "start": format_date_vn(start_date),
            "end": format_date_vn(end_date),
            "data": data,
            "warnings": warnings,
        }, f, ensure_ascii=False, indent=2)

    print(f"\n📊 Summary:")
    for kho, rows in data.items():
        print(f"  {kho}: {len(rows)} rows")
    print(f"\n✅ JSON saved: {out_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
