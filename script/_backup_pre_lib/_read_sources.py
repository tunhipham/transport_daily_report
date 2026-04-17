import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import requests
from io import BytesIO

# 1. Read THIT CA data
print("=" * 60)
print("  THỊT CÁ DATA")
print("=" * 60)
path = r'G:\My Drive\DOCS\DAILY\BÁO CÁO GIAO HÀNG MIỀN ĐÔNG\2026.03 BÁO CÁO GIAO HÀNG KINGFOOD.xlsx'
wb = openpyxl.load_workbook(path, read_only=True)
ws = wb['Sheet1']
header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print("Columns:")
for i, h in enumerate(header):
    print(f"  {i}: {h}")

# Count and sample
cnt = 0
tuyen_vals = set()
dates = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    cnt += 1
    if row[0]:
        dates.add(str(row[0])[:10])
    if row[4]:
        tuyen_vals.add(str(row[4]))
print(f"\nTotal rows: {cnt}")
print(f"Unique dates: {len(dates)}")
print(f"Unique tuyến: {len(tuyen_vals)}")
print(f"Tuyến values: {sorted(tuyen_vals)[:20]}")

# Show sample rows
print("\nSample rows:")
for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
    print(f"  {list(row[:12])}")
wb.close()

# 2. Check KRC source for Tuyen
print("\n" + "=" * 60)
print("  KRC SOURCE (Google Sheets)")
print("=" * 60)
url = "https://docs.google.com/spreadsheets/d/1tWamqjpOI2j2MrYW3Ah6ptmT524CAlQvEP8fCkxfuII/export?format=xlsx"
r = requests.get(url, timeout=60)
wb = openpyxl.load_workbook(BytesIO(r.content), read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}")
ws = wb['KRC']
# Header
for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
    for i, cell in enumerate(row):
        if cell.value:
            print(f"  Col {i}: {cell.value}")
# Sample data
print("\nSample:")
for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
    print(f"  {list(row[:12])}")
wb.close()

# 3. Check KFM/DRY source for Tuyen
print("\n" + "=" * 60)
print("  KFM/DRY SOURCE (Google Sheets)")
print("=" * 60)
url2 = "https://docs.google.com/spreadsheets/d/1LkJFJhOQ8F2WEB3uCk7kA2Phvu8IskVi3YBfVr7pBx0/export?format=xlsx"
r2 = requests.get(url2, timeout=60)
wb2 = openpyxl.load_workbook(BytesIO(r2.content), read_only=True, data_only=True)
print(f"Sheets: {wb2.sheetnames}")
for sn in wb2.sheetnames:
    ws2 = wb2[sn]
    print(f"\n--- Sheet: {sn} ---")
    # Header rows (might be row 1 or 2)
    for row in ws2.iter_rows(min_row=1, max_row=2, values_only=False):
        vals = []
        for i, cell in enumerate(row):
            if cell.value:
                vals.append(f"{i}:{cell.value}")
        if vals:
            print(f"  {vals[:15]}")
    # Sample data
    for row in ws2.iter_rows(min_row=3, max_row=4, values_only=True):
        print(f"  Data: {list(row[:12])}")
wb2.close()

# 4. Check DONG MAT KH sources for Tuyen
print("\n" + "=" * 60)
print("  ĐÔNG MÁT KH SOURCE")
print("=" * 60)
# These are the KH drive files - check if they have Tuyen column
# We'll check a local file if available
import os
data_dir = r'c:\Users\admin\Downloads\transport_daily_report\data'
for f in os.listdir(data_dir):
    if 'dong' in f.lower() or 'mat' in f.lower():
        fp = os.path.join(data_dir, f)
        print(f"\n{f}:")
        wb3 = openpyxl.load_workbook(fp, read_only=True)
        ws3 = wb3.worksheets[0]
        for row in ws3.iter_rows(min_row=1, max_row=2, values_only=False):
            for i, cell in enumerate(row):
                if cell.value:
                    print(f"  Col {i}: {cell.value}")
            print()
        wb3.close()
        break
