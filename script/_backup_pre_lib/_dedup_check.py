import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, os
from collections import Counter

data_dir = r'G:\My Drive\DOCS\DAILY\DS chi tiet chuyen xe\T03.26'
files = sorted([f for f in os.listdir(data_dir) if f.endswith('.xlsx')])

# Sample from first file
f = files[0]
wb = openpyxl.load_workbook(os.path.join(data_dir, f), read_only=True)
ws = wb['Sheet 1']

trip_dest = Counter()
total = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    trip_id = row[0]
    dest = row[9]
    trip_dest[(trip_id, dest)] += 1
    total += 1

print(f"File: {f}")
print(f"Total raw rows: {total}")
print(f"Unique (trip, dest) pairs: {len(trip_dest)}")
print(f"Avg rows per pair: {total/len(trip_dest):.1f}")

most_common = trip_dest.most_common(3)
for (tid, dest), cnt in most_common:
    print(f"  {tid} -> {dest}: {cnt} rows")

# Sample rows
print("\nSample rows:")
for row in ws.iter_rows(min_row=2, max_row=8, values_only=True):
    print(f"  Trip:{row[0]} St:{row[1]} Src:{row[8]} Dst:{row[9]} DstSt:{row[11]} TGden:{row[26]} Thung:{row[18]}")
wb.close()

# All files
print("\n=== All T03.26 files ===")
all_noi_chuyen = Counter()
all_trips = set()
all_trip_dest = set()
all_cancel_trip = set()
all_cancel_dest = set()
for f in files:
    wb = openpyxl.load_workbook(os.path.join(data_dir, f), read_only=True)
    ws = wb['Sheet 1']
    for row in ws.iter_rows(min_row=2, values_only=True):
        nc = str(row[8]) if row[8] else 'N/A'
        all_noi_chuyen[nc] += 1
        all_trips.add(row[0])
        if row[9]:
            all_trip_dest.add((row[0], row[9]))
        if row[1] and 'Hủy' in str(row[1]):
            all_cancel_trip.add(row[0])
        if row[11] and 'Hủy' in str(row[11]):
            all_cancel_dest.add((row[0], row[9]))
    wb.close()

print("Noi chuyen distribution:")
for nc, cnt in all_noi_chuyen.most_common():
    print(f"  {nc}: {cnt} raw rows")
print(f"\nUnique trips: {len(all_trips)}")
print(f"Unique (trip,dest): {len(all_trip_dest)}")
print(f"Cancelled trips: {len(all_cancel_trip)}")
print(f"Cancelled (trip,dest): {len(all_cancel_dest)}")

# Map Noi chuyen to warehouse
print("\nNoi chuyen -> Warehouse mapping:")
print("  KSL -> DRY (KSL = Kho Seedlog = DRY)")
print("  KRC -> KRC")
print("  QCABA -> DONG MAT")
print("  THIT CA -> from separate file")
