# -*- coding: utf-8 -*-
"""Generate W18 weekly plan from master_schedule.json + kiểm kê + NSO châm hàng.
This is the tab "Lịch Tuần" data for W18 (27/04 - 03/05/2026).
"""

import os, sys, json, re
from datetime import datetime, timedelta, date
from io import BytesIO

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(BASE, "script"))
from lib.sources import INVENTORY_SHEET_URL

# Import NSO stores
sys.path.insert(0, os.path.join(BASE, "script", "domains", "nso"))
from generate import STORES as NSO_STORES, parse_date as nso_parse_date

# ─── Config ───
ANCHOR_WEEK = 14
ANCHOR_START = date(2026, 3, 30)  # Monday W14

WEEK_NUM = 18
week_start = ANCHOR_START + timedelta(weeks=WEEK_NUM - ANCHOR_WEEK)  # Monday 27/04
week_dates = [week_start + timedelta(days=i) for i in range(7)]
week_date_strs = [d.strftime("%d/%m/%Y") for d in week_dates]
day_names = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
week_label = f"TUẦN {WEEK_NUM} TỪ {week_dates[0].strftime('%d/%m')} - {week_dates[6].strftime('%d/%m')}"

print(f"📅 Generating W{WEEK_NUM}: {week_label}")
print(f"   Dates: {' | '.join(f'{dn} {ds}' for dn, ds in zip(day_names, week_date_strs))}")

# ─── Load master schedule ───
ms_path = os.path.join(BASE, "data", "master_schedule.json")
with open(ms_path, "r", encoding="utf-8") as f:
    master = json.load(f)

print(f"\n📋 Master schedule: {len(master['stores'])} stores (source: {master['_source']})")

# ─── Parse schedule_ve to weekday indices ───
def parse_schedule_days(schedule_str):
    days = set()
    nums = re.findall(r'\d', schedule_str)
    for n in nums:
        n = int(n)
        if 2 <= n <= 7:
            days.add(n - 2)
    return days

def compute_even_date_schedule(week_dates):
    even_weekdays = set()
    for wd in week_dates:
        if wd and wd.day % 2 == 0 and wd.weekday() != 6:
            even_weekdays.add(wd.weekday())
    return even_weekdays

# ─── Build stores from master schedule ───
stores = []
for s in master["stores"]:
    code = s["code"]
    schedule_ve = s.get("schedule_ve", "")
    shift = s.get("shift", "Đêm")
    
    # Compute delivery days
    if schedule_ve == "Ngày chẵn":
        delivery_weekdays = compute_even_date_schedule(week_dates)
    else:
        delivery_weekdays = parse_schedule_days(schedule_ve)
    
    days = []
    for i, wd in enumerate(week_dates):
        if wd.weekday() in delivery_weekdays:
            days.append(shift)
        else:
            days.append("")
    
    # Build even-date label for A112
    if schedule_ve == "Ngày chẵn" and delivery_weekdays:
        thu_nums = sorted(wd + 2 for wd in delivery_weekdays if wd < 6)
        if thu_nums:
            schedule_ve_display = "Thứ " + "-".join(str(n) for n in thu_nums)
        else:
            schedule_ve_display = "Ngày chẵn"
    else:
        schedule_ve_display = schedule_ve
    
    stores.append({
        "name": s["name"],
        "code": code,
        "schedule_ve": schedule_ve_display,
        "opening_date": "",
        "inventory_date": "",
        "shift": shift,
        "days": days,
    })

print(f"   Built {len(stores)} stores from master schedule")

# ─── Fetch inventory ───
print("\n🔍 Fetching inventory schedule (kiểm kê)...")
inventory = {}
try:
    import requests
    from openpyxl import load_workbook
    r = requests.get(INVENTORY_SHEET_URL, allow_redirects=True, timeout=60)
    r.raise_for_status()
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    ws = wb['Lịch Kiểm kê 2026']
    for row in ws.iter_rows(min_row=10, values_only=False):
        store_id = str(row[3].value or "").strip()
        kiem_ke = row[7].value
        if store_id and kiem_ke:
            dt = None
            if isinstance(kiem_ke, datetime):
                dt = kiem_ke.date()
            elif isinstance(kiem_ke, date):
                dt = kiem_ke
            elif isinstance(kiem_ke, str):
                try:
                    dt = datetime.strptime(kiem_ke, "%d/%m/%Y").date()
                except ValueError:
                    pass
            if dt:
                inventory[store_id] = dt
    wb.close()
    print(f"   Loaded {len(inventory)} entries")
except Exception as e:
    print(f"   ⚠ Could not fetch: {e}")

# Apply inventory cross-check
kk_applied = 0
for store in stores:
    inv_dt = inventory.get(store["code"])
    if not inv_dt:
        continue
    if isinstance(inv_dt, datetime):
        inv_dt = inv_dt.date()
    
    applied = False
    for i, wd in enumerate(week_dates):
        if wd == inv_dt or wd == inv_dt - timedelta(days=1):
            current = store["days"][i]
            if current and current.lower() in ("ngày", "đêm"):
                store["days"][i] = "Kiểm kê"
                applied = True
            elif not current and wd == inv_dt:
                store["days"][i] = "Kiểm kê"
                applied = True
    
    store["inventory_date"] = inv_dt.strftime("%d/%m/%Y")
    if applied:
        kk_applied += 1

print(f"   Applied kiểm kê to {kk_applied} stores")

# ─── NSO Châm hàng ───
print("\n🏪 Checking NSO châm hàng...")

# Import NSO_SCHEDULE from export_weekly_plan.py
sys.path.insert(0, os.path.join(BASE, "script", "dashboard"))
from export_weekly_plan import NSO_SCHEDULE

existing_codes = {s["code"] for s in stores}
nso_injected = 0

for nso in NSO_STORES:
    try:
        opening = nso_parse_date(nso["opening_date"])
    except:
        continue
    
    # Check if original_date rescheduled
    if nso.get("original_date"):
        try:
            orig = nso_parse_date(nso["original_date"])
            orig_in_week = any(week_dates[0] <= orig + timedelta(days=i) <= week_dates[6] for i in range(4))
            new_in_week = any(week_dates[0] <= opening + timedelta(days=i) <= week_dates[6] for i in range(4))
            if orig_in_week and not new_in_week:
                continue
        except:
            pass
    
    cham_days = []
    for i in range(4):
        d = opening + timedelta(days=i)
        if week_dates[0] <= d <= week_dates[6]:
            cham_days.append(d)
    
    if not cham_days:
        continue
    
    code = nso["code"]
    sched = NSO_SCHEDULE.get(code, {})
    schedule_ve = sched.get("schedule_ve", "")
    shift = sched.get("shift", "Đêm")
    
    # Find existing store
    matched = None
    for s in stores:
        if s["code"] == code:
            matched = s
            break
    
    is_injected = False
    if not matched:
        name_sys = nso.get("name_system", "")
        name_full = nso.get("name_full", "")
        store_name = f"{name_sys} - {name_full}" if name_sys else name_full or f"NSO {code}"
        
        matched = {
            "name": store_name,
            "code": code,
            "schedule_ve": schedule_ve,
            "opening_date": opening.strftime("%d/%m/%Y"),
            "inventory_date": "",
            "shift": shift,
            "days": [""] * 7,
        }
        stores.append(matched)
        is_injected = True
        nso_injected += 1
    
    # Set opening_date
    if not matched.get("opening_date"):
        matched["opening_date"] = opening.strftime("%d/%m/%Y")
    
    delivery_weekdays = parse_schedule_days(schedule_ve)
    
    for i, wd in enumerate(week_dates):
        delta = (wd - opening).days
        if 0 <= delta <= 3:
            matched["days"][i] = "Châm hàng"
        elif delta > 3 and is_injected:
            if wd.weekday() in delivery_weekdays:
                matched["days"][i] = shift

nso_cham_count = sum(1 for s in stores if any("châm" in (d or "").lower() for d in s["days"]))
print(f"   Châm hàng: {nso_cham_count} stores ({nso_injected} injected)")

# ─── Stats ───
st_ngay = set()
st_dem = set()
st_cham = set()
st_kk = set()
for s in stores:
    sl = (s.get("shift") or "").lower()
    if sl == "ngày":
        st_ngay.add(s["code"])
    elif sl == "đêm":
        st_dem.add(s["code"])
    for d in s["days"]:
        dl = d.lower() if d else ""
        if "châm" in dl:
            st_cham.add(s["code"])
        elif "kiểm" in dl:
            st_kk.add(s["code"])

stats = {
    "ngay": len(st_ngay),
    "dem": len(st_dem),
    "cham_hang": len(st_cham),
    "kiem_ke": len(st_kk),
}
print(f"\n📊 Stats: {stats['ngay']} ngày, {stats['dem']} đêm, {stats['cham_hang']} châm, {stats['kiem_ke']} kiểm kê")
print(f"   Total stores: {len(stores)}")

# ─── Build W18 data ───
w18 = {
    "label": week_label,
    "total_stores": len(stores),
    "dates": week_date_strs,
    "day_names": day_names,
    "stores": stores,
    "stats": stats,
}

# ─── Merge into weekly_plan.json ───
wp_path = os.path.join(BASE, "docs", "data", "weekly_plan.json")
with open(wp_path, "r", encoding="utf-8") as f:
    wp = json.load(f)

wp["weeks"]["W18"] = w18
if "W18" not in wp["available_weeks"]:
    wp["available_weeks"].append("W18")
    wp["available_weeks"].sort(key=lambda w: int(w[1:]))
wp["_updated"] = datetime.now().strftime("%d/%m/%Y %H:%M")

with open(wp_path, "w", encoding="utf-8") as f:
    json.dump(wp, f, ensure_ascii=False)

fsize = os.path.getsize(wp_path)
print(f"\n✅ weekly_plan.json updated — {len(wp['available_weeks'])} weeks ({fsize:,} bytes)")
print(f"   W18: {week_label} | {len(stores)} stores")
