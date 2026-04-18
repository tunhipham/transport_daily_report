import json, sys
sys.stdout.reconfigure(encoding='utf-8')

d = json.load(open('docs/data/weekly_plan.json', 'r', encoding='utf-8'))
w = d['weeks']['W17']
dates = w['dates']
print(f"W17 dates: {dates}")

for s in w['stores']:
    if s['code'] == 'A121':
        print(f"A121: inv={s['inventory_date']} days={s['days']}")
        break

# Also check what the original Excel has for A121
from openpyxl import load_workbook
wb = load_workbook('output/artifacts/weekly transport plan/Lịch đi hàng ST W17.xlsx', data_only=True)
ws = wb['Lịch về hàng']
for row in ws.iter_rows(min_row=4, values_only=False):
    code = str(row[1].value or "").strip()
    if code == 'A121':
        name = str(row[0].value or "")
        days = [str(row[c].value or "") for c in range(7, 14)]
        print(f"Excel A121: {name}")
        print(f"Excel days: {days}")
        break
wb.close()

# Check inventory date from Google Sheets
import requests
from io import BytesIO
from openpyxl import load_workbook as lw2
sys.path.insert(0, 'script')
from lib.sources import INVENTORY_SHEET_URL
r = requests.get(INVENTORY_SHEET_URL, allow_redirects=True, timeout=60)
wb2 = lw2(BytesIO(r.content), read_only=True, data_only=True)
ws2 = wb2['Lịch Kiểm kê 2026']
for row in ws2.iter_rows(min_row=10, values_only=False):
    sid = str(row[3].value or "").strip()
    if sid == 'A121':
        kk = row[7].value
        print(f"Google Sheet A121: KK={kk} (type={type(kk).__name__})")
        break
wb2.close()
