# -*- coding: utf-8 -*-
"""
export_weekly_plan.py — Parse weekly transport plan Excel files and generate
docs/data/weekly_plan.json for the dashboard.

Sources:
  - Excel: output/artifacts/weekly transport plan/Lịch đi hàng ST W{nn}.xlsx
  - Sheet "Lịch về hàng" → per-week schedule  
  - Google Sheets kiểm kê → cross-check inventory dates (D/D-1)
  - NSO STORES → auto-calculate châm hàng (D→D+3)

Usage:
  python script/dashboard/export_weekly_plan.py
"""

import os, sys, json, glob, re
from datetime import datetime, timedelta, date
from io import BytesIO

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PLAN_DIR = os.path.join(BASE, "output", "artifacts", "weekly transport plan")
DOCS_DATA = os.path.join(BASE, "docs", "data")
os.makedirs(DOCS_DATA, exist_ok=True)

# Import shared sources
sys.path.insert(0, os.path.join(BASE, "script"))
from lib.sources import INVENTORY_SHEET_URL

# Import NSO stores for châm hàng calculation
sys.path.insert(0, os.path.join(BASE, "script", "domains", "nso"))
from generate import STORES as NSO_STORES, parse_date as nso_parse_date

# ═══════════════════════════════════════════════
# NSO SCHEDULE INFO — fixed delivery schedules
# User-confirmed shift: A177 = Ngày, all others = Đêm
# ═══════════════════════════════════════════════
NSO_SCHEDULE = {
    # W16 stores (opening 17-18/04/2026)
    "A148": {"schedule_chia": "Thứ 3-5-7", "schedule_ve": "Thứ 2-4-6", "shift": "Đêm",
             "name_full": "127 Tân Cảng"},
    "A179": {"schedule_chia": "Thứ 3-5-7", "schedule_ve": "Thứ 2-4-6", "shift": "Đêm",
             "name_full": "A1.02 Him Lam Phú An"},
    "A177": {"schedule_chia": "Thứ 2-4-6", "schedule_ve": "Thứ 3-5-7", "shift": "Ngày",
             "name_full": "Sky Garden 2"},
    "A161": {"schedule_chia": "Thứ 2-4-6", "schedule_ve": "Thứ 3-5-7", "shift": "Đêm",
             "name_full": "The Park Residence"},
    # W17 stores — Group 1: chia 3-5-7, về 2-4-6 (opening 23-25/04/2026)
    "A164": {"schedule_chia": "Thứ 3-5-7", "schedule_ve": "Thứ 2-4-6", "shift": "Đêm",
             "name_full": "Opal Boulevard"},
    "A185": {"schedule_chia": "Thứ 3-5-7", "schedule_ve": "Thứ 2-4-6", "shift": "Đêm",
             "name_full": "Vinhomes Grand Park"},
    "A167": {"schedule_chia": "Thứ 3-5-7", "schedule_ve": "Thứ 2-4-6", "shift": "Đêm",
             "name_full": "9 View"},
    "A192": {"schedule_chia": "Thứ 3-5-7", "schedule_ve": "Thứ 2-4-6", "shift": "Đêm",
             "name_full": "Bùi Đình Túy"},
    # W17 stores — Group 2: chia 2-4-6, về 3-5-7
    "A178": {"schedule_chia": "Thứ 2-4-6", "schedule_ve": "Thứ 3-5-7", "shift": "Đêm",
             "name_full": "Celesta Rise"},
    "A191": {"schedule_chia": "Thứ 2-4-6", "schedule_ve": "Thứ 3-5-7", "shift": "Đêm",
             "name_full": "Nguyễn Hữu Cầu"},
    "A176": {"schedule_chia": "Thứ 2-4-6", "schedule_ve": "Thứ 3-5-7", "shift": "Đêm",
             "name_full": "Sunrise Riverside"},
    "A163": {"schedule_chia": "Thứ 2-4-6", "schedule_ve": "Thứ 3-5-7", "shift": "Đêm",
             "name_full": "Celadon City"},
}


def fetch_inventory_schedule():
    """Fetch inventory dates from Google Sheets. Returns dict: store_id → inventory_date."""
    import requests
    from openpyxl import load_workbook
    
    try:
        r = requests.get(INVENTORY_SHEET_URL, allow_redirects=True, timeout=60)
        r.raise_for_status()
        wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
        ws = wb['Lịch Kiểm kê 2026']
        
        inventory = {}  # store_id → datetime
        for row in ws.iter_rows(min_row=10, values_only=False):
            store_id = str(row[3].value or "").strip()  # Col D = ID Mart
            store_name = str(row[2].value or "").strip()  # Col C = Tên Mart  
            kiem_ke = row[7].value  # Col H = Ngày kiểm kê tổng 2026
            if store_id and kiem_ke:
                dt = None
                if isinstance(kiem_ke, datetime):
                    dt = kiem_ke.date()  # datetime → date
                elif isinstance(kiem_ke, date):
                    dt = kiem_ke  # already date
                elif isinstance(kiem_ke, str):
                    try:
                        dt = datetime.strptime(kiem_ke, "%d/%m/%Y").date()
                    except ValueError:
                        pass
                if dt:
                    inventory[store_id] = dt
                    if store_name:
                        inventory[store_name] = dt
        wb.close()
        print(f"  📋 Loaded {len(inventory)} entries from inventory schedule")
        return inventory
    except Exception as e:
        print(f"  ⚠ Could not fetch inventory schedule: {e}")
        return {}


def get_nso_cham_hang(week_dates):
    """Calculate NSO châm hàng for stores opening within this week's date range.
    
    Returns dict: store_code → { opening, name_full, name_system, schedule_info }
    Only includes stores whose opening date (D→D+3) overlaps with week_dates.
    Excludes rescheduled stores whose ORIGINAL opening was in this week but NEW opening is not.
    """
    cham_hang = {}
    
    week_start = min(week_dates)
    week_end = max(week_dates)
    
    for store in NSO_STORES:
        try:
            opening = nso_parse_date(store["opening_date"])
        except:
            continue
        
        # Skip rescheduled stores: if original_date is set and the store's
        # current opening_date doesn't overlap with this week, skip
        if store.get("original_date"):
            try:
                orig = nso_parse_date(store["original_date"])
                # If original was in this week but new is NOT → skip (dời lịch)
                orig_in_week = any(week_start <= orig + timedelta(days=i) <= week_end for i in range(4))
                new_in_week = any(week_start <= opening + timedelta(days=i) <= week_end for i in range(4))
                if orig_in_week and not new_in_week:
                    continue
            except:
                pass
        
        # Check if any of D to D+3 falls within this week
        cham_days = []
        for i in range(4):
            d = opening + timedelta(days=i)
            if week_start <= d <= week_end:
                cham_days.append(d)
        
        if cham_days:
            code = store["code"]
            sched = NSO_SCHEDULE.get(code, {})
            cham_hang[code] = {
                "dates": cham_days,
                "opening": opening,
                "name_full": store.get("name_full", ""),
                "name_system": store.get("name_system", ""),
                "schedule_chia": sched.get("schedule_chia", ""),
                "schedule_ve": sched.get("schedule_ve", ""),
                "shift": sched.get("shift", "Đêm"),
            }
    
    return cham_hang


def _name_matches(nso_name_full, excel_store_name):
    """Check if NSO store name matches an Excel store row name.
    
    NSO name_full: 'A1.02 Him Lam Phú An'
    Excel name:    'KFM_HCM_TDU - A1.02 Him Lam Phú An'
    
    Returns True if enough significant words from name_full appear in the Excel name.
    """
    if not nso_name_full:
        return True  # No name to verify — allow match
    
    excel_lower = excel_store_name.lower()
    # Extract words from NSO name (3+ chars to skip articles/numbers)
    words = [w for w in nso_name_full.lower().split() if len(w) >= 3]
    if not words:
        return True
    
    matches = sum(1 for w in words if w in excel_lower)
    # At least half the significant words must match
    return matches >= max(1, len(words) // 2)


def fmt_date_ddmmyyyy(val):
    """Convert datetime/date to dd/mm/yyyy string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if s == "00:00:00" or s == "0":
        return ""
    return s


def parse_schedule_days(schedule_str):
    """Parse 'Thứ 3-5-7' or 'Thứ 2-4-6' to set of weekday indices (Mon=0)."""
    days = set()
    nums = re.findall(r'\d', schedule_str)
    for n in nums:
        n = int(n)
        if 2 <= n <= 7:
            days.add(n - 2)  # Thứ 2 = Mon = 0, Thứ 7 = Sat = 5
    return days


def parse_excel(filepath):
    """Parse a single weekly plan Excel file."""
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath, data_only=True)
    
    if "Lịch về hàng" not in wb.sheetnames:
        print(f"  ⚠ Sheet 'Lịch về hàng' not found in {filepath}")
        return None
    
    ws = wb["Lịch về hàng"]
    
    week_label = str(ws.cell(2, 8).value or "")
    wn_match = re.search(r'TUẦN\s+(\d+)', week_label, re.IGNORECASE)
    week_num = int(wn_match.group(1)) if wn_match else 0
    week_key = f"W{week_num}"
    
    # R3 cols 8-14: dates
    week_dates = []
    week_date_strs = []
    for c in range(8, 15):
        val = ws.cell(3, c).value
        if isinstance(val, (datetime, date)):
            d = val.date() if isinstance(val, datetime) else val
            week_dates.append(d)
            week_date_strs.append(d.strftime("%d/%m/%Y"))
        else:
            week_date_strs.append(fmt_date_ddmmyyyy(val))
            try:
                d = datetime.strptime(str(val).strip(), "%d/%m/%Y").date()
                week_dates.append(d)
            except:
                week_dates.append(None)
    
    # R4: day names
    day_names = []
    for c in range(8, 15):
        day_names.append(str(ws.cell(4, c).value or ""))
    
    # R5+: store data
    stores = []
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        code = str(ws.cell(r, 2).value or "").strip()
        schedule_chia = str(ws.cell(r, 3).value or "").strip()
        schedule_ve = str(ws.cell(r, 4).value or "").strip()
        opening_raw = ws.cell(r, 5).value
        inventory_raw = ws.cell(r, 6).value
        shift = str(ws.cell(r, 7).value or "").strip()
        
        days = []
        for c in range(8, 15):
            val = ws.cell(r, c).value
            days.append(str(val).strip() if val else "")
        
        # Fix known code mismatches in source Excel
        CODE_CORRECTIONS = {
            # "Sunrise Riverside" is mistakenly coded as A179 in Excel, actual code is A176
            "A179": {"name_contains": "Sunrise Riverside", "correct_code": "A176"},
        }
        if code in CODE_CORRECTIONS:
            fix = CODE_CORRECTIONS[code]
            if fix["name_contains"].lower() in name.lower():
                code = fix["correct_code"]
        
        stores.append({
            "name": name,
            "code": code,
            "schedule_chia": schedule_chia,
            "schedule_ve": schedule_ve,
            "opening_date": fmt_date_ddmmyyyy(opening_raw),
            "inventory_date": fmt_date_ddmmyyyy(inventory_raw),
            "shift": shift,
            "days": days,
        })
    
    wb.close()
    
    return {
        "week_key": week_key,
        "week_num": week_num,
        "label": week_label,
        "total_stores": len(stores),
        "dates": week_date_strs,
        "day_names": day_names,
        "week_dates": week_dates,
        "stores": stores,
    }


def cross_check_inventory(week_data, inventory):
    """Cross-check inventory dates with weekly schedule.
    Updates store days to mark 'Kiểm kê' on D and D-1.
    """
    if not inventory or not week_data:
        return
    
    week_dates = week_data.get("week_dates", [])
    
    for store in week_data["stores"]:
        inv_dt = inventory.get(store["code"]) or inventory.get(store["name"])
        if not inv_dt:
            continue
        # Ensure inv_dt is a date (not datetime) for comparison with week_dates
        if isinstance(inv_dt, datetime):
            inv_dt = inv_dt.date()
        
        # Step 1: Clear any incorrect "Kiểm kê" from Excel source
        for i, wd in enumerate(week_dates):
            if wd is None:
                continue
            current = store["days"][i]
            if current and current.lower().startswith("kiểm"):
                # This day has Kiểm kê — verify it's actually D or D-1
                is_d = (wd == inv_dt)
                is_d1 = (wd == inv_dt - timedelta(days=1))
                if not is_d and not is_d1:
                    # Wrong! Clear it back to empty
                    store["days"][i] = ""
        
        # Step 2: Mark correct D and D-1 as Kiểm kê
        for i, wd in enumerate(week_dates):
            if wd is None:
                continue
            if wd == inv_dt or wd == inv_dt - timedelta(days=1):
                current = store["days"][i]
                if current and current.lower() in ("ngày", "đêm"):
                    store["days"][i] = "Kiểm kê"
                elif not current:
                    if wd == inv_dt:
                        store["days"][i] = "Kiểm kê"
        
        store["inventory_date"] = inv_dt.strftime("%d/%m/%Y")


def apply_nso_cham_hang(week_data, cham_hang):
    """Apply NSO châm hàng logic to the weekly schedule.
    
    For each NSO store opening in this week:
    1. Find matching store in Excel by code + name verification
    2. If not found → inject new row with computed schedule
    3. Mark D to D+3 as 'Châm hàng'
    4. For injected stores: compute post-châm delivery days from schedule
    """
    if not cham_hang or not week_data:
        return
    
    week_dates = week_data.get("week_dates", [])
    existing_codes = {s["code"] for s in week_data["stores"]}
    
    for nso_code, nso in cham_hang.items():
        opening = nso["opening"]
        nso_name = nso.get("name_full", "")
        
        # Find matching store in Excel by code + name
        matched_store = None
        for store in week_data["stores"]:
            if store["code"] == nso_code:
                # Verify name to avoid false positives (e.g., A179 matching wrong store)
                if _name_matches(nso_name, store["name"]):
                    matched_store = store
                    break
        
        is_injected = False
        if not matched_store:
            # Store not in Excel → inject new row
            # Build full store name from NSO data
            name_sys = nso.get("name_system", "")
            name_full = nso.get("name_full", "")
            if name_sys:
                store_name = f"{name_sys} - {name_full}"
            else:
                store_name = name_full or f"NSO {nso_code}"
            
            new_store = {
                "name": store_name,
                "code": nso_code,
                "schedule_chia": nso.get("schedule_chia", ""),
                "schedule_ve": nso.get("schedule_ve", ""),
                "opening_date": opening.strftime("%d/%m/%Y"),
                "inventory_date": "",
                "shift": nso.get("shift", "Đêm"),
                "days": [""] * 7,
            }
            week_data["stores"].append(new_store)
            matched_store = new_store
            is_injected = True
        
        # Apply châm hàng + post-châm schedule
        schedule_ve = nso.get("schedule_ve", "") or matched_store.get("schedule_ve", "")
        shift = nso.get("shift", "") or matched_store.get("shift", "Đêm")
        delivery_weekdays = parse_schedule_days(schedule_ve)
        
        for i, wd in enumerate(week_dates):
            if wd is None:
                continue
            
            delta = (wd - opening).days
            
            if 0 <= delta <= 3:
                # D to D+3: mark as Châm hàng
                matched_store["days"][i] = "Châm hàng"
            elif delta > 3 and is_injected:
                # Post-châm for injected stores: set delivery days
                if wd.weekday() in delivery_weekdays:
                    matched_store["days"][i] = shift
    
    # Cleanup: remove false "Châm hàng" from stores that are NOT valid NSO matches
    # (e.g., Sunrise Riverside A179 in W16 Excel — wrong code match)
    valid_matched = set()
    for nso_code, nso in cham_hang.items():
        nso_name = nso.get("name_full", "")
        for store in week_data["stores"]:
            if store["code"] == nso_code and _name_matches(nso_name, store["name"]):
                valid_matched.add(id(store))
                break
    
    for store in week_data["stores"]:
        if id(store) not in valid_matched:
            for i, d in enumerate(store["days"]):
                if d and "châm" in d.lower():
                    store["days"][i] = ""


def export():
    """Main export function."""
    print("📅 Exporting Weekly Transport Plan data...")
    
    # Find all Excel files
    pattern = os.path.join(PLAN_DIR, "Lịch đi hàng ST W*.xlsx")
    files = sorted(glob.glob(pattern))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    
    if not files:
        print(f"  ⚠ No weekly plan files found in {PLAN_DIR}")
        return False
    
    print(f"  📂 Found {len(files)} files:")
    for f in files:
        print(f"     {os.path.basename(f)}")
    
    # Fetch inventory schedule for cross-check
    print("\n  🔍 Fetching inventory schedule (kiểm kê)...")
    inventory = fetch_inventory_schedule()
    
    # Parse all weeks
    weeks = {}
    available_weeks = []
    
    for filepath in files:
        print(f"\n  📄 Parsing {os.path.basename(filepath)}...")
        week_data = parse_excel(filepath)
        if not week_data:
            continue
        
        wk = week_data["week_key"]
        original_count = len(week_data["stores"])
        
        # Cross-check inventory
        if inventory:
            cross_check_inventory(week_data, inventory)
        
        # Calculate NSO châm hàng
        valid_dates = [d for d in week_data.get("week_dates", []) if d]
        if valid_dates:
            cham_hang = get_nso_cham_hang(valid_dates)
            if cham_hang:
                apply_nso_cham_hang(week_data, cham_hang)
                injected = len(week_data["stores"]) - original_count
                print(f"    🏪 NSO châm hàng: {len(cham_hang)} stores" +
                      (f" ({injected} injected)" if injected else ""))
        
        # Count stats
        n_ngay = n_dem = n_cham = n_kk = 0
        for s in week_data["stores"]:
            for d in s["days"]:
                dl = d.lower() if d else ""
                if dl == "ngày":
                    n_ngay += 1
                elif dl == "đêm":
                    n_dem += 1
                elif "châm" in dl:
                    n_cham += 1
                elif "kiểm" in dl:
                    n_kk += 1
        
        # Clean up for JSON
        stores_clean = []
        for s in week_data["stores"]:
            stores_clean.append({
                "name": s["name"],
                "code": s["code"],
                "schedule_chia": s["schedule_chia"],
                "schedule_ve": s["schedule_ve"],
                "opening_date": s["opening_date"],
                "inventory_date": s["inventory_date"],
                "shift": s["shift"],
                "days": s["days"],
            })
        
        weeks[wk] = {
            "label": week_data["label"],
            "total_stores": len(week_data["stores"]),
            "dates": week_data["dates"],
            "day_names": week_data["day_names"],
            "stores": stores_clean,
            "stats": {
                "ngay": n_ngay,
                "dem": n_dem,
                "cham_hang": n_cham,
                "kiem_ke": n_kk,
            },
        }
        available_weeks.append(wk)
        print(f"    ✅ {wk}: {len(week_data['stores'])} stores | "
              f"{n_ngay} ngày, {n_dem} đêm, {n_cham} châm, {n_kk} kiểm kê")
    
    # Sort weeks
    available_weeks.sort(key=lambda w: int(w[1:]))
    
    # Build output
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    data = {
        "_updated": now_str,
        "available_weeks": available_weeks,
        "weeks": weeks,
    }
    
    out_path = os.path.join(DOCS_DATA, "weekly_plan.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    
    fsize = os.path.getsize(out_path)
    print(f"\n  ✅ {out_path} ({len(available_weeks)} weeks, {fsize:,} bytes)")
    return True


if __name__ == "__main__":
    export()
