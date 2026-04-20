# -*- coding: utf-8 -*-
"""
auto_inventory_watch.py — Monitor kiểm kê schedule changes on Mondays

Polls Google Sheets inventory schedule every hour, detects changes,
re-exports weekly_plan.json, deploys to GitHub Pages, and sends
Telegram notification when changes are found.

Usage:
  python script/dashboard/auto_inventory_watch.py              # One-shot check
  python script/dashboard/auto_inventory_watch.py --watch      # Watch mode (loop every 1h until 17:30)
  python script/dashboard/auto_inventory_watch.py --dry-run    # Check only, no deploy/notify
  python script/dashboard/auto_inventory_watch.py --force      # Force run (ignore day-of-week check)
  python script/dashboard/auto_inventory_watch.py --backup     # Backup: fetch + update + notify (any day, one-shot)

Schedule: runs via Windows Task Scheduler on Mondays only.
          Use --backup for ad-hoc runs outside Monday.
"""

import os, sys, json, hashlib, argparse, subprocess, logging, time, atexit
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
STATE_DIR = os.path.join(BASE, "output", "state")
STATE_PATH = os.path.join(STATE_DIR, "inventory_watch_state.json")
LOCK_PATH = os.path.join(STATE_DIR, "inventory_watch.lock")
LOG_PATH = os.path.join(BASE, "output", "logs", "inventory_watch.log")
TELEGRAM_CONFIG = os.path.join(BASE, "config", "telegram.json")

# Import shared libs
sys.path.insert(0, os.path.join(BASE, "script"))
from lib.sources import INVENTORY_SHEET_URL
from lib.telegram import load_telegram_config, send_telegram_text

# ── Logging ──────────────────────────────────────────────────────────

def setup_logging():
    """Setup logging to both file and console."""
    logger = logging.getLogger("inventory_watch")
    logger.setLevel(logging.INFO)
    # Console handler
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(ch)
    # File handler
    os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
    fh = logging.FileHandler(LOG_PATH, encoding='utf-8')
    fh.setLevel(logging.INFO)
    fh.setFormatter(logging.Formatter('%(asctime)s  %(message)s', datefmt='%Y-%m-%d %H:%M'))
    logger.addHandler(fh)
    return logger

log = setup_logging()

# ══════════════════════════════════════════════════════════════════════
#  Instance locking (prevent duplicate watch processes)
# ══════════════════════════════════════════════════════════════════════

def acquire_lock():
    """Try to acquire lock file. Returns True if successful."""
    os.makedirs(os.path.dirname(LOCK_PATH), exist_ok=True)
    if os.path.exists(LOCK_PATH):
        try:
            with open(LOCK_PATH, "r") as f:
                pid = int(f.read().strip())
            # Check if process is still running
            import ctypes
            kernel32 = ctypes.windll.kernel32
            handle = kernel32.OpenProcess(0x1000, False, pid)  # PROCESS_QUERY_LIMITED_INFORMATION
            if handle:
                kernel32.CloseHandle(handle)
                log.info(f"  ⚠ Another instance running (PID {pid}), exiting.")
                return False
        except (ValueError, OSError):
            pass  # Stale lock file
    with open(LOCK_PATH, "w") as f:
        f.write(str(os.getpid()))
    return True


def release_lock():
    """Release lock file."""
    try:
        if os.path.exists(LOCK_PATH):
            with open(LOCK_PATH, "r") as f:
                pid = int(f.read().strip())
            if pid == os.getpid():
                os.remove(LOCK_PATH)
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════
#  State management
# ══════════════════════════════════════════════════════════════════════

def load_state():
    """Load previous state from JSON file."""
    os.makedirs(STATE_DIR, exist_ok=True)
    if os.path.exists(STATE_PATH):
        with open(STATE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_state(state):
    """Save state to JSON file."""
    os.makedirs(STATE_DIR, exist_ok=True)
    with open(STATE_PATH, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


# ══════════════════════════════════════════════════════════════════════
#  Fetch & diff inventory data
# ══════════════════════════════════════════════════════════════════════

def fetch_inventory():
    """Fetch inventory schedule from Google Sheets.
    Returns dict: store_code → date_str (dd/mm/yyyy).
    """
    import requests
    from io import BytesIO
    from openpyxl import load_workbook

    try:
        r = requests.get(INVENTORY_SHEET_URL, allow_redirects=True, timeout=60)
        r.raise_for_status()
        wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
        ws = wb['Lịch Kiểm kê 2026']

        inventory = {}  # store_code → date_str
        for row in ws.iter_rows(min_row=10, values_only=False):
            store_id = str(row[3].value or "").strip()   # Col D = ID Mart
            kiem_ke = row[7].value                         # Col H = Ngày kiểm kê tổng 2026
            if store_id and kiem_ke:
                dt = None
                if isinstance(kiem_ke, datetime):
                    dt = kiem_ke.date()
                elif hasattr(kiem_ke, 'strftime'):  # date object
                    dt = kiem_ke
                elif isinstance(kiem_ke, str):
                    try:
                        from datetime import date
                        dt = datetime.strptime(kiem_ke, "%d/%m/%Y").date()
                    except ValueError:
                        pass
                if dt:
                    inventory[store_id] = dt.strftime("%d/%m/%Y")
        wb.close()
        log.info(f"  📋 Fetched {len(inventory)} stores with inventory dates")
        return inventory
    except Exception as e:
        log.error(f"  ❌ Could not fetch inventory schedule: {e}")
        return None


def compute_hash(inventory):
    """Hash inventory dict for quick comparison."""
    if not inventory:
        return ""
    normalized = sorted(inventory.items())
    return hashlib.md5(json.dumps(normalized).encode()).hexdigest()


def diff_inventory(old_inv, new_inv):
    """Compare old and new inventory dicts.
    Returns structured diff with added, removed, and changed entries.
    """
    old_inv = old_inv or {}
    new_inv = new_inv or {}

    old_codes = set(old_inv.keys())
    new_codes = set(new_inv.keys())

    added = sorted(new_codes - old_codes)
    removed = sorted(old_codes - new_codes)

    changed = []
    for code in sorted(old_codes & new_codes):
        if old_inv[code] != new_inv[code]:
            changed.append({
                "code": code,
                "old_date": old_inv[code],
                "new_date": new_inv[code],
            })

    return {
        "added": added,
        "removed": removed,
        "changed": changed,
        "has_changes": bool(added or removed or changed),
        "added_details": {c: new_inv[c] for c in added},
        "removed_details": {c: old_inv[c] for c in removed},
    }


def format_diff_text(diff, new_inv):
    """Format diff as human-readable text for logging."""
    lines = []
    if diff["added"]:
        lines.append(f"  🆕 Thêm kiểm kê ({len(diff['added'])}):")
        for code in diff["added"][:10]:
            lines.append(f"     • {code}: {new_inv.get(code, '?')}")
        if len(diff["added"]) > 10:
            lines.append(f"     ... và {len(diff['added']) - 10} stores nữa")
    if diff["removed"]:
        lines.append(f"  🗑️ Xóa kiểm kê ({len(diff['removed'])}):")
        for code in diff["removed"][:10]:
            lines.append(f"     • {code}")
        if len(diff["removed"]) > 10:
            lines.append(f"     ... và {len(diff['removed']) - 10} stores nữa")
    if diff["changed"]:
        lines.append(f"  🔄 Đổi ngày ({len(diff['changed'])}):")
        for ch in diff["changed"][:10]:
            lines.append(f"     • {ch['code']}: {ch['old_date']} → {ch['new_date']}")
        if len(diff["changed"]) > 10:
            lines.append(f"     ... và {len(diff['changed']) - 10} stores nữa")
    return "\n".join(lines) if lines else "  (không có thay đổi)"


def format_telegram_message(diff, new_inv):
    """Format diff as Telegram HTML message."""
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    lines = [f"📋 <b>Lịch Kiểm Kê — Cập nhật {now_str}</b>", ""]

    total_changes = len(diff["added"]) + len(diff["removed"]) + len(diff["changed"])
    lines.append(f"🔄 <b>{total_changes} thay đổi:</b>")

    if diff["changed"]:
        for ch in diff["changed"][:8]:
            lines.append(f"  • {ch['code']}: {ch['old_date']} → {ch['new_date']}")
        if len(diff["changed"]) > 8:
            lines.append(f"  ... +{len(diff['changed']) - 8} stores")

    if diff["added"]:
        lines.append("")
        lines.append(f"🆕 <b>Thêm ({len(diff['added'])}):</b>")
        for code in diff["added"][:5]:
            lines.append(f"  • {code}: {new_inv.get(code, '?')}")
        if len(diff["added"]) > 5:
            lines.append(f"  ... +{len(diff['added']) - 5} stores")

    if diff["removed"]:
        lines.append("")
        lines.append(f"🗑️ <b>Xóa ({len(diff['removed'])}):</b>")
        for code in diff["removed"][:5]:
            lines.append(f"  • {code}")
        if len(diff["removed"]) > 5:
            lines.append(f"  ... +{len(diff['removed']) - 5} stores")

    lines.append("")
    lines.append("✅ Dashboard đã cập nhật")
    lines.append("🔗 https://tunhipham.github.io/transport_daily_report/")

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════
#  Export & Deploy
# ══════════════════════════════════════════════════════════════════════

def run_export():
    """Run export_weekly_plan.py to regenerate weekly_plan.json."""
    script = os.path.join(BASE, "script", "dashboard", "export_weekly_plan.py")
    log.info("  📅 Re-exporting weekly plan...")
    try:
        result = subprocess.run(
            [sys.executable, script],
            capture_output=True, text=True,
            encoding='utf-8', errors='replace',
            timeout=120, cwd=BASE
        )
        if result.returncode == 0:
            log.info("  ✅ Weekly plan re-exported")
            return True
        else:
            log.warning(f"  ⚠ Export returned code {result.returncode}")
            if result.stderr:
                log.warning(f"     {result.stderr[:200]}")
            return False
    except Exception as e:
        log.error(f"  ❌ Export error: {e}")
        return False


def run_deploy():
    """Run deploy.py --domain weekly_plan to push to GitHub Pages."""
    script = os.path.join(BASE, "script", "dashboard", "deploy.py")
    log.info("  🚀 Deploying to GitHub Pages...")
    try:
        result = subprocess.run(
            [sys.executable, script, "--domain", "weekly_plan"],
            capture_output=True, text=True,
            encoding='utf-8', errors='replace',
            timeout=180, cwd=BASE
        )
        if result.returncode == 0:
            log.info("  ✅ Deployed to GitHub Pages")
            return True
        else:
            log.warning(f"  ⚠ Deploy returned code {result.returncode}")
            if result.stdout:
                # Check for "nothing to commit" — that's OK
                if "nothing to commit" in result.stdout or "No changes" in result.stdout:
                    log.info("  ℹ No changes to deploy (data unchanged after export)")
                    return True
            return False
    except Exception as e:
        log.error(f"  ❌ Deploy error: {e}")
        return False


def send_notification(message):
    """Send Telegram notification."""
    bot_token, chat_id = load_telegram_config(TELEGRAM_CONFIG, domain="weekly_plan")
    if not bot_token or not chat_id:
        log.warning("  ⚠ Telegram not configured for weekly_plan domain")
        return False
    mid = send_telegram_text(message, bot_token, chat_id)
    return mid is not None


# ══════════════════════════════════════════════════════════════════════
#  Main check cycle
# ══════════════════════════════════════════════════════════════════════

def run_check(dry_run=False):
    """Run one inventory check cycle.
    Returns: 'changed', 'unchanged', or 'error'.
    """
    now = datetime.now()
    now_str = now.strftime("%d/%m/%Y %H:%M")
    log.info(f"\n{'═'*55}")
    log.info(f"  📋 Inventory Watch — {now_str}")
    log.info(f"{'═'*55}")

    # ── Fetch latest inventory ──
    log.info("\n  🔍 Fetching inventory schedule from Google Sheets...")
    new_inv = fetch_inventory()
    if new_inv is None:
        log.error("  ❌ Fetch failed, skipping this cycle")
        return "error"

    # ── Load previous state ──
    state = load_state()
    prev_inv = state.get("inventory", {})
    prev_hash = state.get("hash", "")

    # ── Compare ──
    new_hash = compute_hash(new_inv)
    log.info(f"  🔑 Hash: {new_hash[:12]}  (prev: {prev_hash[:12] if prev_hash else 'none'})")

    if new_hash == prev_hash and prev_hash:
        log.info("  ✓ Không có thay đổi")
        state["last_check"] = now_str
        state["check_count"] = state.get("check_count", 0) + 1
        save_state(state)
        return "unchanged"

    # ── Changes detected! ──
    diff = diff_inventory(prev_inv, new_inv)

    if not diff["has_changes"] and prev_hash:
        # Hash changed but no meaningful diff (unlikely but possible)
        log.info("  ✓ Hash changed but no meaningful data diff")
        state["hash"] = new_hash
        state["inventory"] = new_inv
        state["last_check"] = now_str
        state["check_count"] = state.get("check_count", 0) + 1
        save_state(state)
        return "unchanged"

    log.info(f"\n  🔔 THAY ĐỔI PHÁT HIỆN!")
    log.info(format_diff_text(diff, new_inv))

    if dry_run:
        log.info("\n  🏃 DRY RUN — skipping export/deploy/notify")
        return "changed"

    # ── Export → Deploy → Notify ──
    export_ok = run_export()
    deploy_ok = False
    if export_ok:
        deploy_ok = run_deploy()

    # Send Telegram notification
    telegram_msg = format_telegram_message(diff, new_inv)
    notify_ok = send_notification(telegram_msg)

    # ── Update state ──
    state["hash"] = new_hash
    state["inventory"] = new_inv
    state["last_check"] = now_str
    state["last_change"] = now_str
    state["check_count"] = state.get("check_count", 0) + 1
    state["change_count"] = state.get("change_count", 0) + 1
    state["last_diff"] = {
        "added": len(diff["added"]),
        "removed": len(diff["removed"]),
        "changed": len(diff["changed"]),
    }
    save_state(state)

    # ── Summary ──
    log.info(f"\n  {'─'*45}")
    log.info(f"  📊 Kết quả:")
    log.info(f"     Export:   {'✅' if export_ok   else '❌'}")
    log.info(f"     Deploy:   {'✅' if deploy_ok   else '❌'}")
    log.info(f"     Telegram: {'✅' if notify_ok   else '❌'}")
    log.info(f"  {'─'*45}")

    return "changed"


# ══════════════════════════════════════════════════════════════════════
#  Watch mode
# ══════════════════════════════════════════════════════════════════════

WATCH_END_HOUR = 17
WATCH_END_MINUTE = 30
POLL_INTERVAL_SEC = 3600  # 1 hour


def watch_mode(dry_run=False):
    """Run in watch mode: loop every 1h until 17:30."""
    if not acquire_lock():
        return
    atexit.register(release_lock)

    log.info(f"\n{'═'*55}")
    log.info(f"  👁️  Inventory Watch — WATCH MODE")
    log.info(f"  📅 Chạy mỗi {POLL_INTERVAL_SEC // 60} phút đến {WATCH_END_HOUR}:{WATCH_END_MINUTE:02d}")
    log.info(f"  🔒 PID: {os.getpid()}")
    log.info(f"{'═'*55}")

    cycle = 0
    while True:
        now = datetime.now()

        # Check end time
        end_time = now.replace(hour=WATCH_END_HOUR, minute=WATCH_END_MINUTE, second=0)
        if now >= end_time:
            log.info(f"\n  ⏰ Đã qua {WATCH_END_HOUR}:{WATCH_END_MINUTE:02d} — stopping watch mode.")
            break

        cycle += 1
        log.info(f"\n  ── Cycle #{cycle} ──")

        try:
            run_check(dry_run=dry_run)
        except Exception as e:
            log.error(f"  ❌ Cycle error: {e}")
            import traceback
            log.error(traceback.format_exc())

        # Calculate sleep time
        now = datetime.now()
        next_run = now + timedelta(seconds=POLL_INTERVAL_SEC)

        # Don't sleep past end time
        if next_run >= end_time:
            remaining = (end_time - now).total_seconds()
            if remaining > 60:
                log.info(f"  💤 Sleeping {int(remaining // 60)} min (last cycle before {WATCH_END_HOUR}:{WATCH_END_MINUTE:02d})...")
                time.sleep(remaining)
            log.info(f"\n  ⏰ Watch mode complete.")
            break
        else:
            log.info(f"  💤 Sleeping {POLL_INTERVAL_SEC // 60} min until {next_run.strftime('%H:%M')}...")
            time.sleep(POLL_INTERVAL_SEC)

    release_lock()


# ══════════════════════════════════════════════════════════════════════
#  Entry point
# ══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Inventory Schedule Watch — Monday auto-update")
    parser.add_argument("--watch", action="store_true",
                        help="Watch mode: poll every 1h until 17:30")
    parser.add_argument("--dry-run", action="store_true",
                        help="Check only, no deploy/notify")
    parser.add_argument("--force", action="store_true",
                        help="Force run (ignore day-of-week check)")
    parser.add_argument("--backup", action="store_true",
                        help="Backup mode: one-shot fetch + update + notify (any day)")
    args = parser.parse_args()

    now = datetime.now()

    # Backup mode: skip all checks, just run
    if args.backup:
        log.info("\n  🔄 BACKUP MODE — manual fetch + update + notify")
        run_check(dry_run=args.dry_run)
        return

    # Day-of-week check (skip unless Monday or --force)
    if now.weekday() != 0 and not args.force:
        day_name = now.strftime("%A")
        log.info(f"  ⏭ Today is {day_name} — inventory watch only runs on Monday.")
        log.info(f"  💡 Use --force or --backup to run anyway.")
        return

    if args.watch:
        watch_mode(dry_run=args.dry_run)
    else:
        run_check(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
