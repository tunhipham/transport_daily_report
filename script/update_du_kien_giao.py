"""
update_du_kien_giao.py — Lấy giờ giao dự kiến từ 5 kho và ghi ra TSV để paste vào sheet "Dự kiến giao".

Usage:
  python script/update_du_kien_giao.py                     # Dùng ngày hôm nay
  python script/update_du_kien_giao.py --date 21/03/2026   # Chỉ định 1 ngày
  python script/update_du_kien_giao.py --date 21/03/2026 --date 22/03/2026   # Nhiều ngày

Sources:
  - KRC:     Google Sheets KRC → col H (giờ đến)
  - KSL:     Google Sheets KFM (sheet DRY) → col H (giờ giao). KSL-Sáng/Tối gộp → KSL
  - THỊT CÁ: KH MEAT (Google Drive) → col "Du kien giao"
  - ĐÔNG MÁT: KH ĐÔNG + KH MÁT (Google Drive) → col "TG DỰ KIẾN" / "Du kien giao"

Output: output/du_kien_giao.tsv (copy vào clipboard luôn nếu chạy trên Windows)
"""
import os, sys, re, csv, argparse
from io import BytesIO
from datetime import datetime, date, time, timedelta
from openpyxl import load_workbook
import requests

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE, "output")

# ── Source URLs ──
KRC_SHEET_URL = "https://docs.google.com/spreadsheets/d/1tWamqjpOI2j2MrYW3Ah6ptmT524CAlQvEP8fCkxfuII/export?format=xlsx"
KFM_SHEET_URL = "https://docs.google.com/spreadsheets/d/1LkJFJhOQ8F2WEB3uCk7kA2Phvu8IskVi3YBfVr7pBx0/export?format=xlsx"

# (name, kho_report, folder_url, default_time_col, default_loai_col)
KH_DRIVE_FOLDERS = [
    ("KH MEAT",      "THỊT CÁ",  "https://drive.google.com/drive/folders/1GIzH8nmCbLhWfpdmxFIn9cHTvQNbnwWr", 22, None),
    ("KH HÀNG ĐÔNG", "ĐÔNG MÁT", "https://drive.google.com/drive/folders/1pQ8coYeV-K0dcHlkvXcJ8KngmH22xp1Z", 17, 18),
    ("KH HÀNG MÁT",  "ĐÔNG MÁT", "https://drive.google.com/drive/folders/1c2zfgcXM8O9ezkOZYj0p4t_ihaJmb98f", 19, 20),
]

# ── Helpers ──

def safe_val(row, idx):
    if idx < len(row):
        return str(row[idx].value or "").strip()
    return ""

def parse_time_str(val):
    """Parse time/timedelta/datetime → 'H:MM' string."""
    if not val:
        return ""
    if isinstance(val, time):
        return f"{val.hour}:{val.minute:02d}"
    if isinstance(val, timedelta):
        total_secs = int(val.total_seconds())
        hours = total_secs // 3600
        mins = (total_secs % 3600) // 60
        return f"{hours}:{mins:02d}"
    if isinstance(val, datetime):
        return f"{val.hour}:{val.minute:02d}"
    s = str(val).strip()
    if s in ("N/A", "n/a", "", "None"):
        return ""
    m = re.match(r'(\d{1,2}):(\d{2})', s)
    if m:
        return f"{int(m.group(1))}:{m.group(2)}"
    return s

def date_to_serial(dd_mm_yyyy):
    parts = dd_mm_yyyy.split("/")
    d = date(int(parts[2]), int(parts[1]), int(parts[0]))
    return (d - date(1899, 12, 30)).days

def date_to_mm_dd_yyyy(dd_mm_yyyy):
    parts = dd_mm_yyyy.split("/")
    return f"{parts[1]}/{parts[0]}/{parts[2]}"

def _list_drive_folder(folder_url):
    import html as _html
    folder_id = re.search(r'folders/([a-zA-Z0-9_-]+)', folder_url).group(1)
    r = requests.get(f"https://drive.google.com/drive/folders/{folder_id}",
                     headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    text = _html.unescape(r.text)
    files, seen = [], set()
    for m in re.finditer(r'"(1[a-zA-Z0-9_-]{25,50})"', text):
        fid = m.group(1)
        if fid in seen:
            continue
        after = text[m.end():m.end() + 500]
        name_match = re.search(r'"([^"]+\.xlsx)"', after)
        if name_match:
            fname = name_match.group(1)
            if "Microsoft" not in fname and "Shared" not in fname:
                seen.add(fid)
                files.append((fid, fname))
    return files

def read_kh_from_drive(folder_url, date_for_file):
    files = _list_drive_folder(folder_url)
    target = next(((fid, fn) for fid, fn in files if date_for_file in fn), None)
    if not target:
        return None
    print(f"    ↳ {target[1]}")
    r = requests.get(f"https://drive.google.com/uc?export=download&confirm=t&id={target[0]}",
                     allow_redirects=True, timeout=60)
    if r.status_code != 200 or len(r.content) < 500:
        return None
    return load_workbook(BytesIO(r.content), read_only=True, data_only=True)


# ── Main ──

def collect_data(target_dates):
    """Return list of row dicts for the given dates."""
    all_rows = []

    # 1. KRC
    print("  → KRC (Google Sheets)...")
    try:
        wb = load_workbook(BytesIO(requests.get(KRC_SHEET_URL, allow_redirects=True, timeout=120).content),
                           read_only=True, data_only=True)
        ws = wb["KRC"]
        cnt = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            dv = str(row[0].value or "").strip()
            if dv in target_dates:
                dd = safe_val(row, 6)
                gt = parse_time_str(row[7].value)
                if dd:
                    s = date_to_serial(dv)
                    all_rows.append(dict(key=f"{s}{dd}KRC", date=date_to_mm_dd_yyyy(dv),
                                         diem_den=dd, gio_den=gt, kho="KRC", loai_hang=""))
                    cnt += 1
        wb.close()
        print(f"    {cnt} rows")
    except Exception as e:
        print(f"    ⚠ ERROR: {e}")

    # 2. KFM → KSL
    print("  → KFM/KSL (Google Sheets)...")
    try:
        wb = load_workbook(BytesIO(requests.get(KFM_SHEET_URL, allow_redirects=True, timeout=120).content),
                           read_only=True, data_only=True)
        ws = next((wb[n] for n in wb.sheetnames if "DRY" in n), wb.worksheets[1])
        cnt = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            dv = str(row[0].value or "").strip()
            if dv in target_dates:
                dd = safe_val(row, 6)
                gt = parse_time_str(row[7].value)
                if dd:
                    s = date_to_serial(dv)
                    all_rows.append(dict(key=f"{s}{dd}KSL", date=date_to_mm_dd_yyyy(dv),
                                         diem_den=dd, gio_den=gt, kho="KSL", loai_hang=""))
                    cnt += 1
        wb.close()
        print(f"    {cnt} rows")
    except Exception as e:
        print(f"    ⚠ ERROR: {e}")

    # 3. KH files
    for folder_name, kho_name, folder_url, def_time_col, def_loai_col in KH_DRIVE_FOLDERS:
        for td in target_dates:
            parts = td.split("/")
            date_for_file = f"{parts[0]}.{parts[1]}.{parts[2]}"
            print(f"  → {folder_name} ({td})...")
            try:
                wb = read_kh_from_drive(folder_url, date_for_file)
                if not wb:
                    print(f"    ⚠ File not found")
                    continue
                ws = wb.worksheets[0]

                # Auto-detect time column by header name
                tcol, lcol = def_time_col, def_loai_col
                for hrow in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                    for i, cell in enumerate(hrow):
                        h = str(cell.value or "").strip().upper()
                        if h in ("DU KIEN GIAO", "DỰ KIẾN GIAO", "TG DỰ KIẾN"):
                            tcol = i
                        if h in ("LOẠI HÀNG", "LOAI HANG"):
                            lcol = i

                serial = date_to_serial(td)
                cnt = 0
                for row in ws.iter_rows(min_row=2, values_only=False):
                    sid = safe_val(row, 2)
                    gt = parse_time_str(row[tcol].value if len(row) > tcol else None)
                    lh = safe_val(row, lcol) if lcol and len(row) > lcol else ""
                    if sid:
                        all_rows.append(dict(key=f"{serial}{sid}{kho_name}", date=date_to_mm_dd_yyyy(td),
                                             diem_den=sid, gio_den=gt, kho=kho_name, loai_hang=lh))
                        cnt += 1
                wb.close()
                print(f"    {cnt} rows")
            except Exception as e:
                print(f"    ⚠ ERROR: {e}")

    return all_rows


def main():
    parser = argparse.ArgumentParser(description="Lấy giờ giao dự kiến → TSV")
    parser.add_argument("--date", action="append",
                        help="Ngày DD/MM/YYYY (có thể dùng nhiều lần). Mặc định: ngày hôm nay")
    args = parser.parse_args()

    if args.date:
        target_dates = args.date
    else:
        target_dates = [datetime.now().strftime("%d/%m/%Y")]

    print("=" * 50)
    print(f"  DỰ KIẾN GIAO — {', '.join(target_dates)}")
    print("=" * 50)

    all_rows = collect_data(target_dates)

    # Summary
    print(f"\n{'─'*50}")
    print(f"TỔNG: {len(all_rows)} rows")
    for kho in ["KRC", "KSL", "THỊT CÁ", "ĐÔNG MÁT"]:
        kr = [r for r in all_rows if r["kho"] == kho]
        print(f"  {kho}: {len(kr)} rows")
    print()

    # Samples
    for kho in ["KRC", "KSL", "THỊT CÁ", "ĐÔNG MÁT"]:
        kr = [r for r in all_rows if r["kho"] == kho]
        if kr:
            print(f"  [{kho}] sample: {kr[0]['diem_den']} → {kr[0]['gio_den']}")

    if not all_rows:
        print("\n⚠ Không có dữ liệu. Kiểm tra lại ngày hoặc sources.")
        return

    # Write TSV
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    tsv_path = os.path.join(OUTPUT_DIR, "du_kien_giao.tsv")
    with open(tsv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        for r in all_rows:
            writer.writerow([r["key"], r["date"], r["diem_den"], r["gio_den"], r["kho"], r["loai_hang"]])

    print(f"\n✅ Saved: {tsv_path}")

    # Auto-copy to clipboard on Windows
    if sys.platform == "win32":
        import subprocess
        try:
            content = open(tsv_path, "r", encoding="utf-8").read()
            proc = subprocess.Popen(["powershell", "-Command",
                                     f'[System.IO.File]::ReadAllText("{tsv_path}", [System.Text.Encoding]::UTF8) | Set-Clipboard'],
                                    stdin=subprocess.PIPE)
            proc.wait()
            print("📋 Đã copy vào clipboard!")
        except Exception:
            print("⚠ Không copy được clipboard. Hãy copy thủ công.")

    print(f"\n👉 Mở sheet 'Dự kiến giao', tìm row trống cuối cùng, paste (Ctrl+V).")


if __name__ == "__main__":
    main()
