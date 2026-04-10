"""
generate_performance_report.py - Transport Performance Report
On-time SLA / On-time Plan / Route Compliance / Trip Completion

Usage:
  python script/generate_performance_report.py --month 03 --year 2026
  python script/generate_performance_report.py  # defaults T03/2026

Output: output/PERFORMANCE_REPORT_T03_2026.html
"""
import os, sys, re, json, warnings
from datetime import datetime, timedelta, time as dtime
from collections import defaultdict, Counter
from calendar import monthrange

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from io import BytesIO
import base64

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
warnings.filterwarnings('ignore')

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT = os.path.join(BASE, "output")

# ── SLA Windows ──
SLA_WINDOWS = {
    "KRC":       (dtime(3, 0), dtime(5, 30)),
    "THỊT CÁ":  (dtime(3, 0), dtime(6, 0)),
    "ĐÔNG MÁT": (dtime(9, 0), dtime(16, 0)),
    "ĐÔNG": (dtime(9, 0), dtime(16, 0)),
    "MÁT": (dtime(9, 0), dtime(16, 0)),
    "KSL-Sáng": (dtime(12, 0), dtime(14, 0)),
    "KSL-Tối":  (dtime(22, 0), dtime(0, 30)),  # overnight
}

KHO_COLORS = {
    "KRC": "#6CA6FF",
    "THỊT CÁ": "#FF9F5A",
    "ĐÔNG MÁT": "#7DD87D",
    "ĐÔNG": "#4CAF50",
    "MÁT": "#81D4FA",
    "KSL-Sáng": "#FFD966",
    "KSL-Tối": "#C49BFF",
}

# Dark theme constants
DARK_BG = "#1E1E2E"
DARK_SURFACE = "#2A2A3C"
DARK_TEXT = "#E0E0E0"
DARK_GRID = "#3A3A4C"

NOI_CHUYEN_MAP = {
    "KSL": "DRY",
    "KRC": "KRC",
    "QCABA": "ĐÔNG MÁT",
}


# ══════════════════════════════════════════════════════════════
# Data Loading
# ══════════════════════════════════════════════════════════════

def parse_arrival_time(s):
    """Parse 'DD/MM/YYYY HH:MM' or time-only to datetime.time."""
    if not s:
        return None
    s = str(s).strip()
    # Full datetime: "02/03/2026 10:02"
    m = re.match(r'\d{2}/\d{2}/\d{4}\s+(\d{1,2}):(\d{2})', s)
    if m:
        return dtime(int(m.group(1)), int(m.group(2)))
    # Time only: "10:02"
    m2 = re.match(r'(\d{1,2}):(\d{2})', s)
    if m2:
        return dtime(int(m2.group(1)), int(m2.group(2)))
    # datetime object
    if hasattr(s, 'hour'):
        return dtime(s.hour, s.minute)
    return None


def parse_arrival_datetime(s):
    """Parse 'DD/MM/YYYY HH:MM' to full datetime."""
    if not s:
        return None
    s = str(s).strip()
    m = re.match(r'(\d{2})/(\d{2})/(\d{4})\s+(\d{1,2}):(\d{2})', s)
    if m:
        return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)),
                        int(m.group(4)), int(m.group(5)))
    return None


def parse_date_str(s):
    """Parse date string to date object. Various formats supported."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ["%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]:
        try:
            return datetime.strptime(s.split(" ")[0] if " " in s and fmt.startswith("%d") else s, fmt).date()
        except ValueError:
            continue
    return None


def get_kho_session(noi_chuyen, arrival_time, depart_time=None):
    """Map warehouse + arrival/depart time to kho name (with KSL session split).
    Never returns 'DRY' — always resolves to KSL-Sáng or KSL-Tối."""
    kho = NOI_CHUYEN_MAP.get(noi_chuyen, noi_chuyen)
    if kho == "DRY":
        # Determine session from arrival_time first, then depart_time
        ref_time = arrival_time or depart_time
        if ref_time:
            if ref_time.hour < 15:
                return "KSL-Sáng"
            else:
                return "KSL-Tối"
        # Final fallback: default to KSL-Sáng
        return "KSL-Sáng"
    return kho


def load_trip_data(month, year):
    """Load and dedup trip data from DS chi tiết chuyến xe."""
    import openpyxl
    data_dir = rf'G:\My Drive\DOCS\DAILY\DS chi tiet chuyen xe\T{month:02d}.{year % 100}'
    
    if not os.path.exists(data_dir):
        print(f"  ⚠ Directory not found: {data_dir}")
        return []
    
    files = sorted([f for f in os.listdir(data_dir) if f.endswith('.xlsx') and not f.startswith('~')])
    print(f"  📁 Found {len(files)} files in {data_dir}")
    
    # Collect all rows, dedup by (trip_id, dest, sub_kho)
    # ĐÔNG and MÁT for the same store are SEPARATE deliveries with
    # different planned times (from KH HÀNG ĐÔNG vs KH HÀNG MÁT)
    seen = {}  # (trip_id, dest, sub_kho) -> row dict
    total_raw = 0
    
    for fname in files:
        fpath = os.path.join(data_dir, fname)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb['Sheet 1']
            for row in ws.iter_rows(min_row=2, values_only=True):
                total_raw += 1
                trip_id = str(row[0] or "").strip()
                trip_status = str(row[1] or "").strip()
                driver = str(row[3] or "").strip()
                depart_date = str(row[5] or "").strip()
                depart_time_raw = str(row[6] or "").strip() if row[6] else ""
                noi_chuyen = str(row[8] or "").strip()
                dest = str(row[9] or "").strip()
                dest_status = str(row[11] or "").strip()
                container_type = str(row[18] or "").strip() if len(row) > 18 else ""
                arrival_raw = str(row[26] or "").strip() if row[26] else ""
                
                if not trip_id or not dest:
                    continue
                
                arrival_time = parse_arrival_time(arrival_raw)
                arrival_dt = parse_arrival_datetime(arrival_raw)
                depart_time = parse_arrival_time(depart_time_raw)
                
                # Parse departure date
                dep_date = parse_date_str(depart_date)
                # If depart_date empty, try from arrival
                if not dep_date and arrival_dt:
                    dep_date = arrival_dt.date()
                
                kho = get_kho_session(noi_chuyen, arrival_time, depart_time)
                
                # Sub-classification for ĐÔNG MÁT
                sub_kho = ""
                if kho == "ĐÔNG MÁT":
                    if "tote" in container_type.lower():
                        sub_kho = "ĐÔNG"
                    else:
                        sub_kho = "MÁT"  # Rổ ABA, Thùng Carton, Bịch nguyên
                
                # Dedup key: include sub_kho so ĐÔNG and MÁT for same
                # (trip, store) are kept as separate rows
                key = (trip_id, dest, sub_kho)
                if key in seen:
                    continue  # Dedup: keep first occurrence
                
                seen[key] = {
                    "trip_id": trip_id,
                    "dest": dest,
                    "kho": kho,
                    "sub_kho": sub_kho,
                    "noi_chuyen": noi_chuyen,
                    "driver": driver,
                    "date": dep_date,
                    "trip_status": trip_status,
                    "dest_status": dest_status,
                    "arrival_time": arrival_time,
                    "arrival_dt": arrival_dt,
                    "arrival_raw": arrival_raw,
                    "container_type": container_type,
                }
            wb.close()
        except Exception as e:
            print(f"    ⚠ Error reading {fname}: {e}")
    
    rows = list(seen.values())
    print(f"  📊 {total_raw} raw rows → {len(rows)} deduped (trip,dest,sub_kho) pairs")
    
    # Filter: remove rows without arrival time
    with_time = [r for r in rows if r["arrival_time"] is not None]
    no_time = len(rows) - len(with_time)
    print(f"  🕐 {len(with_time)} with arrival time, {no_time} without (excluded from on-time)")
    
    return rows


def load_thitca_data(months):
    """Load THỊT CÁ actual data from external file(s)."""
    all_rows = []
    for m in months:
        plan_path = os.path.join(OUTPUT, f"monthly_plan_T{m:02d}.json")
        if not os.path.exists(plan_path):
            continue
        with open(plan_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        tc_actual = data.get("thitca_actual", [])
        rows = []
        for r in tc_actual:
            arrival = parse_arrival_time(r.get("actual_time", ""))
            planned = parse_arrival_time(r.get("planned_time", ""))
            date = parse_date_str(r["date"])
            if not date or not arrival:
                continue
            rows.append({
                "trip_id": f"TC_{r['date']}_{r['store']}_{r['tuyen']}",
                "dest": r["store"],
                "kho": "THỊT CÁ",
                "noi_chuyen": "THỊT CÁ",
                "driver": "",
                "date": date,
                "trip_status": "Hoàn thành",
                "dest_status": "Hoàn thành",
                "arrival_time": arrival,
                "arrival_dt": None,
                "arrival_raw": r.get("actual_time", ""),
                "planned_time": planned,
                "tuyen": r.get("tuyen", ""),
            })
        
        # Dedup by (date, store, tuyen)
        seen = set()
        deduped = []
        for r in rows:
            key = (r["date"], r["dest"], r["tuyen"])
            if key not in seen:
                seen.add(key)
                deduped.append(r)
        
        if len(deduped) < len(rows):
            print(f"  🐟 THỊT CÁ (T{m:02d}): {len(rows)} rows → {len(deduped)} after dedup")
        else:
            print(f"  🐟 THỊT CÁ (T{m:02d}): {len(rows)} rows")
        all_rows.extend(deduped)
    
    print(f"  🐟 THỊT CÁ total: {len(all_rows)} rows")
    return all_rows


def load_plan_data(months):
    """Load monthly plan (planned time + tuyến) from one or more months."""
    plan_lookup = {}
    route_order = defaultdict(list)
    
    for m in months:
        plan_path = os.path.join(OUTPUT, f"monthly_plan_T{m:02d}.json")
        if not os.path.exists(plan_path):
            continue
        with open(plan_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    
        # Forward-fill tuyến: in Google Sheets, tuyến is only on the first row
        # of each route group. Subsequent stops have it blank.
        for kho, rows in data["plan"].items():
            current_tuyen = ""
            current_date = ""
            for r in rows:
                d = r.get("date", "")
                t = r.get("tuyen", "")
                if t:
                    current_tuyen = t
                    current_date = d
                elif d == current_date and current_tuyen:
                    r["tuyen"] = current_tuyen
                else:
                    current_tuyen = ""
                    current_date = d
    
        # Build lookup: (date_str, store, kho) -> planned_time, tuyen
        # For ĐÔNG MÁT: also build (date, store, "ĐÔNG") and (date, store, "MÁT")
        # based on tuyến prefix: HĐ = HÀNG ĐÔNG, KF = HÀNG MÁT
        for kho, rows in data["plan"].items():
            for r in rows:
                date = parse_date_str(r["date"])
                store = r["store"]
                planned = parse_arrival_time(r.get("planned_time", ""))
                tuyen = r.get("tuyen", "")
                if date and store:
                    key = (date, store, kho)
                    if key not in plan_lookup:
                        plan_lookup[key] = {
                            "planned_time": planned,
                            "tuyen": tuyen,
                        }
                    # For ĐÔNG MÁT: add sub_kho specific keys
                    if kho == "ĐÔNG MÁT" and tuyen:
                        if tuyen.startswith("HĐ"):
                            sub_key = (date, store, "ĐÔNG")
                        else:
                            sub_key = (date, store, "MÁT")
                        if sub_key not in plan_lookup:
                            plan_lookup[sub_key] = {
                                "planned_time": planned,
                                "tuyen": tuyen,
                            }
    
        # Build route order: (date, tuyen, kho) -> ordered list of stores
        for kho, rows in data["plan"].items():
            by_route = defaultdict(list)
            for r in rows:
                date = parse_date_str(r["date"])
                tuyen = r.get("tuyen", "")
                store = r["store"]
                planned = parse_arrival_time(r.get("planned_time", ""))
                if date and tuyen and store:
                    by_route[(date, tuyen, kho)].append((planned, store))
            
            for key, store_list in by_route.items():
                if key not in route_order:
                    store_list.sort(key=lambda x: (x[0] or dtime(23, 59)))
                    route_order[key] = [s[1] for s in store_list]

        # Also add THỊT CÁ from thitca_actual
        tc_actual = data.get("thitca_actual", [])
        tc_route_groups = defaultdict(list)
        for r in tc_actual:
            date = parse_date_str(r["date"])
            store = r["store"]
            planned = parse_arrival_time(r.get("planned_time", ""))
            tuyen = r.get("tuyen", "")
            if date and store:
                key = (date, store, "THỊT CÁ")
                if key not in plan_lookup:
                    plan_lookup[key] = {"planned_time": planned, "tuyen": tuyen}
                if tuyen:
                    tc_route_groups[(date, tuyen, "THỊT CÁ")].append((planned, store))
        
        for key, store_list in tc_route_groups.items():
            if key not in route_order:
                store_list.sort(key=lambda x: (x[0] or dtime(23, 59)))
                route_order[key] = [s[1] for s in store_list]
    
    print(f"  📋 Plan lookup: {len(plan_lookup)} entries")
    print(f"  🗺 Route orders: {len(route_order)} routes")
    # Verify route sizes
    sizes = [len(v) for v in route_order.values()]
    if sizes:
        print(f"  📏 Route sizes: avg={sum(sizes)/len(sizes):.1f}, min={min(sizes)}, max={max(sizes)}")
    return plan_lookup, route_order


# ══════════════════════════════════════════════════════════════
# Metric Calculation
# ══════════════════════════════════════════════════════════════

def check_sla(kho, arrival_time):
    """Check if arrival is within SLA. Returns (is_on_time, sla_window)."""
    if kho not in SLA_WINDOWS:
        return None, None
    
    sla_start, sla_end = SLA_WINDOWS[kho]
    
    # KSL-Tối: overnight window 22:00 - 00:30
    # Early delivery (before 22:00) is still on-time
    # Late = past 00:30 the next day
    if kho == "KSL-Tối":
        # Already filtered to >= 15:00 arrival (KSL-Tối session)
        # On-time if: arrival < 00:31 (wraps overnight)
        # Late only if 00:31 <= arrival < 15:00 (very rare edge case)
        if arrival_time.hour >= 15 or (arrival_time.hour == 0 and arrival_time.minute <= 30):
            return True, (sla_start, sla_end)
        return False, (sla_start, sla_end)
    
    # Normal: on-time if arrival <= sla_end (early is also on-time)
    if arrival_time <= sla_end:
        return True, (sla_start, sla_end)
    return False, (sla_start, sla_end)


def check_plan_ontime(arrival_time, planned_time):
    """Check if arrival <= planned time. Returns True if on-time."""
    if not arrival_time or not planned_time:
        return None
    return arrival_time <= planned_time


def check_route_compliance(trip_stores_actual, route_planned_order):
    """
    Compare actual delivery order vs planned order.
    Returns (num_correct, num_wrong) counted per destination.
    If route order matches: all destinations = correct.
    If route order differs: all destinations = wrong.
    """
    # Dedup: keep only first arrival at each store
    seen = set()
    actual_order = []
    for _, s in trip_stores_actual:
        if s not in seen:
            seen.add(s)
            actual_order.append(s)
    
    # Only compare stores that exist in both
    planned_set = set(route_planned_order)
    actual_filtered = [s for s in actual_order if s in planned_set]
    actual_set = set(actual_order)
    planned_filtered = [s for s in route_planned_order if s in actual_set]
    
    num_common = len(actual_filtered)
    if num_common == 0:
        return None  # No common stores to compare
    if num_common == 1:
        return (1, 0)  # 1-store route = correct by default
    
    if actual_filtered == planned_filtered:
        return (num_common, 0)  # All destinations correct
    else:
        return (0, num_common)  # All destinations wrong


def calc_metrics(all_rows, plan_lookup, route_order):
    """Calculate all metrics from trip data."""
    
    metrics = {
        "sla": defaultdict(lambda: defaultdict(lambda: {"on_time": 0, "late": 0})),
        "plan": defaultdict(lambda: defaultdict(lambda: {"on_time": 0, "late": 0})),
        "route": defaultdict(lambda: defaultdict(lambda: {"correct": 0, "wrong": 0})),
        "completion_trip": defaultdict(lambda: defaultdict(lambda: {"complete": 0, "total": 0})),
        "completion_dest": defaultdict(lambda: defaultdict(lambda: {"complete": 0, "total": 0})),
        "trips_per_day": defaultdict(lambda: defaultdict(set)),  # kho -> date -> set of trip_ids/tuyến
        "completion_by_driver": defaultdict(lambda: defaultdict(lambda: {"complete": 0, "total": 0})),
    }
    
    # Track trips for route analysis
    trip_dests = defaultdict(list)  # trip_id -> [(arrival_time, dest, kho)]
    
    for r in all_rows:
        date = r["date"]
        kho = r["kho"]
        arrival = r.get("arrival_time")
        sub_kho = r.get("sub_kho", "")
        # Record metrics under both kho AND sub_kho (ĐÔNG/MÁT) if applicable
        kho_keys = [kho]
        if sub_kho:
            kho_keys.append(sub_kho)
        
        if not date:
            continue
        
        # ── Completion ──
        trip_id = r["trip_id"]
        
        # Destination level
        for k in kho_keys:
            metrics["completion_dest"][k][date]["total"] += 1
            if r["dest_status"] == "Hoàn thành":
                metrics["completion_dest"][k][date]["complete"] += 1
        
        # Trip level (dedup per trip_id per kho per date)
        # We'll aggregate after the loop
        
        # ── Completion by driver ──
        driver = r.get("driver", "")
        if driver:
            for k in kho_keys:
                metrics["completion_by_driver"][k][(date, driver)]["total"] += 1
                if r["dest_status"] == "Hoàn thành":
                    metrics["completion_by_driver"][k][(date, driver)]["complete"] += 1
        
        if not arrival:
            continue
        
        # ── SLA ──
        sla_result, _ = check_sla(kho, arrival)
        if sla_result is not None:
            for k in kho_keys:
                if sla_result:
                    metrics["sla"][k][date]["on_time"] += 1
                else:
                    metrics["sla"][k][date]["late"] += 1
        
        # ── Trips per day ──
        if kho == "THỊT CÁ":
            trip_key = r.get("tuyen", "")
        else:
            trip_key = r["trip_id"]
        if trip_key:
            for k in kho_keys:
                metrics["trips_per_day"][k][date].add(trip_key)
        
        # ── On-time vs Plan ──
        planned = r.get("planned_time")
        if not planned:
            # Look up from plan data
            # Map kho back for lookup
            lookup_kho = kho
            if kho in ("KSL-Sáng", "KSL-Tối"):
                lookup_kho = "DRY"
            # For ĐÔNG MÁT with sub_kho, try sub_kho-specific plan first
            # (KH HÀNG ĐÔNG vs KH HÀNG MÁT have different planned times)
            plan_info = None
            if sub_kho:
                plan_info = plan_lookup.get((date, r["dest"], sub_kho))
            if not plan_info:
                plan_info = plan_lookup.get((date, r["dest"], lookup_kho))
            if plan_info:
                planned = plan_info["planned_time"]
        
        if planned:
            is_ontime = check_plan_ontime(arrival, planned)
            if is_ontime is not None:
                for k in kho_keys:
                    if is_ontime:
                        metrics["plan"][k][date]["on_time"] += 1
                    else:
                        metrics["plan"][k][date]["late"] += 1
        
        # ── Collect for route analysis ──
        tuyen = r.get("tuyen", "")
        if not tuyen:
            lookup_kho = kho
            if kho in ("KSL-Sáng", "KSL-Tối"):
                lookup_kho = "DRY"
            plan_info = plan_lookup.get((date, r["dest"], lookup_kho))
            if plan_info:
                tuyen = plan_info.get("tuyen", "")
        
        if tuyen and arrival:
            trip_dests[(date, tuyen, kho)].append((arrival, r["dest"]))
    
    # ── Trip-level completion (dedup by trip_id) ──
    trip_status_map = {}  # (trip_id, kho) -> (date, status, sub_kho)
    for r in all_rows:
        if not r["date"]:
            continue
        key = (r["trip_id"], r["kho"])
        if key not in trip_status_map:
            trip_status_map[key] = (r["date"], r["trip_status"], r.get("sub_kho", ""))
    
    for (trip_id, kho), (date, status, sub_kho) in trip_status_map.items():
        kkeys = [kho] + ([sub_kho] if sub_kho else [])
        for k in kkeys:
            metrics["completion_trip"][k][date]["total"] += 1
            if status == "Hoàn thành":
                metrics["completion_trip"][k][date]["complete"] += 1
    
    # ── Route compliance ──
    # Build sub_kho lookup for ĐÔNG MÁT destinations
    dest_sub_kho = {}  # (date, dest) -> sub_kho
    for r in all_rows:
        sk = r.get("sub_kho", "")
        if sk and r.get("date"):
            dest_sub_kho[(r["date"], r["dest"])] = sk
    
    for (date, tuyen, kho), actual_stops in trip_dests.items():
        actual_stops.sort(key=lambda x: x[0])
        
        # Find planned order
        lookup_kho = kho
        if kho in ("KSL-Sáng", "KSL-Tối"):
            lookup_kho = "DRY"
        planned_order = route_order.get((date, tuyen, lookup_kho))
        if not planned_order:
            planned_order = route_order.get((date, tuyen, kho))
        if not planned_order:
            continue
        
        result = check_route_compliance(actual_stops, planned_order)
        if result is not None:
            num_correct, num_wrong = result
            metrics["route"][kho][date]["correct"] += num_correct
            metrics["route"][kho][date]["wrong"] += num_wrong
            # Also record for sub_kho
            if kho == "ĐÔNG MÁT":
                for _, dest in actual_stops:
                    sk = dest_sub_kho.get((date, dest), "")
                    if sk:
                        if num_correct > 0:
                            metrics["route"][sk][date]["correct"] += 1
                        else:
                            metrics["route"][sk][date]["wrong"] += 1
    
    return metrics


# ══════════════════════════════════════════════════════════════
# Chart.js Data Preparation
# ══════════════════════════════════════════════════════════════

def prepare_chart_data(metrics, dates, month, year):
    """Prepare all chart data as JSON for Chart.js."""
    labels = [d.strftime('%d/%m') for d in dates]
    khos = list(KHO_COLORS.keys())
    
    charts = {}
    
    # ── Helper: compute average % across khos ──
    def avg_pct(metric_key, val_a="on_time", val_b="late"):
        avgs = []
        for d in dates:
            total_a = sum(metrics[metric_key].get(k, {}).get(d, {}).get(val_a, 0) for k in khos)
            total_b = sum(metrics[metric_key].get(k, {}).get(d, {}).get(val_b, 0) for k in khos)
            total = total_a + total_b
            avgs.append(round(total_a / total * 100, 1) if total > 0 else None)
        return avgs
    
    # ── Chart 1 & 2: On-time (SLA & Plan) ──
    for metric_key, chart_key in [("sla", "sla"), ("plan", "plan")]:
        on_times = []
        lates = []
        for d in dates:
            ot = sum(metrics[metric_key].get(k, {}).get(d, {}).get("on_time", 0) for k in khos)
            lt = sum(metrics[metric_key].get(k, {}).get(d, {}).get("late", 0) for k in khos)
            on_times.append(ot)
            lates.append(lt)
        
        # Per-kho on/late for filtered bars
        kho_bars = {}
        for kho in khos:
            if kho not in metrics[metric_key]:
                continue
            k_ot = []
            k_lt = []
            for d in dates:
                data = metrics[metric_key].get(kho, {}).get(d, {})
                k_ot.append(data.get("on_time", 0))
                k_lt.append(data.get("late", 0))
            kho_bars[kho] = {"on": k_ot, "late": k_lt}
        
        trend_lines = []
        for kho in khos:
            if kho not in metrics[metric_key]:
                continue
            pcts = []
            for d in dates:
                data = metrics[metric_key].get(kho, {}).get(d, {})
                ot = data.get("on_time", 0)
                lt = data.get("late", 0)
                total = ot + lt
                pcts.append(round(ot / total * 100, 1) if total > 0 else None)
            trend_lines.append({"label": kho, "data": pcts, "color": KHO_COLORS[kho]})
        
        charts[chart_key] = {
            "on_times": on_times,
            "lates": lates,
            "kho_bars": kho_bars,
            "trends": trend_lines,
            "avg": avg_pct(metric_key),
        }
    
    # ── Chart 3: Route compliance ──
    corrects = []
    wrongs = []
    for d in dates:
        c = sum(metrics["route"].get(k, {}).get(d, {}).get("correct", 0) for k in khos)
        w = sum(metrics["route"].get(k, {}).get(d, {}).get("wrong", 0) for k in khos)
        corrects.append(c)
        wrongs.append(w)
    
    kho_route_bars = {}
    for kho in khos:
        if kho not in metrics["route"]:
            continue
        k_c = []
        k_w = []
        for d in dates:
            data = metrics["route"].get(kho, {}).get(d, {})
            k_c.append(data.get("correct", 0))
            k_w.append(data.get("wrong", 0))
        kho_route_bars[kho] = {"on": k_c, "late": k_w}
    
    route_trends = []
    for kho in khos:
        if kho not in metrics["route"]:
            continue
        pcts = []
        for d in dates:
            data = metrics["route"].get(kho, {}).get(d, {})
            c = data.get("correct", 0)
            w = data.get("wrong", 0)
            total = c + w
            pcts.append(round(c / total * 100, 1) if total > 0 else None)
        route_trends.append({"label": kho, "data": pcts, "color": KHO_COLORS[kho]})
    
    charts["route"] = {
        "corrects": corrects, "wrongs": wrongs,
        "kho_bars": kho_route_bars,
        "trends": route_trends,
        "avg": avg_pct("route", "correct", "wrong"),
    }
    
    # ── Chart 4: Trips + Completion (stacked bar + trend) ──
    # Bars: completed trips (green) + incomplete trips (red)
    # Trend: % completion rate
    comp_total = []
    comp_complete = []
    for d in dates:
        t = sum(metrics["completion_trip"].get(k, {}).get(d, {}).get("total", 0) for k in khos)
        c = sum(metrics["completion_trip"].get(k, {}).get(d, {}).get("complete", 0) for k in khos)
        comp_total.append(t)
        comp_complete.append(c)
    comp_incomplete = [t - c for t, c in zip(comp_total, comp_complete)]
    
    # Per-kho bars
    kho_comp_bars = {}
    comp_trends = []
    for kho in khos:
        if kho not in metrics["completion_trip"]:
            continue
        k_comp = []
        k_incomp = []
        k_pcts = []
        for d in dates:
            data = metrics["completion_trip"].get(kho, {}).get(d, {})
            t = data.get("total", 0)
            c = data.get("complete", 0)
            k_comp.append(c)
            k_incomp.append(t - c)
            k_pcts.append(round(c / t * 100, 1) if t > 0 else None)
        kho_comp_bars[kho] = {"on": k_comp, "late": k_incomp}
        comp_trends.append({"label": kho, "data": k_pcts, "color": KHO_COLORS[kho]})
    
    # Avg completion %
    comp_avg = []
    for d in dates:
        t = sum(metrics["completion_trip"].get(k, {}).get(d, {}).get("total", 0) for k in khos)
        c = sum(metrics["completion_trip"].get(k, {}).get(d, {}).get("complete", 0) for k in khos)
        comp_avg.append(round(c / t * 100, 1) if t > 0 else None)
    
    charts["trips_completion"] = {
        "on_times": comp_complete,
        "lates": comp_incomplete,
        "kho_bars": kho_comp_bars,
        "trends": comp_trends,
        "avg": comp_avg,
    }
    
    return labels, charts


# ══════════════════════════════════════════════════════════════
# HTML Report
# ══════════════════════════════════════════════════════════════

def generate_weekly_tables(metrics, dates):
    """Generate weekly summary tables for THỊT CÁ, ĐÔNG MÁT, ĐÔNG, MÁT."""
    from datetime import timedelta
    
    DAY_NAMES = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'CN']
    
    if not dates:
        return ""
    
    # Group dates by ISO week (Mon-Sun)
    weeks = {}
    for d in dates:
        iso_year, iso_week, _ = d.isocalendar()
        wk = f"W{iso_week}"
        if wk not in weeks:
            monday = d - timedelta(days=d.weekday())
            weeks[wk] = {"monday": monday, "dates": []}
        weeks[wk]["dates"].append(d)
    
    sorted_weeks = [(k, v) for k, v in sorted(weeks.items(), key=lambda x: x[1]["monday"]) if k not in ("W9", "W10")]
    
    def build_header_and_keys():
        header1 = '<tr><th rowspan="2" class="wt-kho">KHO</th><th rowspan="2" class="wt-metric">Chỉ Tiêu</th>'
        header2 = '<tr>'
        col_keys = []
        for wk_key, wk_data in sorted_weeks:
            monday = wk_data["monday"]
            header1 += f'<th colspan="1" class="wt-wk-total">{wk_key}</th>'
            header2 += f'<td class="wt-wk-label">Tổng</td>'
            col_keys.append((wk_key, None))
            header1 += f'<th colspan="7" class="wt-wk-header">{wk_key} ({monday.strftime("%d/%m")} - {(monday + timedelta(days=6)).strftime("%d/%m/%Y")})</th>'
            for i in range(7):
                day_date = monday + timedelta(days=i)
                header2 += f'<td class="wt-day">{DAY_NAMES[i]}<br>{day_date.strftime("%d/%m")}</td>'
                col_keys.append((wk_key, day_date))
        header1 += '</tr>'
        header2 += '</tr>'
        return header1, header2, col_keys
    
    def compute_values(kho, col_keys, use_sla_ontime=False):
        """Compute metric values per column.
        use_sla_ontime: if True, % On Time = SLA-based; if False = plan-based (arrival vs planned_time)
        """
        data = {
            "Tổng Điểm Giao": [],
            "Đúng & Sớm Kế Hoạch": [],
            "% On Time": [],
            "Tổng Số Chuyến": [],
        }
        
        for wk_key, day_date in col_keys:
            if day_date is None:
                wk_dates = weeks[wk_key]["dates"]
                total_pts = sum(
                    metrics["sla"].get(kho, {}).get(d, {}).get("on_time", 0) +
                    metrics["sla"].get(kho, {}).get(d, {}).get("late", 0)
                    for d in wk_dates
                )
                total_trips = sum(
                    metrics["completion_trip"].get(kho, {}).get(d, {}).get("total", 0)
                    for d in wk_dates
                )
                if use_sla_ontime:
                    on_time_pts = sum(metrics["sla"].get(kho, {}).get(d, {}).get("on_time", 0) for d in wk_dates)
                    ontime_total = total_pts
                else:
                    on_time_pts = sum(metrics["plan"].get(kho, {}).get(d, {}).get("on_time", 0) for d in wk_dates)
                    ontime_total = sum(
                        metrics["plan"].get(kho, {}).get(d, {}).get("on_time", 0) +
                        metrics["plan"].get(kho, {}).get(d, {}).get("late", 0)
                        for d in wk_dates
                    )
            else:
                sla = metrics["sla"].get(kho, {}).get(day_date, {})
                total_pts = sla.get("on_time", 0) + sla.get("late", 0)
                ct = metrics["completion_trip"].get(kho, {}).get(day_date, {})
                total_trips = ct.get("total", 0)
                if use_sla_ontime:
                    on_time_pts = sla.get("on_time", 0)
                    ontime_total = total_pts
                else:
                    plan = metrics["plan"].get(kho, {}).get(day_date, {})
                    on_time_pts = plan.get("on_time", 0)
                    ontime_total = plan.get("on_time", 0) + plan.get("late", 0)
            
            pct = round(on_time_pts / ontime_total * 100, 1) if ontime_total > 0 else ""
            data["Tổng Điểm Giao"].append(total_pts if total_pts > 0 else "")
            data["Đúng & Sớm Kế Hoạch"].append(on_time_pts if ontime_total > 0 else "")
            data["% On Time"].append(f"{pct}%" if pct != "" else "")
            data["Tổng Số Chuyến"].append(total_trips if total_trips > 0 else "")
        
        return data
    
    def compute_sla_row(kho, col_keys):
        """Compute just the SLA on-time row for ĐÔNG MÁT (extra row)."""
        sla_on = []
        sla_pct = []
        for wk_key, day_date in col_keys:
            if day_date is None:
                wk_dates = weeks[wk_key]["dates"]
                total = sum(
                    metrics["sla"].get(kho, {}).get(d, {}).get("on_time", 0) +
                    metrics["sla"].get(kho, {}).get(d, {}).get("late", 0)
                    for d in wk_dates
                )
                ot = sum(metrics["sla"].get(kho, {}).get(d, {}).get("on_time", 0) for d in wk_dates)
            else:
                sla = metrics["sla"].get(kho, {}).get(day_date, {})
                total = sla.get("on_time", 0) + sla.get("late", 0)
                ot = sla.get("on_time", 0)
            pct = round(ot / total * 100, 1) if total > 0 else ""
            sla_on.append(ot if total > 0 else "")
            sla_pct.append(f"{pct}%" if pct != "" else "")
        return sla_on, sla_pct
    
    def format_pct_cell(val, cls):
        """Apply color gradient to percentage cells."""
        if val and val != "":
            pct_val = float(str(val).replace('%', ''))
            if pct_val >= 99:
                bg = 'rgba(34,197,94,0.35)'; fc = '#4ADE80'
            elif pct_val >= 95:
                bg = 'rgba(74,222,128,0.18)'; fc = '#86EFAC'
            elif pct_val >= 90:
                bg = 'rgba(250,204,21,0.28)'; fc = '#EAB308'
            elif pct_val >= 85:
                bg = 'rgba(251,146,60,0.3)'; fc = '#F97316'
            elif pct_val >= 80:
                bg = 'rgba(239,68,68,0.3)'; fc = '#EF4444'
            else:
                bg = 'rgba(220,38,38,0.4)'; fc = '#FCA5A5'
            return f'<td class="{cls}" style="background:{bg};color:{fc};font-weight:600">{val}</td>'
        return f'<td class="{cls}">{val}</td>'
    
    def build_rows(metrics_data, kho, kho_color, col_keys, num_rows):
        rows_html = ""
        for i, (metric_name, values) in enumerate(metrics_data.items()):
            rows_html += '<tr>'
            if i == 0:
                rows_html += f'<td rowspan="{num_rows}" class="wt-kho-cell" style="border-left:3px solid {kho_color}">{kho}</td>'
            is_pct = "%" in metric_name
            rows_html += f'<td class="wt-metric-cell">{metric_name}</td>'
            for j, val in enumerate(values):
                is_total = col_keys[j][1] is None
                cls = "wt-total-cell" if is_total else "wt-day-cell"
                if is_pct:
                    rows_html += format_pct_cell(val, cls)
                else:
                    rows_html += f'<td class="{cls}">{val}</td>'
            rows_html += '</tr>'
        return rows_html
    
    header1, header2, col_keys = build_header_and_keys()
    html = ""
    
    # Table configs: (kho, title, use_sla_for_main, add_sla_extra_row)
    table_configs = [
        ("THỊT CÁ", "THỊT CÁ", True, False),       # SLA-based ontime
        ("ĐÔNG MÁT", "ĐÔNG MÁT", False, True),      # Plan-based + extra SLA row
        ("ĐÔNG", "HÀNG ĐÔNG", False, True),           # Plan-based + extra SLA row
        ("MÁT", "HÀNG MÁT", False, True),            # Plan-based + extra SLA row
    ]
    
    for kho, title, use_sla, add_sla_row in table_configs:
        kho_color = KHO_COLORS.get(kho, "#A8BCFF")
        data = compute_values(kho, col_keys, use_sla_ontime=use_sla)
        
        if add_sla_row:
            # Insert SLA row after "% On Time" (plan-based)
            sla_on_vals, sla_pct_vals = compute_sla_row(kho, col_keys)
            ordered_data = {}
            for k, v in data.items():
                ordered_data[k] = v
                if k == "% On Time":
                    ordered_data["% On Time (SLA)"] = sla_pct_vals
            num_rows = len(ordered_data)
        else:
            ordered_data = data
            num_rows = len(ordered_data)
        
        rows_html = build_rows(ordered_data, title, kho_color, col_keys, num_rows)
        
        html += f"""
    <div class="chart-box" style="overflow-x:auto">
        <h3>📋 {title} — Bảng Tổng Hợp Tuần (On-time SLA)</h3>
        <table class="weekly-table">
            <thead>{header1}{header2}</thead>
            <tbody>{rows_html}</tbody>
        </table>
    </div>"""
    
    return html


def generate_summary_cards(metrics, all_rows):
    """Generate summary statistics for the report header."""
    cards = {}
    
    for kho in KHO_COLORS:
        sla_data = metrics["sla"].get(kho, {})
        total_ot = sum(d.get("on_time", 0) for d in sla_data.values())
        total_lt = sum(d.get("late", 0) for d in sla_data.values())
        total = total_ot + total_lt
        cards[f"sla_{kho}"] = {"total": total, "pct": total_ot / total * 100 if total > 0 else 0}
    
    for kho in KHO_COLORS:
        plan_data = metrics["plan"].get(kho, {})
        total_ot = sum(d.get("on_time", 0) for d in plan_data.values())
        total_lt = sum(d.get("late", 0) for d in plan_data.values())
        total = total_ot + total_lt
        cards[f"plan_{kho}"] = {"total": total, "pct": total_ot / total * 100 if total > 0 else 0}
    
    for kho in KHO_COLORS:
        route_data = metrics["route"].get(kho, {})
        total_c = sum(d.get("correct", 0) for d in route_data.values())
        total_w = sum(d.get("wrong", 0) for d in route_data.values())
        total = total_c + total_w
        cards[f"route_{kho}"] = {"total": total, "pct": total_c / total * 100 if total > 0 else 0}
    
    return cards


def build_html_report(labels_json, charts_json, cards, month, year, weekly_tables_html=""):
    """Build HTML report with Chart.js interactive charts."""
    month_str = f"T{month:02d}/{year}"
    
    def card_html(title, key_prefix):
        rows_html = ""
        for kho in KHO_COLORS:
            data = cards.get(f"{key_prefix}_{kho}", {})
            pct = data.get("pct", 0)
            total = data.get("total", 0)
            color = "#2ECC71" if pct >= 95 else "#F39C12" if pct >= 90 else "#E74C3C"
            rows_html += f"""
            <tr>
                <td class="td-kho">{kho}</td>
                <td class="td-num">{total}</td>
                <td class="td-pct" style="color:{color}">{pct:.1f}%</td>
            </tr>"""
        return f"""
        <div class="card">
            <h3>{title}</h3>
            <table>
                <thead><tr><th>Kho</th><th>Tổng</th><th>Tỷ lệ</th></tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
        </div>"""
    
    kho_names_json = json.dumps(list(KHO_COLORS.keys()), ensure_ascii=False)
    kho_colors_json = json.dumps(KHO_COLORS, ensure_ascii=False)
    
    html = f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width,initial-scale=1">
    <title>Transport Performance Report — {month_str}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-annotation@3.0.1/dist/chartjs-plugin-annotation.min.js"></script>
    <style>
        * {{ margin:0; padding:0; box-sizing:border-box; }}
        body {{ font-family:'Segoe UI',Arial,sans-serif; background:#13131F; color:#E0E0E0; padding:20px; }}
        .container {{ max-width:1400px; margin:0 auto; }}
        h1 {{ text-align:center; color:#A8BCFF; margin-bottom:8px; font-size:24px; }}
        .subtitle {{ text-align:center; color:#888; margin-bottom:24px; font-size:14px; }}
        .cards {{ display:flex; gap:16px; flex-wrap:wrap; margin-bottom:24px; }}
        .card {{
            background:rgba(42,42,60,0.85); backdrop-filter:blur(12px);
            border:1px solid rgba(255,255,255,0.08); border-radius:14px;
            padding:20px; box-shadow:0 4px 20px rgba(0,0,0,0.3); flex:1; min-width:280px;
        }}
        .card h3 {{ margin:0 0 12px; color:#E0E0E0; font-size:15px; }}
        .card table {{ width:100%; border-collapse:collapse; font-size:13px; }}
        .card thead tr {{ background:rgba(255,255,255,0.05); }}
        .card th {{ padding:8px 12px; color:#999; text-align:center; }}
        .card th:first-child {{ text-align:left; }}
        .td-kho {{ padding:6px 12px; border-bottom:1px solid #333; font-weight:600; color:#ccc; }}
        .td-num {{ padding:6px 12px; border-bottom:1px solid #333; text-align:center; color:#aaa; }}
        .td-pct {{ padding:6px 12px; border-bottom:1px solid #333; text-align:center; font-weight:bold; }}
        .chart-box {{
            background:rgba(42,42,60,0.85); backdrop-filter:blur(12px);
            border:1px solid rgba(255,255,255,0.08); border-radius:14px;
            padding:20px; margin-bottom:20px; box-shadow:0 4px 20px rgba(0,0,0,0.3);
        }}
        .chart-box h3 {{ margin:0 0 12px; color:#E0E0E0; display:inline-block; }}
        .chart-row {{ display:flex; gap:20px; }}
        .chart-row > div {{ flex:1; }}
        canvas {{ max-height:400px; }}
        .filter-bar {{ display:inline-flex; gap:6px; margin-left:16px; vertical-align:middle; flex-wrap:wrap; }}
        .fbtn {{
            padding:5px 14px; border-radius:20px; border:1px solid rgba(255,255,255,0.15);
            background:rgba(255,255,255,0.06); color:#aaa; cursor:pointer; font-size:12px;
            font-family:inherit; transition:all 0.2s;
        }}
        .fbtn:hover {{ background:rgba(255,255,255,0.12); color:#ddd; }}
        .fbtn.active {{ background:rgba(168,188,255,0.25); color:#A8BCFF; border-color:#A8BCFF; font-weight:600; }}
        .weekly-table {{ width:100%; border-collapse:collapse; font-size:12px; white-space:nowrap; }}
        .weekly-table th, .weekly-table td {{ padding:6px 10px; border:1px solid rgba(255,255,255,0.1); text-align:center; }}
        .weekly-table thead th {{ background:rgba(42,42,60,1); color:#A8BCFF; font-weight:600; }}
        .weekly-table thead td {{ background:rgba(30,30,48,1); color:#999; font-size:11px; }}
        .wt-kho {{ position:sticky; left:0; z-index:3; background:#1e1e30!important; }}
        .wt-metric {{ position:sticky; left:60px; z-index:3; background:#1e1e30!important; }}
        .wt-kho-cell {{ font-weight:700; color:#E0E0E0; text-align:left!important; padding-left:12px!important; background:#1e1e30!important; position:sticky; left:0; z-index:2; }}
        .wt-metric-cell {{ text-align:left!important; color:#BBB; font-weight:500; white-space:nowrap; background:#1e1e30!important; position:sticky; left:60px; z-index:2; }}
        .wt-total-cell {{ background:rgba(168,188,255,0.08); font-weight:600; color:#E0E0E0; }}
        .wt-day-cell {{ color:#CCC; }}
    </style>
</head>
<body>
<div class="container">
    <h1>🚛 Transport Performance Report</h1>
    <p class="subtitle">{month_str} — On-time · Route Compliance · Trip Completion</p>
    
    <div class="cards">
        {card_html("📊 Đúng Giờ SLA", "sla")}
        {card_html("📅 Đúng Giờ vs Kế Hoạch", "plan")}
        {card_html("🗺 Đúng Tuyến", "route")}
    </div>

    <div class="chart-box"><h3>1. Tỷ Lệ Đúng Giờ SLA</h3><span class="filter-bar" id="filterSLA"></span>
        <p style="color:#777;font-size:12px;margin:4px 0 12px">KRC: 03:00-05:30 · THỊT CÁ: 03:00-06:00 · ĐÔNG MÁT: 09:00-16:00 · KSL-Sáng: 12:00-14:00 · KSL-Tối: 22:00-00:30</p>
        <canvas id="chartSLA"></canvas></div>
    <div class="chart-box"><h3>2. Tỷ Lệ Đúng Giờ vs Kế Hoạch</h3><span class="filter-bar" id="filterPlan"></span><canvas id="chartPlan"></canvas></div>
    <div class="chart-box"><h3>3. Tỷ Lệ Đúng Tuyến</h3><span class="filter-bar" id="filterRoute"></span><canvas id="chartRoute"></canvas></div>
    <div class="chart-box"><h3>4. Số Chuyến & Tỷ Lệ Hoàn Thành</h3><span class="filter-bar" id="filterTripsComp"></span><canvas id="chartTripsComp"></canvas></div>

{weekly_tables_html}

</div>

<script>
const LABELS = {labels_json};
const CHARTS = {charts_json};
const KHO_NAMES = {kho_names_json};
const KHO_COLORS = {kho_colors_json};

Chart.defaults.color = '#AAA';
Chart.defaults.borderColor = 'rgba(255,255,255,0.06)';
Chart.defaults.font.family = "'Segoe UI',Arial,sans-serif";

const TT = {{
    backgroundColor: 'rgba(30,30,46,0.95)', titleColor: '#E0E0E0', bodyColor: '#CCC',
    borderColor: 'rgba(255,255,255,0.15)', borderWidth: 1, cornerRadius: 8, padding: 10,
    bodySpacing: 4, titleFont: {{ weight: 'bold' }},
}};

// ── Filter button factory ──
function makeFilterBtns(containerId, chartIds, rebuildFn) {{
    const bar = document.getElementById(containerId);
    const btns = ['Tổng', ...KHO_NAMES];
    btns.forEach(name => {{
        const b = document.createElement('button');
        b.className = 'fbtn' + (name === 'Tổng' ? ' active' : '');
        b.textContent = name;
        if (name !== 'Tổng') b.style.borderColor = KHO_COLORS[name] + '66';
        b.onclick = () => {{
            bar.querySelectorAll('.fbtn').forEach(x => x.classList.remove('active'));
            b.classList.add('active');
            rebuildFn(name === 'Tổng' ? null : name);
        }};
        bar.appendChild(b);
    }});
}}

// ── Stacked bar + trend (SLA / Plan / Route) ──
const barTrendCharts = {{}};
function buildBarTrend(canvasId, data, lbl1, lbl2, yLabel, refLine, filterKho, yLeftLabel) {{
    if (!yLeftLabel) yLeftLabel = 'Số chuyến';
    if (barTrendCharts[canvasId]) barTrendCharts[canvasId].destroy();
    
    let barOn, barLate;
    if (filterKho && data.kho_bars && data.kho_bars[filterKho]) {{
        barOn = data.kho_bars[filterKho].on;
        barLate = data.kho_bars[filterKho].late;
    }} else {{
        barOn = data.on_times || data.corrects;
        barLate = data.lates || data.wrongs;
    }}
    
    const datasets = [
        {{ label: lbl1, data: barOn, backgroundColor: 'rgba(46,204,113,0.75)', borderRadius: 4, type: 'bar', order: 2, yAxisID: 'y' }},
        {{ label: lbl2, data: barLate, backgroundColor: 'rgba(231,76,60,0.75)', borderRadius: 4, type: 'bar', order: 2, yAxisID: 'y' }},
    ];
    
    if (filterKho) {{
        // Single kho trend
        const t = data.trends.find(t => t.label === filterKho);
        if (t) datasets.push({{
            label: t.label, data: t.data, borderColor: t.color, backgroundColor: t.color + '33',
            type: 'line', order: 1, yAxisID: 'y1', tension: 0.3, pointRadius: 3, pointHoverRadius: 6, borderWidth: 2.5, spanGaps: true, clip: false,
        }});
    }} else {{
        // Average trend
        datasets.push({{
            label: 'Trung bình', data: data.avg, borderColor: '#A8BCFF', backgroundColor: '#A8BCFF33',
            type: 'line', order: 1, yAxisID: 'y1', tension: 0.3, pointRadius: 3, pointHoverRadius: 6, borderWidth: 3, spanGaps: true, clip: false,
        }});
    }}
    
    const ann = refLine ? {{ annotation: {{ annotations: {{ ref: {{
        type: 'line', yMin: refLine, yMax: refLine, yScaleID: 'y1',
        borderColor: '#FF6B6B', borderWidth: 1.5, borderDash: [6, 4],
        label: {{ content: refLine + '%', display: true, position: 'end', backgroundColor: 'rgba(255,107,107,0.85)', color: '#fff', font: {{ size: 11 }}, padding: 4 }}
    }} }} }} }} : {{}};
    
    barTrendCharts[canvasId] = new Chart(document.getElementById(canvasId), {{
        data: {{ labels: LABELS, datasets }},
        options: {{
            responsive: true, interaction: {{ mode: 'index', intersect: false }},
            plugins: {{ tooltip: TT, ...ann }},
            scales: {{
                x: {{ stacked: true, grid: {{ display: false }} }},
                y: {{ stacked: true, title: {{ display: true, text: yLeftLabel }}, grid: {{ color: 'rgba(255,255,255,0.05)' }},
                      ticks: {{ precision: 0 }} }},
                y1: {{ position: 'right', min: 0, max: 100, title: {{ display: true, text: yLabel }}, grid: {{ drawOnChartArea: false }},
                      ticks: {{ stepSize: 20 }} }},
            }},
        }},
    }});
}}

// ── Line chart (Trips / Completion) ──
const lineCharts = {{}};
function buildLineChart(canvasId, lines, yTitle, filterKho, yMin, yMax) {{
    if (lineCharts[canvasId]) lineCharts[canvasId].destroy();
    
    let datasets;
    if (filterKho) {{
        const l = lines.find(l => l.label === filterKho);
        datasets = l ? [{{
            label: l.label, data: l.data, borderColor: l.color, backgroundColor: l.color + '22',
            tension: 0.3, pointRadius: 3, pointHoverRadius: 6, borderWidth: 2.5, fill: false, spanGaps: true,
        }}] : [];
    }} else {{
        // Single total/avg line
        const isPercent = yMax && yMax <= 100;
        const aggData = LABELS.map((_, i) => {{
            let sum = 0, cnt = 0;
            lines.forEach(l => {{ if (l.data[i] != null) {{ sum += l.data[i]; cnt++; }} }});
            if (cnt === 0) return null;
            return isPercent ? Math.round(sum / cnt * 10) / 10 : sum;
        }});
        datasets = [{{
            label: isPercent ? 'Trung bình' : 'Tổng', data: aggData,
            borderColor: '#A8BCFF', backgroundColor: '#A8BCFF22',
            tension: 0.3, pointRadius: 3, pointHoverRadius: 6, borderWidth: 3, fill: false, spanGaps: true,
        }}];
    }}
    
    const scaleOpts = {{ x: {{ grid: {{ display: false }} }}, y: {{ title: {{ display: true, text: yTitle }}, grid: {{ color: 'rgba(255,255,255,0.05)' }}, ticks: {{ precision: 0 }} }} }};
    if (yMin != null) scaleOpts.y.min = yMin;
    if (yMax != null) scaleOpts.y.max = yMax;
    
    lineCharts[canvasId] = new Chart(document.getElementById(canvasId), {{
        type: 'line', data: {{ labels: LABELS, datasets }},
        options: {{ responsive: true, interaction: {{ mode: 'index', intersect: false }}, plugins: {{ tooltip: TT }}, scales: scaleOpts }},
    }});
}}

// ── Init all charts ──
function initSLA(kho) {{ buildBarTrend('chartSLA', CHARTS.sla, 'Đúng SLA', 'Trễ SLA', '% Đúng SLA', 95, kho); }}
function initPlan(kho) {{ buildBarTrend('chartPlan', CHARTS.plan, 'Đúng Kế Hoạch', 'Trễ Kế Hoạch', '% Đúng Kế Hoạch', 95, kho); }}
function initRoute(kho) {{ buildBarTrend('chartRoute', CHARTS.route, 'Đúng Tuyến', 'Sai Tuyến', '% Đúng Tuyến', null, kho, 'Số điểm đến'); }}
function initTripsComp(kho) {{ buildBarTrend('chartTripsComp', CHARTS.trips_completion, 'Hoàn Thành', 'Chưa Hoàn Thành', '% Hoàn Thành', null, kho); }}

makeFilterBtns('filterSLA', ['chartSLA'], initSLA);
makeFilterBtns('filterPlan', ['chartPlan'], initPlan);
makeFilterBtns('filterRoute', ['chartRoute'], initRoute);
makeFilterBtns('filterTripsComp', ['chartTripsComp'], initTripsComp);

initSLA(null); initPlan(null); initRoute(null); initTripsComp(null);
</script>
</body>
</html>"""
    return html


# ══════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Generate Transport Performance Report")
    parser.add_argument("--month", type=int, default=None, help="Single month (legacy)")
    parser.add_argument("--months", type=str, default=None, help="Comma-separated months, e.g. '3,4'")
    parser.add_argument("--year", type=int, default=2026)
    args = parser.parse_args()
    
    year = args.year
    if args.months:
        months = [int(m.strip()) for m in args.months.split(",")]
    elif args.month:
        months = [args.month]
    else:
        months = [3]
    
    month_str = "+".join(f"T{m:02d}" for m in months)
    
    print("=" * 60)
    print(f"  🚛 TRANSPORT PERFORMANCE REPORT — {month_str}/{year}")
    print("=" * 60)
    
    # 1. Load data
    print("\n📥 Loading data...")
    trip_rows = []
    for m in months:
        trip_rows.extend(load_trip_data(m, year))
    thitca_rows = load_thitca_data(months)
    plan_lookup, route_order = load_plan_data(months)
    
    # ── KH MEAT fallback: fill missing THỊT CÁ planned_time from KH Drive files ──
    meat_dir = r"G:\My Drive\DOCS\DAILY\KH MEAT"
    if os.path.exists(meat_dir):
        import openpyxl
        meat_added = 0
        meat_files = sorted([f for f in os.listdir(meat_dir) if f.endswith('.xlsx') and not f.startswith('~')])
        for fname in meat_files:
            try:
                wb = openpyxl.load_workbook(os.path.join(meat_dir, fname), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                # Auto-detect planned_time column (header "Du kien giao" or similar)
                pt_col = 22  # default col 22
                tuyen_col = 11  # default col 11
                for hdr_row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                    for i, cell in enumerate(hdr_row):
                        h = str(cell.value or "").strip().lower()
                        if "du kien giao" in h or "dự kiến giao" in h:
                            pt_col = i
                        if ("tuyen" in h or "tuyến" in h) and "kg" not in h:
                            tuyen_col = i
                for row in ws.iter_rows(min_row=2, values_only=False):
                    date_val = row[0].value
                    if not date_val:
                        continue
                    if hasattr(date_val, 'strftime'):
                        d = date_val.date() if hasattr(date_val, 'date') else date_val
                    else:
                        d = parse_date_str(str(date_val))
                    if not d:
                        continue
                    store = str(row[2].value or "").strip() if len(row) > 2 else ""
                    if not store:
                        continue
                    key = (d, store, "THỊT CÁ")
                    existing = plan_lookup.get(key)
                    if not existing or not existing.get("planned_time"):
                        pt_val = row[pt_col].value if pt_col < len(row) else None
                        planned = parse_arrival_time(str(pt_val)) if pt_val else None
                        tuyen_val = str(row[tuyen_col].value or "").strip() if tuyen_col < len(row) else ""
                        if planned:
                            plan_lookup[key] = {"planned_time": planned, "tuyen": tuyen_val}
                            meat_added += 1
                wb.close()
            except Exception:
                pass
        if meat_added > 0:
            print(f"  🥩 KH MEAT fallback: {meat_added} THỊT CÁ planned_time entries added")
    
    # Combine trip data + THỊT CÁ
    all_rows = trip_rows + thitca_rows
    print(f"\n  📊 Total: {len(all_rows)} rows")
    
    # 2. Calculate metrics
    print("\n📈 Calculating metrics...")
    metrics = calc_metrics(all_rows, plan_lookup, route_order)
    
    # Print summary
    for metric_name in ["sla", "plan", "route"]:
        print(f"\n  {metric_name.upper()}:")
        for kho in KHO_COLORS:
            data = metrics[metric_name].get(kho, {})
            if not data:
                continue
            total_a = sum(d.get(list(d.keys())[0], 0) for d in data.values() if d)
            total_b = sum(d.get(list(d.keys())[1], 0) for d in data.values() if d and len(d) > 1)
            total = total_a + total_b
            pct = total_a / total * 100 if total > 0 else 0
            print(f"    {kho}: {total_a}/{total} ({pct:.1f}%)")
    
    # 3. Prepare chart data
    print("\n🎨 Preparing chart data...")
    from datetime import date as ddate
    
    # Generate all possible dates across all months
    all_possible_dates = []
    for m in months:
        _, days = monthrange(year, m)
        all_possible_dates.extend([ddate(year, m, d) for d in range(1, days + 1)])
    
    # Get dates that have data
    data_dates = set()
    for metric in [metrics["sla"], metrics["plan"], metrics["trips_per_day"]]:
        for kho_data in metric.values():
            if isinstance(kho_data, dict):
                data_dates.update(kho_data.keys())
    dates = sorted([d for d in all_possible_dates if d in data_dates])
    
    if not dates:
        print("  ⚠ No dates with data found!")
        return
    
    print(f"  📅 {len(dates)} days with data ({dates[0]} → {dates[-1]})")
    
    # Use first month for report naming
    labels, chart_data = prepare_chart_data(metrics, dates, months[0], year)
    labels_json = json.dumps(labels, ensure_ascii=False)
    charts_json = json.dumps(chart_data, ensure_ascii=False)
    
    # 4. Generate HTML
    print("\n📝 Generating HTML report...")
    cards = generate_summary_cards(metrics, all_rows)
    weekly_html = generate_weekly_tables(metrics, dates)
    html = build_html_report(labels_json, charts_json, cards, months[0], year, weekly_html)
    
    out_path = os.path.join(OUTPUT, f"PERFORMANCE_REPORT_{month_str}_{year}.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    
    print(f"\n✅ Report saved: {out_path}")
    
    # 5. Export raw data Excel
    print("\n📊 Exporting raw data Excel...")
    xlsx_path = export_raw_excel(all_rows, plan_lookup, route_order, month_str, year)
    print(f"✅ Excel saved: {xlsx_path}")
    
    # 6. Verification
    print("\n🔍 Verification...")
    issues = []
    
    # Check: no "DRY" kho in output
    dry_rows = [r for r in all_rows if r.get("kho") == "DRY"]
    if dry_rows:
        issues.append(f"❌ {len(dry_rows)} rows with kho='DRY' (should be KSL-Sáng/Tối)")
    
    # Check: trip "Hoàn thành" must have planned_time + arrival_time
    missing_plan = []
    missing_arrival = []
    for r in all_rows:
        if r.get("dest_status") != "Hoàn thành":
            continue
        d = r.get("date")
        dest = r.get("dest", "")
        kho = r.get("kho", "")
        
        # Check arrival
        if not r.get("arrival_time"):
            missing_arrival.append((d, dest, kho))
        
        # Check planned_time
        has_plan = bool(r.get("planned_time"))
        if not has_plan:
            lookup_kho = "DRY" if kho in ("KSL-Sáng", "KSL-Tối") else kho
            sub_kho = r.get("sub_kho", "")
            plan_info = None
            if sub_kho:
                plan_info = plan_lookup.get((d, dest, sub_kho), {})
            if not plan_info or not plan_info.get("planned_time"):
                plan_info = plan_lookup.get((d, dest, lookup_kho), {})
            has_plan = bool(plan_info.get("planned_time"))
        if not has_plan:
            missing_plan.append((d, dest, kho))
    
    if missing_arrival:
        issues.append(f"⚠ {len(missing_arrival)} trips 'Hoàn thành' missing arrival_time")
    
    if missing_plan:
        issues.append(f"⚠ {len(missing_plan)} trips 'Hoàn thành' missing planned_time")
        # Group by kho+date for readability
        from collections import Counter
        by_kho = Counter(kho for _, _, kho in missing_plan)
        for kho, cnt in by_kho.most_common():
            dates = sorted(set(str(d) for d, _, k in missing_plan if k == kho))
            issues.append(f"    {kho}: {cnt} rows ({', '.join(dates[:5])}{'...' if len(dates) > 5 else ''})")
    
    if issues:
        print("\n  ⚠ ISSUES FOUND — check data sources:")
        for issue in issues:
            print(f"  {issue}")
    else:
        print("  ✅ All checks passed")
    
    print("=" * 60)


def export_raw_excel(all_rows, plan_lookup, route_order, month_str, year):
    """Export raw per-destination data to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import timedelta
    from datetime import date as ddate
    
    DAY_VN = {0: 'Thứ 2', 1: 'Thứ 3', 2: 'Thứ 4', 3: 'Thứ 5', 4: 'Thứ 6', 5: 'Thứ 7', 6: 'CN'}
    
    # Pre-compute route compliance + rank: (date, tuyen, kho) -> True/False + ranks
    route_compliance = {}
    trip_dests = defaultdict(list)
    for r in all_rows:
        if not r.get("date") or not r.get("arrival_time"):
            continue
        tuyen = r.get("tuyen", "")
        kho = r["kho"]
        sub_kho = r.get("sub_kho", "")
        if not tuyen:
            lookup_kho = "DRY" if kho in ("KSL-Sáng", "KSL-Tối") else kho
            plan = None
            if sub_kho:
                plan = plan_lookup.get((r["date"], r["dest"], sub_kho))
            if not plan:
                plan = plan_lookup.get((r["date"], r["dest"], lookup_kho))
            if not plan:
                plan = plan_lookup.get((r["date"], r["dest"], kho))
            if plan:
                tuyen = plan.get("tuyen", "")
        if tuyen:
            trip_dests[(r["date"], tuyen, kho)].append((r["arrival_time"], r["dest"]))
    
    # Pre-compute: planned rank + actual rank per store per route
    # planned_rank[(date, tuyen, kho, store)] -> rank number
    # actual_rank[(date, tuyen, kho, store)] -> rank number
    planned_rank = {}
    actual_rank = {}
    
    for key, actual_stops in trip_dests.items():
        dt, tuyen, kho = key
        actual_sorted = sorted(actual_stops, key=lambda x: x[0])
        lookup_kho = "DRY" if kho in ("KSL-Sáng", "KSL-Tối") else kho
        planned = route_order.get((dt, tuyen, lookup_kho)) or route_order.get((dt, tuyen, kho))
        if planned:
            result = check_route_compliance(actual_sorted, planned)
            if result is not None:
                num_c, _ = result
                route_compliance[key] = num_c > 0
            
            # Build planned rank: store -> position in planned order
            for i, store in enumerate(planned, 1):
                planned_rank[(dt, tuyen, kho, store)] = i
        
        # Build actual rank: store -> position in actual delivery order (dedup)
        seen_stores = set()
        rank = 0
        for _, store in actual_sorted:
            if store not in seen_stores:
                rank += 1
                seen_stores.add(store)
                actual_rank[(dt, tuyen, kho, store)] = rank
    
    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    
    headers = [
        "Tuần", "Ngày", "Thứ", "Kho", "Phân Loại",
        "Giao Hàng", "Mã Trip / Tuyến", "Điểm Đến",
        "Giờ Đến", "Giờ Kế Hoạch",
        "Đúng Tuyến", "Thứ Tự KH", "Thứ Tự TT",
        "SLA", "Kế Hoạch vs Thực Tế",
    ]
    
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2A2A3C", end_color="2A2A3C", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_bd = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    fill_som = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fill_dung = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    fill_tre = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_bd
    
    sorted_rows = sorted(all_rows, key=lambda r: (
        r.get("date") or ddate(2099, 1, 1),
        r.get("kho", ""),
        r.get("tuyen", ""),
        r.get("arrival_time") or dtime(23, 59),
    ))
    
    row_num = 2
    for r in sorted_rows:
        d = r.get("date")
        if not d:
            continue
        
        kho = r["kho"]
        arrival = r.get("arrival_time")
        dest = r.get("dest", "")
        tuyen = r.get("tuyen", "")
        trip_id = r.get("trip_id", "")
        sub_kho = r.get("sub_kho", "")
        
        # Try sub_kho-specific plan first (ĐÔNG vs MÁT have different planned times)
        plan_info = None
        if sub_kho:
            plan_info = plan_lookup.get((d, dest, sub_kho), {})
        if not plan_info or not plan_info.get("planned_time"):
            plan_info = plan_lookup.get((d, dest, kho), {})
        if not plan_info and kho in ("KSL-Sáng", "KSL-Tối"):
            plan_info = plan_lookup.get((d, dest, "DRY"), {})
        if not plan_info:
            plan_info = {}
        planned_time = r.get("planned_time") or plan_info.get("planned_time")
        if not tuyen:
            tuyen = plan_info.get("tuyen", "")
        
        iso_year, iso_week, _ = d.isocalendar()
        week = f"W{iso_week}"
        day_name = DAY_VN.get(d.weekday(), "")
        has_delivery = "Y" if arrival else "N"
        
        if kho == "THỊT CÁ":
            trip_label = tuyen
        else:
            trip_label = trip_id if trip_id else tuyen
        
        route_key = (d, tuyen, kho) if tuyen else None
        if route_key and route_key in route_compliance:
            route_ok = "Đúng" if route_compliance[route_key] else "Sai"
        else:
            route_ok = ""
        
        # Route ranks
        p_rank = ""
        a_rank = ""
        if tuyen and arrival:
            p_rank = planned_rank.get((d, tuyen, kho, dest), "")
            a_rank = actual_rank.get((d, tuyen, kho, dest), "")
        
        # SLA status
        sla_status = ""
        if arrival and kho in SLA_WINDOWS:
            sla_s, sla_e = SLA_WINDOWS[kho]
            if kho == "KSL-Tối":
                if arrival.hour >= 15:
                    sla_status = "Sớm" if arrival < sla_s else "Đúng"
                elif arrival.hour == 0 and arrival.minute <= 30:
                    sla_status = "Đúng"
                else:
                    sla_status = "Trễ"
            else:
                if arrival < sla_s:
                    sla_status = "Sớm"
                elif arrival <= sla_e:
                    sla_status = "Đúng"
                else:
                    sla_status = "Trễ"
        
        # Plan status
        plan_status = ""
        if arrival and planned_time:
            if arrival < planned_time:
                plan_status = "Sớm"
            elif arrival == planned_time:
                plan_status = "Đúng"
            else:
                plan_status = "Trễ"
        
        arrival_str = arrival.strftime("%H:%M") if arrival else ""
        # Giao=N → clear all KPI columns
        if not arrival:
            planned_str = ""
            route_ok = ""
            p_rank = ""
            a_rank = ""
            sla_status = ""
            plan_status = ""
        else:
            planned_str = planned_time.strftime("%H:%M") if planned_time else ""
        
        values = [
            week, d.strftime("%d/%m/%Y"), day_name, kho, sub_kho,
            has_delivery, trip_label, dest,
            arrival_str, planned_str,
            route_ok, p_rank, a_rank,
            sla_status, plan_status,
        ]
        
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.border = thin_bd
            cell.alignment = Alignment(horizontal="center")
        
        # Color SLA (col N = 14)
        sla_cell = ws.cell(row=row_num, column=14)
        if sla_status == "Sớm":
            sla_cell.fill = fill_som
        elif sla_status == "Đúng":
            sla_cell.fill = fill_dung
        elif sla_status == "Trễ":
            sla_cell.fill = fill_tre
        
        # Color Plan (col O = 15)
        plan_cell = ws.cell(row=row_num, column=15)
        if plan_status == "Sớm":
            plan_cell.fill = fill_som
        elif plan_status == "Đúng":
            plan_cell.fill = fill_dung
        elif plan_status == "Trễ":
            plan_cell.fill = fill_tre
        
        row_num += 1
    
    # Column widths
    widths = {'A': 6, 'B': 13, 'C': 8, 'D': 12, 'E': 10,
              'F': 10, 'G': 22, 'H': 28,
              'I': 10, 'J': 13, 'K': 12, 'L': 10, 'M': 10,
              'N': 8, 'O': 20}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{row_num - 1}"
    
    out_path = os.path.join(OUTPUT, f"RAW_DATA_{month_str}_{year}.xlsx")
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    main()
