"""
du_kien_giao.py - Generate delivery schedule data for "Dự kiến giao" Google Sheet
Usage: python script/du_kien_giao.py [--date DD/MM/YYYY]

Fetches data from KRC, KFM (KSL), KH MEAT/ĐÔNG/MÁT sources and outputs
a TSV file ready to paste into the "Dự kiến giao" sheet tab in
"Dashboard báo cáo ABA" Google Sheet.

Columns: Key | Ngày giao hàng | Điểm đến | Giờ đến dự kiến (+-30') | Kho | Loại hàng
"""
import os, sys, re
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "data")
OUTPUT = os.path.join(BASE, "output")

# ── Online data source URLs ──
KRC_SHEET_URL = "https://docs.google.com/spreadsheets/d/1tWamqjpOI2j2MrYW3Ah6ptmT524CAlQvEP8fCkxfuII/export?format=xlsx"
KFM_SHEET_URL = "https://docs.google.com/spreadsheets/d/1LkJFJhOQ8F2WEB3uCk7kA2Phvu8IskVi3YBfVr7pBx0/export?format=xlsx"

# Google Drive folder URLs for KH data:
# (name, report_kho, tuyen_col, delivery_time_col, loai_hang_col, url)
KH_DRIVE_FOLDERS = [
    ("KH MEAT", "THỊT CÁ", 11, 22, -1, "https://drive.google.com/drive/folders/1GIzH8nmCbLhWfpdmxFIn9cHTvQNbnwWr"),
    ("KH HÀNG ĐÔNG", "ĐÔNG MÁT", 9, 17, 18, "https://drive.google.com/drive/folders/1pQ8coYeV-K0dcHlkvXcJ8KngmH22xp1Z"),
    ("KH HÀNG MÁT", "ĐÔNG MÁT", 9, 19, 20, "https://drive.google.com/drive/folders/1c2zfgcXM8O9ezkOZYj0p4t_ihaJmb98f"),
]


def read_xlsx_from_url(url):
    import requests
    r = requests.get(url, allow_redirects=True, timeout=120)
    r.raise_for_status()
    return load_workbook(BytesIO(r.content), read_only=True, data_only=True)


def _list_drive_folder(folder_url):
    import requests as _req
    import html as _html
    folder_id = re.search(r'folders/([a-zA-Z0-9_-]+)', folder_url).group(1)
    url = f"https://drive.google.com/drive/folders/{folder_id}"
    r = _req.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    text = _html.unescape(r.text)
    files = []
    seen = set()
    for m in re.finditer(r'"(1[a-zA-Z0-9_-]{25,50})"', text):
        fid = m.group(1)
        if fid in seen:
            continue
        after = text[m.end():m.end() + 500]
        name_match = re.search(r'"([^"]+\.xlsx)"', after)
        if name_match:
            fname = name_match.group(1)
            if "Microsoft" not in fname and "Shared" not in fname:
                seen.add(fid)
                files.append((fid, fname))
    return files


def read_kh_from_drive(folder_url, folder_name, date_for_file):
    import requests as _req
    files = _list_drive_folder(folder_url)
    if not files:
        print(f"    ⚠ Không list được folder")
        return None
    target = None
    for fid, fname in files:
        if date_for_file in fname:
            target = (fid, fname)
            break
    if not target:
        return None
    print(f"    ↳ {target[1]}")
    dl_url = f"https://drive.google.com/uc?export=download&confirm=t&id={target[0]}"
    r = _req.get(dl_url, allow_redirects=True, timeout=60)
    if r.status_code != 200 or len(r.content) < 500:
        print(f"    ⚠ Download failed (status={r.status_code})")
        return None
    return load_workbook(BytesIO(r.content), read_only=True, data_only=True)


def safe_val(row, idx):
    if idx < len(row):
        return str(row[idx].value or "").strip()
    return ""


def format_time(time_val):
    """Convert time value to H:MM format."""
    if not time_val:
        return ""
    s = str(time_val).strip()
    # Already in time format like "3:05" or "15:30"
    m = re.match(r'(\d{1,2}):(\d{2})', s)
    if m:
        return f"{int(m.group(1))}:{m.group(2)}"
    # datetime object
    if hasattr(time_val, 'hour'):
        return f"{time_val.hour}:{time_val.minute:02d}"
    # Float (Excel time fraction) e.g. 0.125 = 3:00
    try:
        fval = float(time_val)
        if 0 <= fval < 1:
            total_minutes = round(fval * 24 * 60)
            hours = total_minutes // 60
            minutes = total_minutes % 60
            return f"{hours}:{minutes:02d}"
    except (ValueError, TypeError):
        pass
    return s


def fetch_delivery_data(date_str, date_for_file):
    """Fetch all delivery schedule data from online sources.
    Returns list of dicts with keys: key, date, diem_den, gio_den, kho, loai_hang
    """
    rows = []
    warnings = []
    
    # Format date for output (M/D/YYYY) — sheet uses US format
    parts = date_str.split("/")
    date_out = f"{int(parts[1])}/{int(parts[0])}/{parts[2]}"  # e.g. "3/24/2026"

    # ── 1. KRC ──
    print("  → KRC (Google Sheets)...")
    try:
        wb = read_xlsx_from_url(KRC_SHEET_URL)
        ws = wb["KRC"]
        krc_count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            scv = str(row[0].value or "").strip()
            if scv == date_str:
                diem_den = str(row[6].value or "").strip()  # Col G = Điểm đến
                gio_den = row[7].value                       # Col H = Giờ đến
                if diem_den:
                    gio_str = format_time(gio_den)
                    rows.append({
                        "date": date_out,
                        "diem_den": diem_den,
                        "gio_den": gio_str,
                        "kho": "KRC",
                        "loai_hang": "",
                    })
                    krc_count += 1
        wb.close()
        print(f"    {krc_count} rows")
    except Exception as e:
        warnings.append(f"KRC: lỗi — {e}")

    # ── 2. KFM (KSL) ──
    print("  → KFM (Google Sheets)...")
    try:
        wb = read_xlsx_from_url(KFM_SHEET_URL)
        ws = None
        for name in wb.sheetnames:
            if "DRY" in name:
                ws = wb[name]
                break
        if not ws:
            ws = wb.worksheets[1]

        kfm_count = 0
        for row in ws.iter_rows(min_row=3, values_only=False):
            scv = str(row[0].value or "").strip()
            if scv == date_str:
                diem_den = str(row[6].value or "").strip()   # Col G
                gio_den = row[7].value                        # Col H
                if diem_den:
                    gio_str = format_time(gio_den)
                    rows.append({
                        "date": date_out,
                        "diem_den": diem_den,
                        "gio_den": gio_str,
                        "kho": "KSL",
                        "loai_hang": "",
                    })
                    kfm_count += 1
        wb.close()
        print(f"    {kfm_count} rows")
    except Exception as e:
        warnings.append(f"KFM: lỗi — {e}")

    # ── 3. KH MEAT / ĐÔNG / MÁT ──
    for folder_name, kho_name, tuyen_col, time_col, loai_hang_col, folder_url in KH_DRIVE_FOLDERS:
        print(f"  → {folder_name} (Google Drive)...")
        try:
            wb = read_kh_from_drive(folder_url, folder_name, date_for_file)
            if not wb:
                warnings.append(f"{folder_name}: file {date_for_file} not found")
                continue
            ws = wb.worksheets[0]

            # Auto-detect column positions from header row
            actual_time_col = time_col
            actual_loai_col = loai_hang_col
            for hdr_row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                for i, cell in enumerate(hdr_row):
                    h = str(cell.value or "").strip().lower()
                    if "kien" in h and "giao" in h:  # "Du kien giao" / "Dự kiến giao"
                        actual_time_col = i
                    elif "loại" in h or "loai" in h:  # "LOẠI HÀNG"
                        actual_loai_col = i
            if actual_time_col != time_col or actual_loai_col != loai_hang_col:
                print(f"    ℹ️ Auto-detect: time_col={actual_time_col}, loai_col={actual_loai_col} (default: {time_col}, {loai_hang_col})")

            count = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                diem_den = safe_val(row, 2)   # Col C = Điểm đến
                gio_den = row[actual_time_col].value if actual_time_col < len(row) else None
                # Read loại hàng from file if column specified
                loai_hang_val = ""
                if actual_loai_col >= 0 and actual_loai_col < len(row):
                    loai_hang_val = str(row[actual_loai_col].value or "").strip()
                if diem_den:
                    gio_str = format_time(gio_den) if gio_den else ""
                    rows.append({
                        "date": date_out,
                        "diem_den": diem_den,
                        "gio_den": gio_str,
                        "kho": kho_name,
                        "loai_hang": loai_hang_val,
                    })
                    count += 1
            wb.close()
            print(f"    {count} rows")
        except Exception as e:
            warnings.append(f"{folder_name}: lỗi — {e}")

    return rows, warnings


def generate_tsv(rows, output_path):
    """Write delivery data to TSV file."""
    with open(output_path, "w", encoding="utf-8") as f:
        for r in rows:
            line = f"{r['date']}\t{r['diem_den']}\t{r['gio_den']}\t{r['kho']}\t{r['loai_hang']}"
            f.write(line + "\n")
    return output_path


def main():
    # Parse arguments
    date_str = ""
    if "--date" in sys.argv:
        idx = sys.argv.index("--date")
        date_str = sys.argv[idx + 1]
    if not date_str:
        date_str = datetime.now().strftime("%d/%m/%Y")

    parts = date_str.split("/")
    date_for_file = f"{parts[0]}.{parts[1]}.{parts[2]}"

    print("=" * 60)
    print(f"  DỰ KIẾN GIAO - {date_str}")
    print("=" * 60)

    # Fetch data
    print("\n📋 Fetching delivery schedule data...")
    rows, warnings = fetch_delivery_data(date_str, date_for_file)

    for w in warnings:
        print(f"  ⚠️ {w}")

    # Generate TSV
    os.makedirs(OUTPUT, exist_ok=True)
    date_tag = date_str.replace("/", "")
    tsv_path = os.path.join(OUTPUT, f"du_kien_giao_{date_tag}.tsv")
    generate_tsv(rows, tsv_path)

    print(f"\n📊 Summary:")
    print(f"  Total rows: {len(rows)}")
    # Count by kho
    kho_counts = {}
    for r in rows:
        kho_counts[r["kho"]] = kho_counts.get(r["kho"], 0) + 1
    for kho, count in sorted(kho_counts.items()):
        print(f"    {kho}: {count}")

    print(f"\n✅ TSV saved: {tsv_path}")
    print(f"   Copy and paste into 'Dự kiến giao' sheet in Google Sheets.")
    print("\n" + "=" * 60)
    print("  DONE")
    print("=" * 60)


if __name__ == "__main__":
    main()
