"""Final split test with user-confirmed classification for 18 unclassified barcodes."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from collections import defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(BASE, "script"))
from lib.sources import KH_DONG_LOCAL, KH_MAT_LOCAL, KH_MEAT_LOCAL

DATE_STR = "06/05/2026"
DATE_FILE = "06.05.2026"
DATE_TAG = "06052026"
BACKUP_DIR = os.path.join(BASE, "data", "raw", "daily")
ABA_MASTER = r'G:\My Drive\DOCS\DAILY\ton_aba\data\master_data\Master Data.xlsx'

# User-confirmed overrides for unclassified barcodes
OVERRIDE_DONG = {"8935335402859", "074570022021", "8938527710565"}
OVERRIDE_MAT = {
    "8936113000045", "8936113000083", "11702", "8938562884368", "8938562884375",
    "11541", "8936076278758", "11680", "11542", "8936076278765",
    "11696", "074570052028", "SP001590", "SP001589", "11683",
}

# ── 1. Load barcode classification + overrides ──
print("=" * 72)
print("  ĐÔNG/MÁT SPLIT (FINAL) — 06/05/2026")
print("=" * 72)
print()

wb = load_workbook(ABA_MASTER, read_only=True, data_only=True)
ws = wb.worksheets[0]
barcode_type = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    bc = str(row[1].value or '').strip()
    pl = str(row[4].value or '').strip().upper()
    if bc and pl in ('MÁT', 'ĐÔNG'):
        barcode_type[bc] = pl
wb.close()

# Apply overrides
for bc in OVERRIDE_DONG:
    barcode_type[bc] = 'ĐÔNG'
for bc in OVERRIDE_MAT:
    barcode_type[bc] = 'MÁT'

print("📋 Classification: {} barcodes (+18 overrides applied)".format(len(barcode_type)))

# ── 2. STHI/XE ──
def safe_val(row, idx):
    if idx < len(row):
        return str(row[idx].value or "").strip()
    return ""

REPORT_KHOS = ["KRC", "THỊT CÁ", "ĐÔNG", "MÁT", "KSL-SÁNG", "KSL-TỐI"]

# Read KH for ĐÔNG, MÁT, THỊT CÁ
sthi_dd = defaultdict(set)
sthi_tu = defaultdict(set)

for folder_name, kho_name, tuyen_col, local_dir in [
    ("KH HÀNG ĐÔNG", "ĐÔNG", 9, KH_DONG_LOCAL),
    ("KH HÀNG MÁT", "MÁT", 9, KH_MAT_LOCAL),
    ("KH MEAT", "THỊT CÁ", 11, KH_MEAT_LOCAL),
]:
    if not os.path.isdir(local_dir):
        continue
    matched = None
    for fname in os.listdir(local_dir):
        if DATE_FILE in fname and fname.endswith('.xlsx') and not fname.startswith('~'):
            matched = os.path.join(local_dir, fname)
            break
    if not matched:
        continue
    wb = load_workbook(matched, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2, values_only=False):
        diem_den = safe_val(row, 2)
        tuyen = safe_val(row, tuyen_col)
        if diem_den:
            sthi_dd[kho_name].add(diem_den)
            if tuyen:
                sthi_tu[kho_name].add(tuyen)
    wb.close()

# Read KRC from backup
krc_path = os.path.join(BACKUP_DIR, "krc_{}.xlsx".format(DATE_TAG))
if os.path.exists(krc_path):
    wb = load_workbook(krc_path, read_only=True, data_only=True)
    ws = wb["KRC"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        scv = str(row[0].value or "").strip()
        if scv == DATE_STR:
            diem_den = str(row[6].value or "").strip()
            tuyen = str(row[10].value or "").strip()
            if diem_den and row[7].value:
                sthi_dd["KRC"].add(diem_den)
                if tuyen:
                    sthi_tu["KRC"].add(tuyen)
    wb.close()

# Read KFM for KSL
import re
from datetime import datetime, time as _time

def parse_time_hour(time_text):
    if not time_text:
        return -1
    if isinstance(time_text, datetime):
        return time_text.hour
    if isinstance(time_text, _time):
        return time_text.hour
    s = str(time_text).strip()
    dt_match = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?$', s)
    if dt_match:
        return int(dt_match.group(1))
    m = re.match(r'(\d{1,2}):', s)
    return int(m.group(1)) if m else -1

kfm_path = os.path.join(BACKUP_DIR, "kfm_{}.xlsx".format(DATE_TAG))
if os.path.exists(kfm_path):
    wb = load_workbook(kfm_path, read_only=True, data_only=True)
    kfm_ws = None
    for sname in wb.sheetnames:
        if 'DRY' in sname.upper():
            kfm_ws = wb[sname]
            break
    if not kfm_ws:
        kfm_ws = wb.worksheets[0]
    last_kho = None
    for xl_row in kfm_ws.iter_rows(min_row=3, values_only=False):
        scv = str(xl_row[0].value or "").strip()
        if scv == DATE_STR:
            diem_den = str(xl_row[6].value or "").strip()
            gio_den = str(xl_row[7].value or "").strip()
            gio_di = str(xl_row[4].value or "").strip()
            tuyen = str(xl_row[10].value or "").strip()
            if not diem_den:
                continue
            gio = gio_den or gio_di
            if gio:
                hour = parse_time_hour(gio)
                if hour < 0:
                    continue
                using_gio_di = not gio_den and bool(gio_di)
                sang_cutoff = 15 if using_gio_di else 18
                if 6 <= hour < sang_cutoff:
                    kho = "KSL-SÁNG"
                else:
                    kho = "KSL-TỐI"
                last_kho = kho
                sthi_dd[kho].add(diem_den)
                if tuyen:
                    sthi_tu[kho].add(tuyen)
            elif last_kho:
                sthi_dd[last_kho].add(diem_den)
                if tuyen:
                    sthi_tu[last_kho].add(tuyen)
        else:
            last_kho = None
    wb.close()

# ── 3. PT data with split ──
master_tl = {}
ml_path = os.path.join(BACKUP_DIR, "master_data.xlsx")
if os.path.exists(ml_path):
    wb = load_workbook(ml_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2, values_only=False):
        bc = str(row[0].value or '').strip()
        if not bc:
            continue
        tl_v = row[25].value
        if tl_v is not None:
            try:
                w = float(tl_v)
                if w > 0:
                    master_tl[bc] = w
            except (ValueError, TypeError):
                pass
    wb.close()

KHO_MAP = {
    "KHO ABA MIỀN ĐÔNG": "THỊT CÁ",
    "KHO RAU CỦ": "KRC",
    "Sáng": "KSL-SÁNG", "Tối": "KSL-TỐI",
    "Khách đặt": "KSL-TỐI", "khách đặt": "KSL-TỐI",
    "Socola": "KSL-SÁNG",
    "ĐI SÁNG": "KSL-SÁNG", "ĐI TỐI": "KSL-TỐI",
    "đi sáng": "KSL-SÁNG", "đi tối": "KSL-TỐI",
}

pt_items = defaultdict(float)
pt_tons = defaultdict(float)

tf_path = os.path.join(BACKUP_DIR, "transfer_{}.xlsx".format(DATE_TAG))
wb = load_workbook(tf_path, read_only=True, data_only=True)
ws = wb.worksheets[0]

# First pass: build transfer_tl
transfer_tl = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    code = safe_val(row, 7)
    tl_raw = row[14].value
    if code and tl_raw:
        try:
            tl_val = float(tl_raw)
            if tl_val > 0:
                transfer_tl[code] = tl_val
        except (ValueError, TypeError):
            pass
wb.close()

# Second pass
wb = load_workbook(tf_path, read_only=True, data_only=True)
ws = wb.worksheets[0]
for row in ws.iter_rows(min_row=2, values_only=False):
    ngay = str(row[0].value or '').strip()
    if ngay != DATE_STR:
        continue
    raw_kho = str(row[2].value or '').strip()
    code = str(row[7].value or '').strip()

    # Split ĐÔNG MÁT by barcode classification
    if raw_kho == 'KHO ABA QUÁ CẢNH':
        classification = barcode_type.get(code)
        if classification == 'ĐÔNG':
            report_kho = 'ĐÔNG'
        elif classification == 'MÁT':
            report_kho = 'MÁT'
        else:
            continue  # skip truly unknown (shouldn't happen with overrides)
    else:
        report_kho = KHO_MAP.get(raw_kho)
        if not report_kho:
            continue

    try:
        sl = float(row[10].value or 0)
    except (ValueError, TypeError):
        sl = 0
    try:
        tl = float(row[14].value or 0)
    except (ValueError, TypeError):
        tl = 0
    if tl == 0 and code:
        tl = master_tl.get(code, 0)
    if tl == 0:
        continue
    pt_items[report_kho] += sl
    pt_tons[report_kho] += sl * tl / 1_000_000
wb.close()

# ── 4. Print result ──
print()
fmt = "  {:<10} {:>6} {:>10} {:>6} {:>8} {:>6} {:>6} {:>5} {:>7}"
print(fmt.format("KHO", "STHI", "ITEMS", "XE", "TẤN", "T/XE", "I/ST", "ST/XE", "KG/ST"))
print("  " + "-" * 72)

total_st = total_it = total_xe = 0
total_tn = 0.0
for kho in REPORT_KHOS:
    st = len(sthi_dd.get(kho, set()))
    it = pt_items.get(kho, 0)
    xe = len(sthi_tu.get(kho, set()))
    tn = pt_tons.get(kho, 0)
    txe = tn / xe if xe > 0 else 0
    ist = it / st if st > 0 else 0
    stxe = st / xe if xe > 0 else 0
    kgst = tn * 1000 / st if st > 0 else 0
    print(fmt.format(kho, st, "{:,.0f}".format(it), xe, "{:.2f}".format(tn),
                     "{:.2f}".format(txe), "{:.0f}".format(ist), "{:.1f}".format(stxe),
                     "{:.1f}".format(kgst)))
    total_st += st
    total_it += it
    total_xe += xe
    total_tn += tn

print("  " + "-" * 72)
txe_t = total_tn / total_xe if total_xe > 0 else 0
ist_t = total_it / total_st if total_st > 0 else 0
stxe_t = total_st / total_xe if total_xe > 0 else 0
kgst_t = total_tn * 1000 / total_st if total_st > 0 else 0
print(fmt.format("TOTAL", total_st, "{:,.0f}".format(total_it), total_xe,
                 "{:.2f}".format(total_tn), "{:.2f}".format(txe_t),
                 "{:.0f}".format(ist_t), "{:.1f}".format(stxe_t),
                 "{:.1f}".format(kgst_t)))

# Comparison with old report
print()
print("  📊 So sánh với report cũ (5 kho):")
print("     TOTAL cũ:  603 STHI, 464,609 items, 174 xe, 179.69 tấn")
print("     TOTAL mới: {} STHI, {:,.0f} items, {} xe, {:.2f} tấn".format(
    total_st, total_it, total_xe, total_tn))
print("     STHI tăng vì ĐÔNG có thêm stores riêng (trước gộp chung ĐÔNG MÁT)")
print("     Items/Tấn giữ nguyên (cùng data, chỉ tách kho)")
