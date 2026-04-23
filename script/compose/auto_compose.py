"""
auto_compose.py - Automated mail composition scheduler

Runs periodically (via Windows Task Scheduler or manually), checks each
warehouse's schedule, fetches latest data, detects changes, and
generates/regenerates delivery schedule email HTML.

Usage:
  python script/auto_compose.py              # Normal scheduled run
  python script/auto_compose.py --status     # Show today's compose status
  python script/auto_compose.py --force KRC  # Force compose for specific kho
  python script/auto_compose.py --dry-run    # Check data but don't compose
  python script/auto_compose.py --no-fetch   # Use existing data (skip re-fetch)
  python script/auto_compose.py --watch      # Watch mode: poll Drive for files, auto-compose
  python script/auto_compose.py --watch --poll-interval 5  # Poll every 5 minutes

Schedule config: config/mail_schedule.json
State tracking:  output/auto_compose_state.json
"""
import os, sys, json, hashlib, argparse, subprocess, logging, re, time, atexit
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CONFIG_PATH = os.path.join(BASE, "config", "mail_schedule.json")
STATE_PATH = os.path.join(BASE, "output", "state", "auto_compose_state.json")
OUTPUT = os.path.join(BASE, "output")
LOCK_PATH = os.path.join(BASE, "output", "state", "auto_compose_watch.lock")

# ── Logging ──────────────────────────────────────────────────────────
LOG_PATH = os.path.join(BASE, "output", "logs", "auto_compose.log")

def setup_logging():
    """Setup logging to both file and console."""
    logger = logging.getLogger("auto_compose")
    logger.setLevel(logging.INFO)
    # Console handler
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(ch)
    # File handler
    os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
    fh = logging.FileHandler(LOG_PATH, encoding='utf-8')
    fh.setLevel(logging.INFO)
    fh.setFormatter(logging.Formatter('%(asctime)s  %(message)s', datefmt='%Y-%m-%d %H:%M'))
    logger.addHandler(fh)
    return logger

log = setup_logging()


# ══════════════════════════════════════════════════════════════════════
#  Config & State
# ══════════════════════════════════════════════════════════════════════

def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_state():
    if os.path.exists(STATE_PATH):
        with open(STATE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_state(state):
    with open(STATE_PATH, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def state_key(kho, session):
    """Unique key per kho/session combo."""
    if session:
        return f"{kho}_{session}"
    return kho


# ══════════════════════════════════════════════════════════════════════
#  Date & Time helpers
# ══════════════════════════════════════════════════════════════════════

def get_delivery_date(kho, session, today=None):
    """
    Mail is for D+1 delivery, EXCEPT DRY Tối which is same-day D evening.
    Returns datetime for the delivery date.
    """
    if today is None:
        today = datetime.now()
    if kho == "DRY" and session == "toi":
        return today  # DRY Tối = today's evening
    return today + timedelta(days=1)  # D+1


def format_date_vn(dt):
    return dt.strftime("%d/%m/%Y")


def should_skip_delivery(kho, delivery_date, skip_days):
    """Check if this kho doesn't deliver on delivery_date's day-of-week.
    skip_days: list of weekday ints (Mon=0, Sun=6).
    """
    return delivery_date.weekday() in skip_days


def time_from_str(time_str, ref_date=None):
    """Parse HH:MM string to datetime on ref_date."""
    if ref_date is None:
        ref_date = datetime.now()
    h, m = map(int, time_str.split(":"))
    return ref_date.replace(hour=h, minute=m, second=0, microsecond=0)


def is_in_window(check_time_str, cutoff_str, now=None):
    """Check if current time is within [check_time, cutoff]."""
    if now is None:
        now = datetime.now()
    check = time_from_str(check_time_str, now)
    cutoff = time_from_str(cutoff_str, now)
    return check <= now <= cutoff


def is_at_cutoff(cutoff_str, now=None, window_min=20):
    """Check if within `window_min` minutes before cutoff."""
    if now is None:
        now = datetime.now()
    cutoff = time_from_str(cutoff_str, now)
    diff = (cutoff - now).total_seconds()
    return 0 <= diff <= window_min * 60


def is_past_cutoff(cutoff_str, now=None):
    """Check if past cutoff time."""
    if now is None:
        now = datetime.now()
    cutoff = time_from_str(cutoff_str, now)
    return now > cutoff


# ══════════════════════════════════════════════════════════════════════
#  Google Drive file detection (for --watch mode)
# ══════════════════════════════════════════════════════════════════════

def _list_drive_folder_files(folder_url):
    """List .xlsx files in a public Google Drive folder.
    Returns list of (file_id, filename) tuples.
    """
    import requests as _req
    import html as _html
    m = re.search(r'folders/([a-zA-Z0-9_-]+)', folder_url)
    if not m:
        return []
    folder_id = m.group(1)
    url = f"https://drive.google.com/drive/folders/{folder_id}"
    try:
        r = _req.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        text = _html.unescape(r.text)
    except Exception as e:
        log.warning(f"    ⚠ Drive folder request failed: {e}")
        return []

    files = []
    seen = set()
    for m in re.finditer(r'"(1[a-zA-Z0-9_-]{25,50})"', text):
        fid = m.group(1)
        if fid in seen:
            continue
        after = text[m.end():m.end() + 500]
        name_match = re.search(r'"([^"]+\.xlsx)"', after)
        if name_match:
            fname = name_match.group(1)
            if "Microsoft" not in fname and "Shared" not in fname:
                seen.add(fid)
                files.append((fid, fname))
    return files


def _check_sheet_has_data(sheet_url, date_str):
    """Check if a Google Sheets export has rows for the given date.
    Returns (found: bool, row_count: int).
    """
    import requests
    from io import BytesIO
    from openpyxl import load_workbook
    try:
        r = requests.get(sheet_url, allow_redirects=True, timeout=60)
        r.raise_for_status()
        wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
        # Check all sheets for matching date
        count = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, values_only=False):
                if not row:
                    continue
                cell_val = str(row[0].value or "").strip()
                if cell_val == date_str:
                    count += 1
        wb.close()
        return count > 0, count
    except Exception as e:
        log.warning(f"    ⚠ Sheet check failed: {e}")
        return False, 0


def check_drive_sources(sched, delivery_date_str):
    """Check if Drive/Sheet sources for a kho have data for the delivery date.
    
    Returns dict:
      {
        "ready": bool,        # True if all required sources have files
        "sources": {name: {"found": bool, "detail": str}}
      }
    """
    sources = sched.get("drive_sources", [])
    if not sources:
        return {"ready": False, "sources": {}}

    require_all = sched.get("require_all_sources", True)
    # Convert DD/MM/YYYY → DD.MM.YYYY for Drive file matching
    date_parts = delivery_date_str.split("/")
    date_dot = f"{date_parts[0]}.{date_parts[1]}.{date_parts[2]}" if len(date_parts) == 3 else delivery_date_str

    results = {}
    for src in sources:
        name = src["name"]
        src_type = src.get("type", "drive_folder")

        if src_type == "drive_folder":
            files = _list_drive_folder_files(src["url"])
            found = any(date_dot in fname for _, fname in files)
            matched_file = next((fname for _, fname in files if date_dot in fname), None)
            results[name] = {
                "found": found,
                "detail": matched_file or "not found",
            }
        elif src_type == "sheet":
            found, count = _check_sheet_has_data(src["url"], delivery_date_str)
            results[name] = {
                "found": found,
                "detail": f"{count} rows" if found else "no rows",
            }

    if require_all:
        all_ready = all(r["found"] for r in results.values())
    else:
        all_ready = any(r["found"] for r in results.values())

    return {"ready": all_ready, "sources": results}


# ══════════════════════════════════════════════════════════════════════
#  Data loading & change detection
# ══════════════════════════════════════════════════════════════════════

def get_current_week_info():
    """Determine current week number and start date based on delivery dates.
    
    Uses anchor: W14 starts 30/03/2026 (Monday). Each week = +7 days.
    The week is determined by the LATEST delivery date needed today
    (i.e., D+1 for most khos), so on Sunday 05/04, delivery is Mon 06/04
    which falls in W15.
    """
    # Anchor: W14 starts Monday 30/03/2026
    ANCHOR_WEEK = 14
    ANCHOR_START = datetime(2026, 3, 30)  # Monday
    
    # Use tomorrow (D+1) as reference since most mails are for D+1 delivery
    today = datetime.now()
    tomorrow = today + timedelta(days=1)
    
    # Calculate which week tomorrow falls in
    days_diff = (tomorrow - ANCHOR_START).days
    weeks_diff = days_diff // 7
    
    week_num = ANCHOR_WEEK + weeks_diff
    week_start = ANCHOR_START + timedelta(weeks=weeks_diff)
    
    return f"W{week_num}", format_date_vn(week_start)


def load_plan_data(week, kho, date_str, session=None):
    """Load filtered delivery data from weekly plan JSON.
    Returns list of row dicts, or empty list if not found.
    """
    json_path = os.path.join(OUTPUT, "state", f"weekly_plan_{week}.json")
    if not os.path.exists(json_path):
        return []

    with open(json_path, "r", encoding="utf-8") as f:
        plan = json.load(f)

    kho_key = kho.upper()
    if kho_key not in plan.get("data", {}):
        return []

    rows = plan["data"][kho_key]
    # Filter by date
    filtered = [r for r in rows if r["date"] == date_str]

    # DRY session filter (mirroring compose_mail.py logic)
    # Use gio_den first; if #N/A, fallback to gio_load to determine session
    if kho_key == "DRY" and session:
        SANG_START, SANG_END, TOI_WRAP = 6, 15, 3
        def parse_hour(s):
            if not s: return -1
            try: return int(s.split(":")[0])
            except: return -1

        def get_session_hour(r):
            h = parse_hour(r.get("gio_den", ""))
            if h == -1:
                h = parse_hour(r.get("gio_load", ""))
            return h

        if session == "sang":
            filtered = [r for r in filtered
                        if SANG_START <= get_session_hour(r) < SANG_END
                        or get_session_hour(r) == -1]
        elif session == "toi":
            filtered = [r for r in filtered
                        if get_session_hour(r) >= SANG_END
                        or (0 <= get_session_hour(r) < TOI_WRAP)]

    # Sort by store ID
    filtered.sort(key=lambda r: r["diem_den"])
    return filtered


def compute_data_hash(rows):
    """Hash all meaningful fields to detect any change."""
    normalized = sorted([
        (r.get("diem_den", ""), r.get("gio_den", ""), r.get("date", ""),
         r.get("loai_hang", ""))
        for r in rows
    ])
    return hashlib.md5(json.dumps(normalized, ensure_ascii=False).encode()).hexdigest()


def diff_data(old_rows, new_rows):
    """Compare old and new data, return structured diff.
    Returns dict with added_stores, removed_stores, time_changes.
    """
    old_by_store = {}
    for r in old_rows:
        key = r["diem_den"]
        old_by_store[key] = r.get("gio_den", "")

    new_by_store = {}
    for r in new_rows:
        key = r["diem_den"]
        new_by_store[key] = r.get("gio_den", "")

    added = sorted(set(new_by_store) - set(old_by_store))
    removed = sorted(set(old_by_store) - set(new_by_store))

    time_changes = []
    for store in sorted(set(old_by_store) & set(new_by_store)):
        old_time = old_by_store[store]
        new_time = new_by_store[store]
        if old_time != new_time:
            time_changes.append({
                "store": store,
                "old_time": old_time,
                "new_time": new_time,
            })

    return {
        "added": added,
        "removed": removed,
        "time_changes": time_changes,
        "has_changes": bool(added or removed or time_changes),
    }


def format_diff(diff):
    """Human-readable diff summary."""
    lines = []
    if diff["added"]:
        lines.append(f"  ➕ Thêm: {', '.join(diff['added'])}")
    if diff["removed"]:
        lines.append(f"  ➖ Bớt: {', '.join(diff['removed'])}")
    for tc in diff["time_changes"]:
        lines.append(f"  🔄 {tc['store']}: {tc['old_time']} → {tc['new_time']}")
    return "\n".join(lines) if lines else "  (không có thay đổi)"


# ══════════════════════════════════════════════════════════════════════
#  Fetch data
# ══════════════════════════════════════════════════════════════════════

def fetch_latest_data(week, week_start):
    """Re-fetch all data by running fetch_weekly_plan.py."""
    cmd = [
        sys.executable,
        os.path.join(BASE, "script", "domains", "performance", "fetch_weekly.py"),
        "--week", week,
        "--start", week_start,
    ]
    log.info("  🔄 Fetching latest data from sources...")
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True,
            encoding='utf-8', errors='replace', timeout=600
        )
        if result.returncode == 0:
            log.info("  ✅ Data fetched successfully")
            return True
        else:
            log.warning(f"  ⚠ Fetch returned code {result.returncode}")
            if result.stderr:
                log.warning(f"     {result.stderr[:200]}")
            return False
    except subprocess.TimeoutExpired:
        log.error("  ❌ Fetch timed out (10 min)")
        return False
    except Exception as e:
        log.error(f"  ❌ Fetch error: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════
#  Compose mail
# ══════════════════════════════════════════════════════════════════════

def compose_for_kho(kho, session, date_str, week):
    """Run compose_mail.py for a specific kho and return success + output."""
    cmd = [
        sys.executable,
        os.path.join(BASE, "script", "compose", "compose_mail.py"),
        "--kho", kho,
        "--date", date_str,
        "--week", week,
    ]
    if session:
        cmd.extend(["--session", session])

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True,
            encoding='utf-8', errors='replace', timeout=120
        )
        return result.returncode == 0, result.stdout
    except Exception as e:
        return False, str(e)


def save_kho_output(kho, session):
    """Copy generic output files to kho-specific files for reference."""
    suffix = f"_{kho}"
    if session:
        suffix += f"_{session}"

    files_to_copy = [
        ("_mail_body.html", f"_mail{suffix}_body.html"),
        ("_mail_inject.js", f"_mail{suffix}_inject.js"),
        ("_clip_html.ps1", f"_clip{suffix}_html.ps1"),
    ]

    MAIL_DIR = os.path.join(OUTPUT, "mail")
    os.makedirs(MAIL_DIR, exist_ok=True)
    for src_name, dst_name in files_to_copy:
        src = os.path.join(MAIL_DIR, src_name)
        dst = os.path.join(MAIL_DIR, dst_name)
        if os.path.exists(src):
            try:
                import shutil
                shutil.copy2(src, dst)
            except Exception:
                pass


def inject_to_haraworks(kho, session, date_str, week):
    """Call inject_haraworks.py to paste composed mail into Haraworks CKEditor.
    Returns (success, output_text). NEVER sends — only creates draft.
    """
    cmd = [
        sys.executable,
        os.path.join(BASE, "script", "compose", "inject_haraworks.py"),
        "--kho", kho,
        "--date", date_str,
        "--week", week,
    ]
    if session:
        cmd.extend(["--session", session])

    log.info(f"  🌐 Injecting into Haraworks...")
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True,
            encoding='utf-8', errors='replace', timeout=240
        )
        if result.returncode == 0:
            log.info(f"  ✅ Haraworks injection successful (draft saved)")
            return True, result.stdout
        else:
            log.warning(f"  ⚠ Haraworks injection failed (code {result.returncode})")
            if result.stdout:
                # Show last few lines of output for debugging
                last_lines = result.stdout.strip().split('\n')[-5:]
                for l in last_lines:
                    log.warning(f"     {l}")
            return False, result.stdout
    except subprocess.TimeoutExpired:
        log.error(f"  ❌ Haraworks injection timed out (4 min)")
        return False, "timeout"
    except Exception as e:
        log.error(f"  ❌ Haraworks injection error: {e}")
        return False, str(e)


def maybe_inject_final(kho, session, kho_state, delivery_str, week, dry_run=False, no_inject=False):
    """Inject into Haraworks ONLY when status just became 'final' and hasn't been injected yet.

    Business rule: only inject the FINAL compose (at cutoff or catch-up) to avoid
    updating Haraworks drafts with intermediate data that will change.
    """
    if kho_state.get("status") != "final":
        return kho_state
    if kho_state.get("injected"):
        return kho_state  # Already injected
    if dry_run or no_inject:
        return kho_state

    label = f"{kho}" + (f" {session.title()}" if session else "")
    inject_ok, _ = inject_to_haraworks(kho, session, delivery_str, week)
    if inject_ok:
        kho_state["injected"] = True
        kho_state["inject_time"] = datetime.now().strftime("%H:%M")
        notify(f"📧 {label} → Haraworks",
               f"{kho_state.get('rows', 0)} rows — draft saved, chờ review & gửi")
    else:
        log.warning(f"  ⚠ Inject failed — HTML đã copy clipboard, paste thủ công")
        notify(f"⚠ {label} — Ctrl+V paste thủ công",
               f"Inject CKEditor fail — HTML đã copy clipboard. Mở Haraworks reply → Ctrl+V")
    return kho_state


# ══════════════════════════════════════════════════════════════════════
#  Notification
# ══════════════════════════════════════════════════════════════════════

def notify(title, message, sound=True):
    """Windows notification: console + sound + toast."""
    log.info(f"  🔔 {title}: {message}")

    if sound:
        try:
            import winsound
            winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        except Exception:
            pass

    # Windows balloon tooltip notification
    try:
        ps_script = f'''
Add-Type -AssemblyName System.Windows.Forms
$notify = New-Object System.Windows.Forms.NotifyIcon
$notify.Icon = [System.Drawing.SystemIcons]::Information
$notify.BalloonTipIcon = 'Info'
$notify.BalloonTipTitle = '{title}'
$notify.BalloonTipText = '{message}'
$notify.Visible = $true
$notify.ShowBalloonTip(10000)
Start-Sleep -Seconds 5
$notify.Dispose()
'''
        subprocess.Popen(
            ["powershell", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
        )
    except Exception:
        pass  # Notification is nice-to-have, not critical


# ══════════════════════════════════════════════════════════════════════
#  Status display
# ══════════════════════════════════════════════════════════════════════

def print_status(state, config):
    """Print current compose status for today."""
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    today_state = state.get(today_str, {})

    print(f"\n{'='*60}")
    print(f"  📊 AUTO COMPOSE STATUS — {today_str} ({now.strftime('%H:%M')})")
    print(f"{'='*60}")

    for sched in config["schedules"]:
        kho = sched["kho"]
        session = sched.get("session")
        label = f"{kho}" + (f" {session.title()}" if session else "")
        key = state_key(kho, session)
        kho_state = today_state.get(key, {})

        # Time window status
        if not sched.get("enabled", True):
            window_status = "⏸ DISABLED"
        elif is_in_window(sched["check_time"], sched["cutoff_time"], now):
            window_status = f"🟢 ACTIVE ({sched['check_time']}-{sched['cutoff_time']})"
        elif is_past_cutoff(sched["cutoff_time"], now):
            window_status = f"🔴 PAST CUTOFF ({sched['cutoff_time']})"
        else:
            window_status = f"⏳ WAITING (starts {sched['check_time']})"

        # Compose status
        status = kho_state.get("status", "not_started")
        compose_count = kho_state.get("compose_count", 0)
        last_rows = kho_state.get("rows", 0)

        status_icons = {
            "not_started": "⬜",
            "waiting_data": "🟡",
            "composed": "🟢",
            "final": "✅",
            "skipped": "⏭",
            "backfill": "🔄",
        }
        icon = status_icons.get(status, "❓")

        print(f"\n  {icon} {label}")
        print(f"     Window: {window_status}")
        print(f"     Status: {status}")
        if compose_count > 0:
            print(f"     Composes: {compose_count}x | Last: {kho_state.get('last_compose', 'N/A')} | Rows: {last_rows}")
        if kho_state.get("last_change"):
            print(f"     Last change: {kho_state['last_change']}")

    print(f"\n{'='*60}")


def get_missed_schedules(config, state_today, now):
    """Find schedules past cutoff that need catch-up.
    
    Catches:
    - not_started: never ran at all (script wasn't running during window)
    - waiting_data: ran but no data was available
    - composed: ran and composed, but never got final compose before cutoff
    """
    missed = []
    for sched in config["schedules"]:
        if not sched.get("enabled", True):
            continue
        kho = sched["kho"]
        session = sched.get("session")
        key = state_key(kho, session)
        kho_state = state_today.get(key, {})
        status = kho_state.get("status", "not_started")

        # Catch up if past cutoff AND not yet finalized
        if status in ("not_started", "waiting_data", "composed"):
            if is_past_cutoff(sched["cutoff_time"], now):
                delivery_date = get_delivery_date(kho, session, now)
                skip_days = sched.get("skip_delivery_days", [])
                if not should_skip_delivery(kho, delivery_date, skip_days):
                    missed.append(sched)
    return missed


# ══════════════════════════════════════════════════════════════════════
#  Main orchestration
# ══════════════════════════════════════════════════════════════════════

def process_kho(sched, week, now, state_today, dry_run=False, force=False):
    """Process one warehouse schedule entry. Returns updated kho_state."""
    kho = sched["kho"]
    session = sched.get("session")
    label = f"{kho}" + (f" {session.title()}" if session else "")
    key = state_key(kho, session)

    kho_state = state_today.get(key, {
        "status": "not_started",
        "compose_count": 0,
        "rows": 0,
        "data_hash": "",
        "last_compose": "",
        "last_change": "",
        "prev_rows_snapshot": [],
    })

    # ── Determine delivery date ──
    delivery_date = get_delivery_date(kho, session, now)
    delivery_str = format_date_vn(delivery_date)

    # ── Check day-of-week exclusion ──
    skip_days = sched.get("skip_delivery_days", [])
    if should_skip_delivery(kho, delivery_date, skip_days):
        log.info(f"  ⏭ {label}: skipped — no delivery on {delivery_date.strftime('%A')}")
        kho_state["status"] = "skipped"
        return kho_state

    # ── Check time window (unless forced) ──
    if not force:
        if not is_in_window(sched["check_time"], sched["cutoff_time"], now):
            return kho_state  # Not in window, skip silently
        if kho_state.get("status") == "final":
            return kho_state  # Already finalized

    at_cutoff = is_at_cutoff(sched["cutoff_time"], now)

    log.info(f"\n  {'─'*45}")
    log.info(f"  📧 {label} — delivery {delivery_str}")
    if at_cutoff:
        log.info(f"  ⏰ APPROACHING CUTOFF ({sched['cutoff_time']})")

    # ── Load current data ──
    rows = load_plan_data(week, kho, delivery_str, session)

    if not rows:
        log.info(f"  ⚠ No data found for {delivery_str}")
        kho_state["status"] = "waiting_data"
        return kho_state

    # ── Compute hash & detect changes ──
    current_hash = compute_data_hash(rows)
    prev_hash = kho_state.get("data_hash", "")
    prev_rows = kho_state.get("prev_rows_snapshot", [])

    has_change = current_hash != prev_hash
    first_compose = kho_state["compose_count"] == 0

    # ── Report status ──
    na_count = sum(1 for r in rows if r.get("gio_den", "") in ["#N/A", "", None])
    log.info(f"  📊 {len(rows)} rows | {na_count} #N/A | Hash: {current_hash[:8]}")

    if has_change and prev_rows:
        diff = diff_data(prev_rows, rows)
        if diff["has_changes"]:
            log.info(f"  🔀 Changes detected:")
            log.info(format_diff(diff))
    elif has_change and first_compose:
        log.info(f"  🆕 First data load")

    # ── Decide: compose or skip ──
    should_compose = False
    reason = ""

    if first_compose and rows:
        should_compose = True
        reason = "first compose"
    elif has_change:
        should_compose = True
        reason = "data changed"
    elif at_cutoff and kho_state.get("status") != "final":
        should_compose = True
        reason = "cutoff final check"

    if not should_compose:
        log.info(f"  ✓ No changes since last compose ({kho_state.get('last_compose', 'N/A')})")
        return kho_state

    # ── Compose ──
    log.info(f"  📝 Composing... (reason: {reason})")

    if dry_run:
        log.info(f"  🏃 DRY RUN — would compose {len(rows)} rows")
        return kho_state

    success, output = compose_for_kho(kho, session, delivery_str, week)
    if success:
        # Save kho-specific output files
        save_kho_output(kho, session)

        kho_state["compose_count"] = kho_state.get("compose_count", 0) + 1
        kho_state["rows"] = len(rows)
        kho_state["data_hash"] = current_hash
        kho_state["last_compose"] = now.strftime("%H:%M")
        kho_state["prev_rows_snapshot"] = rows  # Save for next diff
        kho_state["na_count"] = na_count

        if has_change and not first_compose:
            kho_state["last_change"] = now.strftime("%H:%M")

        if at_cutoff:
            kho_state["status"] = "final"
            log.info(f"  ✅ FINAL compose — {len(rows)} rows")
            notify(f"📧 {label} FINAL",
                   f"{len(rows)} rows, delivery {delivery_str}")
        else:
            kho_state["status"] = "composed"
            log.info(f"  ✅ Composed #{kho_state['compose_count']} — {len(rows)} rows")
            if first_compose:
                notify(f"📧 {label} Ready",
                       f"{len(rows)} rows, delivery {delivery_str}")
            elif has_change:
                diff = diff_data(prev_rows, rows)
                changes_summary = []
                if diff["added"]: changes_summary.append(f"+{len(diff['added'])} stores")
                if diff["removed"]: changes_summary.append(f"-{len(diff['removed'])} stores")
                if diff["time_changes"]: changes_summary.append(f"{len(diff['time_changes'])} time changes")
                notify(f"📧 {label} Updated",
                       f"{', '.join(changes_summary)}")
    else:
        log.error(f"  ❌ Compose failed!")
        if output:
            log.error(f"     {output[:300]}")

    return kho_state


def main():
    parser = argparse.ArgumentParser(description="Auto compose mail scheduler")
    parser.add_argument("--dry-run", action="store_true",
                        help="Check data readiness without composing")
    parser.add_argument("--force", type=str, default=None,
                        help="Force compose for specific kho (e.g. KRC, DRY)")
    parser.add_argument("--force-session", type=str, default=None,
                        choices=["sang", "toi"],
                        help="Session for --force DRY")
    parser.add_argument("--status", action="store_true",
                        help="Show current compose status")
    parser.add_argument("--no-fetch", action="store_true",
                        help="Skip data re-fetch, use existing weekly plan JSON")
    parser.add_argument("--reset", action="store_true",
                        help="Reset today's state (start fresh)")
    parser.add_argument("--watch", action="store_true",
                        help="Watch mode: poll Drive folders, auto-compose when files appear")
    parser.add_argument("--poll-interval", type=int, default=None,
                        help="Poll interval in minutes for --watch mode (default from config)")
    parser.add_argument("--no-inject", action="store_true",
                        help="In watch mode, skip auto-inject into Haraworks (only compose HTML)")
    parser.add_argument("--no-auto-inject", action="store_true",
                        help="Skip auto-inject into Haraworks on FINAL compose")
    args = parser.parse_args()

    # ── Load config & state ──
    config = load_config()
    state = load_state()
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")

    # ── Status mode ──
    if args.status:
        print_status(state, config)
        return

    # ── Watch mode ──
    if args.watch:
        poll_min = args.poll_interval or config.get("global", {}).get("watch_poll_interval_min", 10)
        watch_loop(config, state, poll_min, dry_run=args.dry_run, no_inject=args.no_inject)
        return  # watch_loop runs forever

    # ── Reset mode ──
    if args.reset:
        if today_str in state:
            del state[today_str]
            save_state(state)
            log.info(f"🗑 Reset state for {today_str}")
        return

    # ── Ensure today's state exists ──
    if today_str not in state:
        state[today_str] = {}

    # ── Clean up old state (keep last 7 days) ──
    cutoff_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    old_keys = [k for k in state if k < cutoff_date and k != today_str]
    for k in old_keys:
        del state[k]

    # ── Determine week info ──
    week, week_start = get_current_week_info()

    log.info(f"\n{'═'*60}")
    log.info(f"  🤖 AUTO COMPOSE — {now.strftime('%Y-%m-%d %H:%M')}")
    log.info(f"  Week: {week} | Start: {week_start}")
    log.info(f"{'═'*60}")

    # ── Determine which schedules are active ──
    active = []
    for sched in config["schedules"]:
        if not sched.get("enabled", True):
            continue

        kho = sched["kho"]
        session = sched.get("session")
        key = state_key(kho, session)

        # Force mode: only process the forced kho
        if args.force:
            force_kho = args.force.upper()
            if force_kho == "DONG MAT": force_kho = "ĐÔNG MÁT"
            elif force_kho == "THIT CA": force_kho = "THỊT CÁ"
            if kho.upper() != force_kho:
                continue
            if args.force_session and session != args.force_session:
                continue
            active.append(sched)
            continue

        # Check delivery day exclusion
        delivery_date = get_delivery_date(kho, session, now)
        skip_days = sched.get("skip_delivery_days", [])
        if should_skip_delivery(kho, delivery_date, skip_days):
            continue

        # Check time window
        if is_in_window(sched["check_time"], sched["cutoff_time"], now):
            # Also check if already final
            kho_state = state[today_str].get(key, {})
            if kho_state.get("status") != "final":
                active.append(sched)

    # ── Check for missed khos (past cutoff, never composed) ──
    missed = []
    if not args.force:
        missed = get_missed_schedules(config, state[today_str], now)

    if not active and not missed:
        log.info(f"\n  ⏳ No warehouses active right now")

        # Show brief status of what's upcoming
        for sched in config["schedules"]:
            if not sched.get("enabled", True):
                continue
            kho = sched["kho"]
            session = sched.get("session")
            label = f"{kho}" + (f" {session.title()}" if session else "")
            key = state_key(kho, session)
            kho_state = state[today_str].get(key, {})
            status = kho_state.get("status", "not_started")

            if status == "final":
                log.info(f"     ✅ {label}: FINAL ({kho_state.get('rows', '?')} rows)")
            elif status == "backfill":
                log.info(f"     🔄 {label}: catch-up ({kho_state.get('rows', '?')} rows)")
            elif status == "skipped":
                log.info(f"     ⏭ {label}: skipped (no delivery)")
            elif is_past_cutoff(sched["cutoff_time"], now):
                log.info(f"     🔴 {label}: past cutoff ({sched['cutoff_time']})")
            else:
                log.info(f"     ⏳ {label}: waiting (starts {sched['check_time']})")

        save_state(state)
        return

    # ── Fetch latest data (once for all active + missed warehouses) ──
    if not args.no_fetch and not args.dry_run:
        fetch_success = fetch_latest_data(week, week_start)
        if not fetch_success:
            log.warning("  ⚠ Data fetch had issues, proceeding with existing data")

    # ── Process each active kho ──
    for sched in active:
        kho = sched["kho"]
        session = sched.get("session")
        key = state_key(kho, session)

        kho_state = process_kho(
            sched, week, now, state[today_str],
            dry_run=args.dry_run,
            force=bool(args.force),
        )
        # Auto-inject into Haraworks on FINAL compose
        delivery_date = get_delivery_date(kho, session, now)
        delivery_str = format_date_vn(delivery_date)
        kho_state = maybe_inject_final(kho, session, kho_state, delivery_str, week,
                                        dry_run=args.dry_run, no_inject=args.no_auto_inject)
        state[today_str][key] = kho_state

    # ── Catch-up: compose missed khos (past cutoff, not yet final) ──
    if missed:
        missed_labels = [
            f"{s['kho']}" + (f" {s.get('session', '').title()}" if s.get('session') else "")
            for s in missed
        ]
        log.info(f"\n  🔄 CATCH-UP: {len(missed)} missed — {', '.join(missed_labels)}")

        for sched in missed:
            kho = sched["kho"]
            session = sched.get("session")
            key = state_key(kho, session)

            kho_state = process_kho(
                sched, week, now, state[today_str],
                dry_run=args.dry_run,
                force=True,  # Bypass time window
            )
            # Past cutoff catch-up → mark as final (this IS the final compose)
            if kho_state.get("status") in ("composed", "backfill"):
                kho_state["status"] = "final"
            # Auto-inject into Haraworks on FINAL compose
            delivery_date = get_delivery_date(kho, session, now)
            delivery_str = format_date_vn(delivery_date)
            kho_state = maybe_inject_final(kho, session, kho_state, delivery_str, week,
                                            dry_run=args.dry_run, no_inject=args.no_auto_inject)
            state[today_str][key] = kho_state

    # ── Save state ──
    save_state(state)

    # ── Summary ──
    log.info(f"\n{'═'*60}")
    log.info(f"  📋 Summary:")
    for sched in config["schedules"]:
        kho = sched["kho"]
        session = sched.get("session")
        label = f"{kho}" + (f" {session.title()}" if session else "")
        key = state_key(kho, session)
        kho_state = state[today_str].get(key, {})
        status = kho_state.get("status", "not_started")
        rows = kho_state.get("rows", 0)
        count = kho_state.get("compose_count", 0)
        if status == "final":
            log.info(f"     ✅ {label}: FINAL — {rows} rows (composed {count}x)")
        elif status == "backfill":
            log.info(f"     🔄 {label}: catch-up — {rows} rows (composed {count}x)")
        elif status == "composed":
            log.info(f"     🟢 {label}: composed — {rows} rows (composed {count}x)")
        elif status == "skipped":
            log.info(f"     ⏭ {label}: skipped")
        elif status == "waiting_data":
            log.info(f"     🟡 {label}: waiting for data")
        else:
            log.info(f"     ⬜ {label}: {status}")
    log.info(f"{'═'*60}")


# ══════════════════════════════════════════════════════════════════════
#  Instance locking — prevent multiple watch processes
# ══════════════════════════════════════════════════════════════════════

def _is_pid_alive(pid):
    """Check if a process with given PID is still running (Windows)."""
    try:
        result = subprocess.run(
            ["tasklist", "/FI", f"PID eq {pid}", "/NH"],
            capture_output=True, text=True, timeout=5
        )
        return str(pid) in result.stdout
    except Exception:
        return False


def acquire_lock():
    """Try to acquire the lock file. Returns True if acquired, False if another instance is running."""
    os.makedirs(os.path.dirname(LOCK_PATH), exist_ok=True)
    
    if os.path.exists(LOCK_PATH):
        try:
            with open(LOCK_PATH, "r") as f:
                old_pid = int(f.read().strip())
            if _is_pid_alive(old_pid):
                return False  # Another instance is running
            # Stale lock — process died without cleanup
            log.info(f"  🧹 Stale lock (PID {old_pid} dead), taking over")
        except (ValueError, OSError):
            pass  # Corrupted lock file, overwrite

    # Write our PID
    with open(LOCK_PATH, "w") as f:
        f.write(str(os.getpid()))
    return True


def release_lock():
    """Release the lock file."""
    try:
        if os.path.exists(LOCK_PATH):
            with open(LOCK_PATH, "r") as f:
                pid = int(f.read().strip())
            if pid == os.getpid():
                os.remove(LOCK_PATH)
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════
#  Watch mode — poll Drive folders and auto-compose
# ══════════════════════════════════════════════════════════════════════

def watch_loop(config, state, poll_interval_min, dry_run=False, no_inject=False):
    """Continuously poll Google Drive folders for KH files.
    When files for D+1 appear → fetch data → compose mail → inject into Haraworks.
    NEVER sends mail — only creates draft in Haraworks.
    """
    # ── Instance lock — only one watch process at a time ──
    if not acquire_lock():
        log.info(f"  ⏭ Another watch instance is already running (lock: {LOCK_PATH}). Exiting.")
        return
    atexit.register(release_lock)

    log.info(f"\n{'═'*60}")
    log.info(f"  👁 WATCH MODE — polling every {poll_interval_min} min (PID {os.getpid()})")
    if no_inject:
        log.info(f"  ⚠ --no-inject: will NOT auto-inject into Haraworks")
    else:
        log.info(f"  🌐 Auto-inject into Haraworks: ENABLED")
    log.info(f"  ⛔ Auto-send: DISABLED (chỉ tạo draft)")
    log.info(f"  Press Ctrl+C to stop")
    log.info(f"{'═'*60}")

    poll_count = 0

    while True:
        try:
            poll_count += 1
            # Reload state from disk to sync with scheduled runs
            state = load_state()
            now = datetime.now()
            today_str = now.strftime("%Y-%m-%d")

            # Ensure today's state
            if today_str not in state:
                state[today_str] = {}

            # Clean old state (keep 7 days)
            cutoff_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
            for k in [k for k in state if k < cutoff_date and k != today_str]:
                del state[k]

            week, week_start = get_current_week_info()

            log.info(f"\n  ───── Poll #{poll_count} — {now.strftime('%H:%M:%S')} ─────")

            any_composed = False

            for sched in config["schedules"]:
                if not sched.get("enabled", True):
                    continue

                kho = sched["kho"]
                session = sched.get("session")
                label = f"{kho}" + (f" {session.title()}" if session else "")
                key = state_key(kho, session)
                watch_key = f"{key}_watch"

                # Get kho state
                kho_state = state[today_str].get(key, {
                    "status": "not_started",
                    "compose_count": 0,
                    "rows": 0,
                    "data_hash": "",
                    "prev_rows_snapshot": [],
                })
                watch_state = state[today_str].get(watch_key, {
                    "files_detected": False,
                    "last_check": "",
                    "sources_status": {},
                })

                # Skip if already final
                if kho_state.get("status") == "final":
                    continue

                # Check delivery day exclusion
                delivery_date = get_delivery_date(kho, session, now)
                skip_days = sched.get("skip_delivery_days", [])
                if should_skip_delivery(kho, delivery_date, skip_days):
                    continue

                delivery_str = format_date_vn(delivery_date)

                # Check Drive sources
                drive_sources = sched.get("drive_sources", [])
                if not drive_sources:
                    continue

                log.info(f"  🔍 {label} — checking sources for {delivery_str}...")
                result = check_drive_sources(sched, delivery_str)

                for src_name, src_result in result["sources"].items():
                    icon = "✅" if src_result["found"] else "❌"
                    log.info(f"     {icon} {src_name}: {src_result['detail']}")

                watch_state["last_check"] = now.strftime("%H:%M:%S")
                watch_state["sources_status"] = {
                    name: info["found"] for name, info in result["sources"].items()
                }

                if not result["ready"]:
                    log.info(f"     ⏳ Chưa đủ file — chờ poll tiếp")
                    state[today_str][watch_key] = watch_state
                    continue

                # Files detected!
                was_detected_before = watch_state.get("files_detected", False)
                watch_state["files_detected"] = True
                state[today_str][watch_key] = watch_state

                if not was_detected_before:
                    log.info(f"  🆕 {label}: ĐỦ FILE! Fetching & composing...")
                else:
                    # Already detected before — re-check for data changes
                    log.info(f"  🔄 {label}: Re-checking for changes...")

                # Fetch data
                if not dry_run:
                    fetch_success = fetch_latest_data(week, week_start)
                    if not fetch_success:
                        log.warning(f"  ⚠ Fetch failed for {label}, will retry next poll")
                        continue

                    # Compose
                    kho_state = process_kho(
                        sched, week, now, state[today_str],
                        dry_run=False,
                        force=True,  # bypass time window
                    )
                    state[today_str][key] = kho_state

                    if kho_state.get("status") in ("composed", "final"):
                        any_composed = True
                        log.info(f"  ✅ {label}: Composed! {kho_state.get('rows', 0)} rows")
                        log.info(f"     → HTML: output/_mail_{kho}{'_' + session if session else ''}_body.html")

                        # Auto-inject into Haraworks ONLY on FINAL compose
                        kho_state = maybe_inject_final(
                            kho, session, kho_state, delivery_str, week,
                            dry_run=dry_run, no_inject=no_inject
                        )
                        state[today_str][key] = kho_state
                        save_state(state)  # Save immediately after inject
                else:
                    log.info(f"  🏃 DRY RUN — would fetch & compose {label}")

            # Save state after each poll
            save_state(state)

            if any_composed:
                log.info(f"\n  🔔 Mail(s) composed! Check output/*.html")

            # Sleep until next poll
            log.info(f"  💤 Next poll in {poll_interval_min} min...")
            time.sleep(poll_interval_min * 60)

        except KeyboardInterrupt:
            log.info(f"\n  🛑 Watch mode stopped by user")
            save_state(state)
            release_lock()
            break
        except Exception as e:
            log.error(f"  ❌ Watch error: {e}")
            import traceback
            log.error(traceback.format_exc())
            # Continue watching even on error
            log.info(f"  💤 Retrying in {poll_interval_min} min...")
            time.sleep(poll_interval_min * 60)


if __name__ == "__main__":
    main()
