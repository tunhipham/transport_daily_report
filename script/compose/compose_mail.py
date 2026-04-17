"""
compose_mail.py - Generate HTML email body for delivery schedule mails
and copy to Windows clipboard (with Vietnamese diacritics).

Usage:
  python script/compose_mail.py --kho KRC --date 01/04/2026
  python script/compose_mail.py --kho DRY --session sang --date 01/04/2026
  python script/compose_mail.py --kho DRY --session toi --date 31/03/2026
  python script/compose_mail.py --kho "ĐÔNG MÁT" --date 01/04/2026
  python script/compose_mail.py --kho "THỊT CÁ" --date 01/04/2026

Rules:
  - Mail is for D+1 schedule (except DRY Tối which is same-day D evening)
  - KSL (DRY) splits into 2 emails: Sáng (<18h) and Tối (>=18h)
  - Each week = 1 email thread, daily replies into that thread
  - Store IDs (điểm đến) sorted A→Z
  - Output: complete HTML ready to paste into CKEditor
"""
import os, sys, json, argparse, subprocess, re
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTPUT = os.path.join(BASE, "output")

# ── Time boundary for DRY sáng/tối split ──
# Sáng: 6h → 15h (hour 6-14)
# Tối:  15h → 3h sáng hôm sau (hour 15-23, 0-2)
SANG_START = 6
SANG_END = 15   # exclusive: sáng is [6, 15)
TOI_WRAP = 3    # tối includes 0:00 - 2:59 (early morning = still tối)


def parse_time_hour(gio_str):
    """Parse time string like '3:05' or '22:30' and return hour as int."""
    if not gio_str:
        return -1
    parts = gio_str.split(":")
    try:
        return int(parts[0])
    except (ValueError, IndexError):
        return -1


def _format_time_hhmm(gio_str):
    """Ensure time is formatted as HH:MM (e.g. 0:17 → 00:17)."""
    if not gio_str:
        return gio_str
    parts = gio_str.split(":")
    if len(parts) == 2:
        try:
            h, m = int(parts[0]), parts[1]
            return f"{h:02d}:{m}"
        except (ValueError, IndexError):
            pass
    return gio_str


# ── Inventory schedule (kiểm kê) for DRY ──
INVENTORY_SHEET_URL = "https://docs.google.com/spreadsheets/d/1KIXDqGDW60sKNXuHOriT8utPTyhV-pCy11jlf18Zz-0/export?format=xlsx"


def fetch_inventory_schedule():
    """Fetch inventory dates from Google Sheets. Returns dict: store_id → inventory_date (datetime)."""
    import requests
    from io import BytesIO
    from openpyxl import load_workbook

    try:
        r = requests.get(INVENTORY_SHEET_URL, allow_redirects=True, timeout=60)
        r.raise_for_status()
        wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
        ws = wb['Lịch Kiểm kê 2026']

        inventory = {}  # store_id → datetime
        for row in ws.iter_rows(min_row=10, values_only=False):
            store_id = str(row[3].value or "").strip()  # Col D = ID Mart
            kiem_ke = row[7].value  # Col H = Ngày kiểm kê tổng 2026
            if store_id and kiem_ke:
                if hasattr(kiem_ke, 'date'):
                    inventory[store_id] = kiem_ke
                elif isinstance(kiem_ke, str):
                    # Try parsing DD/MM/YYYY
                    try:
                        inventory[store_id] = datetime.strptime(kiem_ke, "%d/%m/%Y")
                    except ValueError:
                        pass
        wb.close()
        print(f"  📋 Loaded {len(inventory)} stores from inventory schedule")
        return inventory
    except Exception as e:
        print(f"  ⚠ Could not fetch inventory schedule: {e}")
        return {}


def get_inventory_flagged_stores(inventory, delivery_date_str):
    """Return set of store IDs whose delivery date falls on inventory D or D-1.
    
    Rule: stores with inventory on date X should not receive goods on X (D) and X-1 (D-1).
    So if delivery_date matches X or X-1 for any store, flag that store.
    """
    flagged = set()
    # Parse delivery date
    try:
        parts = delivery_date_str.split("/")
        delivery_dt = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
    except (ValueError, IndexError):
        return flagged

    for store_id, inv_dt in inventory.items():
        inv_date = inv_dt.date() if hasattr(inv_dt, 'date') else inv_dt
        d_date = delivery_dt.date() if hasattr(delivery_dt, 'date') else delivery_dt
        if isinstance(inv_date, datetime):
            inv_date = inv_date.date()
        if isinstance(d_date, datetime):
            d_date = d_date.date()
        # Flag if delivery is on D (inventory day) or D-1 (day before)
        if d_date == inv_date or d_date == inv_date - timedelta(days=1):
            flagged.add(store_id)

    return flagged


def load_data(week, kho, date_str, session=None):
    """Load and filter data from weekly plan JSON."""
    json_path = os.path.join(OUTPUT, "state", f"weekly_plan_{week}.json")
    if not os.path.exists(json_path):
        print(f"❌ File not found: {json_path}")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        plan = json.load(f)

    kho_key = kho.upper()
    if kho_key not in plan["data"]:
        print(f"❌ Kho '{kho_key}' not found. Available: {list(plan['data'].keys())}")
        sys.exit(1)

    rows = plan["data"][kho_key]

    # Filter by date
    filtered = [r for r in rows if r["date"] == date_str]

    # For DRY, filter by session (sáng/tối)
    # Use gio_den first; if #N/A, fallback to gio_load to determine session
    if kho_key == "DRY" and session:
        def _get_session_hour(r):
            """Return the best available hour for session classification."""
            h = parse_time_hour(r["gio_den"])
            if h == -1:
                # gio_den is #N/A — use gio_load as fallback
                h = parse_time_hour(r.get("gio_load", ""))
            return h

        if session == "sang":
            # Sáng: 6:00 → 14:59, PLUS rows with no time info at all
            filtered = [r for r in filtered
                        if SANG_START <= _get_session_hour(r) < SANG_END
                        or _get_session_hour(r) == -1]
        elif session == "toi":
            # Tối: 15:00 → 2:59 next day (hour >= 15 OR hour < 3), must have valid time
            filtered = [r for r in filtered
                        if _get_session_hour(r) >= SANG_END
                        or (0 <= _get_session_hour(r) < TOI_WRAP)]

    # Sort by điểm đến (store ID) A→Z
    filtered.sort(key=lambda r: r["diem_den"])

    return filtered, plan


def generate_html_krc(rows, date_str):
    """Generate HTML for KRC warehouse email."""
    body = f"""<p>Dear team ST,</p>
<br>
<p>SCM gửi thông tin kế hoạch giao hàng KHO RCQ ngày {date_str}.</p>
<br>
<p>Em cảm ơn!</p>
<br>"""

    table = _make_table(rows, ["Ngày", "Điểm đến", "Giờ đến dự kiến (+-30')"])
    return body + table


def generate_html_dry(rows, date_str, session, flagged_stores=None):
    """Generate HTML for DRY (KSL) warehouse email."""
    session_label = "Sáng" if session == "sang" else "Tối"
    body = f"""<p>Dear team Siêu Thị,</p>
<br>
<p>SCM gửi thông tin kế hoạch giao hàng DC Dry {session_label} {date_str}.</p>
<br>
<p>Em cảm ơn ạ.</p>
<br>"""

    table = _make_table(rows, ["Ngày giao hàng", "Điểm đến", "Giờ đến dự kiến (+-30')"], flagged_stores=flagged_stores)
    return body + table


def generate_html_dong_mat(rows, date_str):
    """Generate HTML for ĐÔNG MÁT warehouse email (has extra 'loại hàng' column)."""
    body = f"""<p>Dear team Siêu Thị,</p>
<br>
<p>SCM gửi thông tin kế hoạch giao hàng Đông Mát ngày {date_str}.</p>
<br>
<p>Em cảm ơn ạ.</p>
<br>"""

    table = _make_table_dong_mat(rows)
    return body + table


def generate_html_thit_ca(rows, date_str):
    """Generate HTML for THỊT CÁ warehouse email."""
    body = f"""<p>Dear team Siêu Thị,</p>
<br>
<p>SCM gửi thông tin kế hoạch giao hàng Thịt Cá ngày {date_str}.</p>
<br>
<p>Em cảm ơn ạ.</p>
<br>"""

    table = _make_table(rows, ["Ngày", "Điểm đến", "Giờ đến dự kiến (+-30')"])
    return body + table


def _normalize_date(date_str):
    """Ensure date is DD/MM/YYYY with leading zeros."""
    parts = date_str.split("/")
    if len(parts) == 3:
        d, m, y = parts
        return f"{int(d):02d}/{int(m):02d}/{y}"
    return date_str


def _safe_date(date_str):
    """Wrap date in span to prevent CKEditor from auto-parsing/reformatting."""
    normalized = _normalize_date(date_str)
    return f'<span style="white-space:nowrap">{normalized}</span>'


def _make_table(rows, headers, flagged_stores=None):
    """Generate an HTML table with standard 3-column format.
    flagged_stores: optional set of store IDs to highlight red (inventory conflict).
    """
    style = 'style="border-collapse:collapse;border:1px solid #000;font-family:Arial,sans-serif;font-size:12px"'
    td_style_base = 'border:1px solid #000;padding:4px 8px;text-align:center'
    th_style = 'style="border:1px solid #000;padding:4px 8px;text-align:center;background:#4472C4;color:white;font-weight:bold"'

    html = f'<table {style}>\n<thead><tr>'
    for h in headers:
        html += f'<th {th_style}>{h}</th>'
    html += '</tr></thead>\n<tbody>\n'

    for i, r in enumerate(rows):
        store_id = r["diem_den"]
        is_flagged = flagged_stores and store_id in flagged_stores

        if is_flagged:
            # Red highlight for inventory conflict
            row_td = f'style="{td_style_base};background:#FF6B6B;color:#000;font-weight:bold"'
        elif i % 2 == 0:
            row_td = f'style="{td_style_base};background:#D9E2F3"'
        else:
            row_td = f'style="{td_style_base}"'

        gio = _format_time_hhmm(r["gio_den"])
        html += f'<tr><td {row_td}>{_safe_date(r["date"])}</td><td {row_td}>{store_id}</td><td {row_td}>{gio}</td></tr>\n'

    html += '</tbody></table>'
    return html


def _make_table_dong_mat(rows):
    """Generate HTML table for ĐÔNG MÁT (4 columns with loại hàng)."""
    headers = ["Ngày", "Điểm đến", "Giờ đến dự kiến (+-30')", "Loại hàng"]
    style = 'style="border-collapse:collapse;border:1px solid #000;font-family:Arial,sans-serif;font-size:12px"'
    td_style_base = 'border:1px solid #000;padding:4px 8px;text-align:center'
    th_style = 'style="border:1px solid #000;padding:4px 8px;text-align:center;background:#4472C4;color:white;font-weight:bold"'

    html = f'<table {style}>\n<thead><tr>'
    for h in headers:
        html += f'<th {th_style}>{h}</th>'
    html += '</tr></thead>\n<tbody>\n'

    for i, r in enumerate(rows):
        if i % 2 == 0:
            row_td = f'style="{td_style_base};background:#D9E2F3"'
        else:
            row_td = f'style="{td_style_base}"'

        loai = r.get("loai_hang", "")
        gio = _format_time_hhmm(r["gio_den"])
        html += f'<tr><td {row_td}>{_safe_date(r["date"])}</td><td {row_td}>{r["diem_den"]}</td><td {row_td}>{gio}</td><td {row_td}>{loai}</td></tr>\n'

    html += '</tbody></table>'
    return html


def copy_html_to_clipboard(html_content, kho=None, session=None):
    """Copy HTML to Windows clipboard using PowerShell.
    
    Saves BOTH generic (_mail_body.html) and kho-specific files
    (e.g. _mail_KRC_body.html) so inject_haraworks.py can pick
    the correct file per warehouse even when multiple kho are
    composed before injecting.
    """
    # Write HTML to generic file (backward compat)
    MAIL_DIR = os.path.join(OUTPUT, "mail")
    os.makedirs(MAIL_DIR, exist_ok=True)
    tmp_path = os.path.join(MAIL_DIR, "_mail_body.html")
    with open(tmp_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    # Also write kho-specific file so inject picks the right one
    if kho:
        suffix = f"_{kho}"
        if session:
            suffix += f"_{session}"
        kho_path = os.path.join(MAIL_DIR, f"_mail{suffix}_body.html")
        with open(kho_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        # Generate kho-specific clip script for CF_HTML clipboard paste
        _generate_clip_ps1(kho_path, kho, session)

    # Also generate a JS snippet that can inject into CKEditor
    js_str = json.dumps(html_content)
    js_code = f"""
// Paste this in browser console (F12) on the Haraworks compose/reply page:
(function() {{
    var html = {js_str};
    var injected = false;

    // Method 1: CKEditor 5 (via __ckeditorInstance on the editable element)
    var editable = document.querySelector('.ck-editor__editable, [role="textbox"][contenteditable="true"]');
    if (editable) {{
        var ck5 = editable.ckeditorInstance;
        if (ck5 && ck5.setData) {{
            ck5.setData(html);
            injected = true;
        }}
    }}

    // Method 2: CKEditor 4 (CKEDITOR global)
    if (!injected && typeof CKEDITOR !== 'undefined') {{
        var instances = CKEDITOR.instances;
        var keys = Object.keys(instances);
        if (keys.length > 0) {{
            instances[keys[0]].setData(html);
            injected = true;
        }}
    }}

    // Method 3: Fallback — innerHTML + force model sync
    if (!injected && editable) {{
        editable.innerHTML = html;
        // Trigger multiple events to force CKEditor to sync its model
        ['input', 'change', 'keyup'].forEach(function(evt) {{
            editable.dispatchEvent(new Event(evt, {{ bubbles: true }}));
        }});
        injected = true;
    }}

    document.title = injected ? 'INJECTED_OK' : 'ERROR: No editor found';
}})();
"""

    js_path = os.path.join(MAIL_DIR, "_mail_inject.js")
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(js_code)

    # Copy plain text to clipboard (fallback)
    try:
        process = subprocess.Popen(
            ["powershell", "-Command", f"Get-Content '{tmp_path}' -Raw | Set-Clipboard"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE
        )
        process.wait(timeout=10)
    except Exception as e:
        print(f"  ⚠ Clipboard copy warning: {e}")

    kho_label = f" ({kho}" + (f" {session}" if session else "") + ")" if kho else ""
    print(f"\n📋 Files generated{kho_label}:")
    print(f"   HTML body: {tmp_path}")
    if kho:
        print(f"   Kho-specific: {kho_path}")
    print(f"   JS inject: {js_path}")
    print(f"   (HTML also copied to clipboard)")


def _generate_clip_ps1(html_path, kho, session=None):
    """Generate a PowerShell script to copy kho-specific HTML as CF_HTML to clipboard."""
    suffix = f"_{kho}"
    if session:
        suffix += f"_{session}"
    MAIL_DIR = os.path.join(OUTPUT, "mail")
    os.makedirs(MAIL_DIR, exist_ok=True)
    ps1_path = os.path.join(MAIL_DIR, f"_clip{suffix}_html.ps1")

    # Escape backslashes for PowerShell
    html_path_escaped = html_path.replace("\\", "\\\\")

    ps1_content = '''# Auto-generated clip script for ''' + kho + (f" {session}" if session else "") + '''
$htmlFile = "''' + html_path_escaped + '''"
Write-Host "Copying: $htmlFile"
Add-Type -AssemblyName System.Windows.Forms
$html = Get-Content -Path $htmlFile -Raw -Encoding UTF8
$prefix = "<html><body><!--StartFragment-->"
$suffix = "<!--EndFragment--></body></html>"
$fullContent = $prefix + $html + $suffix
$startHtml = 105
$startFragment = $startHtml + $prefix.Length
$endFragment = $startFragment + [System.Text.Encoding]::UTF8.GetByteCount($html)
$endHtml = $endFragment + $suffix.Length
$header = "Version:0.9`r`nStartHTML:{0:D10}`r`nEndHTML:{1:D10}`r`nStartFragment:{2:D10}`r`nEndFragment:{3:D10}`r`n" -f $startHtml, $endHtml, $startFragment, $endFragment
$fullClip = $header + $fullContent
$dataObj = New-Object System.Windows.Forms.DataObject
$ms = New-Object System.IO.MemoryStream
$bytes = [System.Text.Encoding]::UTF8.GetBytes($fullClip)
$ms.Write($bytes, 0, $bytes.Length)
$ms.WriteByte(0)
$ms.Position = 0
$dataObj.SetData("HTML Format", $ms)
$dataObj.SetData("Text", $html)
[System.Windows.Forms.Clipboard]::SetDataObject($dataObj, $true)
Write-Host "HTML copied to clipboard in HTML format!"
'''
    with open(ps1_path, "w", encoding="utf-8") as f:
        f.write(ps1_content)


def main():
    parser = argparse.ArgumentParser(description="Generate delivery schedule email HTML")
    parser.add_argument("--kho", required=True, help="Warehouse: KRC, DRY, 'ĐÔNG MÁT', 'THỊT CÁ'")
    parser.add_argument("--date", required=True, help="Date for schedule: DD/MM/YYYY")
    parser.add_argument("--session", choices=["sang", "toi"], help="Session for DRY: sang or toi")
    parser.add_argument("--week", default=None, help="Week number: W14 (auto-detect if not specified)")
    args = parser.parse_args()

    # Auto-detect week from available files
    if not args.week:
        state_dir = os.path.join(OUTPUT, "state")
        os.makedirs(state_dir, exist_ok=True)
        files = [f for f in os.listdir(state_dir) if f.startswith("weekly_plan_W") and f.endswith(".json")]
        if files:
            files.sort()
            args.week = files[-1].replace("weekly_plan_", "").replace(".json", "")
            print(f"  Auto-detected week: {args.week}")
        else:
            print("❌ No weekly plan JSON found. Run fetch_weekly_plan.py first.")
            sys.exit(1)

    kho = args.kho.upper()
    if kho == "DONG MAT":
        kho = "ĐÔNG MÁT"
    elif kho == "THIT CA":
        kho = "THỊT CÁ"

    # Validate DRY requires session
    if kho == "DRY" and not args.session:
        print("❌ DRY requires --session sang or --session toi")
        sys.exit(1)

    print("=" * 60)
    print(f"  📧 COMPOSE MAIL — {kho}")
    if kho == "DRY":
        print(f"  Session: {'Sáng' if args.session == 'sang' else 'Tối'}")
    print(f"  Date: {args.date}")
    print(f"  Week: {args.week}")
    print("=" * 60)

    rows, plan = load_data(args.week, kho, args.date, args.session)
    print(f"\n  📊 {len(rows)} rows (sorted A→Z by store ID)")

    if not rows:
        print(f"\n  ⚠ No data found for {kho} on {args.date}" +
              (f" ({args.session})" if args.session else ""))
        print(f"  Available dates: {sorted(set(r['date'] for r in plan['data'].get(kho, [])))}")
        sys.exit(1)

    # Preview first/last rows
    print(f"  First: {rows[0]['diem_den']} @ {rows[0]['gio_den']}")
    print(f"  Last:  {rows[-1]['diem_den']} @ {rows[-1]['gio_den']}")

    # For DRY: fetch inventory schedule and check conflicts
    flagged_stores = None
    if kho == "DRY":
        print("\n  🔍 Checking inventory schedule (kiểm kê)...")
        inventory = fetch_inventory_schedule()
        if inventory:
            flagged_stores = get_inventory_flagged_stores(inventory, args.date)
            if flagged_stores:
                print(f"  🔴 {len(flagged_stores)} stores flagged (delivery on kiểm kê D or D-1):")
                for sid in sorted(flagged_stores):
                    inv_dt = inventory.get(sid)
                    inv_str = inv_dt.strftime('%d/%m/%Y') if hasattr(inv_dt, 'strftime') else str(inv_dt)
                    print(f"     ⚠ {sid} — kiểm kê: {inv_str}")
            else:
                print("  ✅ No inventory conflicts found")

    # Generate HTML
    if kho == "KRC":
        html = generate_html_krc(rows, args.date)
    elif kho == "DRY":
        html = generate_html_dry(rows, args.date, args.session, flagged_stores=flagged_stores)
    elif kho == "ĐÔNG MÁT":
        html = generate_html_dong_mat(rows, args.date)
    elif kho == "THỊT CÁ":
        html = generate_html_thit_ca(rows, args.date)
    else:
        print(f"❌ Unknown kho: {kho}")
        sys.exit(1)

    copy_html_to_clipboard(html, kho=kho, session=args.session)

    print(f"\n✅ Email body ready!")
    print(f"   → Use JS inject in browser console to paste into CKEditor")
    print(f"   → Or copy HTML from clipboard and paste")
    print("=" * 60)


if __name__ == "__main__":
    main()
