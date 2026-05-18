# -*- coding: utf-8 -*-
"""
Extract transfer data (phiếu chuyển) from StarRocks.
Table: kfm_scm.kf_transfer_items
Fallback: local xlsx files
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from data_pipeline.extractors._base import BaseExtractor
from data_pipeline.config import load_starrocks_config


class StarrocksTransferExtractor(BaseExtractor):
    name = "transfer"

    def extract(self) -> dict:
        """Extract transfer items for the target date from StarRocks."""
        # Parse date from DDMMYYYY tag
        dd = self.date_tag[:2]
        mm = self.date_tag[2:4]
        yyyy = self.date_tag[4:]
        date_str = f"{dd}/{mm}/{yyyy}"
        date_iso = f"{yyyy}-{mm}-{dd}"

        try:
            return self._extract_from_db(date_iso, date_str)
        except Exception as e:
            print(f"    ⚠ StarRocks failed: {e}")
            print(f"    ↪ Falling back to local xlsx...")
            return self._extract_from_file(date_str)

    def _extract_from_db(self, date_iso: str, date_str: str) -> dict:
        """Query StarRocks for transfer items."""
        import pymysql

        cfg = load_starrocks_config()
        conn = pymysql.connect(
            host=cfg["host"],
            port=cfg["port"],
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            charset="utf8mb4",
            connect_timeout=30,
            read_timeout=60,
        )

        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cur:
                # Query transfer items for the target date
                sql = """
                    SELECT 
                        ngay_chuyen AS ngay,
                        kho_xuat,
                        kho_nhan,
                        ma_hang,
                        ten_hang,
                        so_luong,
                        trong_luong,
                        don_vi
                    FROM kf_transfer_items
                    WHERE DATE(ngay_chuyen) = %s
                    ORDER BY kho_xuat, ma_hang
                """
                cur.execute(sql, (date_iso,))
                raw_rows = cur.fetchall()

            # Serialize datetime objects
            rows = []
            for r in raw_rows:
                row = {}
                for k, v in r.items():
                    if hasattr(v, 'isoformat'):
                        row[k] = v.isoformat()
                    else:
                        row[k] = v
                rows.append(row)

            print(f"    ✅ StarRocks: {len(rows)} transfer rows for {date_str}")
            return {
                "rows": rows,
                "row_count": len(rows),
                "source": f"StarRocks kf_transfer_items (date={date_iso})",
            }
        finally:
            conn.close()

    def _extract_from_file(self, date_str: str) -> dict:
        """Fallback: read transfer from local xlsx."""
        from openpyxl import load_workbook
        from lib.sources import TRANSFER_LOCAL

        date_tag = date_str.replace("/", "")
        rows = []

        if not os.path.isdir(TRANSFER_LOCAL):
            print(f"    ⚠ Transfer local dir not found: {TRANSFER_LOCAL}")
            return {"rows": [], "row_count": 0, "source": "file (not found)"}

        matched = None
        for fname in os.listdir(TRANSFER_LOCAL):
            if date_tag in fname and fname.endswith('.xlsx') and not fname.startswith('~'):
                matched = os.path.join(TRANSFER_LOCAL, fname)
                break

        if not matched:
            print(f"    ⚠ No transfer file for {date_tag}")
            return {"rows": [], "row_count": 0, "source": "file (not found)"}

        wb = load_workbook(matched, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=2, values_only=True):
            ngay = str(row[0] or "").strip()
            if ngay != date_str:
                continue
            rows.append({
                "ngay": ngay,
                "kho_xuat": str(row[1] or "").strip(),
                "kho_nhan": str(row[2] or "").strip(),
                "ma_hang": str(row[7] or "").strip(),
                "ten_hang": str(row[8] or "").strip(),
                "so_luong": float(row[10] or 0) if row[10] else 0,
                "trong_luong": float(row[14] or 0) if row[14] else 0,
            })
        wb.close()

        print(f"    📁 File fallback: {len(rows)} rows from {os.path.basename(matched)}")
        return {
            "rows": rows,
            "row_count": len(rows),
            "source": f"file ({os.path.basename(matched)})",
        }
