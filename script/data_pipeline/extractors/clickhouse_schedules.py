# -*- coding: utf-8 -*-
"""
Extract KRC delivery schedule from StarRocks.
Table: kfm_scm.krc_dashboard_delivery_schedule
Fallback: Google Sheets export
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from data_pipeline.extractors._base import BaseExtractor
from data_pipeline.config import load_starrocks_config


class StarrocksScheduleExtractor(BaseExtractor):
    name = "krc_schedule"

    def extract(self) -> dict:
        """Extract KRC delivery schedule for the target date."""
        dd = self.date_tag[:2]
        mm = self.date_tag[2:4]
        yyyy = self.date_tag[4:]
        date_str = f"{dd}/{mm}/{yyyy}"

        try:
            return self._extract_from_db(date_str)
        except Exception as e:
            print(f"    ⚠ StarRocks failed: {e}")
            print(f"    ↪ Falling back to Google Sheets...")
            return self._extract_from_sheets(date_str)

    def _extract_from_db(self, date_str: str) -> dict:
        """Query StarRocks for KRC schedule."""
        import pymysql

        cfg = load_starrocks_config()
        conn = pymysql.connect(
            host=cfg["host"],
            port=cfg["port"],
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            charset="utf8mb4",
            connect_timeout=30,
            read_timeout=60,
        )

        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cur:
                sql = """
                    SELECT 
                        ngay,
                        diem_den,
                        gio_den_dk AS gio_den,
                        tuyen,
                        source AS kho
                    FROM krc_dashboard_delivery_schedule
                    WHERE ngay = %s
                    ORDER BY tuyen, diem_den
                """
                cur.execute(sql, (date_str,))
                rows = cur.fetchall()

            # Serialize
            clean_rows = []
            for r in rows:
                row = {k: (v.isoformat() if hasattr(v, 'isoformat') else v) for k, v in r.items()}
                clean_rows.append(row)

            print(f"    ✅ StarRocks: {len(clean_rows)} KRC schedule rows for {date_str}")
            return {
                "rows": clean_rows,
                "row_count": len(clean_rows),
                "source": f"StarRocks krc_dashboard_delivery_schedule (ngay={date_str})",
            }
        finally:
            conn.close()

    def _extract_from_sheets(self, date_str: str) -> dict:
        """Fallback: read KRC schedule from Google Sheets."""
        from lib.sources import KRC_SHEET_URL
        from openpyxl import load_workbook
        from io import BytesIO
        import requests

        r = requests.get(KRC_SHEET_URL, timeout=120)
        r.raise_for_status()
        wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
        ws = wb["KRC"]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            scv = str(row[0].value or "").strip()
            if scv == date_str:
                diem_den = str(row[6].value or "").strip()
                gio_den = str(row[7].value or "").strip() if row[7].value else ""
                tuyen = str(row[10].value or "").strip()
                if diem_den:
                    rows.append({
                        "ngay": date_str,
                        "diem_den": diem_den,
                        "gio_den": gio_den,
                        "tuyen": tuyen,
                        "kho": "KRC",
                    })
        wb.close()

        print(f"    📊 Sheets fallback: {len(rows)} KRC rows for {date_str}")
        return {
            "rows": rows,
            "row_count": len(rows),
            "source": f"Google Sheets KRC (date={date_str})",
        }
