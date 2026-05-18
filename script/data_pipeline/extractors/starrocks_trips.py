# -*- coding: utf-8 -*-
"""
Extract trip details from StarRocks for performance report.
Table: kfm_scm.kf_trips
Fallback: local xlsx files in DS chi tiet chuyen xe/
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from data_pipeline.extractors._base import BaseExtractor
from data_pipeline.config import load_starrocks_config


class StarrocksTripsExtractor(BaseExtractor):
    name = "trips"

    def __init__(self, date_tag: str, bronze_dir: str, week_start: str = None, week_end: str = None):
        """
        Args:
            date_tag: DDMMYYYY format (used for file naming)
            bronze_dir: bronze storage path
            week_start: YYYY-MM-DD start of week (for weekly scan)
            week_end: YYYY-MM-DD end of week (for weekly scan)
        """
        super().__init__(date_tag, bronze_dir)
        self.week_start = week_start
        self.week_end = week_end

    def extract(self) -> dict:
        """Extract trip data from StarRocks."""
        try:
            return self._extract_from_db()
        except Exception as e:
            print(f"    ⚠ StarRocks failed: {e}")
            print(f"    ↪ Falling back to local xlsx...")
            return self._extract_from_file()

    def _extract_from_db(self) -> dict:
        """Query StarRocks for trip details."""
        import pymysql

        cfg = load_starrocks_config()
        conn = pymysql.connect(
            host=cfg["host"],
            port=cfg["port"],
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            charset="utf8mb4",
            connect_timeout=30,
            read_timeout=120,
        )

        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cur:
                if self.week_start and self.week_end:
                    # Weekly range query (for performance report)
                    sql = """
                        SELECT 
                            trip_id,
                            trip_status,
                            vehicle_number,
                            driver_name,
                            driver_phone,
                            depart_date,
                            depart_time,
                            noi_chuyen,
                            destination,
                            dest_status,
                            container_type,
                            arrival_time,
                            completion_time
                        FROM kf_trips
                        WHERE DATE(depart_date) BETWEEN %s AND %s
                        ORDER BY depart_date, trip_id
                    """
                    cur.execute(sql, (self.week_start, self.week_end))
                else:
                    # Single date query
                    dd = self.date_tag[:2]
                    mm = self.date_tag[2:4]
                    yyyy = self.date_tag[4:]
                    date_iso = f"{yyyy}-{mm}-{dd}"
                    sql = """
                        SELECT 
                            trip_id,
                            trip_status,
                            vehicle_number,
                            driver_name,
                            driver_phone,
                            depart_date,
                            depart_time,
                            noi_chuyen,
                            destination,
                            dest_status,
                            container_type,
                            arrival_time,
                            completion_time
                        FROM kf_trips
                        WHERE DATE(depart_date) = %s
                        ORDER BY trip_id
                    """
                    cur.execute(sql, (date_iso,))

                raw_rows = cur.fetchall()

            # Serialize + identify incomplete trips
            rows = []
            incomplete_count = 0
            for r in raw_rows:
                row = {}
                for k, v in r.items():
                    if hasattr(v, 'isoformat'):
                        row[k] = v.isoformat()
                    else:
                        row[k] = v
                if not row.get("completion_time"):
                    incomplete_count += 1
                    row["_incomplete"] = True
                rows.append(row)

            date_range = f"{self.week_start}→{self.week_end}" if self.week_start else self.date_tag
            print(f"    ✅ StarRocks: {len(rows)} trips ({incomplete_count} incomplete) for {date_range}")
            return {
                "rows": rows,
                "row_count": len(rows),
                "incomplete_count": incomplete_count,
                "source": f"StarRocks kf_trips ({date_range})",
            }
        finally:
            conn.close()

    def _extract_from_file(self) -> dict:
        """Fallback: read from local xlsx trip files."""
        mm = self.date_tag[2:4]
        yy = self.date_tag[6:]
        data_dir = rf'G:\My Drive\DOCS\DAILY\DS chi tiet chuyen xe\T{mm}.{yy}'

        if not os.path.exists(data_dir):
            print(f"    ⚠ Trip directory not found: {data_dir}")
            return {"rows": [], "row_count": 0, "source": "file (not found)"}

        from openpyxl import load_workbook

        files = sorted([f for f in os.listdir(data_dir) if f.endswith('.xlsx') and not f.startswith('~')])
        rows = []
        seen = set()

        for fname in files:
            fpath = os.path.join(data_dir, fname)
            try:
                wb = load_workbook(fpath, read_only=True)
                ws = wb['Sheet 1']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    trip_id = str(row[0] or "").strip()
                    dest = str(row[9] or "").strip()
                    if not trip_id or not dest:
                        continue
                    key = (trip_id, dest)
                    if key in seen:
                        continue
                    seen.add(key)
                    rows.append({
                        "trip_id": trip_id,
                        "trip_status": str(row[1] or "").strip(),
                        "vehicle_number": str(row[2] or "").strip(),
                        "driver_name": str(row[3] or "").strip(),
                        "depart_date": str(row[5] or "").strip(),
                        "noi_chuyen": str(row[8] or "").strip(),
                        "destination": dest,
                        "dest_status": str(row[11] or "").strip(),
                        "container_type": str(row[18] or "").strip() if len(row) > 18 else "",
                        "arrival_time": str(row[26] or "").strip() if row[26] else "",
                    })
                wb.close()
            except Exception as e:
                print(f"    ⚠ Error reading {fname}: {e}")

        print(f"    📁 File fallback: {len(rows)} trips from {len(files)} files")
        return {
            "rows": rows,
            "row_count": len(rows),
            "source": f"file ({data_dir}, {len(files)} files)",
        }
