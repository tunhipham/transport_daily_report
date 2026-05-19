# -*- coding: utf-8 -*-
"""
transfer_adapter.py — DB→xlsx adapter for transfer data
=========================================================
Queries ClickHouse kf_transfer_mart + kf_product_static for a given date,
and writes a 41-column xlsx file at the EXACT same path/format as the legacy
transfer file. This allows generate.py to run unchanged.

Usage:
    python script/data_pipeline/adapters/transfer_adapter.py --date 19/05/2026
    python script/data_pipeline/adapters/transfer_adapter.py --date 19/05/2026 --dry-run
"""
import os, sys, json, argparse
from datetime import datetime

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(BASE, "script"))
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ── Load registry ──
def load_registry():
    with open(os.path.join(BASE, "config", "data_sources.json"), encoding="utf-8") as f:
        return json.load(f)

# ── Column mapping: DB → xlsx (41 cols) ──
# generate.py reads: col0=Ngày, col2=Kho, col7=Mã hàng, col8=Tên hàng, col10=SL chuyển, col14=TL
HEADERS = [
    "Ngày chuyển hàng",          # 0  ← transfer_date
    "Ngày chuyển mong muốn",     # 1  ← expected_transfer_date
    "Chi nhánh chuyển",          # 2  ← from_branch_id → BRANCH_NAME ★
    "Chi nhánh nhận",            # 3  ← to_branch_id (leave as ID for now)
    "Ngày nhận hàng",            # 4  ← received_date
    "Giờ nhận hàng",             # 5
    "Người nhận hàng",           # 6
    "Mã hàng",                   # 7  ← product.base_barcode ★
    "Tên hàng",                  # 8  ← product.name ★
    "Đơn vị tính",               # 9  ← product.unit_name
    "Số lượng chuyển",           # 10 ← transfer_quantity ★
    "Số lượng nhận",             # 11 ← received_quantity
    "SL nhận (Hệ thống)",       # 12
    "SL hoàn/bổ sung",          # 13
    "Giá trị trọng lượng / thể tích",  # 14 ← product.base_net_weight ★
    "Đơn vị trọng lượng / thể tích",   # 15 ← product.base_measure
    "Mã chuyển hàng",           # 16 ← code
    "Đã nhận hàng",             # 17 ← is_creceived → "Có"/"Không"
    "Mã thùng",                 # 18
    "Trạng thái",               # 19 ← status → text
    "Số thùng",                 # 20
    "Ghi chú chuyển (phiếu)",   # 21
    "Ghi chú nhận (phiếu)",     # 22
    "HSD",                      # 23 ← expired_date
    "Ghi chú chuyển (barcode)", # 24
    "Ghi chú nhận (barcode)",   # 25
    "PT chuyển hoàn",           # 26
    "Kho nhận chuyển hoàn",     # 27
    "Chứng từ gốc",             # 28
    "Chứng từ phát sinh",       # 29
    "Cần hậu kiểm",             # 30
    "Đã hậu kiểm",              # 31
    "Hàng cân ký",               # 32
    "Là PT thùng rổ",           # 33
    "Tạo từ claim hàng",        # 34
    "Là PT cân tồn",            # 35
    "Là PT trả tồn",            # 36
    "Là PT claim hàng",         # 37
    "Mã chuyến xe",             # 38
    "Người tạo",                # 39
    "Thời gian tạo",            # 40
]

STATUS_TEXT = {
    1: "Tạo mới", 2: "Đang pick", 3: "Đã pick",
    4: "Đang giao", 5: "Đã nhận", 6: "Đã huỷ",
}


def query_transfer_data(date_iso, registry):
    """Query ClickHouse and return list of row dicts."""
    from data_pipeline.config import load_clickhouse_config
    import requests

    cfg = load_clickhouse_config()
    params = {"user": cfg["user"], "password": cfg["password"], "database": cfg["database"]}

    branch_map = registry["sources"]["transfer"]["branch_map"]
    branch_ids = "','".join(branch_map.keys())

    sql = f"""
        SELECT
            t.code,
            toString(toDate(t.transfer_date)) AS transfer_date_str,
            formatDateTime(t.transfer_date, '%d/%m/%Y') AS transfer_date_vn,
            t.from_branch_id,
            t.to_branch_id,
            t.transfer_quantity,
            t.received_quantity,
            t.status,
            t.is_creceived,
            formatDateTime(t.received_date, '%d/%m/%Y') AS received_date_vn,
            formatDateTime(t.expected_transfer_date, '%d/%m/%Y') AS expected_date_vn,
            formatDateTime(t.expired_date, '%d/%m/%Y') AS expired_date_vn,
            formatDateTime(t.raw_created_at, '%d/%m/%Y %H:%i') AS created_at_vn,
            t.product_id,
            p.name AS product_name,
            p.base_barcode,
            p.unit_name,
            p.base_net_weight,
            p.base_measure
        FROM kf_transfer_mart t
        LEFT JOIN kf_product_static p ON t.product_id = p.id
        WHERE toDate(t.transfer_date) = '{date_iso}'
          AND t.deleted = 0
          AND t.status != 6
          AND t.from_branch_id IN ('{branch_ids}')
        ORDER BY t.from_branch_id, t.code
    """

    print(f"  Querying ClickHouse (date={date_iso})...")
    r = requests.get(cfg["base_url"], params={**params, "query": f"{sql} FORMAT JSONEachRow"}, timeout=120)
    r.raise_for_status()

    rows = []
    for line in r.text.strip().split("\n"):
        if line.strip():
            rows.append(json.loads(line))

    print(f"  → {len(rows):,} rows from DB")
    return rows


def build_xlsx_row(db_row, branch_map):
    """Convert a DB row dict to a 41-element list matching legacy xlsx format."""
    branch_info = branch_map.get(db_row["from_branch_id"], {})
    branch_name = branch_info.get("name", db_row["from_branch_id"])

    row = [""] * len(HEADERS)
    row[0] = db_row["transfer_date_vn"]          # Ngày chuyển hàng
    row[1] = db_row.get("expected_date_vn", "")   # Ngày chuyển mong muốn
    row[2] = branch_name                          # Chi nhánh chuyển ★
    row[3] = db_row["to_branch_id"]               # Chi nhánh nhận (ID)
    row[4] = db_row.get("received_date_vn", "")   # Ngày nhận hàng
    row[7] = db_row["base_barcode"]               # Mã hàng ★
    row[8] = db_row["product_name"]               # Tên hàng ★
    row[9] = db_row["unit_name"]                  # Đơn vị
    row[10] = db_row["transfer_quantity"]          # SL chuyển ★
    row[11] = db_row["received_quantity"]          # SL nhận
    row[14] = db_row["base_net_weight"]            # TL (grams) ★
    row[15] = db_row.get("base_measure", "")      # Đơn vị TL
    row[16] = db_row["code"]                      # Mã chuyển hàng
    row[17] = "Có" if db_row["is_creceived"] else "Không"
    row[19] = STATUS_TEXT.get(db_row["status"], "")
    row[23] = db_row.get("expired_date_vn", "")
    row[40] = db_row.get("created_at_vn", "")
    return row


def write_xlsx(rows, output_path):
    """Write rows to xlsx with headers."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Transfer"

    # Header row
    for i, h in enumerate(HEADERS):
        ws.cell(row=1, column=i+1, value=h)

    # Data rows
    for j, row_data in enumerate(rows, 2):
        for i, val in enumerate(row_data):
            ws.cell(row=j, column=i+1, value=val)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    size = os.path.getsize(output_path)
    print(f"  ✅ Written: {output_path} ({size:,} bytes)")


def main():
    parser = argparse.ArgumentParser(description="DB→xlsx transfer adapter")
    parser.add_argument("--date", required=True, help="Date in DD/MM/YYYY format")
    parser.add_argument("--dry-run", action="store_true", help="Query only, don't write")
    parser.add_argument("--output-dir", default=None,
                        help="Output directory (default: data/raw/daily)")
    args = parser.parse_args()

    print(f"{'='*60}")
    print(f"  📦 Transfer Adapter — {args.date}")
    print(f"{'='*60}")

    # Parse date
    dt = datetime.strptime(args.date, "%d/%m/%Y")
    date_iso = dt.strftime("%Y-%m-%d")
    date_tag = dt.strftime("%d%m%Y")

    registry = load_registry()
    branch_map = registry["sources"]["transfer"]["branch_map"]

    # Query DB
    db_rows = query_transfer_data(date_iso, registry)

    if not db_rows:
        print("  ⚠ No data from DB!")
        sys.exit(1)

    # Build xlsx rows
    xlsx_rows = [build_xlsx_row(r, branch_map) for r in db_rows]

    # Stats
    by_kho = {}
    for r in db_rows:
        kho = branch_map.get(r["from_branch_id"], {}).get("name", "?")
        by_kho[kho] = by_kho.get(kho, 0) + 1
    print(f"\n  By kho:")
    for k, v in sorted(by_kho.items(), key=lambda x: -x[1]):
        print(f"    {k:<25} {v:>8,}")

    if args.dry_run:
        print(f"\n  [DRY RUN] — {len(xlsx_rows)} rows ready, not writing")
        return

    # Write
    out_dir = args.output_dir or os.path.join(BASE, "data", "raw", "daily")
    out_path = os.path.join(out_dir, f"transfer_{date_tag}.xlsx")
    write_xlsx(xlsx_rows, out_path)

    print(f"\n{'='*60}")
    print(f"  ✅ Done — {len(xlsx_rows):,} rows → {os.path.basename(out_path)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
