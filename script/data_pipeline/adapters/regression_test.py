# -*- coding: utf-8 -*-
"""
regression_test.py — Compare DB adapter output vs golden baseline
==================================================================
Runs generate.py's PT reading logic against both:
  1. Original file (from Google Drive backup)
  2. DB-generated file (from adapter)
Then compares KPI outputs.

Usage:
    python script/data_pipeline/adapters/regression_test.py --date 19/05/2026
"""
import os, sys, json
from datetime import datetime
from collections import defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(BASE, "script"))
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import load_workbook


def read_transfer_kpis(filepath, date_str):
    """Read transfer file and compute per-kho stats (same logic as generate.py)."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    kho_items = defaultdict(float)  # kho → count of items
    kho_sl = defaultdict(float)     # kho → sum SL
    total_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=False):
        ngay = str(row[0].value or "").strip()
        if ngay != date_str:
            continue
        raw_kho = str(row[2].value or "").strip()
        if not raw_kho:
            continue
        barcode = str(row[7].value or "").strip()
        sl_raw = row[10].value
        tl_raw = row[14].value
        try:
            sl = float(sl_raw) if sl_raw else 0
        except (ValueError, TypeError):
            sl = 0
        try:
            tl = float(tl_raw) if tl_raw else 0
        except (ValueError, TypeError):
            tl = 0

        kho_items[raw_kho] += 1
        kho_sl[raw_kho] += sl
        total_rows += 1

    wb.close()
    return {"total_rows": total_rows, "by_kho_items": dict(kho_items), "by_kho_sl": dict(kho_sl)}


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    args = parser.parse_args()

    dt = datetime.strptime(args.date, "%d/%m/%Y")
    date_tag = dt.strftime("%d%m%Y")

    original = os.path.join(BASE, "data", "raw", "daily", f"transfer_{date_tag}.xlsx")
    db_file = os.path.join(BASE, "data", "raw", "daily", "db_test", f"transfer_{date_tag}.xlsx")

    print(f"{'='*70}")
    print(f"  🧪 Regression Test — {args.date}")
    print(f"{'='*70}")

    if not os.path.exists(original):
        print(f"  ❌ Original file not found: {original}")
        sys.exit(1)
    if not os.path.exists(db_file):
        print(f"  ❌ DB file not found: {db_file}")
        sys.exit(1)

    print(f"  Original: {os.path.basename(original)} ({os.path.getsize(original):,} bytes)")
    print(f"  DB file:  {os.path.basename(db_file)} ({os.path.getsize(db_file):,} bytes)")

    print(f"\n  Reading original...")
    orig_kpi = read_transfer_kpis(original, args.date)
    print(f"  Reading DB file...")
    db_kpi = read_transfer_kpis(db_file, args.date)

    # Compare
    print(f"\n  {'Metric':<30} {'Original':>12} {'DB':>12} {'Diff':>10} {'Status':>8}")
    print(f"  {'-'*72}")

    all_ok = True

    # Total rows
    diff = db_kpi["total_rows"] - orig_kpi["total_rows"]
    status = "✅" if abs(diff) / max(orig_kpi["total_rows"], 1) < 0.05 else "⚠️"
    if status == "⚠️":
        all_ok = False
    print(f"  {'Total rows':<30} {orig_kpi['total_rows']:>12,} {db_kpi['total_rows']:>12,} {diff:>+10,} {status:>8}")

    # By kho
    all_khos = sorted(set(list(orig_kpi["by_kho_items"].keys()) + list(db_kpi["by_kho_items"].keys())))
    for kho in all_khos:
        orig_items = orig_kpi["by_kho_items"].get(kho, 0)
        db_items = db_kpi["by_kho_items"].get(kho, 0)
        diff = db_items - orig_items
        pct = abs(diff) / max(orig_items, 1) * 100
        status = "✅" if pct < 5 else "⚠️"
        if status == "⚠️":
            all_ok = False
        print(f"  {kho:<30} {orig_items:>12,.0f} {db_items:>12,.0f} {diff:>+10,.0f} {status:>8}")

    # SL comparison
    print(f"\n  {'SL by kho':<30} {'Original':>12} {'DB':>12} {'Diff':>10} {'Status':>8}")
    print(f"  {'-'*72}")
    for kho in all_khos:
        orig_sl = orig_kpi["by_kho_sl"].get(kho, 0)
        db_sl = db_kpi["by_kho_sl"].get(kho, 0)
        diff = db_sl - orig_sl
        pct = abs(diff) / max(orig_sl, 1) * 100
        status = "✅" if pct < 5 else "⚠️"
        if status == "⚠️":
            all_ok = False
        print(f"  {kho:<30} {orig_sl:>12,.1f} {db_sl:>12,.1f} {diff:>+10,.1f} {status:>8}")

    print(f"\n{'='*70}")
    if all_ok:
        print(f"  ✅ PASS — DB output matches original within 5% tolerance")
    else:
        print(f"  ⚠️  DIFF DETECTED — Review above for details")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
