import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

ABA_MASTER = r'G:\My Drive\DOCS\DAILY\ton_aba\data\master_data\Master Data.xlsx'
wb = load_workbook(ABA_MASTER, read_only=True, data_only=True)
ws = wb.worksheets[0]
barcode_type = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    bc = str(row[1].value or '').strip()
    pl = str(row[4].value or '').strip().upper()
    if bc and pl in ('MÁT', 'ĐÔNG'):
        barcode_type[bc] = pl
wb.close()

tf_path = r'g:\My Drive\DOCS\transport_daily_report\data\raw\daily\transfer_06052026.xlsx'
wb = load_workbook(tf_path, read_only=True, data_only=True)
ws = wb.worksheets[0]
unc = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    ngay = str(row[0].value or '').strip()
    raw_kho = str(row[2].value or '').strip()
    if ngay == '06/05/2026' and raw_kho == 'KHO ABA QUÁ CẢNH':
        code = str(row[7].value or '').strip()
        name = str(row[8].value or '').strip()
        try:
            sl = float(row[10].value or 0)
        except (ValueError, TypeError):
            sl = 0
        try:
            tl = float(row[14].value or 0)
        except (ValueError, TypeError):
            tl = 0
        if tl == 0:
            continue
        if code not in barcode_type:
            if code not in unc:
                unc[code] = {'sl': 0, 'tons': 0.0, 'name': name}
            unc[code]['sl'] += sl
            unc[code]['tons'] += sl * tl / 1_000_000
wb.close()

header = "{:<3} {:<18} {:>8} {:>8}   {}".format("No", "Barcode", "Qty", "Tan", "Ten san pham")
print(header)
print("-" * 100)
for i, (bc, info) in enumerate(sorted(unc.items(), key=lambda x: -x[1]['sl']), 1):
    line = "{:<3} {:<18} {:>8,.0f} {:>8.3f}   {}".format(i, bc, info['sl'], info['tons'], info['name'])
    print(line)
print("-" * 100)
total_sl = sum(v['sl'] for v in unc.values())
total_tons = sum(v['tons'] for v in unc.values())
line = "{:<3} {:<18} {:>8,.0f} {:>8.3f}".format("", "TOTAL", total_sl, total_tons)
print(line)
