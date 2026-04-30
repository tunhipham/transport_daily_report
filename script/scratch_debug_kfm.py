"""Debug KFM DRY sheet reading for KSL-TỐI issue."""
import os, sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(BASE, "script"))

BACKUP_DIR = os.path.join(BASE, "data", "raw", "daily")
DATE_STR = "30/04/2026"

def parse_time_hour(time_text):
    if not time_text:
        return -1
    m = re.match(r'(\d{1,2}):', str(time_text).strip())
    return int(m.group(1)) if m else -1

# Load from backup
kfm_path = os.path.join(BACKUP_DIR, "kfm_30042026.xlsx")
print(f"Loading: {kfm_path}")
wb = load_workbook(kfm_path, read_only=True, data_only=True)

# Find DRY sheet
kfm_ws = None
for sname in wb.sheetnames:
    print(f"  Sheet: {sname}")
    if 'DRY' in sname.upper():
        kfm_ws = wb[sname]
        break
if not kfm_ws:
    kfm_ws = wb.worksheets[0]
    print(f"  No DRY sheet found, using first: {wb.sheetnames[0]}")

print(f"\nUsing sheet: {kfm_ws.title}")

# Check headers (row 2)
print("\n=== HEADERS (row 2) ===")
for row in kfm_ws.iter_rows(min_row=2, max_row=2, values_only=False):
    for i, cell in enumerate(row[:15]):
        print(f"  col[{i}] = {cell.value}")

# Read ALL rows for today
print(f"\n=== ALL ROWS for {DATE_STR} ===")
matched = 0
skipped_no_date = 0
skipped_no_dest = 0
skipped_no_time = 0
skipped_hour_neg = 0
sang_count = 0
toi_count = 0
inherited = 0

last_kho = None
for xl_row in kfm_ws.iter_rows(min_row=3, values_only=False):
    scv = str(xl_row[0].value or "").strip()
    
    # Check different date formats
    raw_date = xl_row[0].value
    
    if scv == DATE_STR:
        matched += 1
        diem_den = str(xl_row[6].value or "").strip()
        gio_den_raw = xl_row[7].value
        gio_di_raw = xl_row[4].value
        gio_den = str(gio_den_raw or "").strip()
        gio_di = str(gio_di_raw or "").strip()
        tuyen = str(xl_row[10].value or "").strip()
        
        if not diem_den:
            skipped_no_dest += 1
            continue
            
        gio = gio_den or gio_di
        if gio:
            hour = parse_time_hour(gio)
            if hour < 0:
                skipped_hour_neg += 1
                print(f"  ❌ hour<0: gio_den='{gio_den}' gio_di='{gio_di}' raw_den={repr(gio_den_raw)} raw_di={repr(gio_di_raw)} dest={diem_den}")
                continue
            using_gio_di = not gio_den and bool(gio_di)
            sang_cutoff = 15 if using_gio_di else 18
            if 6 <= hour < sang_cutoff:
                sang_count += 1
                last_kho = "Sáng"
            else:
                toi_count += 1
                last_kho = "Tối"
                print(f"  TỐI: dest={diem_den} tuyen={tuyen} gio_den='{gio_den}' gio_di='{gio_di}' hour={hour}")
        elif last_kho:
            inherited += 1
            if last_kho == "Sáng":
                sang_count += 1
            else:
                toi_count += 1
                print(f"  TỐI(inherited): dest={diem_den} tuyen={tuyen}")
        else:
            skipped_no_time += 1
            print(f"  ❌ no_time: dest={diem_den} tuyen={tuyen} gio_den='{gio_den}' gio_di='{gio_di}'")
    else:
        last_kho = None

# Also check if any dates look different  
print(f"\n=== SAMPLE DATE VALUES (first 20 unique) ===")
dates_seen = set()
for xl_row in kfm_ws.iter_rows(min_row=3, max_row=500, values_only=False):
    raw = xl_row[0].value
    s = str(raw or "").strip()
    if s and s not in dates_seen:
        dates_seen.add(s)
        print(f"  raw={repr(raw)} → str='{s}' match={s == DATE_STR}")
        if len(dates_seen) >= 20:
            break

print(f"\n=== SUMMARY ===")
print(f"  Matched date rows: {matched}")
print(f"  Skipped (no dest): {skipped_no_dest}")
print(f"  Skipped (no time): {skipped_no_time}")
print(f"  Skipped (hour<0): {skipped_hour_neg}")
print(f"  Inherited: {inherited}")
print(f"  Sáng: {sang_count}")
print(f"  Tối: {toi_count}")

wb.close()
