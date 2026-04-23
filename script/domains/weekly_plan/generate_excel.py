# -*- coding: utf-8 -*-
"""
generate_excel.py — Generate weekly transport plan Excel from master_schedule.json

Sources:
  - master_schedule.json  (lịch về/shift cố định)
  - Google Sheets kiểm kê (INVENTORY_SHEET_URL)
  - NSO STORES            (châm hàng D→D+3, skip-first-day)

Output:
  - output/artifacts/weekly transport plan/Lịch đi hàng ST W{nn}.xlsx

Usage:
  python script/domains/weekly_plan/generate_excel.py              # auto W+1
  python script/domains/weekly_plan/generate_excel.py --week 18    # specific week

Skip-first-day rule:
  After D+3, the FIRST delivery day in the daily schedule is SKIPPED.
  Delivery resumes from the SECOND matching day of schedule_ve.
  Example: A164 opens 23/04 (Thu), schedule_ve="Thứ 2-4-6", D+3=26/04 (Sun)
           → Mon 27/04 is first daily match → SKIP
           → Wed 29/04 is second daily match → START delivery here
"""

import os, sys, json, re, argparse
from datetime import datetime, timedelta, date
from io import BytesIO

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Paths
BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
PLAN_DIR = os.path.join(BASE, "output", "artifacts", "weekly transport plan")
os.makedirs(PLAN_DIR, exist_ok=True)

sys.path.insert(0, os.path.join(BASE, "script"))
from lib.sources import INVENTORY_SHEET_URL

# Import NSO stores
sys.path.insert(0, os.path.join(BASE, "script", "domains", "nso"))
from generate import STORES as NSO_STORES, parse_date as nso_parse_date

# Import NSO_SCHEDULE from export_weekly_plan
sys.path.insert(0, os.path.join(BASE, "script", "dashboard"))
from export_weekly_plan import NSO_SCHEDULE

# ═══════════════════════════════════════════════
# WEEK CONFIG
# ═══════════════════════════════════════════════
ANCHOR_WEEK = 14
ANCHOR_START = date(2026, 3, 30)  # Monday W14


def get_week_dates(week_num):
    """Compute Monday→Sunday dates for a given week number."""
    week_start = ANCHOR_START + timedelta(weeks=week_num - ANCHOR_WEEK)
    return [week_start + timedelta(days=i) for i in range(7)]


def auto_detect_week():
    """Auto-detect W+1 (next week)."""
    today = date.today()
    days_since_anchor = (today - ANCHOR_START).days
    current_week = ANCHOR_WEEK + days_since_anchor // 7
    return current_week + 1


# ═══════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════
def parse_schedule_days(schedule_str):
    """Parse 'Thứ 3-5-7' → set of weekday indices (Mon=0)."""
    days = set()
    nums = re.findall(r'\d', schedule_str)
    for n in nums:
        n = int(n)
        if 2 <= n <= 7:
            days.add(n - 2)
    return days


def compute_even_date_schedule(week_dates):
    """A112 Cô Giang: even-date delivery (trừ CN)."""
    even_weekdays = set()
    for wd in week_dates:
        if wd.day % 2 == 0 and wd.weekday() != 6:
            even_weekdays.add(wd.weekday())
    return even_weekdays


def fetch_inventory():
    """Fetch kiểm kê từ Google Sheets."""
    import requests
    from openpyxl import load_workbook

    try:
        r = requests.get(INVENTORY_SHEET_URL, allow_redirects=True, timeout=60)
        r.raise_for_status()
        wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
        ws = wb['Lịch Kiểm kê 2026']
        inv = {}
        for row in ws.iter_rows(min_row=10, values_only=False):
            store_id = str(row[3].value or "").strip()
            kiem_ke = row[7].value
            if store_id and kiem_ke:
                dt = None
                if isinstance(kiem_ke, datetime):
                    dt = kiem_ke.date()
                elif isinstance(kiem_ke, date):
                    dt = kiem_ke
                elif isinstance(kiem_ke, str):
                    try:
                        dt = datetime.strptime(kiem_ke, "%d/%m/%Y").date()
                    except ValueError:
                        pass
                if dt:
                    inv[store_id] = dt
        wb.close()
        print(f"  📋 Kiểm kê: {len(inv)} entries")
        return inv
    except Exception as e:
        print(f"  ⚠ Could not fetch kiểm kê: {e}")
        return {}


# ═══════════════════════════════════════════════
# BUILD STORES
# ═══════════════════════════════════════════════
def build_stores(week_dates, inventory):
    """Build store list from master_schedule.json + kiểm kê."""
    ms_path = os.path.join(BASE, "data", "master_schedule.json")
    with open(ms_path, "r", encoding="utf-8") as f:
        master = json.load(f)

    stores = []
    for s in master["stores"]:
        code = s["code"]
        schedule_ve = s.get("schedule_ve", "")
        shift = s.get("shift", "Đêm")

        # Compute schedule_chia (inverse of schedule_ve)
        if schedule_ve == "Ngày chẵn":
            delivery_weekdays = compute_even_date_schedule(week_dates)
            # Build display label
            thu_nums = sorted(wd + 2 for wd in delivery_weekdays if wd < 6)
            schedule_ve_display = "Thứ " + "-".join(str(n) for n in thu_nums) if thu_nums else "Ngày chẵn"
            # Chia = opposite of về for even-date (approximate)
            schedule_chia = ""
        elif "2-3-5-7" in schedule_ve or "2-4-6" in schedule_ve or "3-5-7" in schedule_ve:
            delivery_weekdays = parse_schedule_days(schedule_ve)
            schedule_ve_display = schedule_ve
            # Chia inverse: 2-4-6 ↔ 3-5-7
            if "2-4-6" in schedule_ve:
                schedule_chia = "Thứ 3-5-7"
            elif "3-5-7" in schedule_ve:
                schedule_chia = "Thứ 2-4-6"
            elif "2-3-5-7" in schedule_ve:
                schedule_chia = "Thứ 2-4-6-7"
            else:
                schedule_chia = ""
        else:
            delivery_weekdays = parse_schedule_days(schedule_ve)
            schedule_ve_display = schedule_ve
            schedule_chia = ""

        # Build 7-day schedule
        days = []
        for i, wd in enumerate(week_dates):
            if wd.weekday() in delivery_weekdays:
                days.append(shift)
            else:
                days.append("")

        # Inventory cross-check
        inv_dt = inventory.get(code)
        inv_date_str = ""
        if inv_dt:
            if isinstance(inv_dt, datetime):
                inv_dt = inv_dt.date()
            inv_date_str = inv_dt.strftime("%d/%m/%Y")
            for i, wd in enumerate(week_dates):
                if wd == inv_dt or wd == inv_dt - timedelta(days=1):
                    current = days[i]
                    if current and current.lower() in ("ngày", "đêm"):
                        days[i] = "Kiểm kê"
                    elif not current and wd == inv_dt:
                        days[i] = "Kiểm kê"

        stores.append({
            "name": s["name"],
            "code": code,
            "schedule_chia": schedule_chia,
            "schedule_ve": schedule_ve_display,
            "opening_date": "",
            "inventory_date": inv_date_str,
            "shift": shift,
            "days": days,
        })

    return stores


# ═══════════════════════════════════════════════
# NSO CHÂM HÀNG (with skip-first-day)
# ═══════════════════════════════════════════════
def apply_nso(stores, week_dates):
    """Apply NSO châm hàng with skip-first-day rule.

    After D+3 (last châm hàng day), the FIRST delivery day matching
    schedule_ve is SKIPPED. Delivery resumes from the SECOND match.
    """
    nso_applied = 0
    nso_injected = 0
    existing_codes = {s["code"] for s in stores}

    for nso in NSO_STORES:
        try:
            opening = nso_parse_date(nso["opening_date"])
        except:
            continue

        # Rescheduled stores
        if nso.get("original_date"):
            try:
                orig = nso_parse_date(nso["original_date"])
                orig_in_week = any(week_dates[0] <= orig + timedelta(days=i) <= week_dates[6] for i in range(4))
                new_in_week = any(week_dates[0] <= opening + timedelta(days=i) <= week_dates[6] for i in range(4))
                if orig_in_week and not new_in_week:
                    continue
            except:
                pass

        # Check if any D→D+3 falls in this week
        cham_days_in_week = []
        for i in range(4):
            d = opening + timedelta(days=i)
            if week_dates[0] <= d <= week_dates[6]:
                cham_days_in_week.append(d)

        # Also check if post-châm days fall in this week (store opened last week)
        d3 = opening + timedelta(days=3)
        post_cham_in_week = any(wd > d3 for wd in week_dates)

        if not cham_days_in_week and not post_cham_in_week:
            continue

        code = nso["code"]
        sched = NSO_SCHEDULE.get(code, {})
        schedule_ve = sched.get("schedule_ve", "")
        shift = sched.get("shift", "Đêm")
        delivery_weekdays = parse_schedule_days(schedule_ve)

        # Find or inject store
        matched = None
        for s in stores:
            if s["code"] == code:
                matched = s
                break

        is_injected = False
        if not matched:
            name_sys = nso.get("name_system", "")
            name_full = nso.get("name_full", "")
            store_name = f"{name_sys} - {name_full}" if name_sys else name_full or f"NSO {code}"
            schedule_chia = sched.get("schedule_chia", "")

            matched = {
                "name": store_name,
                "code": code,
                "schedule_chia": schedule_chia,
                "schedule_ve": schedule_ve,
                "opening_date": opening.strftime("%d/%m/%Y"),
                "inventory_date": "",
                "shift": shift,
                "days": [""] * 7,
            }
            stores.append(matched)
            is_injected = True
            nso_injected += 1

        if not matched.get("opening_date"):
            matched["opening_date"] = opening.strftime("%d/%m/%Y")

        # ─── Skip D+4 rule: only skip if D+4 is a delivery day ───
        # If D+4 falls on a delivery day → skip it (giảm tải 1 ngày)
        # If D+4 is NOT a delivery day → natural gap already exists, no skip
        d4 = opening + timedelta(days=4)
        skip_date = d4 if d4.weekday() in delivery_weekdays else None

        # Apply châm hàng + post-châm with skip
        for i, wd in enumerate(week_dates):
            delta = (wd - opening).days

            if 0 <= delta <= 3:
                matched["days"][i] = "Châm hàng"
            elif delta > 3:
                if skip_date and wd == skip_date:
                    # SKIP D+4 (giảm tải ngay sau châm hàng)
                    matched["days"][i] = ""
                elif wd.weekday() in delivery_weekdays:
                    matched["days"][i] = shift

        nso_applied += 1

    return nso_applied, nso_injected


# ═══════════════════════════════════════════════
# WRITE EXCEL
# ═══════════════════════════════════════════════
def write_excel(stores, week_dates, week_num, week_label):
    """Write Lịch đi hàng ST W{nn}.xlsx matching dashboard styling + export_weekly_plan layout.
    
    Layout must match export_weekly_plan.py expectations:
    - R2 col H (8): week label
    - R3 cols H-N (8-14): dates
    - R4 cols A-G: column headers, cols H-N: day names
    - R5+ cols 1-7: name,code,schedule_chia,schedule_ve,opening,inventory,shift
    - R5+ cols 8-14: days
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Lịch về hàng"

    day_names = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]

    # ═══════════════════════════════════════════════
    # STYLE CONSTANTS (matching dashboard JS export)
    # ═══════════════════════════════════════════════
    thin_border = Border(
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
    )
    thick_border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin'),
    )

    # Title row — dark navy blue
    sTitle = {
        'font': Font(bold=True, size=13, color='FFFFFF'),
        'fill': PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }
    sCount = {
        'font': Font(bold=True, size=12, color='FFFFFF'),
        'fill': PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }
    sWeek = {
        'font': Font(bold=True, size=11, color='FFFFFF'),
        'fill': PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }

    # Date header — dark red
    sDateHdr = {
        'font': Font(bold=True, size=10, color='FFFFFF'),
        'fill': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thick_border,
    }

    # Day name headers — yellow
    sDayHdr = {
        'font': Font(bold=True, size=10),
        'fill': PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thick_border,
    }

    # Column headers — light blue
    sColHdr = {
        'font': Font(bold=True, size=10),
        'fill': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thick_border,
    }

    # Data cells
    sCell = {
        'font': Font(size=10),
        'alignment': Alignment(vertical='center'),
        'border': thin_border,
    }
    sCellC = {
        'font': Font(size=10),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border,
    }
    sCham = {
        'font': Font(bold=True, size=10, color='C65911'),
        'fill': PatternFill(start_color='FBE5D6', end_color='FBE5D6', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border,
    }
    sKK = {
        'font': Font(bold=True, size=10, color='FF0000'),
        'fill': PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border,
    }
    sNgay = {
        'font': Font(size=10),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border,
    }
    sDem = {
        'font': Font(size=10),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border,
    }

    def apply_style(cell, style_dict):
        for attr, val in style_dict.items():
            setattr(cell, attr, val)

    # ═══════════════════════════════════════════════
    # ROW 1: Empty
    # ═══════════════════════════════════════════════
    ws.row_dimensions[1].height = 8

    # ═══════════════════════════════════════════════
    # ROW 2: A2=title, B2=count, C2=empty, D2:J2=week label
    # ═══════════════════════════════════════════════
    ws.row_dimensions[2].height = 28
    c = ws.cell(2, 1, 'LỊCH VỀ HÀNG SIÊU THỊ')
    apply_style(c, sTitle)
    c = ws.cell(2, 2, len(stores))
    apply_style(c, sCount)
    apply_style(ws.cell(2, 3, ''), sWeek)
    # Week label merged D2:J2
    ws.merge_cells('D2:J2')
    c = ws.cell(2, 4, week_label)
    apply_style(c, sWeek)
    for col in range(5, 11):
        apply_style(ws.cell(2, col), sWeek)

    # ═══════════════════════════════════════════════
    # ROW 3: Date headers — D3:J3 = dd/mm/yyyy (no time)
    # ═══════════════════════════════════════════════
    ws.row_dimensions[3].height = 22
    for col in range(1, 4):
        apply_style(ws.cell(3, col, ''), sDateHdr)
    for i, wd in enumerate(week_dates):
        c = ws.cell(3, 4 + i, wd.strftime('%d/%m/%Y'))
        apply_style(c, sDateHdr)

    # ═══════════════════════════════════════════════
    # ROW 4: A=SIÊU THỊ, B=Viết tắt, C=Giờ nhận, D-J=day names
    # ═══════════════════════════════════════════════
    ws.row_dimensions[4].height = 26
    for i, h in enumerate(['SIÊU THỊ', 'Viết tắt', 'Giờ nhận']):
        c = ws.cell(4, 1 + i, h)
        apply_style(c, sColHdr)
    for i, dn in enumerate(day_names):
        c = ws.cell(4, 4 + i, dn)
        apply_style(c, sDayHdr)

    # ═══════════════════════════════════════════════
    # ROW 5+: A=name, B=code, C=shift, D-J=days
    # ═══════════════════════════════════════════════
    for r, s in enumerate(stores, 5):
        c = ws.cell(r, 1, s["name"])
        apply_style(c, sCell)
        c = ws.cell(r, 2, s["code"])
        apply_style(c, sCellC)
        c = ws.cell(r, 3, s["shift"])
        apply_style(c, sCellC)

        for i, day_val in enumerate(s["days"]):
            c = ws.cell(r, 4 + i, day_val if day_val else '')
            vl = (day_val or '').lower()
            if 'châm' in vl or 'cham' in vl:
                apply_style(c, sCham)
            elif 'kiểm' in vl or 'kiem' in vl:
                apply_style(c, sKK)
            elif vl == 'ngày':
                apply_style(c, sNgay)
            elif vl == 'đêm':
                apply_style(c, sDem)
            else:
                apply_style(c, sCellC)

    # ═══════════════════════════════════════════════
    # COLUMN WIDTHS
    # ═══════════════════════════════════════════════
    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    for i in range(7):
        ws.column_dimensions[get_column_letter(4 + i)].width = 14

    # AUTO-FILTER
    ws.auto_filter.ref = f"A4:J{4 + len(stores)}"

    # Save
    filename = f"Lịch đi hàng ST W{week_num}.xlsx"
    filepath = os.path.join(PLAN_DIR, filename)
    wb.save(filepath)
    wb.close()
    return filepath


# ═══════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Generate weekly plan Excel from master_schedule")
    parser.add_argument("--week", type=int, default=None, help="Week number (default: auto W+1)")
    args = parser.parse_args()

    week_num = args.week or auto_detect_week()
    week_dates = get_week_dates(week_num)
    week_label = f"TUẦN {week_num} TỪ {week_dates[0].strftime('%d/%m')} - {week_dates[6].strftime('%d/%m')}"

    print(f"📅 Generating W{week_num}: {week_label}")
    date_strs = ' | '.join(d.strftime("%a %d/%m") for d in week_dates)
    print(f"   {date_strs}")

    # 1. Fetch kiểm kê
    print("\n🔍 Step 1: Fetching kiểm kê...")
    inventory = fetch_inventory()

    # 2. Build stores from master_schedule
    print("\n📋 Step 2: Building stores from master_schedule.json...")
    stores = build_stores(week_dates, inventory)
    print(f"   {len(stores)} stores loaded")

    # 3. NSO châm hàng
    print("\n🏪 Step 3: Applying NSO châm hàng (skip-first-day)...")
    nso_applied, nso_injected = apply_nso(stores, week_dates)
    print(f"   {nso_applied} NSO stores processed ({nso_injected} injected)")

    # 4. Stats
    st_ngay = sum(1 for s in stores if (s.get("shift") or "").lower() == "ngày")
    st_dem = sum(1 for s in stores if (s.get("shift") or "").lower() == "đêm")
    st_cham = sum(1 for s in stores if any("châm" in (d or "").lower() for d in s["days"]))
    st_kk = sum(1 for s in stores if any("kiểm" in (d or "").lower() for d in s["days"]))
    print(f"\n📊 Stats: {st_ngay} ngày, {st_dem} đêm, {st_cham} châm hàng, {st_kk} kiểm kê")
    print(f"   Total: {len(stores)} stores")

    # 5. Write Excel
    print(f"\n💾 Step 5: Writing Excel...")
    filepath = write_excel(stores, week_dates, week_num, week_label)
    fsize = os.path.getsize(filepath)
    print(f"   ✅ {filepath} ({fsize:,} bytes)")

    return filepath


if __name__ == "__main__":
    main()
