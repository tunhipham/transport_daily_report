# -*- coding: utf-8 -*-
"""
fetch_plan_incremental.py — Fetch today's plan data and merge into monthly cache
=================================================================================
Only fetches plan for TODAY (or a specified date) from Google Sheets + KH local,
then merges into the existing monthly_plan_T{mm}.json.

Much faster than full fetch_monthly.py (~30s vs ~3-5 min).

Usage:
    python script/domains/performance/fetch_plan_incremental.py
    python script/domains/performance/fetch_plan_incremental.py --date 27/05/2026
"""
import os, sys, json, re
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
OUTPUT = os.path.join(BASE, "output")

# Reuse sources from fetch_monthly
KRC_SHEET_URL = "https://docs.google.com/spreadsheets/d/1tWamqjpOI2j2MrYW3Ah6ptmT524CAlQvEP8fCkxfuII/export?format=xlsx"
KFM_SHEET_URL = "https://docs.google.com/spreadsheets/d/1LkJFJhOQ8F2WEB3uCk7kA2Phvu8IskVi3YBfVr7pBx0/export?format=xlsx"

KH_DRIVE_FOLDERS = [
    ("KH HÀNG ĐÔNG", "ĐÔNG MÁT"),
    ("KH HÀNG MÁT", "ĐÔNG MÁT"),
]


def format_time(time_val):
    if not time_val:
        return ""
    s = str(time_val).strip()
    m = re.match(r'(\d{1,2}):(\d{2})', s)
    if m:
        return f"{int(m.group(1))}:{m.group(2)}"
    if hasattr(time_val, 'hour'):
        return f"{time_val.hour}:{time_val.minute:02d}"
    try:
        fval = float(time_val)
        if 0 <= fval < 1:
            total_minutes = round(fval * 24 * 60)
            return f"{total_minutes // 60}:{total_minutes % 60:02d}"
    except (ValueError, TypeError):
        pass
    m2 = re.match(r'(\d+):(\d{2}):(\d{2})', s)
    if m2:
        return f"{int(m2.group(1))}:{m2.group(2)}"
    return s


def read_xlsx_from_url(url):
    import requests
    r = requests.get(url, allow_redirects=True, timeout=120, stream=True)
    r.raise_for_status()
    chunks = []
    for chunk in r.iter_content(chunk_size=65536):
        chunks.append(chunk)
    content = b''.join(chunks)
    if len(content) < 500:
        raise ValueError(f"Response too small ({len(content)} bytes)")
    return load_workbook(BytesIO(content), read_only=True, data_only=True)


def fetch_krc_today(date_str):
    """Fetch KRC plan for a specific date."""
    rows = []
    try:
        wb = read_xlsx_from_url(KRC_SHEET_URL)
        ws = None
        for name in wb.sheetnames:
            if name == "KRC":
                ws = wb[name]
                break
        if not ws:
            for name in wb.sheetnames:
                if "KRC" in name:
                    ws = wb[name]
                    break
        if not ws:
            return rows

        for row in ws.iter_rows(min_row=2, values_only=False):
            date_val = str(row[0].value or "").strip()
            if date_val != date_str:
                continue
            store = str(row[6].value or "").strip()
            planned_time = format_time(row[7].value)
            tuyen = str(row[10].value or "").strip() if len(row) > 10 else ""
            if store:
                rows.append({
                    "date": date_str, "store": store,
                    "planned_time": planned_time, "tuyen": tuyen, "kho": "KRC",
                })
        wb.close()
    except Exception as e:
        print(f"    ⚠ KRC: {e}")
    return rows


def fetch_dry_today(date_str):
    """Fetch DRY plan for a specific date."""
    rows = []
    try:
        wb = read_xlsx_from_url(KFM_SHEET_URL)
        ws = None
        for name in wb.sheetnames:
            if "KHÔ" in name or "DRY" in name:
                ws = wb[name]
                break
        if not ws:
            ws = wb.worksheets[1]

        for row in ws.iter_rows(min_row=3, values_only=False):
            date_val = str(row[0].value or "").strip()
            if date_val != date_str:
                continue
            store_6 = str(row[6].value or "").strip() if len(row) > 6 else ""
            store_5 = str(row[5].value or "").strip() if len(row) > 5 else ""
            time_6 = format_time(row[6].value) if len(row) > 6 else ""
            time_7 = format_time(row[7].value) if len(row) > 7 else ""
            tuyen_9 = str(row[9].value or "").strip() if len(row) > 9 else ""
            tuyen_10 = str(row[10].value or "").strip() if len(row) > 10 else ""

            if store_6 and len(store_6) <= 5 and store_6.isalnum():
                store, planned_time, tuyen = store_6, time_7, tuyen_10
            elif store_5 and len(store_5) <= 5 and store_5.isalnum():
                store, planned_time, tuyen = store_5, time_6, tuyen_9
            else:
                continue

            if store:
                rows.append({
                    "date": date_str, "store": store,
                    "planned_time": planned_time, "tuyen": tuyen, "kho": "DRY",
                })
        wb.close()
    except Exception as e:
        print(f"    ⚠ DRY: {e}")
    return rows


def fetch_dongmat_today(date_str):
    """Fetch ĐÔNG MÁT plan for a specific date from local KH files."""
    rows = []
    parts = date_str.split("/")
    dt = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
    date_for_file = dt.strftime("%d.%m.%Y")

    for folder_name, kho_name in KH_DRIVE_FOLDERS:
        local_dir = rf"G:\My Drive\DOCS\DAILY\{folder_name}"
        if not os.path.isdir(local_dir):
            continue
        matched = None
        for fname in os.listdir(local_dir):
            if date_for_file in fname and fname.endswith('.xlsx') and not fname.startswith('~'):
                matched = os.path.join(local_dir, fname)
                break
        if not matched:
            continue

        try:
            wb = load_workbook(matched, read_only=True, data_only=True)
            ws = wb.worksheets[0]

            tuyen_col = 9
            time_col = None
            time_col_backup = None
            for hdr_row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                for i, cell in enumerate(hdr_row):
                    h = str(cell.value or "").strip().lower()
                    if ("tuyen" in h or "tuyến" in h) and "kg" not in h:
                        tuyen_col = i
                    if time_col is None and ("kien" in h or "kiến" in h):
                        if "giao" in h:
                            time_col = i
                        elif ("roi" not in h and "rời" not in h) and time_col_backup is None:
                            time_col_backup = i
            if time_col is None:
                time_col = time_col_backup

            for row in ws.iter_rows(min_row=2, values_only=False):
                store = str(row[2].value or "").strip() if len(row) > 2 else ""
                planned_time = ""
                if time_col is not None and time_col < len(row):
                    planned_time = format_time(row[time_col].value)
                tuyen_raw = row[tuyen_col].value if tuyen_col < len(row) else ""
                tuyen = ""
                if tuyen_raw:
                    tuyen_str = str(tuyen_raw).strip()
                    if tuyen_str:
                        try:
                            int(float(tuyen_str))
                            tuyen = f"HĐ{int(float(tuyen_str))}"
                        except (ValueError, TypeError):
                            tuyen = tuyen_str

                if store:
                    rows.append({
                        "date": date_str, "store": store,
                        "planned_time": planned_time, "tuyen": tuyen, "kho": "ĐÔNG MÁT",
                    })
            wb.close()
            print(f"    ↳ {os.path.basename(matched)}: {len([r for r in rows if r['kho']=='ĐÔNG MÁT'])} rows")
        except Exception as e:
            print(f"    ⚠ {folder_name}: {e}")

    return rows


def merge_into_monthly(date_str, new_rows, month, year):
    """Merge today's plan rows into the monthly plan cache.
    Replaces rows for date_str, keeps everything else intact.
    """
    state_dir = os.path.join(OUTPUT, "state")
    plan_path = os.path.join(state_dir, f"monthly_plan_T{month:02d}.json")

    if os.path.exists(plan_path):
        with open(plan_path, "r", encoding="utf-8") as f:
            plan_data = json.load(f)
    else:
        print(f"  ⚠ monthly_plan_T{month:02d}.json not found — creating new")
        plan_data = {
            "month": f"T{month:02d}", "year": year,
            "start": f"01/{month:02d}/{year}",
            "end": f"28/{month:02d}/{year}",
            "plan": {"KRC": [], "DRY": [], "ĐÔNG MÁT": [], "THỊT CÁ": []},
            "thitca_actual": [],
        }

    # Group new rows by kho
    new_by_kho = {}
    for r in new_rows:
        new_by_kho.setdefault(r["kho"], []).append(r)

    # Replace rows for date_str in each kho
    updated_count = 0
    for kho in ["KRC", "DRY", "ĐÔNG MÁT"]:
        existing = plan_data.get("plan", {}).get(kho, [])
        # Remove old rows for this date
        kept = [r for r in existing if r.get("date") != date_str]
        # Add new rows
        new_kho_rows = new_by_kho.get(kho, [])
        kept.extend(new_kho_rows)
        plan_data["plan"][kho] = kept
        if new_kho_rows:
            updated_count += len(new_kho_rows)
            print(f"    {kho}: +{len(new_kho_rows)} rows for {date_str}")

    # Save
    os.makedirs(state_dir, exist_ok=True)
    with open(plan_path, "w", encoding="utf-8") as f:
        json.dump(plan_data, f, ensure_ascii=False, indent=2)
    print(f"  💾 Merged {updated_count} rows into {plan_path}")

    return plan_data


def fetch_and_merge(date_str=None):
    """Main entry: fetch today's plan and merge into monthly cache."""
    if not date_str:
        tomorrow = datetime.now() + timedelta(days=0)  # today's deliveries
        date_str = tomorrow.strftime("%d/%m/%Y")

    parts = date_str.split("/")
    month = int(parts[1])
    year = int(parts[2])

    print(f"  📋 Incremental plan fetch for {date_str}...")

    krc_rows = fetch_krc_today(date_str)
    print(f"    KRC: {len(krc_rows)} rows")

    dry_rows = fetch_dry_today(date_str)
    print(f"    DRY: {len(dry_rows)} rows")

    dm_rows = fetch_dongmat_today(date_str)
    print(f"    ĐÔNG MÁT: {len(dm_rows)} rows")

    all_rows = krc_rows + dry_rows + dm_rows
    plan_data = merge_into_monthly(date_str, all_rows, month, year)

    return plan_data


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Incremental plan fetch for today")
    parser.add_argument("--date", default=None, help="Date DD/MM/YYYY (default: today)")
    args = parser.parse_args()

    print("=" * 60)
    fetch_and_merge(args.date)
    print("=" * 60)
