# -*- coding: utf-8 -*-
"""
analyze_store_metrics.py — Phân tích SKU/Items/Kg/Rổ-Tote-Kiện cho 5 siêu thị × 4 kho

Usage:
  python script/domains/performance/analyze_store_metrics.py
"""
import os, sys, glob, warnings
from collections import defaultdict
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning)
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
DATA_DIR = os.path.join(BASE, "data", "raw", "daily")
SHARED_DIR = os.path.join(BASE, "data", "shared")
TRIP_DIR = r"G:\My Drive\DOCS\DAILY\DS chi tiet chuyen xe\T04.26"
OUTPUT_DIR = os.path.join(BASE, "output")

# ── Target stores ──
TARGETS_TRANSFER = {
    "KFM_HCM_Q07 - R13 Hưng Vượng 2": "HV2",
    "KFM_HCM_Q07 - 571 Huỳnh Tấn Phát": "HTP",
    "KFM_HCM_Q07 - B0.02 Scenic Valley 2": "SCV",
    "KFM_HCM_TDU - S07.02 Vinhomes Grand Park": "VH4",
    "KFM_HCM_TDU - 222 Lê Văn Thịnh": "LVH",
}
TARGET_CODES = list(TARGETS_TRANSFER.values())  # ['HV2','HTP','SCV','VH4','LVH']

# Warehouse mapping from transfer
WH_MAP_TRANSFER = {
    "KHO ABA QUÁ CẢNH": "ĐÔNG MÁT",
    "KHO ABA MIỀN ĐÔNG": "THỊT CÁ",
    "KHO RAU CỦ": "KRC",
}

# Container type mapping for trip data
CONTAINER_MAP = {
    # ĐÔNG MÁT (QCABA)
    "Rổ ABA đông mát": ("ĐÔNG MÁT", "Rổ"),
    "Tote ABA đông mát": ("ĐÔNG MÁT", "Tote"),
    "Thùng Carton, Bịch nguyên": (None, "Kiện"),  # shared across warehouses
    # KRC
    "KRC Rổ nhựa đen xếp chồng quai đỏ": ("KRC", "Rổ"),
    "Tote đỏ bánh tươi": ("KRC", "Tote"),
    # KSL
    "Seedlog - Thùng tote xanh lá": ("KSL", "Tote"),
    "Pallet có logo, chữ Seedlog": ("KSL", "Pallet"),
    # Others (skip)
    "Kiện đổi trả": (None, None),
    "Hàng lỗi không giao được (Xì, bể, hư hỏng)": (None, None),
}


def parse_date_str(s):
    s = str(s).strip().split(" ")[0]
    for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ════════════════════════════════════════════
# 1. TRANSFER DATA → ĐÔNG MÁT, THỊT CÁ, KRC
# ════════════════════════════════════════════
def load_transfer_data():
    """Load all April transfer files, return daily metrics per (store, kho)."""
    files = sorted(glob.glob(os.path.join(DATA_DIR, "transfer_*042026.xlsx")))
    print(f"📥 Loading {len(files)} transfer files...")

    # daily_metrics[(date, store_code, kho)] = {sku_set, items, kg}
    daily = defaultdict(lambda: {"skus": set(), "items": 0.0, "kg": 0.0})

    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            df = pd.read_excel(fpath)
        except Exception as e:
            print(f"  ⚠ Error reading {fname}: {e}")
            continue

        # Filter target stores & exclude thùng rổ
        mask = (df["Chi nhánh nhận"].isin(TARGETS_TRANSFER)) & (df["Là PT thùng rổ"] != "Có")
        sub = df[mask].copy()

        for _, row in sub.iterrows():
            store = TARGETS_TRANSFER[row["Chi nhánh nhận"]]
            kho = WH_MAP_TRANSFER.get(row["Chi nhánh chuyển"])
            if not kho:
                continue

            date_str = parse_date_str(row["Ngày chuyển hàng"])
            if not date_str:
                continue

            key = (date_str, store, kho)
            daily[key]["skus"].add(row["Mã hàng"])
            daily[key]["items"] += float(row["Số lượng chuyển"] or 0)

            wt_val = row.get("Giá trị trọng lượng / thể tích")
            wt_unit = str(row.get("Đơn vị trọng lượng / thể tích", "")).strip().lower()
            qty = float(row["Số lượng chuyển"] or 0)
            if pd.notna(wt_val) and wt_unit in ("g", "ml"):
                daily[key]["kg"] += float(wt_val) * qty / 1000.0

        print(f"  ✅ {fname}: {len(sub)} rows")

    # Convert sku sets to counts
    result = {}
    for key, val in daily.items():
        result[key] = {"sku": len(val["skus"]), "items": val["items"], "kg": val["kg"]}

    print(f"  📊 Total: {len(result)} (date, store, kho) combinations")
    return result


# ════════════════════════════════════════════
# 2. KSL DATA → from yeu_cau files sheet KF
# ════════════════════════════════════════════
def load_master_weight():
    """Load master data for weight lookup by barcode."""
    md_path = os.path.join(SHARED_DIR, "master_data.xlsx")
    md = pd.read_excel(md_path, usecols=["Barcode", "Net weight", "Giá trị trọng lượng/ thể tích", "Đơn vị trọng lượng/ thể tích"])
    weight_map = {}
    for _, r in md.iterrows():
        bc = str(r["Barcode"]).strip()
        # Net weight is in grams → convert to kg
        nw = r.get("Net weight")
        if pd.notna(nw) and float(nw) > 0:
            weight_map[bc] = float(nw) / 1000.0  # grams → kg
        else:
            val = r.get("Giá trị trọng lượng/ thể tích")
            unit = str(r.get("Đơn vị trọng lượng/ thể tích", "")).strip().lower()
            if pd.notna(val) and unit in ("g", "ml"):
                weight_map[bc] = float(val) / 1000.0  # g/ml → kg per unit
    print(f"  📦 Master weight loaded: {len(weight_map)} barcodes")
    return weight_map


def load_ksl_data(weight_map):
    """Load KSL item-level data from sheet KF."""
    files = sorted(glob.glob(os.path.join(DATA_DIR, "yeu_cau_chuyen_hang_thuong_*042026.xlsx")))
    print(f"\n📥 Loading {len(files)} KSL files (sheet KF)...")

    target_names = set(TARGETS_TRANSFER.keys())
    daily = defaultdict(lambda: {"skus": set(), "items": 0.0, "kg": 0.0})

    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            df = pd.read_excel(fpath, sheet_name="KF", header=0)
        except Exception as e:
            print(f"  ⚠ Error reading KF sheet from {fname}: {e}")
            continue

        if "Nơi nhận" not in df.columns:
            print(f"  ⚠ No 'Nơi nhận' column in {fname}")
            continue

        sub = df[df["Nơi nhận"].isin(target_names)]
        if len(sub) == 0:
            print(f"  ⏭ {fname}: 0 target rows")
            continue

        for _, row in sub.iterrows():
            store = TARGETS_TRANSFER[row["Nơi nhận"]]
            date_str = parse_date_str(row.get("Ngày chuyển mong muốn", ""))
            if not date_str:
                continue

            key = (date_str, store, "KSL")
            bc = str(row.get("Barcode", "")).strip()
            daily[key]["skus"].add(bc)
            qty = float(row.get("Số lượng cần chuyển", 0) or 0)
            daily[key]["items"] += qty

            if bc in weight_map:
                daily[key]["kg"] += weight_map[bc] * qty

        print(f"  ✅ {fname}: {len(sub)} target rows")

    result = {}
    for key, val in daily.items():
        result[key] = {"sku": len(val["skus"]), "items": val["items"], "kg": val["kg"]}

    print(f"  📊 Total KSL: {len(result)} (date, store) combinations")
    return result


# ════════════════════════════════════════════
# 3. TRIP DATA → Rổ/Tote/Kiện
# ════════════════════════════════════════════
def load_trip_data():
    """Load trip detail data for container counts."""
    files = sorted([f for f in os.listdir(TRIP_DIR) if f.endswith('.xlsx') and not f.startswith('~')])
    print(f"\n📥 Loading {len(files)} trip files...")

    # daily_containers[(date, store, kho)] = {"Rổ": x, "Tote": y, "Kiện": z}
    daily = defaultdict(lambda: {"Rổ": 0, "Tote": 0, "Kiện": 0, "Pallet": 0})

    for fname in files:
        fpath = os.path.join(TRIP_DIR, fname)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb["Sheet 1"]
        except Exception as e:
            print(f"  ⚠ Error reading {fname}: {e}")
            continue

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            store_code = str(row[9] or "").strip()
            if store_code not in TARGET_CODES:
                continue

            noi_chuyen = str(row[8] or "").strip()
            container_name = str(row[18] or "").strip()
            qty_raw = row[19]
            date_str = parse_date_str(row[5])

            if not date_str or not container_name:
                continue

            try:
                qty = int(qty_raw) if qty_raw else 0
            except (ValueError, TypeError):
                qty = 0

            # Determine kho and container type
            ct_info = CONTAINER_MAP.get(container_name)
            if not ct_info or ct_info[1] is None:
                continue

            kho_from_ct, ct_type = ct_info

            # "Thùng Carton, Bịch nguyên" is shared - determine kho from noi_chuyen
            if kho_from_ct is None:
                if noi_chuyen == "QCABA":
                    kho = "ĐÔNG MÁT"
                elif noi_chuyen == "KRC":
                    kho = "KRC"
                elif noi_chuyen == "KSL":
                    kho = "KSL"
                else:
                    continue
            else:
                kho = kho_from_ct

            key = (date_str, store_code, kho)
            daily[key][ct_type] += qty
            count += 1

        wb.close()
        print(f"  ✅ {fname}: {count} target rows")

    print(f"  📊 Total trip: {len(daily)} (date, store, kho) combinations")
    return dict(daily)


# ════════════════════════════════════════════
# 4. AGGREGATE → min/max/avg
# ════════════════════════════════════════════
def aggregate_metrics(transfer_data, ksl_data, trip_data):
    """Aggregate daily data into min/max/avg per (store, kho)."""
    # Merge transfer + KSL into one dict
    all_item_data = {}
    all_item_data.update(transfer_data)
    all_item_data.update(ksl_data)

    # Group by (store, kho) → list of daily values
    item_groups = defaultdict(list)  # (store, kho) → [{sku, items, kg}, ...]
    for (date, store, kho), vals in all_item_data.items():
        item_groups[(store, kho)].append(vals)

    trip_groups = defaultdict(list)  # (store, kho) → [{Rổ, Tote, Kiện}, ...]
    for (date, store, kho), vals in trip_data.items():
        trip_groups[(store, kho)].append(vals)

    # Calculate min/max/avg
    results = []
    for store in TARGET_CODES:
        for kho in ["ĐÔNG MÁT", "THỊT CÁ", "KRC", "KSL"]:
            row = {"store": store, "kho": kho}

            items_list = item_groups.get((store, kho), [])
            if items_list:
                row["days_item"] = len(items_list)
                for metric in ["sku", "items", "kg"]:
                    vals = [d[metric] for d in items_list]
                    row[f"{metric}_min"] = min(vals)
                    row[f"{metric}_max"] = max(vals)
                    row[f"{metric}_avg"] = sum(vals) / len(vals)
            else:
                row["days_item"] = 0

            trips_list = trip_groups.get((store, kho), [])
            if trips_list:
                row["days_trip"] = len(trips_list)
                for ct in ["Rổ", "Tote", "Kiện", "Pallet"]:
                    vals = [d.get(ct, 0) for d in trips_list]
                    row[f"{ct}_min"] = min(vals)
                    row[f"{ct}_max"] = max(vals)
                    row[f"{ct}_avg"] = sum(vals) / len(vals)
            else:
                row["days_trip"] = 0

            results.append(row)

    return results


# ════════════════════════════════════════════
# 5. WRITE EXCEL
# ════════════════════════════════════════════
def write_excel(results, output_path):
    """Write results to formatted Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tổng hợp"

    # Styles
    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_fill2 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    hdr_fill3 = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=11)
    center = Alignment(horizontal="center", vertical="center")
    right_al = Alignment(horizontal="right", vertical="center")
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Headers - Row 1 (group headers)
    group_headers = [
        ("", 1, 1, hdr_fill),  # Store
        ("", 2, 2, hdr_fill),  # Kho
        ("", 3, 3, hdr_fill),  # Số ngày
        ("SKU", 4, 6, hdr_fill),
        ("Items", 7, 9, hdr_fill),
        ("Quy đổi Kg", 10, 12, hdr_fill),
        ("", 13, 13, hdr_fill2),  # Số ngày trip
        ("Rổ", 14, 16, hdr_fill2),
        ("Tote", 17, 19, hdr_fill2),
        ("Kiện (Thùng Carton)", 20, 22, hdr_fill2),
        ("Pallet", 23, 25, hdr_fill3),
    ]

    for label, start, end, fill in group_headers:
        if start == end:
            continue
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=label)
        cell.font = hdr_font
        cell.fill = fill
        cell.alignment = hdr_align
        cell.border = thin

    # Row 2 (detail headers)
    detail_headers = [
        "Siêu thị", "Kho", "Số ngày\n(item)",
        "Min", "Max", "TB",
        "Min", "Max", "TB",
        "Min", "Max", "TB",
        "Số ngày\n(trip)",
        "Min", "Max", "TB",
        "Min", "Max", "TB",
        "Min", "Max", "TB",
        "Min", "Max", "TB",
    ]

    fills = [hdr_fill]*3 + [hdr_fill]*9 + [hdr_fill2]*10 + [hdr_fill3]*3
    for i, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = hdr_font
        cell.fill = fills[i-1] if i-1 < len(fills) else hdr_fill
        cell.alignment = hdr_align
        cell.border = thin

    # Data rows
    for row_idx, r in enumerate(results, 3):
        vals = [
            r["store"], r["kho"], r.get("days_item", 0),
            r.get("sku_min", ""), r.get("sku_max", ""), f'{r.get("sku_avg", 0):.0f}' if r.get("days_item") else "",
            r.get("items_min", ""), r.get("items_max", ""), f'{r.get("items_avg", 0):.0f}' if r.get("days_item") else "",
            f'{r.get("kg_min", 0):.1f}' if r.get("days_item") else "",
            f'{r.get("kg_max", 0):.1f}' if r.get("days_item") else "",
            f'{r.get("kg_avg", 0):.1f}' if r.get("days_item") else "",
            r.get("days_trip", 0),
            r.get("Rổ_min", ""), r.get("Rổ_max", ""), f'{r.get("Rổ_avg", 0):.0f}' if r.get("days_trip") else "",
            r.get("Tote_min", ""), r.get("Tote_max", ""), f'{r.get("Tote_avg", 0):.0f}' if r.get("days_trip") else "",
            r.get("Kiện_min", ""), r.get("Kiện_max", ""), f'{r.get("Kiện_avg", 0):.0f}' if r.get("days_trip") else "",
            r.get("Pallet_min", ""), r.get("Pallet_max", ""), f'{r.get("Pallet_avg", 0):.0f}' if r.get("days_trip") else "",
        ]

        is_even = row_idx % 2 == 0
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = thin
            cell.alignment = center if col_idx <= 3 or col_idx == 13 else right_al
            if is_even:
                cell.fill = even_fill

    # Auto-width
    for col in range(1, 26):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 20)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(25)}{len(results) + 2}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")


# ════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  PHÂN TÍCH 5 SIÊU THỊ × 4 KHO — THÁNG 4/2026")
    print("=" * 60)

    # 1. Load master weight data
    print("\n📦 Loading master data for weight lookup...")
    weight_map = load_master_weight()

    # 2. Load transfer data
    print("\n" + "─" * 40)
    transfer = load_transfer_data()

    # 3. Load KSL data
    print("\n" + "─" * 40)
    ksl = load_ksl_data(weight_map)

    # 4. Load trip data
    print("\n" + "─" * 40)
    trip = load_trip_data()

    # 5. Aggregate
    print("\n" + "─" * 40)
    print("📊 Aggregating min/max/avg...")
    results = aggregate_metrics(transfer, ksl, trip)

    # 6. Print summary
    print("\n" + "=" * 60)
    print("  SUMMARY")
    print("=" * 60)
    for r in results:
        if r.get("days_item", 0) > 0 or r.get("days_trip", 0) > 0:
            print(f"  {r['store']} | {r['kho']:10s} | "
                  f"SKU: {r.get('sku_min','')}-{r.get('sku_max','')} (avg {r.get('sku_avg',0):.0f}) | "
                  f"Items: {r.get('items_min','')}-{r.get('items_max','')} (avg {r.get('items_avg',0):.0f}) | "
                  f"Kg: {r.get('kg_avg',0):.1f} | "
                  f"Rổ: {r.get('Rổ_avg',0):.0f} Tote: {r.get('Tote_avg',0):.0f} Kiện: {r.get('Kiện_avg',0):.0f}")

    # 7. Write Excel
    output_path = os.path.join(OUTPUT_DIR, "PHAN_TICH_5ST_4KHO_T04_2026.xlsx")
    print("\n📝 Writing Excel...")
    write_excel(results, output_path)

    print(f"\n{'=' * 60}")
    print(f"  ✅ DONE!")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
