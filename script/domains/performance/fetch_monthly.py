"""
fetch_monthly_plan.py - Fetch planned delivery schedule + Tuyến for a full month.

Fetches planned_time and Tuyến from Google Sheets sources:
- KRC:      sheet "KRC" or "TGGH DỰ KIẾN KRC"
- DRY:      sheet "STS - KHÔ"
- ĐÔNG MÁT: KH Drive folders (col J = Tuyen)
- THỊT CÁ:  sheet "ABA MĐ - THỊT CÁ"

Usage:
  python script/fetch_monthly_plan.py --month 03 --year 2026
  python script/fetch_monthly_plan.py   # defaults to March 2026

Output: output/monthly_plan_T{month}.json
"""
import os, sys, re, json
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
OUTPUT = os.path.join(BASE, "output")

# ── Online sources ──
KRC_SHEET_URL = "https://docs.google.com/spreadsheets/d/1tWamqjpOI2j2MrYW3Ah6ptmT524CAlQvEP8fCkxfuII/export?format=xlsx"
KFM_SHEET_URL = "https://docs.google.com/spreadsheets/d/1LkJFJhOQ8F2WEB3uCk7kA2Phvu8IskVi3YBfVr7pBx0/export?format=xlsx"

# Drive folders per category
KH_DRIVE_FOLDERS = [
    ("KH HÀNG ĐÔNG", "ĐÔNG MÁT", "https://drive.google.com/drive/folders/1pQ8coYeV-K0dcHlkvXcJ8KngmH22xp1Z"),
    ("KH HÀNG MÁT", "ĐÔNG MÁT", "https://drive.google.com/drive/folders/1c2zfgcXM8O9ezkOZYj0p4t_ihaJmb98f"),
]


def read_xlsx_from_url(url):
    import requests
    r = requests.get(url, allow_redirects=True, timeout=120)
    r.raise_for_status()
    return load_workbook(BytesIO(r.content), read_only=True, data_only=True)


def format_time(time_val):
    """Convert time to HH:MM string."""
    if not time_val:
        return ""
    s = str(time_val).strip()
    m = re.match(r'(\d{1,2}):(\d{2})', s)
    if m:
        return f"{int(m.group(1))}:{m.group(2)}"
    if hasattr(time_val, 'hour'):
        return f"{time_val.hour}:{time_val.minute:02d}"
    # Excel time fraction
    try:
        fval = float(time_val)
        if 0 <= fval < 1:
            total_minutes = round(fval * 24 * 60)
            hours = total_minutes // 60
            minutes = total_minutes % 60
            return f"{hours}:{minutes:02d}"
    except (ValueError, TypeError):
        pass
    # timedelta format "H:MM:SS"
    m2 = re.match(r'(\d+):(\d{2}):(\d{2})', s)
    if m2:
        return f"{int(m2.group(1))}:{m2.group(2)}"
    return s


def format_date_vn(dt):
    return dt.strftime("%d/%m/%Y")


def _list_drive_folder(folder_url):
    import requests as _req
    import html as _html
    folder_id = re.search(r'folders/([a-zA-Z0-9_-]+)', folder_url).group(1)
    url = f"https://drive.google.com/drive/folders/{folder_id}"
    r = _req.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    text = _html.unescape(r.text)
    files = []
    seen = set()
    for m in re.finditer(r'"(1[a-zA-Z0-9_-]{25,50})"', text):
        fid = m.group(1)
        if fid in seen:
            continue
        after = text[m.end():m.end() + 500]
        name_match = re.search(r'"([^"]+\.xlsx)"', after)
        if name_match:
            fname = name_match.group(1)
            if "Microsoft" not in fname and "Shared" not in fname:
                seen.add(fid)
                files.append((fid, fname))
    return files


def read_kh_from_drive(folder_url, date_for_file):
    import requests as _req
    files = _list_drive_folder(folder_url)
    if not files:
        return None
    target = None
    for fid, fname in files:
        if date_for_file in fname:
            target = (fid, fname)
            break
    if not target:
        return None
    dl_url = f"https://drive.google.com/uc?export=download&confirm=t&id={target[0]}"
    r = _req.get(dl_url, allow_redirects=True, timeout=60)
    if r.status_code != 200 or len(r.content) < 500:
        return None
    return load_workbook(BytesIO(r.content), read_only=True, data_only=True)


def fetch_krc_plan(date_strs):
    """Fetch KRC planned data: date, store, planned_time, tuyen."""
    print("\n  → KRC planned data...")
    rows = []
    try:
        wb = read_xlsx_from_url(KRC_SHEET_URL)
        # Try "KRC" sheet first, fallback to "TGGH DỰ KIẾN KRC"
        ws = None
        for name in wb.sheetnames:
            if name == "KRC":
                ws = wb[name]
                break
        if not ws:
            for name in wb.sheetnames:
                if "KRC" in name:
                    ws = wb[name]
                    break

        for row in ws.iter_rows(min_row=2, values_only=False):
            date_val = str(row[0].value or "").strip()
            if date_val not in date_strs:
                continue
            store = str(row[6].value or "").strip()  # Col G = Điểm đến
            planned_time = format_time(row[7].value)  # Col H = Giờ đến dự kiến
            tuyen = str(row[10].value or "").strip() if len(row) > 10 else ""  # Col K = Tuyến
            if store:
                rows.append({
                    "date": date_val,
                    "store": store,
                    "planned_time": planned_time,
                    "tuyen": tuyen,
                    "kho": "KRC",
                })
        wb.close()
        print(f"    ✅ {len(rows)} rows")
    except Exception as e:
        print(f"    ⚠ Error: {e}")
    return rows


def fetch_dry_plan(date_strs):
    """Fetch DRY/KSL planned data."""
    print("\n  → DRY/KSL planned data...")
    rows = []
    try:
        wb = read_xlsx_from_url(KFM_SHEET_URL)
        # Find DRY sheet
        ws = None
        for name in wb.sheetnames:
            if "KHÔ" in name or "DRY" in name:
                ws = wb[name]
                print(f"    Using sheet: {name}")
                break
        if not ws:
            ws = wb.worksheets[1]

        # Check header for column positions
        # Standard: Col 0=Date, Col 5=Điểm đến, Col 6=Giờ đến dự kiến, Col 9=Tuyến
        # But some sheets: Col 6=Điểm đến, Col 7=Giờ đến dự kiến, Col 10=Tuyến
        for row in ws.iter_rows(min_row=3, values_only=False):
            date_val = str(row[0].value or "").strip()
            if date_val not in date_strs:
                continue
            # Try col 5 first, then col 6
            store_5 = str(row[5].value or "").strip() if len(row) > 5 else ""
            store_6 = str(row[6].value or "").strip() if len(row) > 6 else ""
            time_6 = format_time(row[6].value) if len(row) > 6 else ""
            time_7 = format_time(row[7].value) if len(row) > 7 else ""
            tuyen_9 = str(row[9].value or "").strip() if len(row) > 9 else ""
            tuyen_10 = str(row[10].value or "").strip() if len(row) > 10 else ""

            # Determine which col layout
            # If col 6 looks like a store code (short alpha), use cols 6/7/10
            # else use cols 5/6/9
            if store_6 and len(store_6) <= 5 and store_6.isalnum():
                store = store_6
                planned_time = time_7
                tuyen = tuyen_10
            elif store_5 and len(store_5) <= 5 and store_5.isalnum():
                store = store_5
                planned_time = time_6
                tuyen = tuyen_9
            else:
                continue

            if store:
                rows.append({
                    "date": date_val,
                    "store": store,
                    "planned_time": planned_time,
                    "tuyen": tuyen,
                    "kho": "DRY",
                })
        wb.close()
        print(f"    ✅ {len(rows)} rows")
    except Exception as e:
        print(f"    ⚠ Error: {e}")
    return rows


def fetch_thitca_plan(date_strs):
    """Fetch THỊT CÁ planned from 'ABA MĐ - THỊT CÁ' sheet in KFM spreadsheet."""
    print("\n  → THỊT CÁ planned data...")
    rows = []
    try:
        wb = read_xlsx_from_url(KFM_SHEET_URL)
        ws = None
        # Prefer "ABA" sheet which has Tuyen column
        for name in wb.sheetnames:
            if "ABA" in name and ("THỊT CÁ" in name or "THIT CA" in name or "THỊT" in name):
                ws = wb[name]
                print(f"    Using sheet: {name}")
                break
        if not ws:
            for name in wb.sheetnames:
                if "THỊT CÁ" in name or "THIT CA" in name:
                    ws = wb[name]
                    print(f"    Using sheet: {name}")
                    break
        if not ws:
            print("    ⚠ Sheet not found")
            return rows

        # Structure: Col 0=Ngày, Col 2=StoreID, Col 3=Tuyen, Col 5=Dự kiến giao, Col 6=TG tới ST
        for row in ws.iter_rows(min_row=3, values_only=False):
            date_val = row[0].value
            if not date_val:
                continue
            # Normalize date
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime("%d/%m/%Y")
            else:
                date_str = str(date_val).strip()
                # Try DD/MM/YYYY format
                m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', date_str)
                if m:
                    date_str = f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
                # Try YYYY-MM-DD format
                m2 = re.match(r'(\d{4})-(\d{2})-(\d{2})', date_str)
                if m2:
                    date_str = f"{m2.group(3)}/{m2.group(2)}/{m2.group(1)}"

            if date_str not in date_strs:
                continue

            store = str(row[2].value or "").strip() if len(row) > 2 else ""
            tuyen = str(row[3].value or "").strip() if len(row) > 3 else ""
            planned_time = format_time(row[5].value) if len(row) > 5 else ""

            if store:
                rows.append({
                    "date": date_str,
                    "store": store,
                    "planned_time": planned_time,
                    "tuyen": tuyen,
                    "kho": "THỊT CÁ",
                })
        wb.close()
        print(f"    ✅ {len(rows)} rows")
    except Exception as e:
        print(f"    ⚠ Error: {e}")
    return rows


def fetch_thitca_external(month, year):
    """Read THỊT CÁ actual data from external file (BÁO CÁO GIAO HÀNG MIỀN ĐÔNG)."""
    print("\n  → THỊT CÁ external (BÁO CÁO GIAO HÀNG)...")
    rows = []
    paths = [
        rf'G:\My Drive\DOCS\DAILY\BÁO CÁO GIAO HÀNG MIỀN ĐÔNG\{year}.{month:02d} BÁO CÁO GIAO HÀNG KINGFOOD.xlsx',
        rf'G:\My Drive\DOCS\DAILY\BÁO CÁO GIAO HÀNG MIỀN ĐÔNG\{month:02d}.{year} BAO CAO GIAO HANG KINGFOOD.xlsx',
        rf'G:\My Drive\DOCS\DAILY\BÁO CÁO GIAO HÀNG MIỀN ĐÔNG\{year}.{month:02d} BAO CAO GIAO HANG KINGFOOD.xlsx',
        rf'G:\My Drive\DOCS\DAILY\BÁO CÁO GIAO HÀNG MIỀN ĐÔNG\{month:02d}.{year} BÁO CÁO GIAO HÀNG KINGFOOD.xlsx',
    ]
    path = None
    for p in paths:
        if os.path.exists(p):
            path = p
            break
    if not path:
        print(f"    ⚠ File not found, tried: {paths[0]}")
        return rows

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        # Auto-detect sheet: try Sheet1, else use first sheet
        # T03 uses "Sheet1", T04 uses "KINGFOOD MEAT"
        ws = None
        if 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
        else:
            ws = wb[wb.sheetnames[0]]
        print(f"    Using sheet: {ws.title}")
        
        # Auto-detect column layout from header
        # T03: A=Ngày B=Mã CH E=Tuyến H=Dự kiến I=TG đến
        # T04: C=Ngày D=Mã CH G=Tuyến J=Dự kiến K=TG đến
        date_col, store_col, tuyen_col, plan_col, actual_col = 0, 1, 4, 7, 8
        for hdr in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, c in enumerate(hdr):
                h = str(c or "").strip().upper()
                if "NGÀY" in h or "NGAY" in h:
                    date_col = i
                elif "MÃ" in h and ("CH" in h or "ĐIỂM" in h or "DIEM" in h):
                    store_col = i
                elif "TUYẾN" in h or "TUYEN" in h:
                    tuyen_col = i
                elif "DỰ KIẾN" in h or "DU KIEN" in h:
                    plan_col = i
                elif "TG ĐẾN" in h or "TG DEN" in h or "ĐẾN CỬA" in h:
                    actual_col = i
        print(f"    Columns: date={date_col}, store={store_col}, tuyen={tuyen_col}, plan={plan_col}, actual={actual_col}")
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = row[date_col] if len(row) > date_col else None
            if not date_val:
                continue
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime("%d/%m/%Y")
            else:
                continue

            store = str(row[store_col] or "").strip() if len(row) > store_col else ""
            tuyen = str(row[tuyen_col] or "").strip() if len(row) > tuyen_col else ""
            planned_time = format_time(row[plan_col]) if len(row) > plan_col and row[plan_col] else ""
            actual_time = format_time(row[actual_col]) if len(row) > actual_col and row[actual_col] else ""

            if store and actual_time:
                rows.append({
                    "date": date_str,
                    "store": store,
                    "tuyen": tuyen,
                    "planned_time": planned_time,
                    "actual_time": actual_time,
                    "kho": "THỊT CÁ",
                })
        wb.close()
        print(f"    ✅ {len(rows)} rows")
    except Exception as e:
        print(f"    ⚠ Error: {e}")
    return rows


def fetch_dongmat_plan(date_strs, month, year):
    """Fetch ĐÔNG MÁT planned data from KH Drive folders — includes Tuyen (col J)."""
    print("\n  → ĐÔNG MÁT planned data (KH Drive)...")
    rows = []
    # Parse dates for file matching
    for date_str in date_strs:
        parts = date_str.split("/")
        dt = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
        date_for_file = dt.strftime("%d.%m.%Y")

        for folder_name, kho_name, folder_url in KH_DRIVE_FOLDERS:
            try:
                wb = read_kh_from_drive(folder_url, date_for_file)
                # Fallback: read from local Google Drive sync if API failed
                if not wb:
                    local_dir = rf"G:\My Drive\DOCS\DAILY\{folder_name}"
                    if os.path.exists(local_dir):
                        for fname in os.listdir(local_dir):
                            if date_for_file in fname and fname.endswith('.xlsx') and not fname.startswith('~'):
                                local_path = os.path.join(local_dir, fname)
                                try:
                                    wb = load_workbook(local_path, read_only=True, data_only=True)
                                    print(f"    ↳ Local fallback: {fname}")
                                except Exception:
                                    pass
                                break
                if not wb:
                    continue
                ws = wb.worksheets[0]

                # Find column indices from header
                tuyen_col = 9  # Col J = index 9 (default)
                time_col = None
                time_col_backup = None  # fallback for "TG DỰ KIẾN" without "giao"
                for hdr_row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                    for i, cell in enumerate(hdr_row):
                        h = str(cell.value or "").strip().lower()
                        # Tuyen: match "Tuyen" or "Tuyến" but NOT "Kg/Tuyến"
                        if ("tuyen" in h or "tuyến" in h) and "kg" not in h:
                            tuyen_col = i
                        # Planned time: prefer "du kien giao" / "dự kiến giao"
                        if time_col is None and ("kien" in h or "kiến" in h):
                            if "giao" in h:
                                time_col = i  # Best match: contains "giao"
                            elif ("roi" not in h and "rời" not in h) and time_col_backup is None:
                                time_col_backup = i  # Backup: "TG DỰ KIẾN" (exclude "rời kho")
                if time_col is None:
                    time_col = time_col_backup

                for row in ws.iter_rows(min_row=2, values_only=False):
                    store = str(row[2].value or "").strip() if len(row) > 2 else ""
                    planned_time = ""
                    if time_col is not None and time_col < len(row):
                        planned_time = format_time(row[time_col].value)
                    tuyen_raw = row[tuyen_col].value if tuyen_col < len(row) else ""
                    tuyen = ""
                    if tuyen_raw:
                        tuyen_str = str(tuyen_raw).strip()
                        if tuyen_str:
                            # Numeric-only tuyến (HÀNG ĐÔNG: 1-6) → prefix "HĐ" to distinguish
                            # from HÀNG MÁT routes (KF01, QC02...)
                            try:
                                int(float(tuyen_str))
                                tuyen = f"HĐ{int(float(tuyen_str))}"
                            except (ValueError, TypeError):
                                tuyen = tuyen_str

                    if store:
                        rows.append({
                            "date": date_str,
                            "store": store,
                            "planned_time": planned_time,
                            "tuyen": tuyen,
                            "kho": "ĐÔNG MÁT",
                        })
                wb.close()
            except Exception as e:
                print(f"    ⚠ {folder_name} {date_str}: {e}")

    print(f"    ✅ {len(rows)} rows")
    return rows


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Fetch monthly plan + Tuyến data")
    parser.add_argument("--month", type=int, default=3)
    parser.add_argument("--year", type=int, default=2026)
    args = parser.parse_args()

    month = args.month
    year = args.year

    # Generate all dates in the month
    from calendar import monthrange
    _, days_in_month = monthrange(year, month)
    dates = [datetime(year, month, d) for d in range(1, days_in_month + 1)]
    date_strs = [format_date_vn(d) for d in dates]

    # Also include US format and other variants for matching
    date_strs_us = [f"{d.month}/{d.day}/{d.year}" for d in dates]
    date_strs_alt = [d.strftime("%d/%m/%Y") for d in dates]

    # Combine all formats for matching
    all_date_strs = set(date_strs + date_strs_us + date_strs_alt)

    print("=" * 60)
    print(f"  FETCH MONTHLY PLAN — T{month:02d}/{year}")
    print(f"  {format_date_vn(dates[0])} → {format_date_vn(dates[-1])}")
    print("=" * 60)

    # Fetch each source
    krc_rows = fetch_krc_plan(all_date_strs)
    dry_rows = fetch_dry_plan(all_date_strs)
    thitca_plan = fetch_thitca_plan(all_date_strs)
    thitca_external = fetch_thitca_external(month, year)

    # ĐÔNG MÁT from Drive (expensive — many HTTP requests)
    dongmat_rows = fetch_dongmat_plan(date_strs, month, year)

    # Combine
    plan_data = {
        "KRC": krc_rows,
        "DRY": dry_rows,
        "ĐÔNG MÁT": dongmat_rows,
        "THỊT CÁ": thitca_plan,
    }

    result = {
        "month": f"T{month:02d}",
        "year": year,
        "start": format_date_vn(dates[0]),
        "end": format_date_vn(dates[-1]),
        "plan": plan_data,
        "thitca_actual": thitca_external,  # THỊT CÁ actual from external file
    }

    # Save
    state_dir = os.path.join(OUTPUT, "state")
    os.makedirs(state_dir, exist_ok=True)
    out_path = os.path.join(state_dir, f"monthly_plan_T{month:02d}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n{'=' * 60}")
    print(f"  📊 Summary:")
    for kho, rows in plan_data.items():
        print(f"    {kho}: {len(rows)} plan rows")
    print(f"    THỊT CÁ external: {len(thitca_external)} rows")
    print(f"\n  ✅ Saved: {out_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
