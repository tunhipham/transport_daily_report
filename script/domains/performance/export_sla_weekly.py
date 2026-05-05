"""
Export SLA On-Time weekly summary to Excel.

Auto-detects ISO weeks from the months provided.
Produces two files:
  1. SLA_ONTIME_W{...}.xlsx          — all kho
  2. SLA_ONTIME_DM_TC_W{...}.xlsx    — ĐÔNG MÁT + THỊT CÁ only

Usage:
  python script/domains/performance/export_sla_weekly.py --months 3,4,5 --year 2026
  python script/domains/performance/export_sla_weekly.py --months 4,5           # year defaults 2026
  python script/domains/performance/export_sla_weekly.py --months 4,5 --weeks 14,15,16,17,18  # override weeks
"""
import os, sys, json, argparse
from datetime import date, timedelta, time as dtime
from calendar import monthrange
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, BASE)

from script.domains.performance.generate import (
    load_trip_data, load_thitca_data, load_plan_data,
    calc_metrics, SLA_WINDOWS, KHO_COLORS,
)

OUTPUT = os.path.join(BASE, "output")
PERF_DIR = os.path.join(OUTPUT, "artifacts", "performance")
os.makedirs(PERF_DIR, exist_ok=True)

DAY_NAMES = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'CN']


def compute_weeks_from_months(months, year):
    """Auto-detect ISO weeks that overlap with the given months.
    Returns sorted list of ISO week numbers."""
    all_dates = []
    for m in months:
        _, days = monthrange(year, m)
        all_dates.extend(date(year, m, d) for d in range(1, days + 1))

    weeks = sorted(set(d.isocalendar()[1] for d in all_dates))
    return weeks


def build_week_info(weeks, year):
    """Build week date ranges for given ISO weeks."""
    info = {}
    for w in weeks:
        monday = date.fromisocalendar(year, w, 1)
        days = [monday + timedelta(days=i) for i in range(7)]
        info[w] = {"monday": monday, "dates": days}
    return info


def make_label(weeks):
    """Build label string like 'W14_W15_W16_W17_W18'."""
    return "_".join(f"W{w}" for w in weeks)


def make_title_range(week_info, weeks):
    """Build date range string like '30/03 - 03/05/2026'."""
    first = week_info[weeks[0]]["monday"]
    last = week_info[weeks[-1]]["monday"] + timedelta(days=6)
    return f"{first.strftime('%d/%m')} - {last.strftime('%d/%m/%Y')}"


def export_sla_excel(metrics, kho_list, filename, title, weeks, week_info, metric_labels=None):
    """Export SLA on-time summary as formatted Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SLA On-Time"

    # ── Styles ──
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_font_w = Font(bold=True, size=11, color="FFFFFF")
    thin_bd = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Color fills for percentage
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")
    fill_orange = PatternFill("solid", fgColor="FCD5B4")
    fill_red = PatternFill("solid", fgColor="FFC7CE")

    def pct_fill(pct):
        if pct >= 95: return fill_green
        if pct >= 90: return fill_yellow
        if pct >= 85: return fill_orange
        return fill_red

    # ── Title row ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(weeks) * 8)
    ws.cell(1, 1, title).font = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    # ── Build header ──
    row = 3
    col = 1
    ws.cell(row, col, "KHO").font = hdr_font_w
    ws.cell(row, col).fill = hdr_fill
    ws.cell(row, col).border = thin_bd
    ws.cell(row, col).alignment = center
    col += 1
    ws.cell(row, col, "Chỉ Tiêu").font = hdr_font_w
    ws.cell(row, col).fill = hdr_fill
    ws.cell(row, col).border = thin_bd
    ws.cell(row, col).alignment = center
    col += 1

    # Week group header (row 2) + day headers (row 3)
    for w in weeks:
        info = week_info[w]
        monday = info["monday"]
        sunday = monday + timedelta(days=6)
        wk_label = f"W{w} ({monday.strftime('%d/%m')} - {sunday.strftime('%d/%m')})"

        # Merge row 2 across 8 columns (Tổng + 7 days)
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 7)
        ws.cell(2, col, wk_label).font = hdr_font_w
        ws.cell(2, col).fill = hdr_fill
        ws.cell(2, col).alignment = Alignment(horizontal="center")
        ws.cell(2, col).border = thin_bd

        # Sub-headers: Tổng + 7 day names
        sub_labels = ["Tổng"] + [f"{DAY_NAMES[i]}\n{info['dates'][i].strftime('%d/%m')}" for i in range(7)]
        for i, lbl in enumerate(sub_labels):
            c = ws.cell(row, col + i, lbl)
            c.font = hdr_font_w
            c.fill = hdr_fill
            c.border = thin_bd
            c.alignment = center

        col += 8

    # Also merge row 2 for KHO and Chỉ Tiêu
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.cell(2, 1, "KHO").font = hdr_font_w
    ws.cell(2, 1).fill = hdr_fill
    ws.cell(2, 1).alignment = center
    ws.cell(2, 1).border = thin_bd
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws.cell(2, 2, "Chỉ Tiêu").font = hdr_font_w
    ws.cell(2, 2).fill = hdr_fill
    ws.cell(2, 2).alignment = center
    ws.cell(2, 2).border = thin_bd

    # ── Data rows ──
    data_row = 4
    if metric_labels is None:
        metric_labels = ["Tổng Điểm Giao", "Đúng Giờ (SLA)", "Trễ (SLA)", "% On Time (SLA)"]

    for kho in kho_list:
        start_data_row = data_row
        for mi, metric_name in enumerate(metric_labels):
            col = 1
            if mi == 0:
                pass  # will merge after
            ws.cell(data_row, 2, metric_name).border = thin_bd
            ws.cell(data_row, 2).alignment = center
            ws.cell(data_row, 2).font = Font(bold=(metric_name.startswith("%")))
            col = 3

            for w in weeks:
                info = week_info[w]
                dates_in_week = info["dates"]

                # Weekly total
                wk_total = sum(
                    metrics["sla"].get(kho, {}).get(d, {}).get("on_time", 0) +
                    metrics["sla"].get(kho, {}).get(d, {}).get("late", 0)
                    for d in dates_in_week
                )
                wk_ontime = sum(
                    metrics["sla"].get(kho, {}).get(d, {}).get("on_time", 0)
                    for d in dates_in_week
                )
                wk_late = wk_total - wk_ontime
                wk_pct = round(wk_ontime / wk_total * 100, 1) if wk_total > 0 else ""

                # Day-level
                day_vals = []
                for d in dates_in_week:
                    sla = metrics["sla"].get(kho, {}).get(d, {})
                    ot = sla.get("on_time", 0)
                    lt = sla.get("late", 0)
                    total = ot + lt
                    pct = round(ot / total * 100, 1) if total > 0 else ""
                    day_vals.append({"total": total, "ontime": ot, "late": lt, "pct": pct})

                # Trips (tuyến) per week/day
                wk_trips = sum(
                    len(metrics["trips_per_day"].get(kho, {}).get(d, set()))
                    for d in dates_in_week
                )
                day_trips = [
                    len(metrics["trips_per_day"].get(kho, {}).get(d, set()))
                    for d in dates_in_week
                ]

                # Write values
                if metric_name == "Tổng Điểm Giao":
                    vals = [wk_total] + [dv["total"] for dv in day_vals]
                elif metric_name == "Đúng Giờ (SLA)":
                    vals = [wk_ontime] + [dv["ontime"] for dv in day_vals]
                elif metric_name == "Trễ (SLA)":
                    vals = [wk_late] + [dv["late"] for dv in day_vals]
                elif metric_name == "% On Time (SLA)":
                    vals = [wk_pct] + [dv["pct"] for dv in day_vals]
                elif metric_name == "Tổng Số Chuyến":
                    vals = [wk_trips] + day_trips

                for i, v in enumerate(vals):
                    cell = ws.cell(data_row, col + i)
                    if v == "" or v == 0:
                        if metric_name.startswith("%"):
                            cell.value = ""
                        else:
                            cell.value = 0 if v == 0 else ""
                    elif metric_name.startswith("%"):
                        cell.value = v / 100
                        cell.number_format = '0.0%'
                        cell.fill = pct_fill(v)
                    else:
                        cell.value = v
                    cell.border = thin_bd
                    cell.alignment = center

                col += 8

            data_row += 1

        # Merge KHO cell
        ws.merge_cells(start_row=start_data_row, start_column=1,
                       end_row=data_row - 1, end_column=1)
        kho_cell = ws.cell(start_data_row, 1, kho)
        kho_cell.font = Font(bold=True, size=11)
        kho_cell.alignment = Alignment(horizontal="center", vertical="center")
        kho_cell.border = thin_bd

        # Add separator border
        for c in range(1, col):
            ws.cell(data_row - 1, c).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='medium'),
            )

    # ── Column widths ──
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    for c in range(3, 3 + len(weeks) * 8):
        ws.column_dimensions[get_column_letter(c)].width = 11

    ws.freeze_panes = "C4"

    out_path = os.path.join(PERF_DIR, filename)
    wb.save(out_path)
    print(f"  ✅ Saved: {out_path}")
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Export SLA On-Time weekly summary to Excel")
    parser.add_argument("--months", type=str, required=True,
                        help="Comma-separated months, e.g. '3,4,5'")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--weeks", type=str, default=None,
                        help="Override ISO weeks, e.g. '14,15,16,17,18'. Default: auto-detect from months.")
    args = parser.parse_args()

    year = args.year
    months = [int(m.strip()) for m in args.months.split(",")]

    # Compute or parse weeks
    if args.weeks:
        weeks = [int(w.strip()) for w in args.weeks.split(",")]
    else:
        weeks = compute_weeks_from_months(months, year)

    wk_info = build_week_info(weeks, year)
    wk_label = make_label(weeks)
    wk_range = make_title_range(wk_info, weeks)

    print("=" * 60)
    print(f"  📊 SLA ON-TIME WEEKLY EXPORT — {', '.join(f'W{w}' for w in weeks)}")
    print(f"  📅 {wk_range}")
    print("=" * 60)

    # Load data
    print("\n📥 Loading data...")
    all_rows = []
    for m in months:
        all_rows.extend(load_trip_data(m, year))
    tc_rows = load_thitca_data(months)
    all_rows.extend(tc_rows)
    plan_lookup, route_order = load_plan_data(months)
    print(f"  📊 Total: {len(all_rows)} rows")

    print("\n📈 Calculating metrics...")
    metrics = calc_metrics(all_rows, plan_lookup, route_order)

    # File 1: All kho
    all_kho = ["KRC", "THỊT CÁ", "ĐÔNG MÁT", "ĐÔNG", "MÁT", "KSL-Sáng", "KSL-Tối"]
    fname1 = f"SLA_ONTIME_{wk_label}.xlsx"
    title1 = f"SLA ON-TIME — {', '.join(f'W{w}' for w in weeks)} ({wk_range})"
    print(f"\n📄 File 1: {fname1} (all kho)")
    export_sla_excel(metrics, all_kho, fname1, title1, weeks, wk_info)

    # File 2: ĐÔNG MÁT + THỊT CÁ only
    dm_tc_kho = ["ĐÔNG MÁT", "ĐÔNG", "MÁT", "THỊT CÁ"]
    fname2 = f"SLA_ONTIME_DM_TC_{wk_label}.xlsx"
    title2 = f"SLA ON-TIME ĐÔNG MÁT & THỊT CÁ — {', '.join(f'W{w}' for w in weeks)} ({wk_range})"
    print(f"\n📄 File 2: {fname2} (ĐÔNG MÁT + THỊT CÁ)")
    export_sla_excel(
        metrics, dm_tc_kho, fname2, title2, weeks, wk_info,
        metric_labels=["Tổng Điểm Giao", "Đúng Giờ (SLA)", "% On Time (SLA)", "Tổng Số Chuyến"]
    )

    print(f"\n✅ Done! Files in: {PERF_DIR}")


if __name__ == "__main__":
    main()
