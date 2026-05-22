"""
capacity_forecast.py — Generate capacity forecast data for KRC and KSL (Dry)
==============================================================================
Reads PO KRC files and yeu_cau_chuyen_hang_thuong files to compute daily
throughput, then outputs capacity_forecast.json for the dashboard.

KRC:  Total tons/day based on PO data   → benchmark 65 Tấn
KSL:  Total items/day (Sáng + Tối)      → benchmark 270,000 items

Usage:
    python script/domains/daily/capacity_forecast.py
"""
import os, sys, json, re, glob
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(BASE, "script"))

# Paths
PO_KRC_DIR = r"G:\My Drive\DOCS\DAILY\po_krc"
YECAU_LOCAL = r"G:\My Drive\DOCS\DAILY\yeu_cau_chuyen_hang_thuong"
ABA_MASTER_PATH = r"G:\My Drive\DOCS\DAILY\ton_aba\data\master_data\Master Data.xlsx"
DOCS_DATA = os.path.join(BASE, "docs", "data")

# ── Import master data URL from shared lib ──
try:
    from lib.sources import MASTER_SHEET_URL
except ImportError:
    MASTER_SHEET_URL = None

# Benchmarks
KRC_BENCHMARK_TONS = 65
KSL_BENCHMARK_ITEMS = 270000
ALERT_THRESHOLD_PCT = 5  # Alert if >5% over benchmark

NOW_STR = datetime.now().strftime("%d/%m/%Y %H:%M")


# ══════════════════════════════════════════════════════════════
# MASTER DATA — barcode → weight (grams)
# ══════════════════════════════════════════════════════════════
def load_master_weights():
    """Load barcode → per-item weight (grams) from ABA Master Data.
    Col B (1) = Mã hàng (barcode), Col J or weight column.
    Also try the online master sheet.
    """
    master = {}
    
    # 1. Try ABA Master Data local file
    if os.path.exists(ABA_MASTER_PATH):
        try:
            print("  → Loading ABA Master Data for weights...")
            wb = load_workbook(ABA_MASTER_PATH, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            for row in ws.iter_rows(min_row=2, values_only=False):
                bc = str(row[1].value or "").strip()
                # Try column with weight data — search multiple potential columns
                if bc:
                    # Check if there's a weight column — ABA master might have KG in different cols
                    for ci in [7, 8, 9, 10]:  # Try common weight columns
                        try:
                            val = row[ci].value
                            if val is not None:
                                w = float(val)
                                if w > 0 and w < 100000:  # Sanity check
                                    master[bc] = w
                                    break
                        except (ValueError, TypeError, IndexError):
                            pass
            wb.close()
            print(f"    {len(master)} barcodes with weight from ABA Master")
        except Exception as e:
            print(f"    ⚠ Error loading ABA Master: {e}")
    
    # 2. Try online master data sheet
    if MASTER_SHEET_URL:
        try:
            print("  → Loading online Master Sheet for weights...")
            from io import BytesIO
            import requests
            r = requests.get(MASTER_SHEET_URL, allow_redirects=True, timeout=120)
            r.raise_for_status()
            wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                bc = str(row[0].value or "").strip()
                if not bc:
                    continue
                tl_value = row[25].value  # Col Z = weight
                if tl_value is not None:
                    try:
                        w = float(tl_value)
                        if w > 0:
                            if bc not in master:
                                master[bc] = w
                                count += 1
                    except (ValueError, TypeError):
                        pass
            wb.close()
            print(f"    +{count} barcodes from online Master Sheet (total: {len(master)})")
        except Exception as e:
            print(f"    ⚠ Error loading online Master Sheet: {e}")
    
    return master


def extract_weight_from_name(product_name):
    """Extract weight in KG from product name as fallback."""
    if not product_name:
        return 0
    text = product_name.upper()
    patterns = [
        (r'(\d+(?:[.,]\d+)?)\s*KG\b', 1.0),
        (r'(\d+(?:[.,]\d+)?)\s*G\b', 0.001),
        (r'(\d+(?:[.,]\d+)?)\s*(?:LÍT|LIT)\b', 1.0),
        (r'(\d+(?:[.,]\d+)?)\s*L\b', 1.0),
        (r'(\d+(?:[.,]\d+)?)\s*ML\b', 0.001),
    ]
    for pattern, multiplier in patterns:
        matches = re.findall(pattern, text)
        if matches:
            try:
                return float(matches[-1].replace(",", ".")) * multiplier
            except ValueError:
                continue
    return 0


# ══════════════════════════════════════════════════════════════
# KRC — Read PO from ClickHouse DB (primary source)
# ══════════════════════════════════════════════════════════════
def read_po_krc_from_db(master_weights=None):
    """Read PO KRC capacity from ClickHouse: kf_purchase_order + kf_receipt_items.

    Fetches row-level data and applies weight fallback (same as transfer PT):
      1. ri.net_weight from receipt_items (grams)
      2. master_weights lookup by barcode (Google Sheets master data)
      3. extract_weight_from_name() from product name

    delivery_date is stored as epoch seconds in kf_purchase_order.

    Returns: dict of {date_str: total_tons}, or None on failure.
    """
    KRC_BRANCH_ID = '5fdc170ebd89c10006f15b7c'
    if master_weights is None:
        master_weights = {}

    try:
        from data_pipeline.config import load_clickhouse_config
        import requests

        cfg = load_clickhouse_config()
        params = {
            'user': cfg['user'],
            'password': cfg['password'],
            'database': cfg['database'],
        }

        # Row-level query: fetch barcode, product_name, qty, net_weight per item
        sql = f"""
        SELECT
            formatDateTime(fromUnixTimestamp(toUInt32(po.delivery_date)), '%d/%m/%Y') AS del_date,
            ri.product_barcode AS barcode,
            ri.product_name AS product_name,
            ri.qty AS qty,
            ri.net_weight AS net_weight
        FROM kf_receipt_items ri
        INNER JOIN kf_purchase_order po
            ON ri.purchase_code = po.code
            AND po.branch_id = '{KRC_BRANCH_ID}'
            AND po.deleted = 0
            AND po.status IN (5, 7)
        WHERE ri.branch_id = '{KRC_BRANCH_ID}'
        FORMAT JSONEachRow
        """

        print("  → Querying PO KRC from ClickHouse DB (row-level)...")
        r = requests.get(
            cfg['base_url'],
            params={**params, 'query': sql},
            timeout=120,
        )
        r.raise_for_status()

        daily_tons = defaultdict(float)
        total_rows = 0
        rows_with_weight = 0
        rows_master_fallback = 0
        rows_name_fallback = 0
        rows_no_weight = 0

        for line in r.text.strip().split('\n'):
            if not line.strip():
                continue
            obj = json.loads(line)
            total_rows += 1

            date_str = obj.get('del_date', '')
            if not date_str:
                continue

            qty = float(obj.get('qty', 0))
            if qty <= 0:
                continue

            net_weight = float(obj.get('net_weight', 0))  # grams
            barcode = str(obj.get('barcode', '')).strip()
            product_name = str(obj.get('product_name', '')).strip()

            # Weight resolution: same priority as transfer (PT)
            if net_weight > 0:
                # Source 1: net_weight from receipt_items (grams)
                weight_grams = net_weight
                rows_with_weight += 1
            elif barcode and barcode in master_weights:
                # Source 2: master data (Google Sheets Col Z — grams)
                weight_grams = master_weights[barcode]
                rows_master_fallback += 1
            else:
                # Source 3: extract from product name
                weight_kg = extract_weight_from_name(product_name)
                if weight_kg > 0:
                    weight_grams = weight_kg * 1000  # convert KG → grams
                    rows_name_fallback += 1
                else:
                    rows_no_weight += 1
                    continue

            # qty * weight_grams / 1,000,000 = tons
            tons = qty * weight_grams / 1_000_000
            daily_tons[date_str] += tons

        # Round results
        for d in daily_tons:
            daily_tons[d] = round(daily_tons[d], 4)

        # Remove zero-ton dates
        daily_tons = defaultdict(float, {d: t for d, t in daily_tons.items() if t > 0})

        date_count = len(daily_tons)
        print(f"    ✅ DB: {date_count} dates, {total_rows:,} rows processed")
        print(f"       Weight sources: DB={rows_with_weight:,}, master={rows_master_fallback:,}, "
              f"name={rows_name_fallback:,}, skip={rows_no_weight:,}")
        if date_count > 0:
            sorted_dates = sorted(daily_tons.keys(),
                                  key=lambda d: datetime.strptime(d, "%d/%m/%Y"))
            for d in sorted_dates[-5:]:
                print(f"       {d}: {daily_tons[d]:.2f} tấn")
        return daily_tons

    except Exception as e:
        print(f"    ⚠ DB query failed: {e}")
        return None


# ══════════════════════════════════════════════════════════════
# KRC — Read PO files (fallback) → total tons/day
# ══════════════════════════════════════════════════════════════
def read_po_krc_local(master_weights):
    """Read all PO KRC local Excel files and aggregate total weight (tons) per delivery date.

    Columns:
      Col P (15): Ngày giao hàng dự kiến (delivery date)
      Col S (18): Barcode
      Col T (19): Tên sản phẩm (product name)
      Col V (21): Số lượng (quantity)
      Col W (22): Khối lượng (kg)

    Returns: dict of {date_str: total_tons}
    """
    daily_tons = defaultdict(float)
    missing_barcodes = {}

    if not os.path.isdir(PO_KRC_DIR):
        print(f"  ⚠ PO KRC directory not found: {PO_KRC_DIR}")
        return daily_tons

    po_files = [f for f in os.listdir(PO_KRC_DIR)
                if f.endswith('.xlsx') and not f.startswith('~') and f != 'desktop.ini']

    if not po_files:
        print("  ⚠ No PO KRC files found")
        return daily_tons

    print(f"  → Reading {len(po_files)} PO KRC file(s) [local fallback]...")

    for fname in po_files:
        filepath = os.path.join(PO_KRC_DIR, fname)
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            file_count = 0

            for row in ws.iter_rows(min_row=2, values_only=False):
                # Col P (15): delivery date
                date_val = row[15].value
                if not date_val:
                    continue

                # Parse date
                if isinstance(date_val, datetime):
                    delivery_date = date_val.strftime("%d/%m/%Y")
                elif isinstance(date_val, str):
                    delivery_date = date_val.strip()
                    # Normalize date format
                    for fmt in ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"]:
                        try:
                            dt = datetime.strptime(delivery_date, fmt)
                            delivery_date = dt.strftime("%d/%m/%Y")
                            break
                        except ValueError:
                            continue
                else:
                    continue

                barcode = str(row[18].value or "").strip()  # Col S
                product_name = str(row[19].value or "").strip()  # Col T
                quantity = 0
                try:
                    quantity = float(row[21].value or 0)  # Col V
                except (ValueError, TypeError):
                    pass

                weight_kg = 0
                try:
                    weight_kg = float(row[22].value or 0)  # Col W
                except (ValueError, TypeError):
                    pass

                if weight_kg > 0:
                    # Weight already in KG, convert to tons
                    daily_tons[delivery_date] += weight_kg / 1000
                    file_count += 1
                elif barcode and barcode in master_weights:
                    # Lookup from master data (weight in grams per item)
                    w_grams = master_weights[barcode]
                    total_kg = quantity * w_grams / 1000
                    daily_tons[delivery_date] += total_kg / 1000  # Convert to tons
                    file_count += 1
                else:
                    # Try extract weight from product name
                    w_kg = extract_weight_from_name(product_name)
                    if w_kg > 0 and quantity > 0:
                        total_kg = quantity * w_kg
                        daily_tons[delivery_date] += total_kg / 1000
                        file_count += 1
                    elif barcode:
                        if barcode not in missing_barcodes:
                            missing_barcodes[barcode] = product_name

            wb.close()
            print(f"    ↳ {fname}: {file_count} rows processed")
        except Exception as e:
            print(f"    ⚠ Error reading {fname}: {e}")

    if missing_barcodes:
        print(f"    ⚠ {len(missing_barcodes)} barcodes missing weight data (top 5):")
        for i, (bc, name) in enumerate(list(missing_barcodes.items())[:5]):
            print(f"      {bc}: {name[:40]}")

    return daily_tons


def read_po_krc(master_weights):
    """Read PO KRC data: DB first, fallback to local files, merge results.

    Priority: DB dates override local file dates.
    Local file dates fill gaps where DB has no data.
    """
    db_data = read_po_krc_from_db(master_weights)
    local_data = read_po_krc_local(master_weights)

    if db_data is not None and len(db_data) > 0:
        # Merge: DB takes priority, local fills gaps
        merged = dict(db_data)
        local_only = 0
        for d, tons in local_data.items():
            if d not in merged:
                merged[d] = tons
                local_only += 1
        if local_only > 0:
            print(f"    📎 Merged: {len(db_data)} DB dates + {local_only} local-only dates = {len(merged)} total")
        else:
            print(f"    📎 Using DB data ({len(db_data)} dates), local had no extra dates")
        return merged
    elif local_data:
        print(f"    📎 Using local files only ({len(local_data)} dates)")
        return local_data
    else:
        print(f"    ⚠ No KRC data from DB or local files")
        return defaultdict(float)


# ══════════════════════════════════════════════════════════════
# KSL — Read yeu_cau files → total items/day (Sáng + Tối)
# ══════════════════════════════════════════════════════════════
def read_ksl_items():
    """Read yeu_cau_chuyen_hang_thuong files and aggregate total items per date.
    
    The yeu_cau files contain KSL (Dry) warehouse transfer requests.
    We sum up 'Số lượng cần chuyển' for all items across all files for each date.
    
    Returns: dict of {date_str: total_items}
    """
    daily_items = defaultdict(float)
    
    if not os.path.isdir(YECAU_LOCAL):
        print(f"  ⚠ Yeu cau directory not found: {YECAU_LOCAL}")
        return daily_items
    
    yc_files = [f for f in os.listdir(YECAU_LOCAL)
                if f.endswith('.xlsx') and not f.startswith('~') and 'yeu_cau_chuyen_hang_thuong' in f]
    
    if not yc_files:
        print("  ⚠ No yeu_cau files found")
        return daily_items
    
    print(f"  → Reading {len(yc_files)} yeu_cau file(s) for KSL items...")
    
    for fname in yc_files:
        filepath = os.path.join(YECAU_LOCAL, fname)
        try:
            # Extract date from filename: yeu_cau_chuyen_hang_thuong_DDMMYYYY...
            date_match = re.search(r'(\d{8})', fname)
            if not date_match:
                continue
            date_tag = date_match.group(1)
            # Parse DDMMYYYY
            try:
                dt = datetime.strptime(date_tag, "%d%m%Y")
                date_str = dt.strftime("%d/%m/%Y")
            except ValueError:
                continue
            
            wb = load_workbook(filepath, read_only=True, data_only=True)
            
            # Find KF sheet or use first sheet
            ws = None
            for name in wb.sheetnames:
                if name == 'KF':
                    ws = wb[name]
                    break
            if not ws:
                ws = wb.worksheets[0]
            
            # Auto-detect columns
            col_idx = {}
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                for i, cell in enumerate(row):
                    if cell.value:
                        col_idx[str(cell.value).strip()] = i
            
            i_sl = col_idx.get('Số lượng cần chuyển', 17)
            
            file_total = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                try:
                    sl_val = row[i_sl].value if i_sl < len(row) else None
                    if sl_val is not None:
                        sl = float(sl_val)
                        if sl > 0:
                            file_total += sl
                except (ValueError, TypeError, IndexError):
                    pass
            
            daily_items[date_str] += file_total
            wb.close()
            print(f"    ↳ {fname}: {file_total:,.0f} items → {date_str}")
        except Exception as e:
            print(f"    ⚠ Error reading {fname}: {e}")
    
    return daily_items


# ══════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════
def _load_existing_forecast():
    """Load existing capacity_forecast.json to use as cache."""
    out_path = os.path.join(DOCS_DATA, "capacity_forecast.json")
    if not os.path.exists(out_path):
        return None
    try:
        with open(out_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def read_ksl_items_incremental(existing_dates=None):
    """Read only NEW yeu_cau files (dates not in existing_dates).
    
    existing_dates: set of date strings already in cache (e.g. {'09/03/2026', ...})
    If None, reads ALL files (full rebuild).
    """
    daily_items = defaultdict(float)
    
    if not os.path.isdir(YECAU_LOCAL):
        print(f"  ⚠ Yeu cau directory not found: {YECAU_LOCAL}")
        return daily_items
    
    yc_files = [f for f in os.listdir(YECAU_LOCAL)
                if f.endswith('.xlsx') and not f.startswith('~') and 'yeu_cau_chuyen_hang_thuong' in f]
    
    if not yc_files:
        print("  ⚠ No yeu_cau files found")
        return daily_items
    
    # Filter to only new files
    new_files = []
    skipped = 0
    for fname in yc_files:
        date_match = re.search(r'(\d{8})', fname)
        if not date_match:
            continue
        date_tag = date_match.group(1)
        try:
            dt = datetime.strptime(date_tag, "%d%m%Y")
            date_str = dt.strftime("%d/%m/%Y")
        except ValueError:
            continue
        
        if existing_dates and date_str in existing_dates:
            skipped += 1
            continue
        new_files.append((fname, date_str))
    
    if skipped > 0:
        print(f"  ⏭ Skipped {skipped} cached dates")
    
    if not new_files:
        print(f"  ✅ No new yeu_cau files to process")
        return daily_items
    
    print(f"  → Reading {len(new_files)} NEW yeu_cau file(s)...")
    
    for fname, date_str in new_files:
        filepath = os.path.join(YECAU_LOCAL, fname)
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            
            ws = None
            for name in wb.sheetnames:
                if name == 'KF':
                    ws = wb[name]
                    break
            if not ws:
                ws = wb.worksheets[0]
            
            col_idx = {}
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                for i, cell in enumerate(row):
                    if cell.value:
                        col_idx[str(cell.value).strip()] = i
            
            i_sl = col_idx.get('Số lượng cần chuyển', 17)
            
            file_total = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                try:
                    sl_val = row[i_sl].value if i_sl < len(row) else None
                    if sl_val is not None:
                        sl = float(sl_val)
                        if sl > 0:
                            file_total += sl
                except (ValueError, TypeError, IndexError):
                    pass
            
            daily_items[date_str] += file_total
            wb.close()
            print(f"    ↳ {fname}: {file_total:,.0f} items → {date_str}")
        except Exception as e:
            print(f"    ⚠ Error reading {fname}: {e}")
    
    return daily_items


def export_capacity_forecast():
    """Generate and export capacity_forecast.json (incremental mode).
    
    - KRC: always re-queries DB (fast ~1s), filter to current month
    - KSL: loads cached data from existing JSON, only reads NEW yeu_cau files
    """
    print(f"\n{'='*60}")
    print(f"  📊 Capacity Forecast Export — {NOW_STR}")
    print(f"{'='*60}\n")
    
    # Load existing cache
    existing = _load_existing_forecast()
    
    # ── KRC: always re-query DB (fast) ──
    print("\n📦 KRC — PO Data:")
    master_weights = load_master_weights()
    krc_data = read_po_krc(master_weights)
    
    # KRC: filter to current month onwards
    krc_cutoff = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    all_dates_krc = sorted(
        [d for d in krc_data.keys() if datetime.strptime(d, "%d/%m/%Y") >= krc_cutoff],
        key=lambda d: datetime.strptime(d, "%d/%m/%Y")
    )
    
    krc_forecast = []
    for d in all_dates_krc:
        tons = round(krc_data[d], 2)
        exceeds = tons > KRC_BENCHMARK_TONS * (1 + ALERT_THRESHOLD_PCT / 100)
        pct_of_cap = round(tons / KRC_BENCHMARK_TONS * 100, 1) if KRC_BENCHMARK_TONS > 0 else 0
        dt = datetime.strptime(d, "%d/%m/%Y")
        krc_forecast.append({
            "date": d,
            "iso": dt.strftime("%Y-%m-%d"),
            "tons": tons,
            "pct_capacity": pct_of_cap,
            "exceeds_alert": exceeds,
        })
    
    # ── KSL: incremental — only read new files ──
    print("\n📦 KSL — Yeu Cau Data:")
    
    # Build set of dates already in cache
    cached_ksl = {}
    if existing and "ksl" in existing and existing["ksl"].get("data"):
        for entry in existing["ksl"]["data"]:
            cached_ksl[entry["date"]] = entry["items"]
        print(f"  📎 Cache: {len(cached_ksl)} dates loaded from existing JSON")
    
    # Read only new files
    new_ksl = read_ksl_items_incremental(existing_dates=set(cached_ksl.keys()) if cached_ksl else None)
    
    # Merge: cached + new (new overwrites if same date)
    merged_ksl = dict(cached_ksl)
    for d, items in new_ksl.items():
        merged_ksl[d] = items
    
    all_dates_ksl = sorted(merged_ksl.keys(), key=lambda d: datetime.strptime(d, "%d/%m/%Y"))
    
    ksl_forecast = []
    for d in all_dates_ksl:
        items = round(merged_ksl[d])
        exceeds = items > KSL_BENCHMARK_ITEMS * (1 + ALERT_THRESHOLD_PCT / 100)
        pct_of_cap = round(items / KSL_BENCHMARK_ITEMS * 100, 1) if KSL_BENCHMARK_ITEMS > 0 else 0
        dt = datetime.strptime(d, "%d/%m/%Y")
        ksl_forecast.append({
            "date": d,
            "iso": dt.strftime("%Y-%m-%d"),
            "items": items,
            "pct_capacity": pct_of_cap,
            "exceeds_alert": exceeds,
        })
    
    data = {
        "_updated": NOW_STR,
        "krc": {
            "benchmark_tons": KRC_BENCHMARK_TONS,
            "alert_threshold_pct": ALERT_THRESHOLD_PCT,
            "data": krc_forecast,
        },
        "ksl": {
            "benchmark_items": KSL_BENCHMARK_ITEMS,
            "alert_threshold_pct": ALERT_THRESHOLD_PCT,
            "data": ksl_forecast,
        },
    }
    
    os.makedirs(DOCS_DATA, exist_ok=True)
    out_path = os.path.join(DOCS_DATA, "capacity_forecast.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    new_count = len(new_ksl)
    cached_count = len(cached_ksl)
    print(f"\n{'='*60}")
    print(f"  ✅ Output: {out_path}")
    print(f"     KRC: {len(krc_forecast)} days, benchmark {KRC_BENCHMARK_TONS}T")
    print(f"     KSL: {len(ksl_forecast)} days ({cached_count} cached + {new_count} new), benchmark {KSL_BENCHMARK_ITEMS:,} items")
    
    krc_alerts = sum(1 for d in krc_forecast if d["exceeds_alert"])
    ksl_alerts = sum(1 for d in ksl_forecast if d["exceeds_alert"])
    if krc_alerts or ksl_alerts:
        print(f"     ⚠ Alerts: KRC={krc_alerts} days, KSL={ksl_alerts} days exceed {ALERT_THRESHOLD_PCT}% over capacity")
    
    print(f"{'='*60}")
    return True


if __name__ == "__main__":
    export_capacity_forecast()

