"""
generate_report.py - Read all data sources and output summary report
Usage: python script/generate_report.py [--date DD/MM/YYYY] [--send]

ALL sources fetched online from Google Sheets/Drive (no local data/ files).
"""
import os, sys, json, re, csv
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(BASE, "script"))
BACKUP_DIR = os.path.join(BASE, "data", "raw", "daily")

# ── Online data source URLs (from shared lib) ──
from lib.sources import (
    KRC_SHEET_URL,
    KFM_SHEET_URL as KFM_XLSX_URL,
    MASTER_SHEET_URL,
    TRANSFER_FOLDER_URL,
    YECAU_FOLDER_URL,
    KH_MEAT_FOLDER_URL,
    KH_DONG_LOCAL,
    KH_MAT_LOCAL,
    TRANSFER_LOCAL,
)

KH_DRIVE_FOLDERS = [
    ("KH MEAT", "THỊT CÁ", 11, KH_MEAT_FOLDER_URL),
]
# KH ĐÔNG + MÁT: read from local Google Drive sync (Drive API folder IDs no longer valid)
KH_LOCAL_FOLDERS = [
    ("KH HÀNG ĐÔNG", "ĐÔNG MÁT", 9, KH_DONG_LOCAL),
    ("KH HÀNG MÁT", "ĐÔNG MÁT", 9, KH_MAT_LOCAL),
]

# ── Kho mapping (PT raw kho → report kho) ──
KHO_MAP = {
    "KHO ABA MIỀN ĐÔNG": "THỊT CÁ",
    "KHO ABA QUÁ CẢNH": "ĐÔNG MÁT",
    "KHO RAU CỦ": "KRC",
    "Sáng": "KSL-SÁNG", "Tối": "KSL-TỐI",
    "Khách đặt": "KSL-TỐI", "khách đặt": "KSL-TỐI",
    "Socola": "KSL-SÁNG",
    "ĐI SÁNG": "KSL-SÁNG", "ĐI TỐI": "KSL-TỐI",
    "đi sáng": "KSL-SÁNG", "đi tối": "KSL-TỐI",
}

REPORT_KHOS = ["KRC", "THỊT CÁ", "ĐÔNG MÁT", "KSL-SÁNG", "KSL-TỐI"]
KHO_COLORS = {"KRC": "#4caf50", "THỊT CÁ": "#e53935", "ĐÔNG MÁT": "#1e88e5",
              "KSL-SÁNG": "#ff9800", "KSL-TỐI": "#9c27b0"}


# ──────────────────────────────────────────
#  Helper functions
# ──────────────────────────────────────────

def parse_time_hour(time_text):
    if not time_text:
        return -1
    m = re.match(r'(\d{1,2}):', str(time_text).strip())
    return int(m.group(1)) if m else -1


def safe_val(row, idx):
    if idx < len(row):
        return str(row[idx].value or "").strip()
    return ""


def _save_backup(name, content_bytes):
    """Save raw bytes to data/ folder as backup."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    path = os.path.join(BACKUP_DIR, name)
    with open(path, "wb") as f:
        f.write(content_bytes)
    print(f"    💾 Backup: {name}")


def _load_backup(name):
    """Load backup file from data/ folder. Returns bytes or None."""
    path = os.path.join(BACKUP_DIR, name)
    if os.path.exists(path):
        with open(path, "rb") as f:
            return f.read()
    return None


def read_xlsx_from_url(url, backup_name=None):
    import requests
    import time as _time
    max_retries = 3
    last_err = None
    for attempt in range(max_retries):
        try:
            r = requests.get(url, allow_redirects=True, timeout=120, stream=True)
            r.raise_for_status()
            chunks = []
            for chunk in r.iter_content(chunk_size=65536):
                chunks.append(chunk)
            content = b''.join(chunks)
            if len(content) < 500:
                raise ValueError(f"Response too small ({len(content)} bytes)")
            if backup_name:
                _save_backup(backup_name, content)
            return load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            last_err = e
            if attempt < max_retries - 1:
                wait = 3 * (attempt + 1)
                print(f"    ⚠ Attempt {attempt+1}/{max_retries} failed: {e}")
                print(f"      Retry in {wait}s...")
                _time.sleep(wait)
    # All retries failed — try backup
    if backup_name:
        cached = _load_backup(backup_name)
        if cached:
            print(f"    ⚠ Online lỗi sau {max_retries} lần thử, dùng backup: {backup_name}")
            return load_workbook(BytesIO(cached), read_only=True, data_only=True)
    raise last_err


def read_csv_from_url(url, retries=2, backup_name=None):
    """Read a Google Sheets gviz/tq CSV URL and return list of rows (list of strings)."""
    import requests
    from io import StringIO
    import time
    for attempt in range(retries):
        try:
            r = requests.get(url, allow_redirects=True, timeout=120)
            r.raise_for_status()
            if backup_name:
                _save_backup(backup_name, r.content)
            reader = csv.reader(StringIO(r.text))
            return list(reader)
        except Exception as e:
            if attempt < retries - 1:
                print(f"    ⚠ Retry {attempt+1}/{retries}: {e}")
                time.sleep(2)
            else:
                if backup_name:
                    cached = _load_backup(backup_name)
                    if cached:
                        print(f"    ⚠ Online lỗi, dùng backup: {backup_name}")
                        reader = csv.reader(StringIO(cached.decode('utf-8')))
                        return list(reader)
                raise


def _list_drive_folder(folder_url):
    import requests as _req
    import html as _html
    folder_id = re.search(r'folders/([a-zA-Z0-9_-]+)', folder_url).group(1)
    url = f"https://drive.google.com/drive/folders/{folder_id}"
    r = _req.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    text = _html.unescape(r.text)
    files = []
    seen = set()
    for m in re.finditer(r'"(1[a-zA-Z0-9_-]{25,50})"', text):
        fid = m.group(1)
        if fid in seen:
            continue
        after = text[m.end():m.end() + 500]
        name_match = re.search(r'"([^"]+\.xlsx?)"', after)
        if name_match:
            fname = name_match.group(1)
            if "Microsoft" not in fname and "Shared" not in fname:
                seen.add(fid)
                files.append((fid, fname))
    return files


def _download_drive_file(file_id, backup_name=None):
    import requests as _req
    dl_url = f"https://drive.google.com/uc?export=download&confirm=t&id={file_id}"
    try:
        r = _req.get(dl_url, allow_redirects=True, timeout=60)
        if r.status_code == 200 and len(r.content) > 500:
            if backup_name:
                _save_backup(backup_name, r.content)
            return load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    except Exception as e:
        if backup_name:
            cached = _load_backup(backup_name)
            if cached:
                print(f"    ⚠ Online lỗi, dùng backup: {backup_name}")
                return load_workbook(BytesIO(cached), read_only=True, data_only=True)
        raise
    # Download returned empty/bad response — try backup
    if backup_name:
        cached = _load_backup(backup_name)
        if cached:
            print(f"    ⚠ Download trống, dùng backup: {backup_name}")
            return load_workbook(BytesIO(cached), read_only=True, data_only=True)
    return None


def read_kh_from_drive(folder_url, folder_name, date_for_file, backup_name=None):
    files = _list_drive_folder(folder_url)
    if not files:
        print(f"    ⚠ Không list được folder")
        # Try backup if folder listing fails
        if backup_name:
            cached = _load_backup(backup_name)
            if cached:
                print(f"    ⚠ Dùng backup: {backup_name}")
                return load_workbook(BytesIO(cached), read_only=True, data_only=True)
        return None
    target = None
    for fid, fname in files:
        if date_for_file in fname:
            target = (fid, fname)
            break
    if not target:
        return None
    print(f"    ↳ {target[1]}")
    return _download_drive_file(target[0], backup_name=backup_name)


def extract_weight_grams(product_name):
    text = product_name.upper()
    patterns = [
        (r'(\d+(?:[.,]\d+)?)\s*KG\b', 1000),
        (r'(\d+(?:[.,]\d+)?)\s*G\b', 1),
        (r'(\d+(?:[.,]\d+)?)\s*(?:LÍT|LIT)\b', 1000),
        (r'(\d+(?:[.,]\d+)?)\s*L\b', 1000),
        (r'(\d+(?:[.,]\d+)?)\s*ML\b', 1),
    ]
    for pattern, multiplier in patterns:
        matches = re.findall(pattern, text)
        if matches:
            try:
                return float(matches[-1].replace(",", ".")) * multiplier
            except ValueError:
                continue
    return 0


# ──────────────────────────────────────────
#  Load master data (ONLINE from Google Sheets)
# ──────────────────────────────────────────

def load_master_data():
    """Load barcode → per-item weight (grams) from Google Sheets.
    Col A (0) = Barcode cơ bản, Col Z (25) = Giá trị trọng lượng / thể tích.
    """
    master_tl = {}
    print(f"  → Master data (Google Sheets)...")
    try:
        wb = read_xlsx_from_url(MASTER_SHEET_URL, backup_name="master_data.xlsx")
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=2, values_only=False):
            barcode = str(row[0].value or "").strip()
            if not barcode:
                continue
            tl_value = row[25].value  # Col Z
            if tl_value is not None:
                try:
                    w = float(tl_value)
                    if w > 0:
                        master_tl[barcode] = w
                except (ValueError, TypeError):
                    pass
        wb.close()
        print(f"    {len(master_tl)} barcodes loaded")
    except Exception as e:
        print(f"    ⚠ Error: {e}")
    return master_tl


# ──────────────────────────────────────────
#  Read STHI + XE data
# ──────────────────────────────────────────

def read_sthi_data(date_str, date_for_file, date_tag=None):
    rows = []
    warnings = []

    # 1. KRC
    krc_count = 0
    print("  → KRC (Google Sheets)...")
    try:
        wb = read_xlsx_from_url(KRC_SHEET_URL, backup_name=f"krc_{date_tag}.xlsx" if date_tag else None)
        ws = wb["KRC"]
        for row in ws.iter_rows(min_row=2, values_only=False):
            scv = str(row[0].value or "").strip()
            if scv == date_str:
                diem_den = str(row[6].value or "").strip()
                gio_den = row[7].value
                tuyen = str(row[10].value or "").strip()
                if diem_den and gio_den:
                    rows.append({"kho": "KRC", "diem_den": diem_den, "tuyen": tuyen})
                    krc_count += 1
        wb.close()
        print(f"    {krc_count} rows")
    except Exception as e:
        warnings.append(f"KRC: lỗi tải — {e}")
    if krc_count == 0:
        warnings.append("KRC: 0 rows — Google Sheet chưa cập nhật?")

    # 2. KFM (read DRY sheet online via XLSX — CSV gviz endpoint truncates large sheets)
    sang_count = toi_count = 0
    print("  → KFM (Google Sheets XLSX)...")
    try:
        kfm_wb = read_xlsx_from_url(KFM_XLSX_URL, backup_name=f"kfm_{date_tag}.xlsx" if date_tag else None)
        # Find the DRY sheet
        kfm_ws = None
        for sname in kfm_wb.sheetnames:
            if 'DRY' in sname.upper():
                kfm_ws = kfm_wb[sname]
                break
        if not kfm_ws:
            kfm_ws = kfm_wb.worksheets[0]
        # Row 1 = title, Row 2 = headers, Row 3+ = data
        # col0=Ngày, col4=Giờ đi, col6=Điểm đến, col7=Giờ đến, col10=Tuyến
        last_kho = None  # Track kho from previous row for multi-stop routes
        for xl_row in kfm_ws.iter_rows(min_row=3, values_only=False):
            scv = str(xl_row[0].value or "").strip()
            if scv == date_str:
                diem_den = str(xl_row[6].value or "").strip()
                gio_den = str(xl_row[7].value or "").strip()
                gio_di = str(xl_row[4].value or "").strip()
                tuyen = str(xl_row[10].value or "").strip()
                if not diem_den:
                    continue
                # Use giờ đến if available, otherwise fallback to giờ đi
                gio = gio_den or gio_di
                if gio:
                    hour = parse_time_hour(gio)
                    if hour < 0:
                        continue
                    using_gio_di = not gio_den and bool(gio_di)
                    # Giờ đến < 18h = Sáng; Giờ đi < 15h = Sáng (đi 15h+ = Tối)
                    sang_cutoff = 15 if using_gio_di else 18
                    if 6 <= hour < sang_cutoff:
                        kho = "KSL-Sáng"
                        sang_count += 1
                    else:
                        kho = "KSL-Tối"
                        toi_count += 1
                    last_kho = kho
                    rows.append({"kho": kho, "diem_den": diem_den, "tuyen": tuyen})
                elif last_kho:
                    # Multi-stop route: sub-leg has no time, inherit from previous
                    if last_kho == "KSL-Sáng":
                        sang_count += 1
                    else:
                        toi_count += 1
                    rows.append({"kho": last_kho, "diem_den": diem_den, "tuyen": tuyen})
            else:
                last_kho = None  # Reset on date change
        kfm_wb.close()
        print(f"    Sáng: {sang_count}, Tối: {toi_count}")
    except Exception as e:
        warnings.append(f"KFM: lỗi tải — {e}")

    # 3. KH
    for folder_name, kho_name, tuyen_col, folder_url in KH_DRIVE_FOLDERS:
        print(f"  → {folder_name} (Google Drive)...")
        try:
            bk_name = f"{folder_name.replace(' ', '_').lower()}_{date_tag}.xlsx" if date_tag else None
            wb = read_kh_from_drive(folder_url, folder_name, date_for_file, backup_name=bk_name)
            if not wb:
                warnings.append(f"{folder_name}: file {date_for_file} not found")
                continue
            ws = wb.worksheets[0]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                diem_den = safe_val(row, 2)
                tuyen = safe_val(row, tuyen_col)
                if diem_den:
                    rows.append({"kho": kho_name, "diem_den": diem_den, "tuyen": tuyen})
                    count += 1
            wb.close()
            print(f"    {count} rows")
        except Exception as e:
            warnings.append(f"{folder_name}: lỗi — {e}")

    # 3b. KH from local Google Drive sync (ĐÔNG + MÁT)
    for folder_name, kho_name, tuyen_col, local_dir in KH_LOCAL_FOLDERS:
        print(f"  → {folder_name} (Google Drive)...")
        try:
            if not os.path.isdir(local_dir):
                warnings.append(f"{folder_name}: folder local không tìm thấy")
                continue
            # Find file matching date pattern
            matched = None
            for fname in os.listdir(local_dir):
                if date_for_file in fname and fname.endswith('.xlsx') and not fname.startswith('~'):
                    matched = os.path.join(local_dir, fname)
                    break
            if not matched:
                warnings.append(f"{folder_name}: file {date_for_file} not found")
                continue
            print(f"    ↳ {os.path.basename(matched)}")
            wb = load_workbook(matched, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                diem_den = safe_val(row, 2)
                tuyen = safe_val(row, tuyen_col)
                if diem_den:
                    rows.append({"kho": kho_name, "diem_den": diem_den, "tuyen": tuyen})
                    count += 1
            wb.close()
            print(f"    {count} rows")
        except Exception as e:
            warnings.append(f"{folder_name}: lỗi — {e}")

    return rows, warnings


# ──────────────────────────────────────────
#  Read PT data (ONLINE from Google Drive)
# ──────────────────────────────────────────

def read_pt_data(date_str, master_tl):
    """Read transfer + yeu_cau from Google Drive folders (online)."""
    rows = []
    warnings = []

    # Build date tag for filename matching: DD/MM/YYYY → DDMMYYYY
    parts = date_str.split("/")
    date_tag = f"{parts[0]}{parts[1]}{parts[2]}"  # e.g. 23032026

    # Build barcode → TL lookup from transfer
    transfer_tl = {}

    # 1. Transfer — from Google Drive folder
    # Check local sync paths first (Google Drive desktop sync), then online
    TRANSFER_LOCAL_PATHS = [
        os.path.join(BACKUP_DIR, f"transfer_{date_tag}.xlsx"),
        os.path.join(TRANSFER_LOCAL, f"transfer_{date_tag}.xlsx"),
    ]
    transfer_wb = None
    transfer_source = None

    # Try local files first (prefer largest file)
    for local_path in TRANSFER_LOCAL_PATHS:
        if os.path.exists(local_path) and os.path.getsize(local_path) > 10000:
            try:
                candidate = load_workbook(local_path, read_only=True, data_only=True)
                if transfer_wb is None:
                    transfer_wb = candidate
                    transfer_source = local_path
                else:
                    # Keep the larger file (more complete data)
                    if os.path.getsize(local_path) > os.path.getsize(transfer_source):
                        transfer_wb.close()
                        transfer_wb = candidate
                        transfer_source = local_path
                    else:
                        candidate.close()
            except Exception:
                pass

    print(f"  → Transfer (tag={date_tag})...")
    if transfer_wb is None:
        # Fallback: download from Google Drive
        try:
            tf_files = _list_drive_folder(TRANSFER_FOLDER_URL)
            target = None
            for fid, fname in tf_files:
                if date_tag in fname and fname.startswith("transfer"):
                    target = (fid, fname)
                    break
            if target:
                print(f"    ↳ Online: {target[1]}")
                transfer_wb = _download_drive_file(target[0], backup_name=f"transfer_{date_tag}.xlsx")
                transfer_source = "online"
            else:
                warnings.append(f"Transfer: file transfer_{date_tag} not found")
        except Exception as e:
            warnings.append(f"Transfer: error — {e}")
    else:
        print(f"    ↳ Local: {os.path.basename(transfer_source)} ({os.path.getsize(transfer_source):,} bytes)")
        # Also save to backup dir if source is not already there
        bk_path = os.path.join(BACKUP_DIR, f"transfer_{date_tag}.xlsx")
        if transfer_source != bk_path:
            import shutil
            shutil.copy2(transfer_source, bk_path)
            print(f"    💾 Backup: transfer_{date_tag}.xlsx")

    if transfer_wb:
        try:
            ws = transfer_wb.worksheets[0]
            for row in ws.iter_rows(min_row=2, values_only=False):
                code = safe_val(row, 7)  # Mã hàng
                tl_raw = row[14].value    # TL
                if code and tl_raw:
                    try:
                        tl_val = float(tl_raw)
                        if tl_val > 0:
                            transfer_tl[code] = tl_val
                    except (ValueError, TypeError):
                        pass

                ngay = str(row[0].value or "").strip()
                if ngay == date_str:
                    raw_kho = str(row[2].value or "").strip()
                    report_kho = KHO_MAP.get(raw_kho)
                    if not report_kho:
                        continue
                    sl_raw = row[10].value
                    try:
                        sl = float(sl_raw) if sl_raw else 0
                    except (ValueError, TypeError):
                        sl = 0
                    try:
                        tl = float(tl_raw) if tl_raw else 0
                    except (ValueError, TypeError):
                        tl = 0
                    if tl == 0 and code:
                        tl = master_tl.get(code, 0)
                    if tl == 0:
                        continue
                    rows.append({"kho": report_kho, "sl": sl, "tl_grams": tl})
            transfer_wb.close()
            print(f"    {len(rows)} PT rows from transfer")
        except Exception as e:
            warnings.append(f"Transfer: error reading — {e}")
    else:
        warnings.append(f"Transfer: no file available for {date_tag}")

    # 2. Yeu cau — from Google Drive folder (may have _nso files too)
    print(f"  → Yeu cau (Google Drive, tag={date_tag})...")
    yc_row_count = 0
    try:
        yc_files = _list_drive_folder(YECAU_FOLDER_URL)
        # Find ALL files matching this date (main + _nso variants)
        targets = []
        for fid, fname in yc_files:
            if date_tag in fname and "yeu_cau_chuyen_hang_thuong" in fname:
                targets.append((fid, fname))

        if not targets:
            warnings.append(f"Yeu cau: file yeu_cau_{date_tag} not found")
        
        for target_fid, target_fname in targets:
            print(f"    ↳ {target_fname}")
            # Use filename-based backup (supports _nso variants)
            bk_name = target_fname.replace(' ', '_')
            wb = _download_drive_file(target_fid, backup_name=bk_name)
            if not wb:
                warnings.append(f"Yeu cau: download failed for {target_fname}")
                continue
            
            ws = None
            for name in wb.sheetnames:
                if name == 'KF':
                    ws = wb[name]
                    break
            if not ws:
                ws = wb.worksheets[0]

            # Auto-detect columns by header name
            col_idx = {}
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
                for i, cell in enumerate(row):
                    if cell.value:
                        col_idx[str(cell.value).strip()] = i

            i_barcode = col_idx.get('Barcode', 2)
            i_name = col_idx.get('Tên sản phẩm', 3)
            i_sl = col_idx.get('Số lượng cần chuyển', 17)
            i_plo = col_idx.get('PLO ghi chú', 23)

            for row in ws.iter_rows(min_row=2, values_only=False):
                code = safe_val(row, i_barcode)
                if not code:
                    continue
                product_name = safe_val(row, i_name)
                raw_kho = safe_val(row, i_plo)
                report_kho = KHO_MAP.get(raw_kho)
                if not report_kho:
                    continue
                sl_raw = row[i_sl].value if i_sl < len(row) else None
                try:
                    sl = float(sl_raw) if sl_raw else 0
                except (ValueError, TypeError):
                    sl = 0
                tl = master_tl.get(code, 0)
                if tl == 0:
                    tl = transfer_tl.get(code, 0)
                if tl == 0:
                    tl = extract_weight_grams(product_name)
                if tl == 0:
                    continue
                rows.append({"kho": report_kho, "sl": sl, "tl_grams": tl})
                yc_row_count += 1
            wb.close()
        if yc_row_count > 0:
            print(f"    {yc_row_count} PT rows from yeu_cau ({len(targets)} file(s))")
    except Exception as e:
        warnings.append(f"Yeu cau: error — {e}")

    return rows, warnings


# ──────────────────────────────────────────
#  Calculate summary
# ──────────────────────────────────────────

def calculate_summary(sthi_rows, pt_rows, date_str):
    sthi_diem_den = defaultdict(set)
    sthi_tuyen = defaultdict(set)
    for r in sthi_rows:
        kho = r["kho"]
        if kho == "KSL-Sáng": kho = "KSL-SÁNG"
        elif kho == "KSL-Tối": kho = "KSL-TỐI"
        sthi_diem_den[kho].add(r["diem_den"])
        if r["tuyen"]:
            sthi_tuyen[kho].add(r["tuyen"])

    pt_items = defaultdict(float)
    pt_tons = defaultdict(float)
    for r in pt_rows:
        kho = r["kho"]
        pt_items[kho] += r["sl"]
        pt_tons[kho] += r["sl"] * r["tl_grams"] / 1_000_000

    total_sthi = sum(len(sthi_diem_den.get(k, set())) for k in REPORT_KHOS)
    total_items = sum(pt_items.get(k, 0) for k in REPORT_KHOS)
    total_xe = sum(len(sthi_tuyen.get(k, set())) for k in REPORT_KHOS)
    total_tons = sum(pt_tons.get(k, 0) for k in REPORT_KHOS)

    # Print console summary
    print()
    print(f"  {'KHO':<10} {'STHI':>6} {'ITEMS':>10} {'XE':>6} {'TẤN':>8} {'T/XE':>6} {'I/ST':>6} {'ST/XE':>5} {'KG/ST':>7}")
    print("  " + "-" * 72)
    for kho in REPORT_KHOS:
        st = len(sthi_diem_den.get(kho, set()))
        it = pt_items.get(kho, 0)
        xe = len(sthi_tuyen.get(kho, set()))
        tn = pt_tons.get(kho, 0)
        txe = tn / xe if xe > 0 else 0
        ist = it / st if st > 0 else 0
        stxe = st / xe if xe > 0 else 0
        kgst = tn * 1000 / st if st > 0 else 0
        print(f"  {kho:<10} {st:>6} {it:>10,.0f} {xe:>6} {tn:>8.2f} {txe:>6.2f} {ist:>6.0f} {stxe:>5.1f} {kgst:>7.1f}")
    print("  " + "-" * 72)
    txe_t = total_tons / total_xe if total_xe > 0 else 0
    ist_t = total_items / total_sthi if total_sthi > 0 else 0
    stxe_t = total_sthi / total_xe if total_xe > 0 else 0
    kgst_t = total_tons * 1000 / total_sthi if total_sthi > 0 else 0
    print(f"  {'TOTAL':<10} {total_sthi:>6} {total_items:>10,.0f} {total_xe:>6} {total_tons:>8.2f} {txe_t:>6.2f} {ist_t:>6.0f} {stxe_t:>5.1f} {kgst_t:>7.1f}")

    return {
        "date": date_str,
        "khos": {kho: {
            "sl_sthi": len(sthi_diem_den.get(kho, set())),
            "sl_items": pt_items.get(kho, 0),
            "sl_xe": len(sthi_tuyen.get(kho, set())),
            "san_luong_tan": pt_tons.get(kho, 0),
        } for kho in REPORT_KHOS},
        "total_sthi": total_sthi,
        "total_items": total_items,
        "total_xe": total_xe,
        "total_tons": total_tons,
    }


# ──────────────────────────────────────────
#  History
# ──────────────────────────────────────────

HISTORY_FILE = os.path.join(BASE, "output", "state", "history.json")

def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_history(history):
    os.makedirs(os.path.dirname(HISTORY_FILE), exist_ok=True)
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def update_history(result):
    history = load_history()
    # Remove existing entry for this date
    history = [h for h in history if h["date"] != result["date"]]
    history.append({
        "date": result["date"],
        "total_sthi": result["total_sthi"],
        "total_items": result["total_items"],
        "total_xe": result["total_xe"],
        "total_tons": result["total_tons"],
        "khos": {k: {
            "san_luong_tan": v["san_luong_tan"],
            "sl_items": v.get("sl_items", 0),
            "sl_xe": v.get("sl_xe", 0),
            "sl_sthi": v.get("sl_sthi", 0),
        } for k, v in result["khos"].items()},
    })
    # Sort by date and keep last 30
    history.sort(key=lambda x: datetime.strptime(x["date"], "%d/%m/%Y"))
    history = history[-30:]
    save_history(history)
    return history


# ──────────────────────────────────────────
#  Auto Commentary — Trend charts for Items & Xe
# ──────────────────────────────────────────

def _build_trend_svg(history, result, metric_key, total_key, title, fmt_val):
    """Build an SVG trend chart for a given metric, broken down by kho.
    - metric_key: key inside khos dict, e.g. 'sl_items', 'sl_xe'
    - total_key:  key on top-level, e.g. 'total_items', 'total_xe'
    - fmt_val:    callable to format values for labels, e.g. lambda v: f'{v:,.0f}'
    """
    chart_w = 760
    chart_h = 340
    pad_l = 65
    pad_r = 25
    pad_t = 25
    pad_b = 70
    plot_w = chart_w - pad_l - pad_r
    plot_h = chart_h - pad_t - pad_b

    n = len(history)
    if n == 0:
        return ""

    total_vals = [h.get(total_key, 0) for h in history]
    max_val = max(total_vals) * 1.15 if max(total_vals) > 0 else 1

    svg = [f'<svg viewBox="0 0 {chart_w} {chart_h}" xmlns="http://www.w3.org/2000/svg" style="width:100%;height:auto">']

    # Grid lines
    for i in range(5):
        y = pad_t + plot_h * i / 4
        v = max_val * (4 - i) / 4
        svg.append(f'<line x1="{pad_l}" y1="{y:.0f}" x2="{chart_w-pad_r}" y2="{y:.0f}" stroke="#3a3f4a" stroke-width="0.5"/>')
        svg.append(f'<text x="{pad_l-5}" y="{y+4:.0f}" text-anchor="end" font-size="13" font-weight="700" fill="#b0b5c0">{fmt_val(v)}</text>')

    inner_margin = plot_w / n * 0.5 if n > 1 else 0
    def xpos(i):
        return pad_l + inner_margin + (i * (plot_w - 2 * inner_margin) / max(n - 1, 1)) if n > 1 else pad_l + plot_w / 2
    def ypos(v):
        return pad_t + plot_h * (1 - v / max_val) if max_val > 0 else pad_t + plot_h

    # Bars for total
    bar_w = (plot_w - 2 * inner_margin) / n * 0.5 if n > 1 else 30
    for i, tv in enumerate(total_vals):
        x = xpos(i) - bar_w / 2
        y = ypos(tv)
        h = pad_t + plot_h - y
        svg.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{h:.1f}" rx="2" fill="rgba(255,255,255,0.45)" stroke="rgba(255,255,255,0.60)" stroke-width="0.8"/>')

    # Kho lines — only draw if history has per-kho data for this metric
    has_kho_data = any(
        h["khos"].get(kho, {}).get(metric_key, 0) > 0
        for h in history for kho in REPORT_KHOS
    )
    if has_kho_data:
        for kho in REPORT_KHOS:
            color = KHO_COLORS[kho]
            kho_vals = [h["khos"].get(kho, {}).get(metric_key, 0) for h in history]
            pts = [(xpos(i), ypos(v)) for i, v in enumerate(kho_vals)]
            if len(pts) > 1:
                polyline = " ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
                svg.append(f'<polyline points="{polyline}" fill="none" stroke="{color}" stroke-width="2" stroke-linejoin="round" stroke-linecap="round" opacity="0.9"/>')
            for x, y in pts:
                svg.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="2.5" fill="{color}" stroke="#1e2029" stroke-width="1"/>')

    # Total dashed line + labels
    total_pts = [(xpos(i), ypos(v)) for i, v in enumerate(total_vals)]
    if len(total_pts) > 1:
        polyline = " ".join(f"{x:.1f},{y:.1f}" for x, y in total_pts)
        svg.append(f'<polyline points="{polyline}" fill="none" stroke="rgba(255,255,255,0.85)" stroke-width="2" stroke-dasharray="6,3" stroke-linejoin="round"/>')
    for i, (x, y) in enumerate(total_pts):
        v = total_vals[i]
        svg.append(f'<text x="{x:.1f}" y="{y-8:.1f}" text-anchor="middle" font-size="13" font-weight="800" fill="#ffffff">{fmt_val(v)}</text>')

    # X-axis labels
    for i, h in enumerate(history):
        x = xpos(i)
        label = h["date"][:5]
        svg.append(f'<text x="{x:.1f}" y="{pad_t+plot_h+18:.0f}" text-anchor="middle" font-size="12" font-weight="700" fill="#b0b5c0">{label}</text>')

    # Legend
    leg_y = chart_h - 18
    leg_items = list(REPORT_KHOS) + ["TOTAL"]
    leg_colors = [KHO_COLORS[k] for k in REPORT_KHOS] + ["rgba(255,255,255,0.7)"]
    total_leg_w = len(leg_items) * 100
    leg_start = (chart_w - total_leg_w) / 2
    for j, (lk, lc) in enumerate(zip(leg_items, leg_colors)):
        lx = leg_start + j * 100
        if lk == "TOTAL":
            svg.append(f'<line x1="{lx:.0f}" y1="{leg_y}" x2="{lx+16:.0f}" y2="{leg_y}" stroke="{lc}" stroke-width="1.5" stroke-dasharray="4,2"/>')
        else:
            svg.append(f'<line x1="{lx:.0f}" y1="{leg_y}" x2="{lx+16:.0f}" y2="{leg_y}" stroke="{lc}" stroke-width="2"/>')
            svg.append(f'<circle cx="{lx+8:.0f}" cy="{leg_y}" r="2.5" fill="{lc}"/>')
        svg.append(f'<text x="{lx+20:.0f}" y="{leg_y+5:.0f}" font-size="12" font-weight="700" fill="#c8ccd0">{lk}</text>')

    svg.append('</svg>')
    return "\n".join(svg)


def _fmt_delta_inline(today_v, compare_v):
    """Format delta as colored inline text for the comparison note."""
    if compare_v is None or compare_v == 0:
        return "—"
    pct = (today_v - compare_v) / compare_v * 100
    if pct > 0:
        return f'<span style="color:#4caf50;font-weight:700">▲ +{pct:.1f}%</span>'
    elif pct < 0:
        return f'<span style="color:#ef5350;font-weight:700">▼ {pct:.1f}%</span>'
    return '<span style="color:#8a8f9a">— 0%</span>'


def generate_commentary(result, history):
    """Generate HTML for Items + Xe trend charts with comparison notes, plus Tấn note."""
    from datetime import datetime as _dt

    today_date = _dt.strptime(result["date"], "%d/%m/%Y")

    # Find yesterday and LFL
    yesterday = None
    lfl = None
    lfl_date_target = (today_date - timedelta(days=7)).strftime("%d/%m/%Y")
    yesterday_date_target = (today_date - timedelta(days=1)).strftime("%d/%m/%Y")

    for h in history:
        if h["date"] == result["date"]:
            continue
        if h["date"] == yesterday_date_target:
            yesterday = h
        if h["date"] == lfl_date_target:
            lfl = h

    if not history:
        return {"tan_note": "", "extra_charts": ""}

    # — Tấn comparison note (for the existing trend chart) —
    tan_note_parts = []
    if yesterday:
        tan_note_parts.append(f'vs Hôm qua ({yesterday["date"][:5]}): {_fmt_delta_inline(result["total_tons"], yesterday["total_tons"])}')
    if lfl:
        tan_note_parts.append(f'vs LFL ({lfl["date"][:5]}): {_fmt_delta_inline(result["total_tons"], lfl["total_tons"])}')
    tan_note_html = ""
    if tan_note_parts:
        tan_note = " &nbsp;·&nbsp; ".join(tan_note_parts)
        tan_note_html = f'<div class="cm-note">Hôm nay: <b>{result["total_tons"]:.2f}</b> tấn &nbsp;·&nbsp; {tan_note}</div>'

    # — Items trend chart —
    items_svg = _build_trend_svg(
        history, result, "sl_items", "total_items",
        "TREND SỐ LƯỢNG ITEMS THEO KHO",
        lambda v: f"{v/1000:.0f}K" if v >= 1000 else f"{v:.0f}",
    )
    items_note_parts = []
    if yesterday:
        items_note_parts.append(f'vs Hôm qua ({yesterday["date"][:5]}): {_fmt_delta_inline(result["total_items"], yesterday["total_items"])}')
    if lfl:
        items_note_parts.append(f'vs LFL ({lfl["date"][:5]}): {_fmt_delta_inline(result["total_items"], lfl["total_items"])}')
    items_note = " &nbsp;·&nbsp; ".join(items_note_parts) if items_note_parts else ""

    # — Xe trend chart —
    xe_svg = _build_trend_svg(
        history, result, "sl_xe", "total_xe",
        "TREND SỐ LƯỢNG XE THEO KHO",
        lambda v: f"{v:.0f}",
    )
    xe_note_parts = []
    if yesterday:
        xe_note_parts.append(f'vs Hôm qua ({yesterday["date"][:5]}): {_fmt_delta_inline(result["total_xe"], yesterday["total_xe"])}')
    if lfl:
        xe_note_parts.append(f'vs LFL ({lfl["date"][:5]}): {_fmt_delta_inline(result["total_xe"], lfl["total_xe"])}')
    xe_note = " &nbsp;·&nbsp; ".join(xe_note_parts) if xe_note_parts else ""

    extra_charts = f"""<div class="chart-box trend-box">
      <div class="chart-title">TREND SỐ LƯỢNG ITEMS THEO KHO</div>
      {items_svg}
      <div class="cm-note">Hôm nay: <b>{result['total_items']:,.0f}</b> items &nbsp;·&nbsp; {items_note}</div>
    </div>
    <div class="chart-box trend-box">
      <div class="chart-title">TREND SỐ LƯỢNG XE THEO KHO</div>
      {xe_svg}
      <div class="cm-note">Hôm nay: <b>{result['total_xe']:,}</b> xe &nbsp;·&nbsp; {xe_note}</div>
    </div>"""

    return {"tan_note": tan_note_html, "extra_charts": extra_charts}


# ──────────────────────────────────────────
#  Weekly chart helpers
# ──────────────────────────────────────────

def _aggregate_week_data(daily_entries, monday, sunday):
    """Aggregate daily entries for Mon-Sun into weekly totals."""
    week_entries = [h for h in daily_entries
                    if monday <= datetime.strptime(h["date"], "%d/%m/%Y") <= sunday]
    if not week_entries:
        return None
    result = {
        "total_sthi": 0, "total_items": 0, "total_xe": 0, "total_tons": 0,
        "days_count": len(week_entries),
        "khos": {kho: {"san_luong_tan": 0, "sl_items": 0, "sl_xe": 0}
                 for kho in REPORT_KHOS},
    }
    for entry in week_entries:
        result["total_sthi"] += entry.get("total_sthi", 0)
        result["total_items"] += entry.get("total_items", 0)
        result["total_xe"] += entry.get("total_xe", 0)
        result["total_tons"] += entry.get("total_tons", 0)
        for kho in REPORT_KHOS:
            kdata = entry.get("khos", {}).get(kho, {})
            result["khos"][kho]["san_luong_tan"] += kdata.get("san_luong_tan", 0)
            result["khos"][kho]["sl_items"] += kdata.get("sl_items", 0)
            result["khos"][kho]["sl_xe"] += kdata.get("sl_xe", 0)
    return result


def _build_weekly_history(history):
    """Group daily entries by ISO week and return list of aggregated weekly dicts."""
    week_groups = defaultdict(list)
    for h in history:
        dt = datetime.strptime(h["date"], "%d/%m/%Y")
        iso_year, iso_week, _ = dt.isocalendar()
        week_groups[(iso_year, iso_week)].append(h)
    weekly_history = []
    for (iso_year, iso_week), entries in sorted(week_groups.items()):
        jan4 = datetime(iso_year, 1, 4)
        start_of_w1 = jan4 - timedelta(days=jan4.isoweekday() - 1)
        monday = start_of_w1 + timedelta(weeks=iso_week - 1)
        sunday = monday + timedelta(days=6)
        agg = _aggregate_week_data(entries, monday, sunday)
        if agg:
            agg["week_label"] = f"W{iso_week}"
            agg["date_range"] = f"{monday.strftime('%d/%m')}\u2013{sunday.strftime('%d/%m')}"
            weekly_history.append(agg)
    return weekly_history


def _build_weekly_trend_svg(weekly_history, metric_key, total_key, fmt_fn):
    """Build SVG trend chart where x-axis = weeks."""
    chart_w, chart_h = 760, 340
    pad_l, pad_r, pad_t, pad_b = 65, 25, 25, 70
    plot_w = chart_w - pad_l - pad_r
    plot_h = chart_h - pad_t - pad_b
    n = len(weekly_history)
    if n == 0:
        return ""
    total_vals = [w[total_key] for w in weekly_history]
    max_val = max(total_vals) * 1.15 if max(total_vals) > 0 else 1
    svg = [f'<svg viewBox="0 0 {chart_w} {chart_h}" xmlns="http://www.w3.org/2000/svg" style="width:100%;height:auto">']
    for i in range(5):
        y = pad_t + plot_h * i / 4
        v = max_val * (4 - i) / 4
        svg.append(f'<line x1="{pad_l}" y1="{y:.0f}" x2="{chart_w-pad_r}" y2="{y:.0f}" stroke="#3a3f4a" stroke-width="0.5"/>')
        svg.append(f'<text x="{pad_l-5}" y="{y+4:.0f}" text-anchor="end" font-size="13" font-weight="700" fill="#b0b5c0">{fmt_fn(v)}</text>')
    inner_margin = plot_w / n * 0.5 if n > 1 else 0
    def xpos(i):
        return pad_l + inner_margin + (i * (plot_w - 2 * inner_margin) / max(n - 1, 1)) if n > 1 else pad_l + plot_w / 2
    def ypos(v):
        return pad_t + plot_h * (1 - v / max_val) if max_val > 0 else pad_t + plot_h
    bar_w = min((plot_w - 2 * inner_margin) / n * 0.55, 80) if n > 1 else 60
    for i, tv in enumerate(total_vals):
        x = xpos(i) - bar_w / 2
        y = ypos(tv)
        h = pad_t + plot_h - y
        svg.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w:.1f}" height="{h:.1f}" rx="3" fill="rgba(255,255,255,0.45)" stroke="rgba(255,255,255,0.60)" stroke-width="0.8"/>')
    has_kho = any(w.get("khos", {}).get(k, {}).get(metric_key, 0) > 0 for w in weekly_history for k in REPORT_KHOS)
    if has_kho:
        for kho in REPORT_KHOS:
            color = KHO_COLORS[kho]
            kv = [w.get("khos", {}).get(kho, {}).get(metric_key, 0) for w in weekly_history]
            pts = [(xpos(i), ypos(v)) for i, v in enumerate(kv)]
            if len(pts) > 1:
                svg.append(f'<polyline points="{" ".join(f"{x:.1f},{y:.1f}" for x,y in pts)}" fill="none" stroke="{color}" stroke-width="2" stroke-linejoin="round" stroke-linecap="round" opacity="0.9"/>')
            for x, y in pts:
                svg.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="2.5" fill="{color}" stroke="#1e2029" stroke-width="1"/>')
    total_pts = [(xpos(i), ypos(v)) for i, v in enumerate(total_vals)]
    if len(total_pts) > 1:
        svg.append(f'<polyline points="{" ".join(f"{x:.1f},{y:.1f}" for x,y in total_pts)}" fill="none" stroke="rgba(255,255,255,0.85)" stroke-width="2" stroke-dasharray="6,3" stroke-linejoin="round"/>')
    for i, (x, y) in enumerate(total_pts):
        svg.append(f'<text x="{x:.1f}" y="{y-8:.1f}" text-anchor="middle" font-size="13" font-weight="800" fill="#ffffff">{fmt_fn(total_vals[i])}</text>')
    for i, w in enumerate(weekly_history):
        x = xpos(i)
        svg.append(f'<text x="{x:.1f}" y="{pad_t+plot_h+18:.0f}" text-anchor="middle" font-size="13" font-weight="800" fill="#f0c060">{w["week_label"]}</text>')
        svg.append(f'<text x="{x:.1f}" y="{pad_t+plot_h+32:.0f}" text-anchor="middle" font-size="10" font-weight="600" fill="#8a8f9a">{w["date_range"]}</text>')
    leg_y = chart_h - 10
    leg_items = list(REPORT_KHOS) + ["TOTAL"]
    leg_colors = [KHO_COLORS[k] for k in REPORT_KHOS] + ["rgba(255,255,255,0.7)"]
    leg_start = (chart_w - len(leg_items) * 100) / 2
    for j, (lk, lc) in enumerate(zip(leg_items, leg_colors)):
        lx = leg_start + j * 100
        if lk == "TOTAL":
            svg.append(f'<line x1="{lx:.0f}" y1="{leg_y}" x2="{lx+16:.0f}" y2="{leg_y}" stroke="{lc}" stroke-width="1.5" stroke-dasharray="4,2"/>')
        else:
            svg.append(f'<line x1="{lx:.0f}" y1="{leg_y}" x2="{lx+16:.0f}" y2="{leg_y}" stroke="{lc}" stroke-width="2"/>')
            svg.append(f'<circle cx="{lx+8:.0f}" cy="{leg_y}" r="2.5" fill="{lc}"/>')
        svg.append(f'<text x="{lx+20:.0f}" y="{leg_y+5:.0f}" font-size="12" font-weight="700" fill="#c8ccd0">{lk}</text>')
    svg.append('</svg>')
    return "\n".join(svg)


# ──────────────────────────────────────────
#  HTML report + image export (ENHANCED)
# ──────────────────────────────────────────

def build_report_html(result, history, weekly_history=None):
    """Build interactive HTML dashboard with date filters and Chart.js charts."""
    import json as _json
    date = result["date"]

    # ── Embed data ──
    embed_history = []
    for h in history:
        entry = {
            "date": h["date"],
            "total_sthi": h.get("total_sthi", 0),
            "total_items": round(h.get("total_items", 0), 1),
            "total_xe": h.get("total_xe", 0),
            "total_tons": round(h.get("total_tons", 0), 4),
            "khos": {}
        }
        for kho in REPORT_KHOS:
            kd = h.get("khos", {}).get(kho, {})
            entry["khos"][kho] = {
                "san_luong_tan": round(kd.get("san_luong_tan", 0), 4),
                "sl_items": round(kd.get("sl_items", 0), 1),
                "sl_xe": kd.get("sl_xe", 0),
                "sl_sthi": kd.get("sl_sthi", 0),
            }
        # Distribute total_sthi proportionally by sl_xe for old entries
        total_sthi_val = h.get("total_sthi", 0)
        has_sthi = any(entry["khos"][k].get("sl_sthi", 0) > 0 for k in REPORT_KHOS)
        if not has_sthi and total_sthi_val > 0:
            total_xe_val = sum(entry["khos"][k]["sl_xe"] for k in REPORT_KHOS)
            if total_xe_val > 0:
                for kho in REPORT_KHOS:
                    entry["khos"][kho]["sl_sthi"] = round(total_sthi_val * entry["khos"][kho]["sl_xe"] / total_xe_val)
        embed_history.append(entry)

    current_full = {
        "date": result["date"],
        "total_sthi": result["total_sthi"],
        "total_items": round(result["total_items"], 1),
        "total_xe": result["total_xe"],
        "total_tons": round(result["total_tons"], 4),
        "khos": {}
    }
    for kho in REPORT_KHOS:
        d = result["khos"][kho]
        current_full["khos"][kho] = {
            "sl_sthi": d["sl_sthi"],
            "sl_items": round(d["sl_items"], 1),
            "sl_xe": d["sl_xe"],
            "san_luong_tan": round(d["san_luong_tan"], 4),
        }

    kho_config = {"list": REPORT_KHOS, "colors": KHO_COLORS}
    history_json = _json.dumps(embed_history, ensure_ascii=False)
    current_json = _json.dumps(current_full, ensure_ascii=False)
    kho_json = _json.dumps(kho_config, ensure_ascii=False)

    # ── Weekly parts (individual for grid alignment) ──
    wk_donut_html = ""
    wk_tan_html = ""
    wk_items_html = ""
    wk_xe_html = ""
    if weekly_history and len(weekly_history) > 0:
        cw = weekly_history[-1]
        pw = weekly_history[-2] if len(weekly_history) >= 2 else None
        w_stops, w_labels, w_cum = [], "", 0
        if cw["total_tons"] > 0:
            for kho in REPORT_KHOS:
                pct = cw["khos"][kho]["san_luong_tan"] / cw["total_tons"] * 100
                color = KHO_COLORS.get(kho, '#666')
                w_stops.append(f"{color} {w_cum:.1f}% {w_cum+pct:.1f}%")
                w_labels += f'<div class="leg-item"><span class="leg-color" style="background:{color}"></span>{kho} {pct:.1f}%</div>\n'
                w_cum += pct
        w_grad = ", ".join(w_stops) if w_stops else "#555 0% 100%"

        wk_tan_svg = _build_weekly_trend_svg(weekly_history, "san_luong_tan", "total_tons", lambda v: f"{v:.0f}")
        wk_items_svg = _build_weekly_trend_svg(weekly_history, "sl_items", "total_items", lambda v: f"{v/1000:.0f}K" if v >= 1000 else f"{v:.0f}")
        wk_xe_svg = _build_weekly_trend_svg(weekly_history, "sl_xe", "total_xe", lambda v: f"{v:.0f}")

        wk_tan_note = wk_items_note = wk_xe_note = ""
        if pw:
            wk_tan_note = f'<div class="cm-note">{cw["week_label"]}: <b>{cw["total_tons"]:,.2f}</b> t\u1ea5n &nbsp;\u00b7&nbsp; vs {pw["week_label"]}: {_fmt_delta_inline(cw["total_tons"], pw["total_tons"])}</div>'
            wk_items_note = f'<div class="cm-note">{cw["week_label"]}: <b>{cw["total_items"]:,.0f}</b> items &nbsp;\u00b7&nbsp; vs {pw["week_label"]}: {_fmt_delta_inline(cw["total_items"], pw["total_items"])}</div>'
            wk_xe_note = f'<div class="cm-note">{cw["week_label"]}: <b>{cw["total_xe"]:,}</b> xe &nbsp;\u00b7&nbsp; vs {pw["week_label"]}: {_fmt_delta_inline(cw["total_xe"], pw["total_xe"])}</div>'

        wk_donut_html = f"""<div class="chart-box">
  <div class="chart-title">% \u0110\u00d3NG G\u00d3P S\u1ea2N L\u01af\u1ee2NG ({cw["week_label"]})</div>
  <div class="donut-wrap">
    <div class="donut" style="background: conic-gradient({w_grad})"><div class="dhole"><span class="dhole-val">{cw["total_tons"]:,.1f}</span><span class="dhole-lbl">T\u1ea5n</span></div></div>
    <div class="legend">{w_labels}</div>
  </div>
</div>"""
        wk_tan_html = f"""<div class="chart-box trend-box"><div class="chart-title">TREND S\u1ea2N L\u01af\u1ee2NG (T\u1ea4N)</div><div class="svg-wrap">{wk_tan_svg}</div>{wk_tan_note}</div>"""
        wk_items_html = f"""<div class="chart-box trend-box"><div class="chart-title">TREND S\u1ed0 L\u01af\u1ee2NG ITEMS</div><div class="svg-wrap">{wk_items_svg}</div>{wk_items_note}</div>"""
        wk_xe_html = f"""<div class="chart-box trend-box"><div class="chart-title">TREND S\u1ed0 L\u01af\u1ee2NG XE</div><div class="svg-wrap">{wk_xe_svg}</div>{wk_xe_note}</div>"""


    # ── CSS ──
    css = """
* { margin:0; padding:0; box-sizing:border-box; }
body { background:#1e2029; color:#d8dbe0; display:flex; flex-direction:column; align-items:center; padding:28px; font-family:'Segoe UI',Arial,sans-serif; }
.title { background:linear-gradient(135deg,#1f5c28,#2d7a3a); color:#fff; font-size:28px; font-weight:bold; padding:18px 54px; text-align:center; border-radius:10px; margin-bottom:16px; box-shadow:0 3px 12px rgba(0,0,0,.4); letter-spacing:.5px; }
.filter-row { display:flex; align-items:center; gap:10px; margin-bottom:16px; flex-wrap:wrap; justify-content:center; }
.filter-row label { color:#8a8f9a; font-weight:700; font-size:14px; white-space:nowrap; }
.filter-row select { background:#282c38; color:#e8eaef; border:1px solid #3a3f4a; border-radius:8px; padding:8px 14px; font-size:15px; font-weight:700; cursor:pointer; outline:none; transition:border-color .2s; }
.filter-row select:hover { border-color:#4a5a6a; }
.filter-row select:focus { border-color:#38b854; box-shadow:0 0 0 2px rgba(56,184,84,.2); }
.range-btn { background:#282c38; color:#b0b5c0; border:1px solid #3a3f4a; border-radius:8px; padding:7px 16px; font-size:14px; font-weight:700; cursor:pointer; transition:all .2s; }
.range-btn:hover { background:#333842; border-color:#4a5a6a; color:#e8eaef; }
.range-btn.active { background:#2d7a3a; border-color:#38b854; color:#fff; }
.cards { display:flex; gap:18px; margin-bottom:20px; flex-wrap:wrap; justify-content:center; }
.card { background:#282c38; border:1px solid #3a3f4a; border-radius:12px; padding:16px 36px; text-align:center; min-width:170px; box-shadow:0 2px 8px rgba(0,0,0,.3); transition:transform .15s; }
.card:hover { transform:translateY(-2px); }
.card-val { font-size:34px; font-weight:800; color:#38b854; }
.card-lbl { font-size:15px; color:#8a8f9a; margin-top:4px; text-transform:uppercase; letter-spacing:.5px; }
.report { border-collapse:collapse; font-size:18px; box-shadow:0 2px 12px rgba(0,0,0,.4); border-radius:8px; overflow:hidden; margin-bottom:20px; }
.report th,.report td { border:1px solid #3a3f4a; padding:13px 20px; text-align:center; white-space:nowrap; }
.report thead th { background:#2a3a2a; color:#6aba6a; font-weight:700; font-size:16px; padding:12px 20px; line-height:1.3; text-transform:uppercase; letter-spacing:.3px; }
.report thead th.kpi-header { background:#1e3050; color:#7ab8f5; }
.report thead .group-header { font-size:18px; letter-spacing:.5px; padding:10px 20px; }
.report tbody td { background:#252830; color:#d8dbe0; font-size:18px; font-weight:500; }
.report tbody tr:nth-child(even) td { background:#2a2d38; }
.report tbody td.kho { font-weight:700; text-align:left; padding-left:18px; color:#e8eaef; font-size:18px; }
.report tbody td.kpi { background:#1e2840; color:#7ab8f5; }
.report tbody tr:nth-child(even) td.kpi { background:#222a45; }
.report .total-row td { background:#2a3a2a; font-weight:700; color:#38b854; font-size:18px; border-top:2px solid #2d7a3a; }
.report .total-row td.kpi { background:#1e3050; color:#7ab8f5; }
.number { font-variant-numeric:tabular-nums; }
.dot { display:inline-block; width:12px; height:12px; border-radius:50%; margin-right:8px; vertical-align:middle; }
.chart-box { border:1px solid #3a3f4a; border-radius:12px; padding:20px; background:#252830; box-shadow:0 2px 8px rgba(0,0,0,.3); margin-bottom:16px; }
.chart-title { font-size:17px; font-weight:700; color:#f0c060; text-align:center; margin-bottom:14px; text-transform:uppercase; letter-spacing:.5px; }
.donut-wrap { display:flex; align-items:center; gap:22px; justify-content:center; }
.donut { width:180px; height:180px; border-radius:50%; position:relative; box-shadow:0 0 20px rgba(0,0,0,.3); flex-shrink:0; }
.dhole { position:absolute; top:38px; left:38px; width:104px; height:104px; border-radius:50%; background:#252830; display:flex; flex-direction:column; align-items:center; justify-content:center; }
.dhole-val { font-size:28px; font-weight:800; color:#fff; }
.dhole-lbl { font-size:14px; color:#b0b5c0; text-transform:uppercase; font-weight:700; }
.legend { font-size:17px; color:#e8eaef; font-weight:700; }
.leg-item { margin:7px 0; display:flex; align-items:center; gap:9px; white-space:nowrap; }
.leg-color { width:16px; height:16px; border-radius:3px; flex-shrink:0; }
.cm-note { font-size:16px; color:#d0d4da; text-align:center; padding:8px 0 2px; font-weight:500; line-height:1.5; }
.trend-box { margin-bottom:0; }
.charts-grid { display:grid; grid-template-columns:1fr 1fr; gap:16px 20px; width:100%; max-width:1650px; align-items:stretch; }
.column-header { font-size:22px; font-weight:800; color:#f0c060; text-align:center; text-transform:uppercase; letter-spacing:1px; padding:10px 0 4px; border-bottom:2px solid #3a3f4a; }
.ccw { position:relative; height:300px; }
.svg-wrap { overflow:hidden; }
.svg-wrap svg { max-width:100%; height:auto; display:block; }
.chart-box.trend-box { overflow:hidden; }
"""

    # ── JavaScript ──
    js = r"""
var HISTORY=JSON.parse(document.getElementById('hData').textContent);
var CURRENT=JSON.parse(document.getElementById('cData').textContent);
var CFG=JSON.parse(document.getElementById('kData').textContent);
var KHO_LIST=CFG.list,KHO_COLORS=CFG.colors;

function pDate(s){var p=s.split('/');return new Date(+p[2],+p[1]-1,+p[0]);}
function fN(v){return Math.round(v).toLocaleString();}
function fPct(a,b){
  if(!b||b===0)return'\u2014';
  var p=(a-b)/b*100;
  if(p>0)return'<span style="color:#4caf50;font-weight:700">\u25b2 +'+p.toFixed(1)+'%</span>';
  if(p<0)return'<span style="color:#ef5350;font-weight:700">\u25bc '+p.toFixed(1)+'%</span>';
  return'<span style="color:#8a8f9a">\u2014 0%</span>';
}

/* ===== DATE PICKER ===== */
function initDP(){
  var el=document.getElementById('datePickerRow');
  var dates=HISTORY.map(function(h){return h.date;});
  var o='';
  for(var i=dates.length-1;i>=0;i--){
    var d=dates[i],cur=d===CURRENT.date;
    o+='<option value="'+d+'"'+(cur?' selected':'')+'>'+d+(cur?' (H\u00f4m nay)':'')+'</option>';
  }
  el.innerHTML='<label>\ud83d\udcc5 CH\u1eccN NG\u00c0Y:</label><select id="dpS" onchange="onDP(this.value)">'+o+'</select>';
}
function onDP(d){
  var e=null;
  for(var i=0;i<HISTORY.length;i++){if(HISTORY[i].date===d){e=HISTORY[i];break;}}
  if(!e)return;
  rCards(e);rTable(e,d===CURRENT.date?CURRENT:null);rDonut(e);
}

/* ===== CARDS ===== */
function rCards(e){
  document.getElementById('cardsContainer').innerHTML=
    '<div class="card"><div class="card-val">'+e.total_tons.toFixed(2)+'</div><div class="card-lbl">T\u1ed4NG T\u1ea4N</div></div>'+
    '<div class="card"><div class="card-val">'+e.total_xe+'</div><div class="card-lbl">T\u1ed4NG XE</div></div>'+
    '<div class="card"><div class="card-val">'+e.total_sthi+'</div><div class="card-lbl">T\u1ed4NG SI\u00caU TH\u1eca</div></div>'+
    '<div class="card"><div class="card-val">'+fN(e.total_items)+'</div><div class="card-lbl">T\u1ed4NG ITEMS</div></div>';
}

/* ===== TABLE ===== */
function rTable(entry,full){
  var c=document.getElementById('tableContainer');
  var h='<table class="report"><thead><tr>'+
    '<th rowspan="2">NG\u00c0Y</th><th rowspan="2">KHO</th>'+
    '<th colspan="4" class="group-header">CH\u1ec8 TI\u00caU CH\u00cdNH</th>'+
    '<th colspan="4" class="group-header kpi-header">CH\u1ec8 S\u1ed0 HI\u1ec6U SU\u1ea4T</th>'+
    '</tr><tr>'+
    '<th>S\u1ed0 L\u01af\u1ee2NG<br>SI\u00caU TH\u1eca</th><th>S\u1ed0 L\u01af\u1ee2NG<br>ITEMS</th>'+
    '<th>S\u1ed0 L\u01af\u1ee2NG<br>XE</th><th>S\u1ea2N L\u01af\u1ee2NG<br>(T\u1ea4N)</th>'+
    '<th class="kpi-header">T\u1ea4N/XE</th><th class="kpi-header">ITEMS<br>/SI\u00caU TH\u1eca</th>'+
    '<th class="kpi-header">SI\u00caU TH\u1eca<br>/XE</th><th class="kpi-header">KG<br>/SI\u00caU TH\u1eca</th>'+
    '</tr></thead><tbody>';
  var tST=0,tIT=0,tXE=0,tTN=0;
  for(var ki=0;ki<KHO_LIST.length;ki++){
    var kho=KHO_LIST[ki];
    var kd=entry.khos[kho]||{};
    var fd=full?(full.khos[kho]||{}):{};
    var st=fd.sl_sthi||kd.sl_sthi||0,it=kd.sl_items||0,xe=kd.sl_xe||0,tn=kd.san_luong_tan||0;
    var col=KHO_COLORS[kho]||'#666';
    var txe=xe>0?(tn/xe).toFixed(2):'\u2014';
    var ist=st>0?fN(it/st):'\u2014';
    var stxe=(st>0&&xe>0)?(st/xe).toFixed(1):'\u2014';
    var kgst=st>0?(tn*1000/st).toFixed(1):'\u2014';
    h+='<tr><td>'+entry.date+'</td>'+
      '<td class="kho"><span class="dot" style="background:'+col+'"></span>'+kho+'</td>'+
      '<td class="number">'+(st>0?st.toLocaleString():'\u2014')+'</td>'+
      '<td class="number">'+fN(it)+'</td>'+
      '<td class="number">'+xe.toLocaleString()+'</td>'+
      '<td class="number">'+tn.toFixed(2)+'</td>'+
      '<td class="number kpi">'+txe+'</td>'+
      '<td class="number kpi">'+ist+'</td>'+
      '<td class="number kpi">'+stxe+'</td>'+
      '<td class="number kpi">'+kgst+'</td></tr>';
    tST+=st;tIT+=it;tXE+=xe;tTN+=tn;
  }
  if(tST===0&&entry.total_sthi>0)tST=entry.total_sthi;
  var tt=tXE>0?(tTN/tXE).toFixed(2):'\u2014';
  var ti=tST>0?fN(tIT/tST):'\u2014';
  var ts=(tST>0&&tXE>0)?(tST/tXE).toFixed(1):'\u2014';
  var tk=tST>0?(tTN*1000/tST).toFixed(1):'\u2014';
  h+='<tr class="total-row"><td colspan="2">TOTAL</td>'+
    '<td class="number">'+(tST>0?tST.toLocaleString():'\u2014')+'</td>'+
    '<td class="number">'+fN(tIT)+'</td>'+
    '<td class="number">'+tXE.toLocaleString()+'</td>'+
    '<td class="number">'+tTN.toFixed(2)+'</td>'+
    '<td class="number kpi">'+tt+'</td>'+
    '<td class="number kpi">'+ti+'</td>'+
    '<td class="number kpi">'+ts+'</td>'+
    '<td class="number kpi">'+tk+'</td></tr>';
  h+='</tbody></table>';
  c.innerHTML=h;
}

/* ===== DONUT ===== */
function rDonut(entry){
  var box=document.getElementById('donutBox');
  var tot=entry.total_tons;
  if(tot<=0){box.innerHTML='<div class="chart-title">% \u0110\u00d3NG G\u00d3P S\u1ea2N L\u01af\u1ee2NG</div><p style="text-align:center;color:#8a8f9a;padding:20px">Kh\u00f4ng c\u00f3 d\u1eef li\u1ec7u</p>';return;}
  var stops=[],legs='',cum=0;
  for(var i=0;i<KHO_LIST.length;i++){
    var kho=KHO_LIST[i];
    var pct=((entry.khos[kho]||{}).san_luong_tan||0)/tot*100;
    var cl=KHO_COLORS[kho];
    stops.push(cl+' '+cum.toFixed(1)+'% '+(cum+pct).toFixed(1)+'%');
    legs+='<div class="leg-item"><span class="leg-color" style="background:'+cl+'"></span>'+kho+' '+pct.toFixed(1)+'%</div>';
    cum+=pct;
  }
  box.innerHTML='<div class="chart-title">% \u0110\u00d3NG G\u00d3P S\u1ea2N L\u01af\u1ee2NG</div>'+
    '<div class="donut-wrap"><div class="donut" style="background:conic-gradient('+stops.join(', ')+')"><div class="dhole"><span class="dhole-val">'+tot.toFixed(1)+'</span><span class="dhole-lbl">T\u1ea5n</span></div></div>'+
    '<div class="legend">'+legs+'</div></div>';
}

/* ===== RANGE CONTROLS ===== */
var cFrom,cTo;
function initRC(){
  var el=document.getElementById('chartControls');
  var dates=HISTORY.map(function(h){return h.date;});
  var n=dates.length;
  var di=n>14?n-14:0;
  cFrom=dates[di];cTo=dates[n-1];
  var fo='',to2='';
  for(var i=0;i<n;i++){
    fo+='<option value="'+dates[i]+'"'+(dates[i]===cFrom?' selected':'')+'>'+dates[i]+'</option>';
    to2+='<option value="'+dates[i]+'"'+(dates[i]===cTo?' selected':'')+'>'+dates[i]+'</option>';
  }
  el.innerHTML='<label>\ud83d\udcc8 KHO\u1ea2NG TH\u1edcI GIAN:</label>'+
    '<button class="range-btn" onclick="setR(7,this)">7 ng\u00e0y</button>'+
    '<button class="range-btn active" onclick="setR(14,this)">14 ng\u00e0y</button>'+
    '<button class="range-btn" onclick="setR(30,this)">30 ng\u00e0y</button>'+
    '<span style="color:#555;margin:0 8px">|</span>'+
    '<label>T\u1eeb:</label><select id="cfS" onchange="onRC()">'+fo+'</select>'+
    '<label style="margin-left:8px">\u0110\u1ebfn:</label><select id="ctS" onchange="onRC()">'+to2+'</select>';
}
function setR(days,btn){
  var dates=HISTORY.map(function(h){return h.date;});
  var n=dates.length;
  cFrom=dates[Math.max(0,n-days)];cTo=dates[n-1];
  document.getElementById('cfS').value=cFrom;
  document.getElementById('ctS').value=cTo;
  var btns=document.querySelectorAll('.range-btn');
  for(var i=0;i<btns.length;i++)btns[i].classList.remove('active');
  if(btn)btn.classList.add('active');
  updCharts();
}
function onRC(){
  cFrom=document.getElementById('cfS').value;
  cTo=document.getElementById('ctS').value;
  var btns=document.querySelectorAll('.range-btn');
  for(var i=0;i<btns.length;i++)btns[i].classList.remove('active');
  updCharts();
}
function getF(){
  var f=pDate(cFrom),t=pDate(cTo);
  return HISTORY.filter(function(h){var d=pDate(h.date);return d>=f&&d<=t;});
}

/* ===== CHART.JS ===== */
var chT,chI,chX;
function mkDS(data,mK,tK){
  var ds=[];
  ds.push({type:'bar',label:'TOTAL',data:data.map(function(h){return h[tK];}),
    backgroundColor:'rgba(255,255,255,0.18)',borderColor:'rgba(255,255,255,0.35)',
    borderWidth:1,borderRadius:3,order:2,
    barPercentage:0.6,categoryPercentage:0.8});
  for(var i=0;i<KHO_LIST.length;i++){
    var kho=KHO_LIST[i];
    ds.push({type:'line',label:kho,data:data.map(function(h){return(h.khos[kho]||{})[mK]||0;}),
      borderColor:KHO_COLORS[kho],backgroundColor:KHO_COLORS[kho],
      borderWidth:2,pointRadius:3,pointHoverRadius:6,tension:0.15,fill:false,order:1});
  }
  ds.push({type:'line',label:'TOTAL (tổng)',data:data.map(function(h){return h[tK];}),
    borderColor:'rgba(255,255,255,0.8)',borderWidth:2,borderDash:[6,3],
    pointRadius:0,tension:0.15,fill:false,order:0});
  return ds;
}
var cOpts={
  responsive:true,maintainAspectRatio:false,
  interaction:{mode:'index',intersect:false},
  plugins:{
    legend:{position:'bottom',labels:{color:'#c8ccd0',font:{size:12,weight:'bold'},usePointStyle:true,padding:14,
      filter:function(item){return item.text!=='TOTAL (t\u1ed5ng)';}}},
    tooltip:{backgroundColor:'rgba(30,32,41,0.95)',titleColor:'#f0c060',bodyColor:'#d8dbe0',
      borderColor:'#3a3f4a',borderWidth:1,padding:12,
      callbacks:{label:function(ctx){
        if(ctx.dataset.label==='TOTAL (t\u1ed5ng)')return null;
        var v=ctx.raw;
        if(typeof v==='number'&&v>1000)return ctx.dataset.label+': '+fN(v);
        if(typeof v==='number')return ctx.dataset.label+': '+v.toFixed(2);
        return ctx.dataset.label+': '+v;
      }}}
  },
  scales:{
    x:{ticks:{color:'#b0b5c0',font:{size:12,weight:'bold'},maxRotation:45,autoSkip:true,maxTicksLimit:20},grid:{color:'rgba(58,63,74,0.5)'}},
    y:{ticks:{color:'#b0b5c0',font:{size:13,weight:'bold'}},grid:{color:'rgba(58,63,74,0.5)'},beginAtZero:true}
  }
};
function deepClone(o){return JSON.parse(JSON.stringify(o));}
function initCharts(){
  var data=getF();
  var labels=data.map(function(h){return h.date.substring(0,5);});
  chT=new Chart(document.getElementById('chartTon'),{type:'bar',data:{labels:labels,datasets:mkDS(data,'san_luong_tan','total_tons')},options:deepClone(cOpts)});
  chI=new Chart(document.getElementById('chartItems'),{type:'bar',data:{labels:labels,datasets:mkDS(data,'sl_items','total_items')},options:deepClone(cOpts)});
  chX=new Chart(document.getElementById('chartXe'),{type:'bar',data:{labels:labels,datasets:mkDS(data,'sl_xe','total_xe')},options:deepClone(cOpts)});
  updNotes(data);
}
function updCharts(){
  var data=getF();
  var labels=data.map(function(h){return h.date.substring(0,5);});
  var cfgs=[['san_luong_tan','total_tons'],['sl_items','total_items'],['sl_xe','total_xe']];
  var chs=[chT,chI,chX];
  for(var i=0;i<3;i++){
    chs[i].data.labels=labels;
    chs[i].data.datasets=mkDS(data,cfgs[i][0],cfgs[i][1]);
    chs[i].update();
  }
  updNotes(data);
}
function updNotes(data){
  var ids=['noteT','noteI','noteX'];
  if(data.length<2){for(var i=0;i<3;i++)document.getElementById(ids[i]).innerHTML='';return;}
  var last=data[data.length-1],prev=data[data.length-2];
  var ld=pDate(last.date),lfd=new Date(ld.getTime()-7*86400000);
  var ls=String(lfd.getDate()).padStart(2,'0')+'/'+String(lfd.getMonth()+1).padStart(2,'0')+'/'+lfd.getFullYear();
  var lfl=null;
  for(var i=0;i<HISTORY.length;i++){if(HISTORY[i].date===ls){lfl=HISTORY[i];break;}}
  var ms=[
    {id:'noteT',v:last.total_tons,f:function(v){return v.toFixed(2);},u:'t\u1ea5n',pv:prev.total_tons,lv:lfl?lfl.total_tons:null},
    {id:'noteI',v:last.total_items,f:function(v){return fN(v);},u:'items',pv:prev.total_items,lv:lfl?lfl.total_items:null},
    {id:'noteX',v:last.total_xe,f:function(v){return v.toString();},u:'xe',pv:prev.total_xe,lv:lfl?lfl.total_xe:null}
  ];
  for(var i=0;i<ms.length;i++){
    var m=ms[i];
    var s=last.date.substring(0,5)+': <b>'+m.f(m.v)+'</b> '+m.u;
    s+=' &nbsp;\u00b7&nbsp; vs H\u00f4m qua ('+prev.date.substring(0,5)+'): '+fPct(m.v,m.pv);
    if(lfl)s+=' &nbsp;\u00b7&nbsp; vs LFL ('+lfl.date.substring(0,5)+'): '+fPct(m.v,m.lv);
    document.getElementById(m.id).innerHTML=s;
  }
}

/* ===== INIT ===== */
document.addEventListener('DOMContentLoaded',function(){
  initDP();
  var cur=null;
  for(var i=0;i<HISTORY.length;i++){if(HISTORY[i].date===CURRENT.date){cur=HISTORY[i];break;}}
  if(!cur)cur=HISTORY[HISTORY.length-1];
  rCards(cur);rTable(cur,CURRENT);rDonut(cur);
  initRC();initCharts();
});
"""

    # ── Assemble HTML ──
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Báo cáo xuất kho - {date}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>{css}</style></head><body>
<script type="application/json" id="hData">{history_json}</script>
<script type="application/json" id="cData">{current_json}</script>
<script type="application/json" id="kData">{kho_json}</script>
<div class="title">BÁO CÁO SẢN LƯỢNG VÀ LƯU LƯỢNG XE XUẤT KHO</div>
<div class="filter-row" id="datePickerRow"></div>
<div class="cards" id="cardsContainer"></div>
<div id="tableContainer"></div>
<div class="filter-row" id="chartControls"></div>
<div class="charts-grid">
  <div class="column-header">THEO NGÀY</div>
  <div class="column-header">THEO TUẦN</div>
  <div class="chart-box" id="donutBox"></div>
  {wk_donut_html}
  <div class="chart-box trend-box">
    <div class="chart-title">TREND SẢN LƯỢNG THEO KHO (TẤN)</div>
    <div class="ccw"><canvas id="chartTon"></canvas></div>
    <div class="cm-note" id="noteT"></div>
  </div>
  {wk_tan_html}
  <div class="chart-box trend-box">
    <div class="chart-title">TREND SỐ LƯỢNG ITEMS THEO KHO</div>
    <div class="ccw"><canvas id="chartItems"></canvas></div>
    <div class="cm-note" id="noteI"></div>
  </div>
  {wk_items_html}
  <div class="chart-box trend-box">
    <div class="chart-title">TREND SỐ LƯỢNG XE THEO KHO</div>
    <div class="ccw"><canvas id="chartXe"></canvas></div>
    <div class="cm-note" id="noteX"></div>
  </div>
  {wk_xe_html}
</div>
<script>{js}</script>
</body></html>"""



def _section_css():
    """Shared CSS for section HTML pages (same dark theme)."""
    return """
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body {
    background: #1e2029;
    display: inline-flex; flex-direction: column; align-items: center;
    padding: 28px 28px 36px; font-family: 'Segoe UI', Arial, sans-serif;
  }
  .title {
    background: linear-gradient(135deg, #1f5c28, #2d7a3a);
    color: #fff; font-size: 28px; font-weight: bold;
    letter-spacing: 0.5px; padding: 18px 54px;
    text-align: center; border-radius: 10px; margin-bottom: 20px;
    box-shadow: 0 3px 12px rgba(0,0,0,0.4);
  }
  .cards {
    display: flex; gap: 18px; margin-bottom: 20px;
  }
  .card {
    background: #282c38; border: 1px solid #3a3f4a; border-radius: 12px;
    padding: 16px 36px; text-align: center; min-width: 170px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.3);
  }
  .card-val { font-size: 34px; font-weight: 800; color: #38b854; }
  .card-lbl { font-size: 15px; color: #8a8f9a; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.5px; }
  .report {
    border-collapse: collapse; font-size: 18px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.4); border-radius: 8px;
    overflow: hidden; margin-bottom: 20px;
  }
  .report th, .report td {
    border: 1px solid #3a3f4a; padding: 13px 20px;
    text-align: center; white-space: nowrap;
  }
  .report thead th {
    background: #2a3a2a; color: #6aba6a;
    font-weight: 700; font-size: 16px;
    padding: 12px 20px; line-height: 1.3;
    text-transform: uppercase; letter-spacing: 0.3px;
  }
  .report thead th.kpi-header {
    background: #1e3050; color: #7ab8f5;
  }
  .report thead .group-header {
    font-size: 18px; letter-spacing: 0.5px; padding: 10px 20px;
  }
  .report tbody td {
    background: #252830; color: #d8dbe0;
    font-size: 18px; font-weight: 500;
  }
  .report tbody tr:nth-child(even) td { background: #2a2d38; }
  .report tbody td.kho {
    font-weight: 700; text-align: left; padding-left: 18px;
    color: #e8eaef; font-size: 18px;
  }
  .report tbody td.kpi {
    background: #1e2840; color: #7ab8f5;
  }
  .report tbody tr:nth-child(even) td.kpi { background: #222a45; }
  .report .total-row td {
    background: #2a3a2a; font-weight: 700;
    color: #38b854; font-size: 18px;
    border-top: 2px solid #2d7a3a;
  }
  .report .total-row td.kpi {
    background: #1e3050; color: #7ab8f5;
  }
  .number { font-variant-numeric: tabular-nums; }
  .dot {
    display: inline-block; width: 12px; height: 12px;
    border-radius: 50%; margin-right: 8px; vertical-align: middle;
  }
  .chart-box {
    border: 1px solid #3a3f4a; border-radius: 12px; padding: 20px;
    background: #252830; box-shadow: 0 2px 8px rgba(0,0,0,0.3);
  }
  .chart-title {
    font-size: 17px; font-weight: 700; color: #f0c060;
    text-align: center; margin-bottom: 14px;
    text-transform: uppercase; letter-spacing: 0.5px;
  }
  .donut-wrap {
    display: flex; align-items: center; gap: 22px;
  }
  .donut {
    width: 180px; height: 180px; border-radius: 50%;
    position: relative; box-shadow: 0 0 20px rgba(0,0,0,0.3);
  }
  .dhole {
    position: absolute; top: 38px; left: 38px;
    width: 104px; height: 104px; border-radius: 50%;
    background: #252830; display: flex; flex-direction: column;
    align-items: center; justify-content: center;
  }
  .dhole-val { font-size: 28px; font-weight: 800; color: #ffffff; }
  .dhole-lbl { font-size: 14px; color: #b0b5c0; text-transform: uppercase; font-weight: 700; }
  .legend { font-size: 17px; color: #e8eaef; font-weight: 700; }
  .leg-item { margin: 7px 0; display: flex; align-items: center; gap: 9px; white-space: nowrap; }
  .leg-color { width: 16px; height: 16px; border-radius: 3px; flex-shrink: 0; }
  .cm-note {
    font-size: 16px; color: #d0d4da; text-align: center;
    padding: 8px 0 2px; font-weight: 500; line-height: 1.5;
  }
  .section-split {
    display: flex; gap: 20px; align-items: stretch;
  }
  .section-half {
    flex: 1; display: flex; flex-direction: column; gap: 10px;
  }
  .section-half .chart-box { flex: 1; }
  .column-header {
    font-size: 22px; font-weight: 800; color: #f0c060;
    text-align: center; text-transform: uppercase;
    letter-spacing: 1px; padding: 10px 0 4px;
    border-bottom: 2px solid #3a3f4a;
  }
  .section-title {
    font-size: 22px; font-weight: 800; color: #f0c060;
    text-align: center; text-transform: uppercase;
    letter-spacing: 1px; margin-bottom: 16px;
  }
"""


def _wrap_section(body_content):
    """Wrap section body content with HTML boilerplate + shared CSS."""
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>{_section_css()}</style></head>
<body>
{body_content}
</body></html>"""


def build_section_htmls(result, history, weekly_history=None):
    """Build 5 separate HTML strings, one for each section of the report."""
    date = result["date"]
    total = result
    commentary_data = generate_commentary(result, history)
    tan_note_html = commentary_data["tan_note"]
    commentary_extra = commentary_data["extra_charts"]

    # ── Section 1: Title + Cards + Table ──
    cards_html = f"""
    <div class="cards">
      <div class="card"><div class="card-val">{total['total_tons']:.2f}</div><div class="card-lbl">TỔNG TẤN</div></div>
      <div class="card"><div class="card-val">{total['total_xe']}</div><div class="card-lbl">TỔNG XE</div></div>
      <div class="card"><div class="card-val">{total['total_sthi']}</div><div class="card-lbl">TỔNG SIÊU THỊ</div></div>
      <div class="card"><div class="card-val">{total['total_items']:,.0f}</div><div class="card-lbl">TỔNG ITEMS</div></div>
    </div>"""

    rows_html = ""
    for kho in REPORT_KHOS:
        d = result["khos"][kho]
        st = d['sl_sthi']; it = d['sl_items']; xe = d['sl_xe']; tn = d['san_luong_tan']
        txe = tn / xe if xe > 0 else 0
        ist = it / st if st > 0 else 0
        stxe = st / xe if xe > 0 else 0
        kgst = tn * 1000 / st if st > 0 else 0
        color = KHO_COLORS.get(kho, '#666')
        rows_html += f"""    <tr>
      <td>{date}</td>
      <td class="kho"><span class="dot" style="background:{color}"></span>{kho}</td>
      <td class="number">{st:,}</td>
      <td class="number">{it:,.0f}</td>
      <td class="number">{xe:,}</td>
      <td class="number">{tn:.2f}</td>
      <td class="number kpi">{txe:.2f}</td>
      <td class="number kpi">{ist:,.0f}</td>
      <td class="number kpi">{stxe:.1f}</td>
      <td class="number kpi">{kgst:,.1f}</td>
    </tr>\n"""

    txe_t = total['total_tons'] / total['total_xe'] if total['total_xe'] > 0 else 0
    ist_t = total['total_items'] / total['total_sthi'] if total['total_sthi'] > 0 else 0
    stxe_t = total['total_sthi'] / total['total_xe'] if total['total_xe'] > 0 else 0
    kgst_t = total['total_tons'] * 1000 / total['total_sthi'] if total['total_sthi'] > 0 else 0

    section1 = _wrap_section(f"""
<div class="title">BÁO CÁO SẢN LƯỢNG VÀ LƯU LƯỢNG XE XUẤT KHO</div>
{cards_html}
<table class="report">
  <thead>
    <tr>
      <th rowspan="2">NGÀY</th><th rowspan="2">KHO</th>
      <th colspan="4" class="group-header">CHỈ TIÊU CHÍNH</th>
      <th colspan="4" class="group-header kpi-header">CHỈ SỐ HIỆU SUẤT</th>
    </tr>
    <tr>
      <th>SỐ LƯỢNG<br>SIÊU THỊ</th><th>SỐ LƯỢNG<br>ITEMS</th>
      <th>SỐ LƯỢNG<br>XE</th><th>SẢN LƯỢNG<br>(TẤN)</th>
      <th class="kpi-header">TẤN/XE</th><th class="kpi-header">ITEMS<br>/SIÊU THỊ</th>
      <th class="kpi-header">SIÊU THỊ<br>/XE</th><th class="kpi-header">KG<br>/SIÊU THỊ</th>
    </tr>
  </thead>
  <tbody>
{rows_html}    <tr class="total-row">
      <td colspan="2">TOTAL</td>
      <td class="number">{total['total_sthi']:,}</td>
      <td class="number">{total['total_items']:,.0f}</td>
      <td class="number">{total['total_xe']:,}</td>
      <td class="number">{total['total_tons']:.2f}</td>
      <td class="number kpi">{txe_t:.2f}</td>
      <td class="number kpi">{ist_t:,.0f}</td>
      <td class="number kpi">{stxe_t:.1f}</td>
      <td class="number kpi">{kgst_t:,.1f}</td>
    </tr>
  </tbody>
</table>
""")

    # ── Donut data (daily) ──
    donut_segments = []
    if total['total_tons'] > 0:
        cumulative = 0
        for kho in REPORT_KHOS:
            pct = result["khos"][kho]["san_luong_tan"] / total["total_tons"] * 100
            donut_segments.append((kho, pct, cumulative, KHO_COLORS.get(kho, '#666')))
            cumulative += pct
    donut_labels = ""
    stops = []
    for kho, pct, cum, color in donut_segments:
        stops.append(f"{color} {cum:.1f}% {cum+pct:.1f}%")
        donut_labels += f'<div class="leg-item"><span class="leg-color" style="background:{color}"></span>{kho} {pct:.1f}%</div>\n'
    donut_gradient = ", ".join(stops) if stops else "#555 0% 100%"

    daily_donut = f"""<div class="section-half">
      <div class="column-header">THEO NGÀY</div>
      <div class="chart-box">
        <div class="chart-title">% ĐÓNG GÓP SẢN LƯỢNG</div>
        <div class="donut-wrap">
          <div class="donut" style="background: conic-gradient({donut_gradient})"><div class="dhole"><span class="dhole-val" id="dV">{total['total_tons']:.1f}</span><span class="dhole-lbl">Tấn</span></div></div>
          <div class="legend">
{donut_labels}          </div>
        </div>
      </div>
    </div>"""

    # ── Donut data (weekly) ──
    weekly_donut = "<div class=\"section-half\"></div>"
    if weekly_history and len(weekly_history) > 0:
        cw = weekly_history[-1]
        w_stops, w_labels, w_cum = [], "", 0
        if cw["total_tons"] > 0:
            for kho in REPORT_KHOS:
                pct = cw["khos"][kho]["san_luong_tan"] / cw["total_tons"] * 100
                color = KHO_COLORS.get(kho, '#666')
                w_stops.append(f"{color} {w_cum:.1f}% {w_cum+pct:.1f}%")
                w_labels += f'<div class="leg-item"><span class="leg-color" style="background:{color}"></span>{kho} {pct:.1f}%</div>\n'
                w_cum += pct
        w_grad = ", ".join(w_stops) if w_stops else "#555 0% 100%"
        weekly_donut = f"""<div class="section-half">
      <div class="column-header">THEO TUẦN</div>
      <div class="chart-box">
        <div class="chart-title">% ĐÓNG GÓP SẢN LƯỢNG ({cw["week_label"]})</div>
        <div class="donut-wrap">
          <div class="donut" style="background: conic-gradient({w_grad})"><div class="dhole"><span class="dhole-val">{cw["total_tons"]:,.1f}</span><span class="dhole-lbl">Tấn</span></div></div>
          <div class="legend">
{w_labels}          </div>
        </div>
      </div>
    </div>"""

    section2 = _wrap_section(f"""
<div class="section-title">% ĐÓNG GÓP SẢN LƯỢNG THEO KHO</div>
<div class="section-split">
  {daily_donut}
  {weekly_donut}
</div>
""")

    # ── Section 3: Trend sản lượng (tấn) ──
    daily_trend_tan = ""
    if history and len(history) > 0:
        trend_svg = _build_trend_svg(
            history[-14:], result, "san_luong_tan", "total_tons",
            "TREND SẢN LƯỢNG", lambda v: f"{v:.0f}")
        daily_trend_tan = f"""<div class="section-half">
      <div class="column-header">THEO NGÀY</div>
      <div class="chart-box trend-box">
        <div class="chart-title">TREND SẢN LƯỢNG THEO KHO (TẤN)</div>
        {trend_svg}
        {tan_note_html}
      </div>
    </div>"""

    weekly_trend_tan = "<div class=\"section-half\"></div>"
    if weekly_history and len(weekly_history) > 0:
        cw = weekly_history[-1]
        pw = weekly_history[-2] if len(weekly_history) >= 2 else None
        wk_tan = _build_weekly_trend_svg(weekly_history, "san_luong_tan", "total_tons", lambda v: f"{v:.0f}")
        wk_tan_note = ""
        if pw:
            wk_tan_note = f'<div class="cm-note">{cw["week_label"]}: <b>{cw["total_tons"]:,.2f}</b> tấn &nbsp;·&nbsp; vs {pw["week_label"]}: {_fmt_delta_inline(cw["total_tons"], pw["total_tons"])}</div>'
        weekly_trend_tan = f"""<div class="section-half">
      <div class="column-header">THEO TUẦN</div>
      <div class="chart-box trend-box">
        <div class="chart-title">TREND SẢN LƯỢNG (TẤN)</div>
        {wk_tan}
        {wk_tan_note}
      </div>
    </div>"""

    section3 = _wrap_section(f"""
<div class="section-title">TREND SẢN LƯỢNG</div>
<div class="section-split">
  {daily_trend_tan}
  {weekly_trend_tan}
</div>
""")

    # ── Section 4: Trend items ──
    daily_items_svg = _build_trend_svg(
        history[-14:], result, "sl_items", "total_items",
        "TREND SỐ LƯỢNG ITEMS",
        lambda v: f"{v/1000:.0f}K" if v >= 1000 else f"{v:.0f}")

    # items comparison note
    from datetime import datetime as _dt
    today_date = _dt.strptime(result["date"], "%d/%m/%Y")
    yesterday_target = (today_date - timedelta(days=1)).strftime("%d/%m/%Y")
    lfl_target = (today_date - timedelta(days=7)).strftime("%d/%m/%Y")
    yesterday = lfl = None
    for h in history:
        if h["date"] == result["date"]: continue
        if h["date"] == yesterday_target: yesterday = h
        if h["date"] == lfl_target: lfl = h

    items_note_parts = []
    if yesterday:
        items_note_parts.append(f'vs Hôm qua ({yesterday["date"][:5]}): {_fmt_delta_inline(result["total_items"], yesterday["total_items"])}')
    if lfl:
        items_note_parts.append(f'vs LFL ({lfl["date"][:5]}): {_fmt_delta_inline(result["total_items"], lfl["total_items"])}')
    items_note = " &nbsp;·&nbsp; ".join(items_note_parts) if items_note_parts else ""
    items_note_html = f'<div class="cm-note">Hôm nay: <b>{result["total_items"]:,.0f}</b> items &nbsp;·&nbsp; {items_note}</div>' if items_note else ""

    daily_items = f"""<div class="section-half">
      <div class="column-header">THEO NGÀY</div>
      <div class="chart-box trend-box">
        <div class="chart-title">TREND SỐ LƯỢNG ITEMS THEO KHO</div>
        {daily_items_svg}
        {items_note_html}
      </div>
    </div>"""

    weekly_items = "<div class=\"section-half\"></div>"
    if weekly_history and len(weekly_history) > 0:
        cw = weekly_history[-1]
        pw = weekly_history[-2] if len(weekly_history) >= 2 else None
        wk_items_svg = _build_weekly_trend_svg(weekly_history, "sl_items", "total_items", lambda v: f"{v/1000:.0f}K" if v >= 1000 else f"{v:.0f}")
        wk_items_note = ""
        if pw:
            wk_items_note = f'<div class="cm-note">{cw["week_label"]}: <b>{cw["total_items"]:,.0f}</b> items &nbsp;·&nbsp; vs {pw["week_label"]}: {_fmt_delta_inline(cw["total_items"], pw["total_items"])}</div>'
        weekly_items = f"""<div class="section-half">
      <div class="column-header">THEO TUẦN</div>
      <div class="chart-box trend-box">
        <div class="chart-title">TREND SỐ LƯỢNG ITEMS</div>
        {wk_items_svg}
        {wk_items_note}
      </div>
    </div>"""

    section4 = _wrap_section(f"""
<div class="section-title">TREND SỐ LƯỢNG ITEMS</div>
<div class="section-split">
  {daily_items}
  {weekly_items}
</div>
""")

    # ── Section 5: Trend xe ──
    daily_xe_svg = _build_trend_svg(
        history[-14:], result, "sl_xe", "total_xe",
        "TREND SỐ LƯỢNG XE",
        lambda v: f"{v:.0f}")

    xe_note_parts = []
    if yesterday:
        xe_note_parts.append(f'vs Hôm qua ({yesterday["date"][:5]}): {_fmt_delta_inline(result["total_xe"], yesterday["total_xe"])}')
    if lfl:
        xe_note_parts.append(f'vs LFL ({lfl["date"][:5]}): {_fmt_delta_inline(result["total_xe"], lfl["total_xe"])}')
    xe_note = " &nbsp;·&nbsp; ".join(xe_note_parts) if xe_note_parts else ""
    xe_note_html = f'<div class="cm-note">Hôm nay: <b>{result["total_xe"]:,}</b> xe &nbsp;·&nbsp; {xe_note}</div>' if xe_note else ""

    daily_xe = f"""<div class="section-half">
      <div class="column-header">THEO NGÀY</div>
      <div class="chart-box trend-box">
        <div class="chart-title">TREND SỐ LƯỢNG XE THEO KHO</div>
        {daily_xe_svg}
        {xe_note_html}
      </div>
    </div>"""

    weekly_xe = "<div class=\"section-half\"></div>"
    if weekly_history and len(weekly_history) > 0:
        cw = weekly_history[-1]
        pw = weekly_history[-2] if len(weekly_history) >= 2 else None
        wk_xe_svg = _build_weekly_trend_svg(weekly_history, "sl_xe", "total_xe", lambda v: f"{v:.0f}")
        wk_xe_note = ""
        if pw:
            wk_xe_note = f'<div class="cm-note">{cw["week_label"]}: <b>{cw["total_xe"]:,}</b> xe &nbsp;·&nbsp; vs {pw["week_label"]}: {_fmt_delta_inline(cw["total_xe"], pw["total_xe"])}</div>'
        weekly_xe = f"""<div class="section-half">
      <div class="column-header">THEO TUẦN</div>
      <div class="chart-box trend-box">
        <div class="chart-title">TREND SỐ LƯỢNG XE</div>
        {wk_xe_svg}
        {wk_xe_note}
      </div>
    </div>"""

    section5 = _wrap_section(f"""
<div class="section-title">TREND SỐ LƯỢNG XE</div>
<div class="section-split">
  {daily_xe}
  {weekly_xe}
</div>
""")

    return [section1, section2, section3, section4, section5]


def export_section_images(section_htmls, output_dir, date_tag):
    """Export 5 section HTMLs as PNG images using Playwright."""
    from playwright.sync_api import sync_playwright

    suffixes = ["1_BANG", "2_DONGGOP", "3_SANLUONG", "4_ITEMS", "5_XE"]
    labels = ["Bảng KPI", "% Đóng góp", "Trend Sản lượng", "Trend Items", "Trend Xe"]
    paths = []

    with sync_playwright() as p:
        browser = p.chromium.launch()
        context = browser.new_context(
            viewport={"width": 1750, "height": 1200},
            device_scale_factor=3,
        )

        for i, (html, suffix, label) in enumerate(zip(section_htmls, suffixes, labels)):
            temp_file = os.path.join(output_dir, f"_section_{suffix}.html")
            with open(temp_file, "w", encoding="utf-8") as f:
                f.write(html)

            page = context.new_page()
            page.goto(f"file:///{temp_file.replace(os.sep, '/')}")
            page.wait_for_load_state("networkidle")

            body = page.query_selector("body")
            box = body.bounding_box()

            out_path = os.path.join(output_dir, f"BAO_CAO_{date_tag}_{suffix}.png")
            page.screenshot(
                path=out_path,
                clip={"x": 0, "y": 0, "width": box["width"], "height": box["height"]},
            )
            page.close()
            paths.append(out_path)
            print(f"    ✅ {label}: {os.path.basename(out_path)}")

            try:
                os.remove(temp_file)
            except OSError:
                pass

        browser.close()
    return paths


def export_report_image(html_content, output_path):
    from playwright.sync_api import sync_playwright

    html_file = os.path.join(os.path.dirname(output_path), "_report_temp.html")
    with open(html_file, "w", encoding="utf-8") as f:
        f.write(html_content)

    with sync_playwright() as p:
        browser = p.chromium.launch()
        context = browser.new_context(
            viewport={"width": 1750, "height": 2400},
            device_scale_factor=3,
        )
        page = context.new_page()
        page.goto(f"file:///{html_file.replace(os.sep, '/')}")
        page.wait_for_load_state("networkidle")

        body = page.query_selector("body")
        box = body.bounding_box()

        page.screenshot(
            path=output_path,
            clip={"x": 0, "y": 0, "width": box["width"], "height": box["height"]},
        )
        browser.close()

    try:
        os.remove(html_file)
    except OSError:
        pass
    return output_path


# ──────────────────────────────────────────
#  Telegram (via shared lib)
# ──────────────────────────────────────────

from lib.telegram import (
    load_telegram_config as _load_tg_config,
    send_telegram_photo as _send_tg_photo,
    send_telegram_document as _send_tg_doc,
    delete_messages_by_tag,
    track_sent_message,
    load_sent_messages as _load_sent,
    save_sent_messages as _save_sent,
)

TELEGRAM_CONFIG = os.path.join(BASE, "config", "telegram.json")
SENT_MSGS_FILE = os.path.join(BASE, "output", "state", "sent_messages.json")

def load_telegram_config():
    return _load_tg_config(TELEGRAM_CONFIG, domain="daily")

def _load_sent_messages():
    return _load_sent(SENT_MSGS_FILE)

def _save_sent_messages(data):
    _save_sent(SENT_MSGS_FILE, data)

def delete_telegram_messages(date_tag):
    """Delete all previously sent Telegram messages for a given date_tag."""
    bot_token, chat_id = load_telegram_config()
    if not bot_token or not chat_id:
        return
    delete_messages_by_tag(SENT_MSGS_FILE, date_tag, bot_token, chat_id)

def send_telegram_photo(image_path, caption=""):
    """Send photo to Telegram. Returns message_id on success, None on failure."""
    bot_token, chat_id = load_telegram_config()
    if not bot_token or not chat_id:
        print("  ❌ Telegram chưa cấu hình!")
        return None
    return _send_tg_photo(image_path, caption, bot_token, chat_id, fallback_document=True)

def send_telegram_document(file_path, caption=""):
    """Send document to Telegram. Returns message_id on success, None on failure."""
    bot_token, chat_id = load_telegram_config()
    if not bot_token or not chat_id:
        return None
    return _send_tg_doc(file_path, caption, bot_token, chat_id)


# ──────────────────────────────────────────
#  Filter report (--week / --range)
# ──────────────────────────────────────────

def _filter_history_entries(history, mode, value):
    """Filter history entries by mode.
    mode='week': value='W15' or '15' → entries in ISO week 15 (of current year or matching year)
    mode='range': value=('DD/MM/YYYY', 'DD/MM/YYYY') → entries within date range (inclusive)
    Returns (filtered_entries, label_str).
    """
    filtered = []
    label = ""

    if mode == "week":
        # Parse week number: W15, w15, 15
        wk_str = value.upper().lstrip("W")
        try:
            target_week = int(wk_str)
        except ValueError:
            print(f"  ❌ Invalid week: {value}")
            return [], ""

        for h in history:
            dt = datetime.strptime(h["date"], "%d/%m/%Y")
            _, iso_week, _ = dt.isocalendar()
            if iso_week == target_week:
                filtered.append(h)

        if filtered:
            dates = [datetime.strptime(h["date"], "%d/%m/%Y") for h in filtered]
            min_d = min(dates).strftime("%d/%m")
            max_d = max(dates).strftime("%d/%m")
            label = f"TUẦN W{target_week} ({min_d}–{max_d})"
        else:
            label = f"TUẦN W{target_week}"

    elif mode == "range":
        start_str, end_str = value
        try:
            start_dt = datetime.strptime(start_str, "%d/%m/%Y")
            end_dt = datetime.strptime(end_str, "%d/%m/%Y")
        except ValueError:
            print(f"  ❌ Invalid date range: {start_str} - {end_str}")
            return [], ""

        for h in history:
            dt = datetime.strptime(h["date"], "%d/%m/%Y")
            if start_dt <= dt <= end_dt:
                filtered.append(h)

        label = f"{start_str} – {end_str}"

    filtered.sort(key=lambda x: datetime.strptime(x["date"], "%d/%m/%Y"))
    return filtered, label


def _aggregate_entries(entries):
    """Aggregate multiple history entries into a combined result dict."""
    agg = {
        "total_sthi": 0, "total_items": 0, "total_xe": 0, "total_tons": 0,
        "khos": {kho: {"sl_sthi": 0, "sl_items": 0, "sl_xe": 0, "san_luong_tan": 0}
                 for kho in REPORT_KHOS},
    }
    for entry in entries:
        agg["total_sthi"] += entry.get("total_sthi", 0)
        agg["total_items"] += entry.get("total_items", 0)
        agg["total_xe"] += entry.get("total_xe", 0)
        agg["total_tons"] += entry.get("total_tons", 0)
        for kho in REPORT_KHOS:
            kdata = entry.get("khos", {}).get(kho, {})
            agg["khos"][kho]["san_luong_tan"] += kdata.get("san_luong_tan", 0)
            agg["khos"][kho]["sl_items"] += kdata.get("sl_items", 0)
            agg["khos"][kho]["sl_xe"] += kdata.get("sl_xe", 0)
            # sl_sthi not in history — estimate from total_sthi proportionally
    return agg


def build_filter_report_html(entries, agg, label, full_history):
    """Build dark-themed HTML report for filtered period with per-day rows + aggregate."""
    num_days = len(entries)

    # ── Summary cards ──
    cards_html = f"""
    <div class="cards">
      <div class="card"><div class="card-val">{agg['total_tons']:.2f}</div><div class="card-lbl">TỔNG TẤN</div></div>
      <div class="card"><div class="card-val">{agg['total_xe']:,}</div><div class="card-lbl">TỔNG XE</div></div>
      <div class="card"><div class="card-val">{agg['total_sthi']:,}</div><div class="card-lbl">TỔNG SIÊU THỊ</div></div>
      <div class="card"><div class="card-val">{agg['total_items']:,.0f}</div><div class="card-lbl">TỔNG ITEMS</div></div>
      <div class="card card-days"><div class="card-val">{num_days}</div><div class="card-lbl">SỐ NGÀY</div></div>
    </div>"""

    # ── Per-day table rows (grouped by date, each date has 5 kho rows) ──
    rows_html = ""
    for entry in entries:
        d = entry["date"]
        first_kho = True
        day_sthi = entry.get("total_sthi", 0)
        day_items = entry.get("total_items", 0)
        day_xe = entry.get("total_xe", 0)
        day_tons = entry.get("total_tons", 0)

        for kho in REPORT_KHOS:
            kdata = entry.get("khos", {}).get(kho, {})
            tn = kdata.get("san_luong_tan", 0)
            it = kdata.get("sl_items", 0)
            xe = kdata.get("sl_xe", 0)
            color = KHO_COLORS.get(kho, '#666')

            date_cell = f'<td rowspan="5" class="date-cell">{d}</td>' if first_kho else ""
            rows_html += f"""    <tr>
      {date_cell}
      <td class="kho"><span class="dot" style="background:{color}"></span>{kho}</td>
      <td class="number">{it:,.0f}</td>
      <td class="number">{xe:,}</td>
      <td class="number">{tn:.2f}</td>
    </tr>\n"""
            first_kho = False

        # Day subtotal row
        txe = day_tons / day_xe if day_xe > 0 else 0
        rows_html += f"""    <tr class="day-total-row">
      <td colspan="2">Tổng ngày</td>
      <td class="number">{day_items:,.0f}</td>
      <td class="number">{day_xe:,}</td>
      <td class="number">{day_tons:.2f}</td>
    </tr>\n"""

    # ── Aggregate total per kho ──
    agg_rows_html = ""
    for kho in REPORT_KHOS:
        kd = agg["khos"][kho]
        tn = kd["san_luong_tan"]
        it = kd["sl_items"]
        xe = kd["sl_xe"]
        color = KHO_COLORS.get(kho, '#666')
        agg_rows_html += f"""    <tr>
      <td class="kho"><span class="dot" style="background:{color}"></span>{kho}</td>
      <td class="number">{it:,.0f}</td>
      <td class="number">{xe:,}</td>
      <td class="number">{tn:.2f}</td>
    </tr>\n"""

    txe_t = agg['total_tons'] / agg['total_xe'] if agg['total_xe'] > 0 else 0
    kgst_t = agg['total_tons'] * 1000 / agg['total_sthi'] if agg['total_sthi'] > 0 else 0

    # ── Avg per day ──
    avg_tons = agg['total_tons'] / num_days if num_days > 0 else 0
    avg_xe = agg['total_xe'] / num_days if num_days > 0 else 0
    avg_items = agg['total_items'] / num_days if num_days > 0 else 0
    avg_sthi = agg['total_sthi'] / num_days if num_days > 0 else 0

    # ── Donut chart ──
    donut_segments = []
    if agg['total_tons'] > 0:
        cumulative = 0
        for kho in REPORT_KHOS:
            pct = agg["khos"][kho]["san_luong_tan"] / agg["total_tons"] * 100
            donut_segments.append((kho, pct, cumulative, KHO_COLORS.get(kho, '#666')))
            cumulative += pct
    donut_labels = ""
    stops = []
    for kho, pct, cum, color in donut_segments:
        stops.append(f"{color} {cum:.1f}% {cum+pct:.1f}%")
        donut_labels += f'<div class="leg-item"><span class="leg-color" style="background:{color}"></span>{kho} {pct:.1f}%</div>\n'
    donut_gradient = ", ".join(stops) if stops else "#555 0% 100%"

    # ── Trend chart for filtered period (bar chart, per-day totals) ──
    trend_svg = ""
    if len(entries) > 1:
        trend_svg = _build_trend_svg(
            entries, entries[-1], "san_luong_tan", "total_tons",
            "TREND SẢN LƯỢNG", lambda v: f"{v:.0f}")

    # ── Trend items chart ──
    items_svg = ""
    if len(entries) > 1:
        items_svg = _build_trend_svg(
            entries, entries[-1], "sl_items", "total_items",
            "TREND ITEMS",
            lambda v: f"{v/1000:.0f}K" if v >= 1000 else f"{v:.0f}")

    # ── Trend xe chart ──
    xe_svg = ""
    if len(entries) > 1:
        xe_svg = _build_trend_svg(
            entries, entries[-1], "sl_xe", "total_xe",
            "TREND XE", lambda v: f"{v:.0f}")

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    background: #1e2029;
    display: inline-flex; flex-direction: column; align-items: center;
    padding: 28px 28px 36px; font-family: 'Segoe UI', Arial, sans-serif;
    min-width: 900px;
  }}
  .title {{
    background: linear-gradient(135deg, #1a4a6e, #2678a0);
    color: #fff; font-size: 26px; font-weight: bold;
    letter-spacing: 0.5px; padding: 18px 48px;
    text-align: center; border-radius: 10px; margin-bottom: 8px;
    box-shadow: 0 3px 12px rgba(0,0,0,0.4);
  }}
  .subtitle {{
    color: #8a8f9a; font-size: 15px; text-align: center;
    margin-bottom: 20px; font-weight: 600;
    letter-spacing: 0.3px;
  }}
  .cards {{
    display: flex; gap: 14px; margin-bottom: 20px; flex-wrap: wrap;
    justify-content: center;
  }}
  .card {{
    background: #282c38; border: 1px solid #3a3f4a; border-radius: 12px;
    padding: 14px 28px; text-align: center; min-width: 148px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.3);
  }}
  .card-val {{ font-size: 30px; font-weight: 800; color: #38b854; }}
  .card-days .card-val {{ color: #f0c060; }}
  .card-lbl {{ font-size: 13px; color: #8a8f9a; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.5px; }}
  .avg-row {{
    display: flex; gap: 14px; margin-bottom: 20px; flex-wrap: wrap;
    justify-content: center;
  }}
  .avg-card {{
    background: #1e2840; border: 1px solid #304060; border-radius: 10px;
    padding: 10px 22px; text-align: center; min-width: 130px;
  }}
  .avg-val {{ font-size: 22px; font-weight: 800; color: #7ab8f5; }}
  .avg-lbl {{ font-size: 12px; color: #6a7a9a; margin-top: 2px; text-transform: uppercase; }}
  .report {{
    border-collapse: collapse; font-size: 16px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.4); border-radius: 8px;
    overflow: hidden; margin-bottom: 20px;
  }}
  .report th, .report td {{
    border: 1px solid #3a3f4a; padding: 10px 16px;
    text-align: center; white-space: nowrap;
  }}
  .report thead th {{
    background: #2a3a2a; color: #6aba6a;
    font-weight: 700; font-size: 14px;
    padding: 10px 16px; line-height: 1.3;
    text-transform: uppercase; letter-spacing: 0.3px;
  }}
  .report tbody td {{
    background: #252830; color: #d8dbe0;
    font-size: 16px; font-weight: 500;
  }}
  .report tbody tr:nth-child(even) td {{ background: #2a2d38; }}
  .report tbody td.kho {{
    font-weight: 700; text-align: left; padding-left: 14px;
    color: #e8eaef; font-size: 16px;
  }}
  .report tbody td.date-cell {{
    background: #1e2840; color: #7ab8f5; font-weight: 700;
    font-size: 16px; vertical-align: middle;
  }}
  .report .day-total-row td {{
    background: #2a3040; font-weight: 600; color: #8ab0d0;
    font-size: 15px; border-top: 1px solid #4a5a7a;
  }}
  .report .total-row td {{
    background: #2a3a2a; font-weight: 700;
    color: #38b854; font-size: 17px;
    border-top: 2px solid #2d7a3a;
  }}
  .number {{ font-variant-numeric: tabular-nums; }}
  .dot {{
    display: inline-block; width: 10px; height: 10px;
    border-radius: 50%; margin-right: 6px; vertical-align: middle;
  }}
  .charts {{
    display: flex; flex-direction: column; gap: 20px; align-items: stretch;
    width: 100%; max-width: 850px;
  }}
  .charts-row {{
    display: flex; gap: 20px; align-items: flex-start;
  }}
  .chart-box {{
    border: 1px solid #3a3f4a; border-radius: 12px; padding: 20px;
    background: #252830; box-shadow: 0 2px 8px rgba(0,0,0,0.3);
    flex: 1;
  }}
  .chart-title {{
    font-size: 16px; font-weight: 700; color: #f0c060;
    text-align: center; margin-bottom: 14px;
    text-transform: uppercase; letter-spacing: 0.5px;
  }}
  .donut-wrap {{
    display: flex; align-items: center; gap: 22px;
  }}
  .donut {{
    width: 160px; height: 160px; border-radius: 50%;
    position: relative; box-shadow: 0 0 20px rgba(0,0,0,0.3);
    flex-shrink: 0;
  }}
  .dhole {{
    position: absolute; top: 32px; left: 32px;
    width: 96px; height: 96px; border-radius: 50%;
    background: #252830; display: flex; flex-direction: column;
    align-items: center; justify-content: center;
  }}
  .dhole-val {{ font-size: 24px; font-weight: 800; color: #ffffff; }}
  .dhole-lbl {{ font-size: 13px; color: #b0b5c0; text-transform: uppercase; font-weight: 700; }}
  .legend {{ font-size: 15px; color: #e8eaef; font-weight: 700; }}
  .leg-item {{ margin: 5px 0; display: flex; align-items: center; gap: 8px; white-space: nowrap; }}
  .leg-color {{ width: 14px; height: 14px; border-radius: 3px; flex-shrink: 0; }}
  .agg-table {{
    border-collapse: collapse; font-size: 16px; width: 100%;
    margin-top: 10px;
  }}
  .agg-table th, .agg-table td {{
    border: 1px solid #3a3f4a; padding: 8px 14px;
    text-align: center; white-space: nowrap;
  }}
  .agg-table thead th {{
    background: #1e3050; color: #7ab8f5;
    font-weight: 700; font-size: 13px;
    text-transform: uppercase;
  }}
  .agg-table tbody td {{
    background: #252830; color: #d8dbe0; font-weight: 500;
  }}
  .agg-table .total-row td {{
    background: #2a3a2a; font-weight: 700; color: #38b854;
    border-top: 2px solid #2d7a3a;
  }}
</style></head>
<body>
<div class="title">BÁO CÁO TỔNG HỢP · {label}</div>
<div class="subtitle">{num_days} ngày · Dữ liệu từ history.json</div>
{cards_html}
<div class="avg-row">
  <div class="avg-card"><div class="avg-val">{avg_tons:.2f}</div><div class="avg-lbl">TB Tấn/Ngày</div></div>
  <div class="avg-card"><div class="avg-val">{avg_xe:.0f}</div><div class="avg-lbl">TB Xe/Ngày</div></div>
  <div class="avg-card"><div class="avg-val">{avg_items:,.0f}</div><div class="avg-lbl">TB Items/Ngày</div></div>
  <div class="avg-card"><div class="avg-val">{avg_sthi:.0f}</div><div class="avg-lbl">TB Siêu Thị/Ngày</div></div>
  <div class="avg-card"><div class="avg-val">{txe_t:.2f}</div><div class="avg-lbl">Tấn/Xe (tổng)</div></div>
</div>

<!-- Aggregate per-kho table -->
<div class="chart-box" style="margin-bottom:20px; max-width:600px; width:100%;">
  <div class="chart-title">TỔNG HỢP THEO KHO</div>
  <table class="agg-table">
    <thead><tr><th>KHO</th><th>ITEMS</th><th>XE</th><th>TẤN</th></tr></thead>
    <tbody>
{agg_rows_html}    <tr class="total-row">
      <td>TOTAL</td>
      <td class="number">{agg['total_items']:,.0f}</td>
      <td class="number">{agg['total_xe']:,}</td>
      <td class="number">{agg['total_tons']:.2f}</td>
    </tr>
    </tbody>
  </table>
</div>

<div class="charts">
  <div class="charts-row">
    <div class="chart-box">
      <div class="chart-title">% ĐÓNG GÓP SẢN LƯỢNG</div>
      <div class="donut-wrap">
        <div class="donut" style="background: conic-gradient({donut_gradient})"><div class="dhole"><span class="dhole-val">{agg['total_tons']:.1f}</span><span class="dhole-lbl">Tấn</span></div></div>
        <div class="legend">
{donut_labels}        </div>
      </div>
    </div>
  </div>
  {"<div class='chart-box'><div class='chart-title'>TREND SẢN LƯỢNG (TẤN)</div>" + trend_svg + "</div>" if trend_svg else ""}
  {"<div class='chart-box'><div class='chart-title'>TREND ITEMS</div>" + items_svg + "</div>" if items_svg else ""}
  {"<div class='chart-box'><div class='chart-title'>TREND XE</div>" + xe_svg + "</div>" if xe_svg else ""}
</div>

<!-- Per-day detail table -->
<div style="margin-top:24px; width:100%; max-width:850px;">
  <div class="chart-box" style="padding:16px;">
    <div class="chart-title">CHI TIẾT TỪNG NGÀY</div>
    <table class="report" style="width:100%;">
      <thead><tr><th>NGÀY</th><th>KHO</th><th>ITEMS</th><th>XE</th><th>TẤN</th></tr></thead>
      <tbody>
{rows_html}      <tr class="total-row">
        <td colspan="2">TỔNG CỘNG ({num_days} ngày)</td>
        <td class="number">{agg['total_items']:,.0f}</td>
        <td class="number">{agg['total_xe']:,}</td>
        <td class="number">{agg['total_tons']:.2f}</td>
      </tr>
      </tbody>
    </table>
  </div>
</div>

</body></html>"""


def run_filter_report(mode, value):
    """Run filter report mode (--week or --range). No online fetch, no Telegram."""
    history = load_history()
    if not history:
        print("  ❌ Không có dữ liệu history.json!")
        return

    entries, label = _filter_history_entries(history, mode, value)
    if not entries:
        print(f"  ❌ Không tìm thấy dữ liệu cho {label}")
        print(f"  📊 History có {len(history)} entries: {history[0]['date']} → {history[-1]['date']}")
        return

    agg = _aggregate_entries(entries)

    # Print console summary
    print(f"\n  📊 FILTER: {label} — {len(entries)} ngày")
    print()
    print(f"  {'KHO':<10} {'ITEMS':>10} {'XE':>6} {'TẤN':>8}")
    print("  " + "-" * 40)
    for kho in REPORT_KHOS:
        kd = agg["khos"][kho]
        print(f"  {kho:<10} {kd['sl_items']:>10,.0f} {kd['sl_xe']:>6} {kd['san_luong_tan']:>8.2f}")
    print("  " + "-" * 40)
    print(f"  {'TOTAL':<10} {agg['total_items']:>10,.0f} {agg['total_xe']:>6} {agg['total_tons']:>8.2f}")

    # Build HTML
    html = build_filter_report_html(entries, agg, label, history)

    # Save
    output_dir = os.path.join(BASE, "output", "artifacts", "daily")
    os.makedirs(output_dir, exist_ok=True)
    tag = label.replace(" ", "_").replace("/", "").replace("–", "-").replace("(", "").replace(")", "")
    html_path = os.path.join(output_dir, f"BAO_CAO_FILTER_{tag}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n  ✅ HTML: {html_path}")

    # Open in browser
    import webbrowser
    webbrowser.open(f"file:///{html_path.replace(os.sep, '/')}")
    print(f"  🌐 Đã mở trình duyệt để review")


# ──────────────────────────────────────────
#  Main
# ──────────────────────────────────────────

def main():
    # ── Check for filter modes first ──
    if "--week" in sys.argv:
        idx = sys.argv.index("--week")
        week_val = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else ""
        print("=" * 60)
        print(f"  FILTER REPORT — Week {week_val}")
        print("=" * 60)
        run_filter_report("week", week_val)
        print("\n" + "=" * 60)
        print("  DONE")
        print("=" * 60)
        return

    if "--range" in sys.argv:
        idx = sys.argv.index("--range")
        if idx + 2 < len(sys.argv):
            start_date = sys.argv[idx + 1]
            end_date = sys.argv[idx + 2]
        else:
            print("  ❌ --range cần 2 tham số: DD/MM/YYYY DD/MM/YYYY")
            return
        print("=" * 60)
        print(f"  FILTER REPORT — {start_date} → {end_date}")
        print("=" * 60)
        run_filter_report("range", (start_date, end_date))
        print("\n" + "=" * 60)
        print("  DONE")
        print("=" * 60)
        return

    # ── Normal daily report mode ──
    # Parse arguments
    date_str = ""
    if "--date" in sys.argv:
        idx = sys.argv.index("--date")
        date_str = sys.argv[idx + 1]
    if not date_str:
        date_str = datetime.now().strftime("%d/%m/%Y")

    send_telegram = "--send" in sys.argv

    parts = date_str.split("/")
    date_for_file = f"{parts[0]}.{parts[1]}.{parts[2]}"
    date_tag = f"{parts[0]}{parts[1]}{parts[2]}"

    print("=" * 60)
    print(f"  DAILY REPORT - {date_str}")
    print("=" * 60)

    # Step 1: Read STHI+XE
    print("\n📋 Reading STHI+XE data...")
    sthi_rows, sthi_warnings = read_sthi_data(date_str, date_for_file, date_tag=date_tag)
    for w in sthi_warnings:
        print(f"  ⚠️ {w}")

    # Step 2: Load master data (online)
    print("\n📋 Loading master data (online)...")
    master_tl = load_master_data()

    # Step 3: Read PT (online)
    print("\n📦 Reading PT data (online)...")
    pt_rows, pt_warnings = read_pt_data(date_str, master_tl)
    for w in pt_warnings:
        print(f"  {w}")

    # Step 4: Calculate and print summary
    result = calculate_summary(sthi_rows, pt_rows, date_str)

    if sthi_warnings or pt_warnings:
        print("\n⚠️  WARNINGS:")
        for w in sthi_warnings + pt_warnings:
            print(f"  • {w}")

    # Step 5: Update history
    history = update_history(result)
    print(f"\n📊 History: {len(history)} entries")

    # Step 5b: Build weekly history
    weekly_history = _build_weekly_history(history)
    print(f"📅 Weekly history: {len(weekly_history)} weeks")

    # Step 6: Export section images (5 PNGs)
    output_dir = os.path.join(BASE, "output", "artifacts", "daily")
    os.makedirs(output_dir, exist_ok=True)
    date_tag = date_str.replace("/", "")

    print(f"\n🖼️  Rendering 5 section images...")
    section_htmls = build_section_htmls(result, history, weekly_history)
    section_paths = export_section_images(section_htmls, output_dir, date_tag)

    # Save full HTML version for browser viewing
    html = build_report_html(result, history, weekly_history)
    html_path = os.path.join(output_dir, f"BAO_CAO_{date_tag}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✅ HTML: {html_path}")

    # Step 7: Send to Telegram (only if --send flag is passed)
    if send_telegram:
        print(f"\n📤 Sending to Telegram...")

        # Delete old messages for this date if any
        delete_telegram_messages(date_tag)

        caption = f"📊 Báo cáo xuất kho {date_str} — Tổng: {result['total_tons']:.2f} tấn, {result['total_xe']} xe, {result['total_sthi']} ST"
        section_labels = ["📋 Bảng KPI", "🍩 % Đóng góp", "📈 Trend Sản lượng", "📦 Trend Items", "🚛 Trend Xe"]
        sent_msg_ids = []
        for img_path, sec_label in zip(section_paths, section_labels):
            mid = send_telegram_photo(img_path, f"{caption}\n{sec_label}")
            if mid:
                sent_msg_ids.append(mid)
        mid = send_telegram_document(html_path, f"📋 Báo cáo HTML {date_str} — mở bằng trình duyệt để xem chi tiết")
        if mid:
            sent_msg_ids.append(mid)

        # Save sent message IDs for future deletion
        if sent_msg_ids:
            sent_data = _load_sent_messages()
            sent_data[date_tag] = sent_msg_ids
            _save_sent_messages(sent_data)
            print(f"  💾 Đã lưu {len(sent_msg_ids)} message IDs (có thể xóa khi gửi lại)")
    else:
        print(f"\n📌 Review report:")
        for sp in section_paths:
            print(f"   • {sp}")
        print(f"   • {html_path}")
        print(f"   Để gửi Telegram, chạy lại với: python script/generate_report.py --date {date_str} --send")

    print("\n" + "=" * 60)
    print("  DONE")
    print("=" * 60)


if __name__ == "__main__":
    main()
