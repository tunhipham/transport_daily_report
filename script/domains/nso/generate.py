"""
NSO Dashboard Generator
========================
Generates:
1. nso_dashboard.html  — Interactive dashboard for NSO tracking
2. Khai_truong_NSO.xlsx — Opening schedule Excel export (for sharing)
3. Lich_cham_hang_NSO.xlsx — Replenishment schedule Excel export (dates mm/dd/yyyy)

Usage:
    python generate_nso_dashboard.py

Update STORES list below when new stores are added or dates change.
"""

import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Navigate up to repo root: script/domains/nso/ → transport_daily_report/
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(REPO_ROOT, "script"))
# Original data location (from shared lib)
from lib.sources import NSO_DATA_DIR as DATA_DIR
OUTPUT_DIR = os.path.join(REPO_ROOT, 'output', 'artifacts', 'nso')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# =====================================================
# VERSION RULES (from image — rule châm hàng theo version)
# =====================================================
VERSION_RULES = {
    2000: {"D": 3500, "D1": 5500, "D2": 3500, "D3": 3000, "total": 15500},
    1500: {"D": 2800, "D1": 4200, "D2": 2800, "D3": 2300, "total": 12100},
    1000: {"D": 2500, "D1": 4000, "D2": 2500, "D3": 2100, "total": 11100},
    700:  {"D": 1800, "D1": 3000, "D2": 1800, "D3": 1500, "total": 8100},
}
DONG_MAT_KG = 400  # cố định cho mọi version

# =====================================================
# STORE DATA
# =====================================================
# name_system : tên chuẩn hệ thống (DSST col B). None = chưa có trên hệ thống
# name_full   : tên đầy đủ trên hệ thống hoặc mail
# name_mail   : tên từ email (fallback)
# code        : brand_id (DSST col C)
# opening_date: ngày khai trương dự kiến dd/mm/yyyy
# version     : Version_ST_FMCG (DSST col H). None = chưa xác định
# original_date: ngày KT cũ nếu bị dời (dd/mm/yyyy). None = không dời
STORES = [
    {
        "code": "A161",
        "name_system": "KFM_HCM_NBE",
        "name_full": "A2.SH28-30 The Park Residence",
        "name_mail": "The Park Residence",
        "opening_date": "17/04/2026",
        "version": 1500,
        "original_date": None,
    },
    {
        "code": "A148",
        "name_system": "KFM_HCM_BTH",
        "name_full": "127 Tân Cảng",
        "name_mail": "127 Tân Cảng",
        "opening_date": "17/04/2026",
        "version": 1500,
        "original_date": None,
    },
    {
        "code": "A179",
        "name_system": "KFM_HCM_TDU",
        "name_full": "A1.02 Him Lam Phú An",
        "name_mail": "Him Lam Phú An",
        "opening_date": "17/04/2026",
        "version": 700,
        "original_date": None,
    },
    {
        "code": "A177",
        "name_system": "KFM_HCM_Q07",
        "name_full": "1016/21 Chung Cư Sky Garden 2-R1-2",
        "name_mail": "Sky Garden 2",
        "opening_date": "18/04/2026",
        "version": 700,
        "original_date": None,
    },
    {
        "code": "A164",
        "name_system": "KFM_BDU_DAN",
        "name_full": "B2.01.05 Tháp B Opal Boulevard",
        "name_mail": "Opal Boulevard",
        "opening_date": "23/04/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": None,
    },
    {
        "code": "A185",
        "name_system": "KFM_HCM_TDU",
        "name_full": "BS15 Vinhomes Grand Park",
        "name_mail": "VHGP Q9 - BS1501",
        "opening_date": "25/04/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": "23/04/2026",  # Dời từ 23/04 sang 25/04
    },
    {
        "code": "A178",
        "name_system": "KFM_HCM_NBE",
        "name_full": "G16 Celesta Rise",
        "name_mail": "Celesta Rise G16",
        "opening_date": "24/04/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": None,
    },
    {
        "code": "A191",
        "name_system": "KFM_HCM_Q01",
        "name_full": "35 Nguyễn Hữu Cầu",
        "name_mail": "35-37 Nguyễn Hữu Cầu",
        "opening_date": "24/04/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": None,
    },
    {
        "code": "A176",
        "name_system": "KFM_HCM_NBE",
        "name_full": "A.1.05 Sunrise Riverside",
        "name_mail": "Sunrise Riverside",
        "opening_date": "24/04/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": None,
    },
    {
        "code": "A163",
        "name_system": "KFM_HCM_TPH",
        "name_full": "S2.0.28 Block A6 Alnata Plus Celadon City",
        "name_mail": "Celadon Boulevard",
        "opening_date": "25/04/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": None,
    },
    {
        "code": "A167",
        "name_system": "KFM_HCM_TDU",
        "name_full": "TM1.15 Chung Cư 9 View",
        "name_mail": "Shophouse 9 View",
        "opening_date": "25/04/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": None,
    },
    {
        "code": "A192",
        "name_system": "KFM_HCM_BTH",
        "name_full": "61 Bùi Đình Túy",
        "name_mail": "61 Bùi Đình Túy",
        "opening_date": "25/04/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": None,
    },
    {
        "code": "A186",
        "name_system": "KFM_HCM_TPH",
        "name_full": "H.38 Melody Residences Âu Cơ",
        "name_mail": "Melody Residence",
        "opening_date": "07/05/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": "30/04/2026",  # Dời trễ 1 tuần
    },
    {
        "code": "A181",
        "name_system": "KFM_BDU_DAN",
        "name_full": "Đường M-KP Nhị Đồng - Dĩ An",
        "name_mail": "Masterise Center Point",
        "opening_date": "07/05/2026",
        "version": None,  # TODO: cần bổ sung version từ DSST col H
        "original_date": "30/04/2026",  # Dời trễ 1 tuần
    },
    {
        "code": "A175",
        "name_system": None,  # Chưa tìm thấy trên Quai
        "name_full": "Golden Mansion 119 Phổ Quang",
        "name_mail": "Golden Mansion 119 Phổ Quang",
        "opening_date": "15/05/2026",
        "version": None,
        "original_date": None,
    },
    {
        "code": "A189",
        "name_system": None,  # Chưa tìm thấy trên Quai
        "name_full": "91-93 Nguyễn Sơn",
        "name_mail": "91-93 Nguyễn Sơn",
        "opening_date": "15/05/2026",
        "version": None,
        "original_date": None,
    },
    {
        "code": "A171",
        "name_system": None,  # Chưa tìm thấy trên Quai
        "name_full": "819-821 Phạm Thế Hiển - Q8",
        "name_mail": "819-821 Phạm Thế Hiển",
        "opening_date": "08/05/2026",
        "version": None,
        "original_date": None,
    },
    {
        "code": "A190",
        "name_system": None,  # Chưa tìm thấy trên Quai
        "name_full": "Botanica Premier - PNH",
        "name_mail": "BOTANICA PREMIER",
        "opening_date": "09/05/2026",
        "version": None,
        "original_date": None,
    },
    {
        "code": "A188",
        "name_system": None,  # Chưa tìm thấy trên Quai
        "name_full": "Chung cư Thuận Việt - Q11",
        "name_mail": "CC Thuận Việt",
        "opening_date": "09/05/2026",
        "version": None,
        "original_date": None,
    },
    {
        "code": "A184",
        "name_system": None,  # Chưa tìm thấy trên Quai
        "name_full": "Saigon Mia - L1-08",
        "name_mail": "Saigon Mia - L1-08",
        "opening_date": "15/05/2026",
        "version": None,
        "original_date": None,
    },
]

# =====================================================
# HELPER FUNCTIONS
# =====================================================
DAY_NAMES = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "CN"]

def parse_date(s):
    return datetime.strptime(s, "%d/%m/%Y").date()

def fmt_ddmm(d):
    return d.strftime("%d/%m/%Y")

def fmt_mmdd(d):
    return d.strftime("%m/%d/%Y")

def day_name(d):
    return DAY_NAMES[d.weekday()]

def get_display_name(store):
    """System name preferred, fallback to mail name."""
    if store["name_system"]:
        return f'{store["name_system"]} - {store["name_full"]}'
    return store["name_mail"]

def get_display_name_code(store):
    """Display name with code for replenishment card headers."""
    return f'{get_display_name(store)} ({store["code"]})'

def get_short_label(store):
    """Short name for calendar cells: code - name_mail."""
    return f'{store["code"]} - {store["name_mail"]}'

def get_status(store, today):
    """Trả về dict(type, text, css) hoặc None nếu hết hạn (D+3 xong)."""
    d = parse_date(store["opening_date"])
    delta = (today - d).days

    # Dời lịch check
    if store.get("original_date"):
        orig = parse_date(store["original_date"])
        if delta > 3:
            return None  # đã khai trương xong → loại
        direction = "Dời lịch" if d > orig else "Đôn lịch"
        return {
            "type": "reschedule",
            "text": f'{direction}: {store["original_date"]} → {store["opening_date"]}',
            "css": "badge-reschedule",
        }

    if delta < 0:
        return {"type": "upcoming", "text": "Sắp khai trương", "css": "badge-upcoming"}
    elif 0 <= delta <= 3:
        return {"type": "opening", "text": "Đang khai trương", "css": "badge-opening"}
    else:
        return None  # loại khỏi list

def week_range(today):
    """Mon-Sun of current week."""
    mon = today - timedelta(days=today.weekday())
    sun = mon + timedelta(days=6)
    return mon, sun

def replenishment_stores(stores, today):
    """Stores opening in current week that have version info."""
    mon, sun = week_range(today)
    out = []
    for s in stores:
        d = parse_date(s["opening_date"])
        if mon <= d <= sun and s["version"] and s["version"] in VERSION_RULES:
            out.append(s)
    return out

def build_schedule(store):
    """Build 7-day replenishment schedule for a store."""
    if not store["version"] or store["version"] not in VERSION_RULES:
        return None
    rules = VERSION_RULES[store["version"]]
    d = parse_date(store["opening_date"])

    days = []
    for i in range(7):
        dt = d + timedelta(days=i)
        if i == 0:
            ksl = rules["D"]
        elif i == 1:
            ksl = rules["D1"]
        elif i == 2:
            ksl = rules["D2"]
        elif i == 3:
            ksl = rules["D3"]
        else:
            ksl = "Daily"

        # Đông Mát 400kg on D+1 (as per source sheet pattern)
        dm = DONG_MAT_KG if i == 1 else 0

        days.append({
            "date": dt,
            "label": f"D" if i == 0 else f"D+{i}",
            "ksl": ksl,
            "dm": dm,
        })

    return {"store": store, "days": days, "total_ksl": rules["total"], "total_dm": DONG_MAT_KG}


# =====================================================
# GENERATE HTML DASHBOARD
# =====================================================
def generate_dashboard():
    today = date.today()
    mon, sun = week_range(today)

    # Active stores (not past D+3)
    active = []
    for s in STORES:
        st = get_status(s, today)
        if st:
            active.append((s, st))

    rep_stores = replenishment_stores(STORES, today)
    schedules = [build_schedule(s) for s in rep_stores]
    schedules = [x for x in schedules if x]

    # Stats
    n_total = len(active)
    n_upcoming = sum(1 for _, st in active if st["type"] == "upcoming")
    n_opening = sum(1 for _, st in active if st["type"] == "opening")
    n_resched = sum(1 for _, st in active if st["type"] == "reschedule")

    # --- Build calendar data (JSON for JS) ---
    import json as _json
    cal_events = []
    for s, st in active:
        d = parse_date(s["opening_date"])
        cal_events.append({
            "date": d.isoformat(),
            "label": get_short_label(s),
            "code": s["code"],
            "type": st["type"],
        })
    cal_json = _json.dumps(cal_events, ensure_ascii=False)

    # --- Build table rows (with data-date for JS filtering) ---
    trows = ""
    for idx, (s, st) in enumerate(active, 1):
        name = get_display_name(s)
        # ISO date for JS filtering
        iso_date = parse_date(s["opening_date"]).isoformat()
        trows += f'''
            <tr class="fade-in store-row" data-date="{iso_date}" data-status="{st["type"]}" style="animation-delay:{idx*0.04}s">
                <td class="tc stt">{idx}</td>
                <td>{name}</td>
                <td class="tc"><span class="code">{s["code"]}</span></td>
                <td class="tc">{s["opening_date"]}</td>
                <td class="tc"><span class="badge {st["css"]}">{st["text"]}</span></td>
            </tr>'''

    # --- Build replenishment cards ---
    rep_html = ""
    if schedules:
        for sched in schedules:
            s = sched["store"]
            label = get_display_name_code(s)
            ver = s["version"]

            dh = ""   # day headers
            dn = ""   # day names
            kc = ""   # ksl cells
            dc = ""   # dm cells
            for d in sched["days"]:
                dt = d["date"]
                dh += f'<th>{fmt_ddmm(dt)}</th>'
                dn += f'<td class="tc dn">{day_name(dt)}</td>'
                if isinstance(d["ksl"], int):
                    kc += f'<td class="tc kv">{d["ksl"]:,}</td>'
                else:
                    kc += f'<td class="tc dv">{d["ksl"]}</td>'
                if d["dm"] > 0:
                    dc += f'<td class="tc dmv">{d["dm"]}</td>'
                else:
                    dc += f'<td class="tc"></td>'

            rep_html += f'''
            <div class="rcard fade-in">
                <div class="rhead">
                    <h3>📦 {label}</h3>
                    <span class="vbadge">Version {ver:,}</span>
                </div>
                <div class="tw">
                    <table class="rtable">
                        <thead><tr>
                            <th class="cl">Chuyển</th>
                            {dh}
                            <th>Total</th>
                        </tr></thead>
                        <tbody>
                            <tr class="dnr"><td class="cl">Thứ</td>{dn}<td></td></tr>
                            <tr><td class="cl">KSL (22:00-24:00PM) items</td>{kc}<td class="tc tv">{sched["total_ksl"]:,}</td></tr>
                            <tr><td class="cl">Đông Mát ITL(13:00-14:00PM) Kg</td>{dc}<td class="tc tv">{sched["total_dm"]}</td></tr>
                        </tbody>
                    </table>
                </div>
            </div>'''
    else:
        rep_html = '<div class="nodata"><p>⏳ Không có siêu thị khai trương trong tuần này cần lịch châm hàng.</p></div>'

    # --- Version reference ---
    vrows = ""
    for ver in sorted(VERSION_RULES.keys(), reverse=True):
        r = VERSION_RULES[ver]
        vrows += f'<tr><td class="tc">{ver:,}</td><td class="tc">{r["D"]:,}</td><td class="tc">{r["D1"]:,}</td><td class="tc">{r["D2"]:,}</td><td class="tc">{r["D3"]:,}</td><td class="tc tv">{r["total"]:,}</td></tr>'

    html = f'''<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>NSO Dashboard — Lịch Khai Trương &amp; Châm Hàng</title>
<meta name="description" content="Dashboard theo dõi lịch khai trương và châm hàng NSO">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --bg:#0a0e1a;--bg2:#111827;--bgc:rgba(17,24,39,.85);--bgh:rgba(30,41,59,.9);
  --t1:#f1f5f9;--t2:#94a3b8;--t3:#64748b;--bd:rgba(148,163,184,.1);
  --ac:#6366f1;--acl:#818cf8;--g1:linear-gradient(135deg,#6366f1,#8b5cf6);
  --g2:linear-gradient(135deg,#06b6d4,#3b82f6);
  --grn:#10b981;--org:#f59e0b;--red:#ef4444;--pur:#a855f7;--blu:#3b82f6
}}
body{{font-family:'Inter',-apple-system,sans-serif;background:var(--bg);color:var(--t1);line-height:1.6;min-height:100vh}}
.ctn{{max-width:1440px;margin:0 auto;padding:2rem}}

/* Header */
.hdr{{text-align:center;margin-bottom:2.5rem;padding:1.5rem 0}}
.hdr h1{{font-size:2.4rem;font-weight:800;background:var(--g1);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:.4rem}}
.hdr .sub{{color:var(--t2);font-size:1rem;font-weight:300}}
.hdr .di{{color:var(--t3);font-size:.85rem;margin-top:.4rem}}

/* Stats */
.sg{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:1.25rem;margin-bottom:2.5rem}}
.sc{{background:var(--bgc);border:1px solid var(--bd);border-radius:16px;padding:1.25rem;text-align:center;backdrop-filter:blur(10px);transition:transform .3s,border-color .3s}}
.sc:hover{{transform:translateY(-4px);border-color:var(--ac)}}
.sc .sv{{font-size:2.4rem;font-weight:800;margin-bottom:.2rem}}
.sc .sl{{color:var(--t2);font-size:.8rem;font-weight:600;text-transform:uppercase;letter-spacing:.05em}}
.sc:nth-child(1) .sv{{color:var(--blu)}}
.sc:nth-child(2) .sv{{color:var(--grn)}}
.sc:nth-child(3) .sv{{color:var(--org)}}
.sc:nth-child(4) .sv{{color:var(--pur)}}

/* Section */
.sec{{margin-bottom:2.5rem}}
.sh{{display:flex;align-items:center;gap:.6rem;margin-bottom:1.2rem;padding-bottom:.6rem;border-bottom:2px solid var(--bd)}}
.sh h2{{font-size:1.35rem;font-weight:700}}
.sh .si{{font-size:1.4rem}}

/* Tables */
.tw{{overflow-x:auto;border-radius:12px;border:1px solid var(--bd)}}
table{{width:100%;border-collapse:collapse;font-size:.88rem}}
thead th{{background:rgba(99,102,241,.15);color:var(--acl);font-weight:600;padding:.75rem .8rem;text-align:left;white-space:nowrap;border-bottom:2px solid var(--bd)}}
tbody td{{padding:.65rem .8rem;border-bottom:1px solid var(--bd);transition:background .2s}}
tbody tr:hover td{{background:var(--bgh)}}
.tc{{text-align:center}}

/* Badges */
.code{{background:rgba(99,102,241,.2);color:var(--acl);padding:.2rem .6rem;border-radius:8px;font-weight:600;font-size:.82rem}}
.badge{{display:inline-block;padding:.25rem .7rem;border-radius:20px;font-size:.78rem;font-weight:600;white-space:nowrap}}
.badge-upcoming{{background:rgba(16,185,129,.15);color:var(--grn);border:1px solid rgba(16,185,129,.3)}}
.badge-opening{{background:rgba(245,158,11,.15);color:var(--org);border:1px solid rgba(245,158,11,.3);animation:pulse 2s infinite}}
.badge-reschedule{{background:rgba(168,85,247,.15);color:var(--pur);border:1px solid rgba(168,85,247,.3)}}
@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.65}}}}

/* Replenishment */
.rcard{{background:var(--bgc);border:1px solid var(--bd);border-radius:16px;padding:1.25rem;margin-bottom:1.25rem;backdrop-filter:blur(10px)}}
.rhead{{display:flex;align-items:center;justify-content:space-between;margin-bottom:.8rem;flex-wrap:wrap;gap:.4rem}}
.rhead h3{{font-size:1.05rem;font-weight:700}}
.vbadge{{background:var(--g2);color:#fff;padding:.2rem .6rem;border-radius:8px;font-size:.78rem;font-weight:600}}
.rtable th{{background:rgba(6,182,212,.12);color:#67e8f9;padding:.5rem .6rem;font-size:.8rem}}
.rtable td{{padding:.45rem .6rem}}
.cl{{font-weight:600;color:var(--t2);white-space:nowrap;min-width:220px}}
.ds{{font-size:.68rem;font-weight:400;color:var(--t3)}}
.dnr td{{color:var(--t3);font-size:.78rem;font-style:italic}}
.dn{{}}
.kv{{color:#fbbf24;font-weight:700}}
.dv{{color:var(--t3);font-style:italic}}
.dmv{{color:#67e8f9;font-weight:600}}
.tv{{color:var(--acl);font-weight:800}}
.nodata{{text-align:center;padding:2.5rem;color:var(--t3);background:var(--bgc);border-radius:16px;border:1px dashed var(--bd)}}

/* Version ref */
.vref{{background:var(--bgc);border:1px solid var(--bd);border-radius:16px;padding:1.25rem}}
.vref h3{{margin-bottom:.8rem;font-weight:600;color:var(--t2)}}

/* Animations */
.fade-in{{animation:fadeUp .5s ease forwards;opacity:0}}
@keyframes fadeUp{{from{{opacity:0;transform:translateY(18px)}}to{{opacity:1;transform:translateY(0)}}}}

/* Filter bar */
.fbar{{display:flex;align-items:center;gap:1rem;margin-bottom:1.5rem;flex-wrap:wrap;padding:1rem 1.25rem;background:var(--bgc);border:1px solid var(--bd);border-radius:14px;backdrop-filter:blur(10px)}}
.fbar label{{color:var(--t2);font-size:.82rem;font-weight:600;white-space:nowrap}}
.fbar input[type="date"]{{background:rgba(30,41,59,.8);border:1px solid var(--bd);color:var(--t1);padding:.45rem .7rem;border-radius:8px;font-family:inherit;font-size:.85rem;outline:none;transition:border-color .2s}}
.fbar input[type="date"]:focus{{border-color:var(--ac)}}
.fbar input[type="date"]::-webkit-calendar-picker-indicator{{filter:invert(1) brightness(.7)}}
.fbtn{{background:var(--g1);color:#fff;border:none;padding:.45rem 1rem;border-radius:8px;font-family:inherit;font-size:.82rem;font-weight:600;cursor:pointer;transition:opacity .2s}}
.fbtn:hover{{opacity:.85}}
.fbtn.sec-btn{{background:transparent;border:1px solid var(--bd);color:var(--t2)}}
.fbtn.sec-btn:hover{{border-color:var(--ac);color:var(--t1)}}
.fcount{{color:var(--t3);font-size:.78rem;margin-left:auto}}

/* Calendar */
.cal-wrap{{background:var(--bgc);border:1px solid var(--bd);border-radius:16px;padding:1.25rem;backdrop-filter:blur(10px);margin-bottom:1.25rem}}
.cal-nav{{display:flex;align-items:center;justify-content:center;gap:1rem;margin-bottom:1rem}}
.cal-nav button{{background:transparent;border:1px solid var(--bd);color:var(--t2);width:36px;height:36px;border-radius:8px;font-size:1.1rem;cursor:pointer;transition:all .2s;display:flex;align-items:center;justify-content:center}}
.cal-nav button:hover{{border-color:var(--ac);color:var(--t1);background:rgba(99,102,241,.1)}}
.cal-title{{font-size:1.15rem;font-weight:700;color:var(--t1);min-width:180px;text-align:center}}
.cal-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:2px}}
.cal-hd{{background:rgba(99,102,241,.15);color:var(--acl);font-weight:600;font-size:.78rem;padding:.5rem .3rem;text-align:center;border-radius:6px}}
.cal-cell{{min-height:80px;background:rgba(30,41,59,.4);border-radius:6px;padding:.3rem .35rem;position:relative;transition:background .2s}}
.cal-cell:hover{{background:rgba(30,41,59,.7)}}
.cal-cell.empty{{background:transparent;min-height:40px}}
.cal-cell.today{{border:2px solid var(--ac)}}
.cal-day{{font-size:.75rem;font-weight:600;color:var(--t3);margin-bottom:2px}}
.cal-day.sun{{color:var(--red)}}
.cal-ev{{font-size:.65rem;line-height:1.3;padding:2px 4px;border-radius:4px;margin-bottom:2px;font-weight:500;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;cursor:default}}
.cal-ev.upcoming{{background:rgba(16,185,129,.15);color:var(--grn)}}
.cal-ev.opening{{background:rgba(245,158,11,.15);color:var(--org)}}
.cal-ev.reschedule{{background:rgba(168,85,247,.15);color:var(--pur)}}

/* Footer */
.ftr{{text-align:center;padding:1.5rem 0;color:var(--t3);font-size:.78rem;border-top:1px solid var(--bd);margin-top:1.5rem}}

@media(max-width:768px){{
  .ctn{{padding:1rem}}.hdr h1{{font-size:1.7rem}}
  .sg{{grid-template-columns:repeat(2,1fr);gap:.8rem}}
  .sc .sv{{font-size:1.8rem}}
  .fbar{{flex-direction:column;align-items:stretch}}
  .fcount{{margin-left:0}}
  .cal-cell{{min-height:60px}}
  .cal-ev{{font-size:.58rem}}
}}
</style>
</head>
<body>
<div class="ctn">
  <!-- Header -->
  <div class="hdr fade-in">
    <h1>🏪 NSO Dashboard</h1>
    <p class="sub">Theo dõi Lịch Khai Trương &amp; Châm Hàng — New Store Opening</p>
    <p class="di">Cập nhật: {fmt_ddmm(today)} ({day_name(today)}) &nbsp;|&nbsp; Tuần: {fmt_ddmm(mon)} → {fmt_ddmm(sun)}</p>
  </div>

  <!-- Stats -->
  <div class="sg">
    <div class="sc fade-in" style="animation-delay:.1s"><div class="sv">{n_total}</div><div class="sl">Tổng Store Active</div></div>
    <div class="sc fade-in" style="animation-delay:.15s"><div class="sv">{n_upcoming}</div><div class="sl">Sắp Khai Trương</div></div>
    <div class="sc fade-in" style="animation-delay:.2s"><div class="sv">{n_opening}</div><div class="sl">Đang Khai Trương</div></div>
    <div class="sc fade-in" style="animation-delay:.25s"><div class="sv">{n_resched}</div><div class="sl">Dời Lịch</div></div>
  </div>

  <!-- Calendar View -->
  <div class="sec">
    <div class="sh fade-in"><span class="si">📅</span><h2>Lịch Tháng</h2></div>
    <div class="cal-wrap fade-in" style="animation-delay:.1s">
      <div class="cal-nav">
        <button onclick="calPrev()" title="Tháng trước">◀</button>
        <div class="cal-title" id="calTitle"></div>
        <button onclick="calNext()" title="Tháng sau">▶</button>
      </div>
      <div class="cal-grid" id="calGrid"></div>
    </div>
  </div>

  <!-- Opening Schedule -->
  <div class="sec">
    <div class="sh fade-in"><span class="si">📋</span><h2>Lịch Khai Trương NSO</h2></div>

    <!-- Date Range Filter -->
    <div class="fbar fade-in" style="animation-delay:.1s">
      <label for="filterFrom">📅 Từ ngày</label>
      <input type="date" id="filterFrom">
      <label for="filterTo">→ Đến ngày</label>
      <input type="date" id="filterTo">
      <button class="fbtn" onclick="applyFilter()">Lọc</button>
      <button class="fbtn sec-btn" onclick="resetFilter()">Reset</button>
      <span class="fcount" id="filterCount">Hiển thị: {n_total}/{n_total} stores</span>
    </div>

    <div class="tw fade-in" style="animation-delay:.15s">
      <table id="storeTable">
        <thead><tr>
          <th class="tc" style="width:50px">STT</th>
          <th>Tên Siêu Thị</th>
          <th class="tc" style="width:75px">Code</th>
          <th class="tc" style="width:115px">Ngày KT Dự Kiến</th>
          <th class="tc" style="width:280px">Trạng Thái</th>
        </tr></thead>
        <tbody>{trows}</tbody>
      </table>
    </div>
  </div>

  <!-- Replenishment -->
  <div class="sec">
    <div class="sh fade-in"><span class="si">📦</span><h2>Lịch Châm Hàng — Tuần {fmt_ddmm(mon)} → {fmt_ddmm(sun)}</h2></div>
    {rep_html}
  </div>

  <!-- Version Rules -->
  <div class="sec">
    <div class="sh fade-in"><span class="si">📊</span><h2>Bảng Quy Tắc Châm Hàng Theo Version</h2></div>
    <div class="vref fade-in">
      <div class="tw">
        <table>
          <thead><tr>
            <th class="tc">Version</th><th class="tc">D (Ngày KT)</th><th class="tc">D+1</th>
            <th class="tc">D+2</th><th class="tc">D+3</th><th class="tc">TOTAL (4 DAYS)</th>
          </tr></thead>
          <tbody>{vrows}</tbody>
        </table>
      </div>
    </div>
  </div>

  <div class="ftr">
    <p>NSO Dashboard — Generated {fmt_ddmm(today)} &nbsp;|&nbsp; Nguồn: Lịch châm hàng + Khai trương NSO</p>
    <p>Store tự động ẩn sau D+3. Chạy lại script khi có cập nhật lịch.</p>
  </div>
</div>
<script>
function parseDate(s){{return new Date(s);}}
function applyFilter(){{
  const from=document.getElementById('filterFrom').value;
  const to=document.getElementById('filterTo').value;
  const rows=document.querySelectorAll('.store-row');
  let shown=0,total=rows.length;
  let up=0,op=0,rs=0;
  rows.forEach((r,i)=>{{
    const d=r.getAttribute('data-date');
    const st=r.getAttribute('data-status');
    let show=true;
    if(from&&d<from)show=false;
    if(to&&d>to)show=false;
    r.style.display=show?'':'none';
    if(show){{
      shown++;
      r.querySelector('.stt').textContent=shown;
      if(st==='upcoming')up++;
      if(st==='opening')op++;
      if(st==='reschedule')rs++;
    }}
  }});
  document.getElementById('filterCount').textContent='Hiển thị: '+shown+'/'+total+' stores';
  const cards=document.querySelectorAll('.sc .sv');
  if(cards.length>=4){{cards[0].textContent=shown;cards[1].textContent=up;cards[2].textContent=op;cards[3].textContent=rs;}}
}}
function resetFilter(){{
  document.getElementById('filterFrom').value='';
  document.getElementById('filterTo').value='';
  const rows=document.querySelectorAll('.store-row');
  let total=rows.length,up=0,op=0,rs=0;
  rows.forEach((r,i)=>{{
    r.style.display='';
    r.querySelector('.stt').textContent=i+1;
    const st=r.getAttribute('data-status');
    if(st==='upcoming')up++;
    if(st==='opening')op++;
    if(st==='reschedule')rs++;
  }});
  document.getElementById('filterCount').textContent='Hiển thị: '+total+'/'+total+' stores';
  const cards=document.querySelectorAll('.sc .sv');
  if(cards.length>=4){{cards[0].textContent=total;cards[1].textContent=up;cards[2].textContent=op;cards[3].textContent=rs;}}
}}

// === Calendar ===
const calEvents = {cal_json};
const MONTH_NAMES = ['Tháng 1','Tháng 2','Tháng 3','Tháng 4','Tháng 5','Tháng 6','Tháng 7','Tháng 8','Tháng 9','Tháng 10','Tháng 11','Tháng 12'];
const DAY_HEADERS = ['Thứ 2','Thứ 3','Thứ 4','Thứ 5','Thứ 6','Thứ 7','CN'];
let calYear = new Date().getFullYear();
let calMonth = new Date().getMonth();

function toLocalISO(dt) {{
  const y = dt.getFullYear();
  const m = String(dt.getMonth() + 1).padStart(2, '0');
  const d = String(dt.getDate()).padStart(2, '0');
  return y + '-' + m + '-' + d;
}}

function renderCal() {{
  const grid = document.getElementById('calGrid');
  const title = document.getElementById('calTitle');
  title.textContent = MONTH_NAMES[calMonth] + ' / ' + calYear;
  
  let html = '';
  DAY_HEADERS.forEach((d,i) => {{
    html += '<div class="cal-hd' + (i===6?' sun':'') + '">' + d + '</div>';
  }});
  
  const first = new Date(calYear, calMonth, 1);
  const last = new Date(calYear, calMonth + 1, 0);
  let startDay = (first.getDay() + 6) % 7;
  
  const todayStr = toLocalISO(new Date());
  
  for (let i = 0; i < startDay; i++) {{
    html += '<div class="cal-cell empty"></div>';
  }}
  
  for (let d = 1; d <= last.getDate(); d++) {{
    const dt = new Date(calYear, calMonth, d);
    const iso = toLocalISO(dt);
    const isSun = dt.getDay() === 0;
    const isToday = iso === todayStr;
    
    html += '<div class="cal-cell' + (isToday ? ' today' : '') + '">';
    html += '<div class="cal-day' + (isSun ? ' sun' : '') + '">' + d + '</div>';
    
    calEvents.forEach(ev => {{
      if (ev.date === iso) {{
        html += '<div class="cal-ev ' + ev.type + '" title="' + ev.label + '">' + ev.label + '</div>';
      }}
    }});
    
    html += '</div>';
  }}
  
  grid.innerHTML = html;
}}

function calPrev() {{ calMonth--; if (calMonth < 0) {{ calMonth = 11; calYear--; }} renderCal(); }}
function calNext() {{ calMonth++; if (calMonth > 11) {{ calMonth = 0; calYear++; }} renderCal(); }}

// Init calendar on load
renderCal();
</script>
</body>
</html>'''

    out = os.path.join(OUTPUT_DIR, 'nso_dashboard.html')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  ✅ Dashboard : {out}")


# =====================================================
# GENERATE EXCEL: KHAI TRƯƠNG NSO
# =====================================================
def generate_khai_truong_xlsx():
    """Format giống sheet Khai Trương NSO — milestones D-10 → D-2."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Khai Trương NSO"

    today = date.today()
    active = [s for s in STORES if get_status(s, today)]
    if not active:
        ws.cell(row=1, column=1, value="Không có store nào active")
        wb.save(os.path.join(OUTPUT_DIR, 'Khai_truong_NSO.xlsx'))
        return

    # Styles
    hfont = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    dfont = Font(name='Arial', size=10, bold=True, color='4F46E5')
    sfont = Font(name='Arial', size=10, bold=True)
    nfont = Font(name='Arial', size=10)
    rfont = Font(name='Arial', size=10, bold=True, color='FF0000')
    cen = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    milestones = [
        ("Gửi số chia hàng đợt 1 cho DC (Từ tồn có sẵn) - 70%-80%", "Trước 17:30PM", -10, nfont),
        ("Gửi số chia hàng đợt 2 cho DC - 20%-30%", "Trước 12:00AM", -6, nfont),
        ("Ngày KSL chốt phiếu chuyển khai trương (đợt 1 + đợt 2 + bổ sung phát sinh)", "Tối", -5, nfont),
        ("", "Trước 6:00AM", -4, Font(name='Arial', size=10, italic=True)),
        ("OPS xuống kiểm hàng trước khai trương D-3", "Sáng", -3, rfont),
        ("DC chuyển hàng về siêu thị D-2", "Tối", -2, rfont),
    ]

    ws.column_dimensions['A'].width = 55

    # Row 1: Opening dates
    for i, s in enumerate(active):
        col = 2 + i * 3
        d = parse_date(s["opening_date"])
        c = ws.cell(row=1, column=col, value=fmt_ddmm(d))
        c.font = dfont; c.alignment = cen; c.border = bdr
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)

    # Row 3: Store names
    for i, s in enumerate(active):
        col = 2 + i * 3
        c = ws.cell(row=3, column=col, value=get_display_name_code(s))
        c.font = sfont; c.alignment = lft; c.border = bdr
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+2)

    # Row 4: Sub-headers (Thời gian | Ngày | Thứ) for each store
    for i in range(len(active)):
        col = 2 + i * 3
        for j, lbl in enumerate(["Thời gian", "Ngày", "Thứ"]):
            c = ws.cell(row=4, column=col+j, value=lbl)
            c.font = hfont; c.fill = hfill; c.alignment = cen; c.border = bdr

    # Rows 5+: Milestones
    for mi, (mname, mtime, offset, mfont) in enumerate(milestones):
        row = 5 + mi
        c = ws.cell(row=row, column=1, value=mname)
        c.font = mfont; c.alignment = lft; c.border = bdr

        for si, s in enumerate(active):
            col = 2 + si * 3
            d = parse_date(s["opening_date"])
            md = d + timedelta(days=offset)
            dn = day_name(md)

            ws.cell(row=row, column=col, value=mtime).font = nfont
            ws.cell(row=row, column=col).alignment = cen
            ws.cell(row=row, column=col).border = bdr

            ws.cell(row=row, column=col+1, value=fmt_ddmm(md)).font = nfont
            ws.cell(row=row, column=col+1).alignment = cen
            ws.cell(row=row, column=col+1).border = bdr

            ws.cell(row=row, column=col+2, value=dn).font = nfont
            ws.cell(row=row, column=col+2).alignment = cen
            ws.cell(row=row, column=col+2).border = bdr

    # Column widths
    for i in range(len(active)):
        ws.column_dimensions[get_column_letter(2+i*3)].width = 16
        ws.column_dimensions[get_column_letter(3+i*3)].width = 14
        ws.column_dimensions[get_column_letter(4+i*3)].width = 10

    out = os.path.join(OUTPUT_DIR, 'Khai_truong_NSO.xlsx')
    wb.save(out)
    print(f"  ✅ Excel KT  : {out}")


# =====================================================
# GENERATE EXCEL: LỊCH CHÂM HÀNG
# =====================================================
def generate_lich_cham_hang_xlsx():
    """Format giống sheet Lịch châm hàng — dates mm/dd/yyyy."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lịch châm hàng"

    today = date.today()
    rep = replenishment_stores(STORES, today)

    if not rep:
        ws.cell(row=1, column=1, value="Không có store khai trương trong tuần cần lịch châm hàng")
        wb.save(os.path.join(OUTPUT_DIR, 'Lich_cham_hang_NSO.xlsx'))
        print(f"  ⚠️  Excel CH  : Không có store trong tuần")
        return

    # Styles
    hfont = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
    sfont = Font(name='Arial', size=10, bold=True)
    nfont = Font(name='Arial', size=10)
    vfont = Font(name='Arial', size=10, bold=True, color='B45309')
    difont = Font(name='Arial', size=10, italic=True, color='6B7280')
    tfont = Font(name='Arial', size=10, bold=True, color='4F46E5')
    dmfont = Font(name='Arial', size=10, bold=True, color='0E7490')
    cen = Alignment(horizontal='center', vertical='center')
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    row = 1
    for store in rep:
        sched = build_schedule(store)
        if not sched:
            continue
        label = get_display_name_code(store)

        # Row 1: Header
        headers = ["Store", "Ngày"] + [fmt_mmdd(d["date"]) for d in sched["days"]] + ["Total"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hfont; c.fill = hfill; c.alignment = cen; c.border = bdr
        row += 1

        # Row 2: Chuyển (day names)
        ws.cell(row=row, column=2, value="Chuyển").font = nfont
        ws.cell(row=row, column=2).alignment = cen
        ws.cell(row=row, column=2).border = bdr
        for ci, d in enumerate(sched["days"], 3):
            c = ws.cell(row=row, column=ci, value=day_name(d["date"]))
            c.font = nfont; c.alignment = cen; c.border = bdr
        row += 1

        # Row 3: KSL items
        ws.cell(row=row, column=1, value=label).font = sfont
        ws.cell(row=row, column=1).alignment = lft
        ws.cell(row=row, column=1).border = bdr
        ws.cell(row=row, column=2, value="KSL (22:00-24:00PM) items").font = nfont
        ws.cell(row=row, column=2).alignment = lft
        ws.cell(row=row, column=2).border = bdr
        for ci, d in enumerate(sched["days"], 3):
            c = ws.cell(row=row, column=ci)
            if isinstance(d["ksl"], int):
                c.value = d["ksl"]; c.font = vfont
            else:
                c.value = d["ksl"]; c.font = difont
            c.alignment = cen; c.border = bdr
        # Total
        c = ws.cell(row=row, column=10, value=sched["total_ksl"])
        c.font = tfont; c.alignment = cen; c.border = bdr
        c.number_format = '#,##0'
        row += 1

        # Row 4: Đông Mát
        ws.cell(row=row, column=2, value="Đông Mát ITL(13:00-14:00PM) Kg").font = nfont
        ws.cell(row=row, column=2).alignment = lft
        ws.cell(row=row, column=2).border = bdr
        for ci, d in enumerate(sched["days"], 3):
            c = ws.cell(row=row, column=ci)
            if d["dm"] > 0:
                c.value = d["dm"]; c.font = dmfont
            c.alignment = cen; c.border = bdr
        # Total
        c = ws.cell(row=row, column=10, value=sched["total_dm"])
        c.font = tfont; c.alignment = cen; c.border = bdr
        row += 2  # blank separator

    # Column widths
    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 32
    for ci in range(3, 11):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    out = os.path.join(OUTPUT_DIR, 'Lich_cham_hang_NSO.xlsx')
    wb.save(out)
    print(f"  ✅ Excel CH  : {out}")


# =====================================================
# SCREENSHOT CALENDAR
# =====================================================
def screenshot_calendar():
    """Capture the calendar section of the dashboard as PNG using Selenium."""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    import time

    dashboard_path = os.path.join(OUTPUT_DIR, 'nso_dashboard.html')
    png_path = os.path.join(OUTPUT_DIR, 'nso_calendar.png')

    opts = Options()
    opts.add_argument('--headless=new')
    opts.add_argument('--no-sandbox')
    opts.add_argument('--disable-gpu')
    opts.add_argument('--window-size=1400,2000')
    opts.add_argument('--force-device-scale-factor=1.5')

    try:
        driver = webdriver.Chrome(options=opts)
        driver.get(f'file:///{dashboard_path.replace(chr(92), "/")}')
        time.sleep(1.5)  # Wait for animations + calendar render

        # Find the calendar wrapper and scroll to it
        cal = driver.find_element(By.CSS_SELECTOR, '.cal-wrap')
        driver.execute_script("arguments[0].scrollIntoView(true);", cal)
        time.sleep(0.3)
        cal.screenshot(png_path)
        driver.quit()
        print(f"  ✅ Calendar  : {png_path}")
        return png_path
    except Exception as e:
        print(f"  ⚠️  Screenshot: Lỗi — {e}")
        try:
            driver.quit()
        except:
            pass
        return None


# =====================================================
# TELEGRAM INTEGRATION (via shared lib)
# =====================================================
import json
from lib.telegram import (
    load_telegram_config as _load_tg_config,
    send_telegram_photo as _send_tg_photo,
    send_telegram_document as _send_tg_doc,
    delete_telegram_message as _del_tg_msg,
    load_sent_messages as _load_sent,
    save_sent_messages as _save_sent,
)

TELEGRAM_CONFIG = os.path.join(REPO_ROOT, 'config', 'telegram.json')
SENT_MSG_FILE = os.path.join(REPO_ROOT, 'output', 'state', 'nso', 'sent_messages.json')

def load_telegram_config():
    return _load_tg_config(TELEGRAM_CONFIG, domain="nso")

def load_sent_messages():
    return _load_sent(SENT_MSG_FILE)

def save_sent_messages(data):
    _save_sent(SENT_MSG_FILE, data)

def delete_old_nso_messages(bot_token, chat_id):
    sent = load_sent_messages()
    old_ids = sent.get('nso_message_ids', [])
    if old_ids:
        print(f"  🗑️  Xóa {len(old_ids)} tin nhắn NSO cũ...")
        for mid in old_ids:
            _del_tg_msg(mid, bot_token, chat_id)
    sent['nso_message_ids'] = []
    save_sent_messages(sent)

def send_telegram_document(caption, file_path, bot_token, chat_id):
    mid = _send_tg_doc(file_path, caption, bot_token, chat_id)
    if mid:
        sent = load_sent_messages()
        ids = sent.get('nso_message_ids', [])
        ids.append(mid)
        sent['nso_message_ids'] = ids
        save_sent_messages(sent)
    return mid

def send_telegram_photo(photo_path, bot_token, chat_id):
    mid = _send_tg_photo(photo_path, "", bot_token, chat_id)
    if mid:
        sent = load_sent_messages()
        ids = sent.get('nso_message_ids', [])
        ids.append(mid)
        sent['nso_message_ids'] = ids
        save_sent_messages(sent)
    return mid

def build_telegram_message():
    """Build summary caption for Telegram notification."""
    today = date.today()
    mon_this, sun_this = week_range(today)
    mon_next = mon_this + timedelta(days=7)
    sun_next = sun_this + timedelta(days=7)

    active = [(s, get_status(s, today)) for s in STORES]
    active = [(s, st) for s, st in active if st]

    this_week = []
    for s, st in active:
        d = parse_date(s["opening_date"])
        if mon_this <= d <= sun_this:
            this_week.append(s)

    next_week = []
    for s, st in active:
        d = parse_date(s["opening_date"])
        if mon_next <= d <= sun_next:
            next_week.append(s)

    lines = []
    lines.append("🏪 <b>NSO Dashboard — Thông báo Thứ 3</b>")
    lines.append(f"📅 {fmt_ddmm(today)} ({day_name(today)})")
    lines.append(f"📋 Tổng active: {len(active)} stores")
    lines.append("")

    n_tw = len(this_week)
    lines.append(f"📌 <b>TUẦN NÀY — {n_tw} store ({fmt_ddmm(mon_this)} → {fmt_ddmm(sun_this)}):</b>")
    if this_week:
        for s in this_week:
            name = get_display_name(s)
            lines.append(f"  • <b>{name}</b> [{s['code']}] — {s['opening_date']}")
    else:
        lines.append("  ⏳ Không có store khai trương")
    lines.append("")

    n_nw = len(next_week)
    lines.append(f"📌 <b>TUẦN SAU — {n_nw} store ({fmt_ddmm(mon_next)} → {fmt_ddmm(sun_next)}):</b>")
    if next_week:
        for s in next_week:
            name = get_display_name(s)
            lines.append(f"  • <b>{name}</b> [{s['code']}] — {s['opening_date']}")
    else:
        lines.append("  ⏳ Không có store khai trương")

    return "\n".join(lines)


# =====================================================
# MAIN
# =====================================================
if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='NSO Dashboard Generator')
    parser.add_argument('--send-telegram', action='store_true',
                        help='Gửi thông báo tóm tắt qua Telegram sau khi generate')
    args = parser.parse_args()

    print("=" * 60)
    print("  NSO Dashboard Generator")
    print("=" * 60)
    print(f"  Hôm nay : {fmt_ddmm(date.today())} ({day_name(date.today())})")
    mon, sun = week_range(date.today())
    print(f"  Tuần    : {fmt_ddmm(mon)} → {fmt_ddmm(sun)}")
    print()

    generate_dashboard()
    generate_khai_truong_xlsx()
    generate_lich_cham_hang_xlsx()

    if args.send_telegram:
        print()
        bot_token, chat_id = load_telegram_config()
        delete_old_nso_messages(bot_token, chat_id)
        cal_png = screenshot_calendar()
        if cal_png:
            send_telegram_photo(cal_png, bot_token, chat_id)
        msg = build_telegram_message()
        dashboard_file = os.path.join(OUTPUT_DIR, 'nso_dashboard.html')
        send_telegram_document(msg, dashboard_file, bot_token, chat_id)

    print()
    print("  🎉 Hoàn tất!")
    print(f"  → Mở file: {os.path.join(OUTPUT_DIR, 'nso_dashboard.html')}")
    print()

