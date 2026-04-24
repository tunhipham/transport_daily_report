# -*- coding: utf-8 -*-
"""
nso_master.py — NSO Master Data Manager
=========================================
Single source of truth for NSO store data.
Reads/writes data/nso/nso_master.xlsx with change tracking.

Usage:
    from nso_master import NsoMaster
    master = NsoMaster()
    stores = master.load()
    master.update_store("A185", opening_date="25/04/2026", source="Manual")
    master.add_store({...}, source="Mail scan")
    master.save()
"""

import os, json, sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
MASTER_DIR = os.path.join(REPO_ROOT, "data", "nso")
MASTER_FILE = os.path.join(MASTER_DIR, "nso_master.xlsx")
OUTPUT_DIR = os.path.join(REPO_ROOT, "output", "state", "nso")
OLD_JSON = os.path.join(REPO_ROOT, "data", "nso_stores.json")

# Column order for Stores sheet
STORE_COLS = ["stt", "code", "name_system", "name_full", "name_mail",
              "opening_date", "version", "original_date", "status"]


class NsoMaster:
    """Manages NSO master Excel with change history."""

    def __init__(self):
        os.makedirs(MASTER_DIR, exist_ok=True)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        self.stores = []
        self.history = []
        self._loaded = False

    def load(self):
        """Load stores from master Excel. Falls back to old JSON."""
        if os.path.exists(MASTER_FILE):
            self._load_excel()
        elif os.path.exists(OLD_JSON):
            self._load_json()
        else:
            print("  ⚠ No master data found")
            self.stores = []
        self._loaded = True
        return self.stores

    def _load_excel(self):
        """Load from nso_master.xlsx."""
        import openpyxl
        wb = openpyxl.load_workbook(MASTER_FILE)

        # Map display headers → internal keys
        HEADER_MAP = {
            "STT": "stt", "Code": "code", "Tên hệ thống": "name_system",
            "Tên đầy đủ": "name_full", "Tên mail": "name_mail",
            "Ngày KT": "opening_date", "Version": "version",
            "Ngày gốc": "original_date", "Status": "status",
        }

        # Stores sheet
        ws = wb["Stores"]
        raw_headers = [c.value for c in ws[1]]
        headers = [HEADER_MAP.get(h, h) for h in raw_headers]
        self.stores = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            d = {}
            for i, h in enumerate(headers):
                if i < len(row) and h != "stt" and h != "status":
                    val = row[i]
                    # Convert version to int
                    if h == "version" and val:
                        try:
                            val = int(val)
                        except (ValueError, TypeError):
                            pass
                    d[h] = val if val else None
            self.stores.append(d)

        # History sheet
        if "History" in wb.sheetnames:
            ws_h = wb["History"]
            H_MAP = {
                "Thời gian": "timestamp", "Code": "code", "Tên": "name",
                "Thay đổi": "action", "Giá trị cũ": "old_value",
                "Giá trị mới": "new_value", "Nguồn": "source",
            }
            raw_h = [c.value for c in ws_h[1]]
            h_headers = [H_MAP.get(h, h) for h in raw_h]
            self.history = []
            for row in ws_h.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                d = {}
                for i, h in enumerate(h_headers):
                    if i < len(row):
                        d[h] = row[i] or ""
                self.history.append(d)

        wb.close()

    def _load_json(self):
        """Migrate from old nso_stores.json."""
        print("  📦 Migrating from nso_stores.json...")
        with open(OLD_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.stores = data if isinstance(data, list) else []
        self.history = []

    def save(self):
        """Save to nso_master.xlsx."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # ── Sheet: Stores ──
        ws = wb.active
        ws.title = "Stores"

        # Headers
        headers = ["STT", "Code", "Tên hệ thống", "Tên đầy đủ", "Tên mail",
                    "Ngày KT", "Version", "Ngày gốc", "Status"]
        keys = STORE_COLS

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data
        for i, store in enumerate(self.stores):
            row = i + 2
            ws.cell(row=row, column=1, value=i + 1).border = thin_border
            ws.cell(row=row, column=2, value=store.get("code") or "").border = thin_border
            ws.cell(row=row, column=3, value=store.get("name_system") or "").border = thin_border
            ws.cell(row=row, column=4, value=store.get("name_full") or "").border = thin_border
            ws.cell(row=row, column=5, value=store.get("name_mail") or "").border = thin_border
            ws.cell(row=row, column=6, value=store.get("opening_date") or "").border = thin_border
            ws.cell(row=row, column=7, value=store.get("version") or "").border = thin_border
            ws.cell(row=row, column=8, value=store.get("original_date") or "").border = thin_border

            # Status
            status = self._get_status(store)
            cell = ws.cell(row=row, column=9, value=status)
            cell.border = thin_border
            if "Dời" in status:
                cell.font = Font(color="9B59B6")
            elif "Đang" in status:
                cell.font = Font(color="E67E22")
            elif "Sắp" in status:
                cell.font = Font(color="27AE60")

        # Column widths
        widths = [5, 8, 22, 35, 30, 12, 9, 12, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Freeze pane
        ws.freeze_panes = "A2"

        # ── Sheet: History ──
        ws_h = wb.create_sheet("History")
        h_headers = ["Thời gian", "Code", "Tên", "Thay đổi", "Giá trị cũ", "Giá trị mới", "Nguồn"]
        h_keys = ["timestamp", "code", "name", "action", "old_value", "new_value", "source"]

        for col, h in enumerate(h_headers, 1):
            cell = ws_h.cell(row=1, column=col, value=h)
            cell.fill = PatternFill(start_color="7D3C98", end_color="7D3C98", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for i, entry in enumerate(self.history):
            row = i + 2
            for col, key in enumerate(h_keys, 1):
                cell = ws_h.cell(row=row, column=col, value=entry.get(key) or "")
                cell.border = thin_border

        h_widths = [18, 8, 30, 15, 15, 15, 12]
        for i, w in enumerate(h_widths, 1):
            ws_h.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws_h.freeze_panes = "A2"

        # Save
        wb.save(MASTER_FILE)
        print(f"  💾 Master: {MASTER_FILE} ({len(self.stores)} stores, {len(self.history)} history entries)")

        # Also save JSON for backward compat
        json_path = os.path.join(MASTER_DIR, "nso_stores.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(self.stores, f, ensure_ascii=False, indent=2)

    def save_output(self, scan_summary=None):
        """Save scan output to output/nso/ for review."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import shutil

        now = datetime.now()
        ts = now.strftime("%Y%m%d_%H%M")

        # Copy master to output
        out_master = os.path.join(OUTPUT_DIR, "nso_master.xlsx")
        if os.path.exists(MASTER_FILE):
            shutil.copy2(MASTER_FILE, out_master)

        # Scan summary text
        if scan_summary:
            summary_path = os.path.join(OUTPUT_DIR, "scan_summary.txt")
            with open(summary_path, "w", encoding="utf-8") as f:
                f.write(f"NSO Scan — {now.strftime('%d/%m/%Y %H:%M')}\n")
                f.write("=" * 50 + "\n\n")
                f.write(scan_summary)
            print(f"  📄 Summary: {summary_path}")

        print(f"  📁 Output: {out_master}")

    def _get_status(self, store):
        """Get status text for a store."""
        from datetime import date
        try:
            parts = store["opening_date"].split("/")
            d = date(int(parts[2]), int(parts[1]), int(parts[0]))
        except (KeyError, ValueError, IndexError):
            return "—"

        today = date.today()
        orig = store.get("original_date")

        if orig and orig != store["opening_date"]:
            return f"Dời: {orig} → {store['opening_date']}"
        elif d <= today:
            return "Đang khai trương"
        else:
            return "Sắp khai trương"

    def _log(self, code, name, action, old_val, new_val, source):
        """Add entry to history."""
        self.history.append({
            "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "code": code or "",
            "name": name or "",
            "action": action,
            "old_value": str(old_val) if old_val else "",
            "new_value": str(new_val) if new_val else "",
            "source": source,
        })

    def update_store(self, code, source="Manual", **fields):
        """Update a store by code. Logs changes."""
        for store in self.stores:
            if store.get("code") == code:
                name = store.get("name_full") or store.get("name_mail") or ""
                for key, new_val in fields.items():
                    old_val = store.get(key)
                    if old_val != new_val:
                        self._log(code, name, f"Update {key}", old_val, new_val, source)
                        store[key] = new_val
                return True
        return False

    def add_store(self, store_data, source="Mail scan"):
        """Add a new store. Logs addition."""
        name = store_data.get("name_mail") or store_data.get("name_full") or ""
        code = store_data.get("code") or ""
        self._log(code, name, "Thêm mới", "", store_data.get("opening_date", ""), source)
        self.stores.append(store_data)

    def merge_mail(self, mail_stores, dsst_lookup, source="Mail scan"):
        """Merge mail data into master. Returns summary string."""
        from fetch_nso_mail import _name_match, _normalize
        import re

        added = []
        updated_dates = []
        enriched = []

        for ms in mail_stores:
            # Find match
            matched = None
            for cs in self.stores:
                if _name_match(ms["name_mail"], cs):
                    matched = cs
                    break

            if matched:
                # Check date change
                if matched["opening_date"] != ms["opening_date"]:
                    old = matched["opening_date"]
                    if not matched.get("original_date"):
                        matched["original_date"] = old
                    matched["opening_date"] = ms["opening_date"]
                    self._log(
                        matched.get("code") or "", matched.get("name_full") or ms["name_mail"],
                        "Dời lịch", old, ms["opening_date"], source
                    )
                    updated_dates.append(matched.get("code") or ms["name_mail"][:20])
            else:
                # New store
                new_store = {
                    "code": None, "name_system": None,
                    "name_full": ms["name_mail"], "name_mail": ms["name_mail"],
                    "opening_date": ms["opening_date"], "version": None,
                    "original_date": None,
                }
                self.stores.append(new_store)
                self._log("", ms["name_mail"], "Thêm mới", "", ms["opening_date"], source)
                added.append(ms["name_mail"][:25])

        # Enrich with DSST
        noise = {"chung", "cư", "siêu", "thị", "mới", "bổ", "sung", "kfm", "hcm"}
        for store in self.stores:
            code = store.get("code")
            if code:
                dsst = dsst_lookup.get(code, {})
                if dsst.get("name_system") and not store.get("name_system"):
                    store["name_system"] = dsst["name_system"]
                    enriched.append(code)
                if dsst.get("version") and not store.get("version"):
                    store["version"] = dsst["version"]
                continue

            # Fuzzy match
            mail_name = (store.get("name_mail") or store.get("name_full") or "").lower()
            if not mail_name:
                continue
            keywords = [w for w in re.split(r'[\s\-/,\.]+', mail_name) if len(w) > 1 and w not in noise]
            best_match, best_score = None, 0
            for dsst_code, dsst_info in dsst_lookup.items():
                dsst_name = (dsst_info.get("branch_name") or dsst_info.get("name_full") or "").lower()
                if not dsst_name:
                    continue
                score = sum(1 for kw in keywords if kw in dsst_name)
                if score > best_score and score >= 2:
                    best_score = score
                    best_match = (dsst_code, dsst_info)
            if best_match:
                dsst_code, dsst_info = best_match
                old_code = store.get("code")
                store["code"] = dsst_code
                store["name_system"] = dsst_info.get("name_system")
                if not store.get("name_full"):
                    store["name_full"] = dsst_info.get("name_full")
                store["version"] = dsst_info.get("version")
                if old_code != dsst_code:
                    self._log(dsst_code, store.get("name_mail", ""), "DSST match",
                              old_code or "—", dsst_code, "DSST")
                    enriched.append(dsst_code)

        # Build summary
        lines = []
        lines.append(f"Mail stores: {len(mail_stores)}")
        lines.append(f"Master stores: {len(self.stores)}")
        lines.append(f"New: {len(added)}")
        lines.append(f"Date updates: {len(updated_dates)}")
        lines.append(f"DSST enriched: {len(enriched)}")
        if added:
            lines.append(f"\nNew stores:")
            for a in added:
                lines.append(f"  + {a}")
        if updated_dates:
            lines.append(f"\nDate changes:")
            for u in updated_dates:
                lines.append(f"  ~ {u}")
        if enriched:
            lines.append(f"\nDSST matches:")
            for e in enriched:
                lines.append(f"  → {e}")

        return "\n".join(lines), added, updated_dates


# ═══════════════════════════════════════════════════════════
# CLI: Initialize master from existing JSON
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    master = NsoMaster()
    master.load()
    print(f"  Loaded {len(master.stores)} stores")
    master.save()
    master.save_output()
    print("  Done!")
