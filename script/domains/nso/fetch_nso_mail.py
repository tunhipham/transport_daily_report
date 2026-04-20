# -*- coding: utf-8 -*-
"""
fetch_nso_mail.py — Scan Haraworks inbox for NSO schedule updates
==================================================================
Opens Haraworks, finds "Fwd: Cập nhật NSO - 2026" email,
parses the table, cross-references DSST sheet, and updates nso_stores.json.

Usage:
    python script/domains/nso/fetch_nso_mail.py              # Full run
    python script/domains/nso/fetch_nso_mail.py --dry-run     # Parse only, no write
    python script/domains/nso/fetch_nso_mail.py --force       # Skip day-of-week check

Schedule: Monday + Tuesday at logon (Windows Task Scheduler)
"""
import os, sys, json, time, argparse, re
from datetime import datetime, date

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(REPO_ROOT, "script"))

from lib.sources import DSST_SHEET_URL, DSST_GID
from domains.nso.generate import JSON_STORES_PATH, load_stores, save_stores

HARAWORKS_INBOX = "https://ic.haraworks.vn/internal_mail/inbox"
MAIL_SUBJECT_PATTERN = "Cập nhật NSO"
MAIL_SENDER_NAME = "Lương Thị Hương Giang"

LOGIN_USER = "SC012433"
LOGIN_PASS = "@Ptn240400"


# ══════════════════════════════════════════════════════════════
#  SELENIUM SETUP (reuse Edge profile from inject_haraworks)
# ══════════════════════════════════════════════════════════════

def _kill_stale_drivers():
    """Kill leftover msedgedriver.exe from previous sessions."""
    import subprocess as _sp
    try:
        _sp.run(["taskkill", "/f", "/im", "msedgedriver.exe"],
                capture_output=True, text=True, timeout=5)
    except Exception:
        pass
    lock_path = os.path.join(os.path.expanduser("~"), ".edge_automail", "lockfile")
    if os.path.exists(lock_path):
        try:
            os.remove(lock_path)
        except Exception:
            pass


def setup_driver():
    """Setup Edge WebDriver with shared AutoMail profile."""
    from selenium import webdriver
    from selenium.webdriver.edge.options import Options as EdgeOptions

    _kill_stale_drivers()
    auto_profile_dir = os.path.join(os.path.expanduser("~"), ".edge_automail")
    os.makedirs(auto_profile_dir, exist_ok=True)

    options = EdgeOptions()
    options.page_load_strategy = 'eager'
    options.add_argument(f"--user-data-dir={auto_profile_dir}")
    options.add_argument("--profile-directory=Default")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument("--disable-infobars")
    options.add_argument("--window-size=1600,900")
    options.add_argument("--log-level=3")

    try:
        driver = webdriver.Edge(options=options)
    except Exception as e:
        print(f"  ⚠ Edge failed ({e}), trying Chrome...")
        from selenium.webdriver.chrome.options import Options as ChromeOptions
        chrome_options = ChromeOptions()
        chrome_options.page_load_strategy = 'eager'
        chrome_auto_dir = os.path.join(REPO_ROOT, ".chrome_automail")
        os.makedirs(chrome_auto_dir, exist_ok=True)
        chrome_options.add_argument(f"--user-data-dir={chrome_auto_dir}")
        chrome_options.add_argument("--profile-directory=Default")
        chrome_options.add_argument("--window-size=1600,900")
        chrome_options.add_argument("--log-level=3")
        driver = webdriver.Chrome(options=chrome_options)

    driver.implicitly_wait(10)
    driver.set_page_load_timeout(30)
    return driver


# ══════════════════════════════════════════════════════════════
#  LOGIN
# ══════════════════════════════════════════════════════════════

def _is_logged_in(driver):
    url = driver.current_url
    source = driver.page_source
    return (
        "internal_mail" in url or
        "Hộp Thư Đến" in source or
        "Tạo Thư Mới" in source or
        "Kingfoodmart" in source
    )


def ensure_logged_in(driver):
    """Auto-login to Haraworks if needed."""
    from selenium.webdriver.common.by import By

    print("  🔑 Checking login...")
    try:
        driver.get(HARAWORKS_INBOX)
    except Exception:
        pass
    time.sleep(3)

    if _is_logged_in(driver):
        print("  ✅ Already logged in")
        return True

    print("  🔐 Logging in...")

    # Step 1: Dismiss
    try:
        dismissed = driver.execute_script("""
            var els = document.querySelectorAll('button, a, span, div[role="button"]');
            for (var i = 0; i < els.length; i++) {
                var text = (els[i].textContent || '').trim().toLowerCase();
                if (text === 'dismiss' || text === 'bỏ qua' || text === 'đóng') {
                    if (els[i].offsetParent !== null) { els[i].click(); return true; }
                }
            }
            return false;
        """)
        if dismissed:
            print("  ✅ Dismissed")
        time.sleep(2)
    except Exception:
        pass

    # Step 2: Sign in with password
    try:
        clicked = driver.execute_script("""
            var els = document.querySelectorAll('button, a, span, div[role="button"]');
            for (var i = 0; i < els.length; i++) {
                var text = (els[i].textContent || '').trim().toLowerCase();
                if (text.includes('sign in with password') || text.includes('đăng nhập bằng mật khẩu')) {
                    if (els[i].offsetParent !== null) { els[i].click(); return true; }
                }
            }
            return false;
        """)
        if clicked:
            print("  ✅ Sign in with password clicked")
        time.sleep(2)
    except Exception:
        pass

    # Step 3: Check if form already has credentials (pre-filled)
    # Try clicking submit directly first
    try:
        submit_clicked = driver.execute_script("""
            var submit = document.querySelector('button[type="submit"], input[type="submit"]');
            if (submit && submit.offsetParent !== null) {
                submit.click();
                return true;
            }
            return false;
        """)
        if submit_clicked:
            print("  ✅ Submit clicked (pre-filled credentials)")
            time.sleep(5)
            if _is_logged_in(driver):
                print("  ✅ Login successful!")
                return True
    except Exception:
        pass

    # Step 3b: Enter credentials if not pre-filled
    try:
        user_selectors = [
            "input[name='Username']", "input[name='username']", "input[name='email']",
            "input[type='email']", "input[type='text']"
        ]
        for sel in user_selectors:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            for el in els:
                if el.is_displayed():
                    el.clear()
                    el.send_keys(LOGIN_USER)
                    print(f"  ✅ Username entered")
                    break
            else:
                continue
            break

        els = driver.find_elements(By.CSS_SELECTOR, "input[type='password']")
        for el in els:
            if el.is_displayed():
                el.clear()
                el.send_keys(LOGIN_PASS)
                print("  ✅ Password entered")
                break

        time.sleep(1)
        driver.execute_script("""
            var btn = document.querySelector('button[type="submit"]');
            if (btn) btn.click();
        """)
        time.sleep(5)
    except Exception as e:
        print(f"  ⚠ Credential entry error: {e}")

    # Verify
    for i in range(12):
        if _is_logged_in(driver):
            print("  ✅ Login successful!")
            return True
        time.sleep(5)

    print("  ❌ Login failed")
    return False


# ══════════════════════════════════════════════════════════════
#  FIND AND PARSE NSO MAIL
# ══════════════════════════════════════════════════════════════

def find_nso_mail(driver):
    """Find the NSO update email in inbox and navigate to its detail page."""
    from selenium.webdriver.common.by import By

    print(f"  🔍 Searching for '{MAIL_SUBJECT_PATTERN}'...")
    driver.get(HARAWORKS_INBOX)
    time.sleep(4)

    # Extract the detail URL from the inbox list
    detail_url = driver.execute_script("""
        var subject = arguments[0];
        // Search all links for the NSO email
        var links = document.querySelectorAll('a[href*="detail"]');
        for (var i = 0; i < links.length; i++) {
            var row = links[i].closest('tr') || links[i].parentElement;
            var text = (row ? row.textContent : links[i].textContent) || '';
            if (text.indexOf(subject) >= 0) {
                return links[i].href;
            }
        }
        // Fallback: search text nodes
        var all = document.querySelectorAll('td, div, span');
        for (var i = 0; i < all.length; i++) {
            var t = (all[i].textContent || '').trim();
            if (t.indexOf(subject) >= 0 && t.length < 500) {
                // Walk up to find the clickable row/link
                var el = all[i];
                for (var j = 0; j < 10; j++) {
                    if (!el.parentElement) break;
                    el = el.parentElement;
                    var link = el.querySelector('a[href*="detail"]');
                    if (link) return link.href;
                }
            }
        }
        return null;
    """, MAIL_SUBJECT_PATTERN)

    if not detail_url:
        # Fallback: click and wait for navigation
        print("  ⚠ No detail URL found, trying click...")
        clicked = driver.execute_script("""
            var subject = arguments[0];
            var all = document.querySelectorAll('td, div, span, b, strong');
            for (var i = 0; i < all.length; i++) {
                var t = (all[i].textContent || '').trim();
                if (t.indexOf(subject) >= 0 && t.length < 300) {
                    all[i].click();
                    return true;
                }
            }
            return false;
        """, MAIL_SUBJECT_PATTERN)
        if clicked:
            time.sleep(5)
            if "detail" in driver.current_url:
                print(f"  ✅ Opened mail: {driver.current_url}")
                return True
        print("  ❌ NSO mail not found")
        return False

    # Navigate directly to the detail page
    print(f"  📧 Found mail URL: {detail_url}")
    driver.get(detail_url)
    time.sleep(5)

    if "detail" in driver.current_url:
        print(f"  ✅ Opened mail detail")
        return True

    print("  ⚠ URL didn't change to detail")
    return False


def parse_nso_table(driver):
    """Parse NSO store list from the email body.

    Actual email format (plain text inside CKEditor):
        160. The Park Residence - Nguyễn Hữu Thọ
            Ngày nhận mặt bằng: 15/03/2026
            Ngày khai trương: 17/04/2026 - Thứ 6
    """
    print("  📋 Parsing NSO email...")

    # Wait for CKEditor to render
    time.sleep(3)

    # Extract innerText from CKEditor content
    ck_text = driver.execute_script(
        'var ck = document.querySelector(".ck-content");'
        'return ck ? ck.innerText : null;'
    )

    if not ck_text:
        print("  ❌ No CKEditor content found")
        return []

    print(f"  📄 Email text: {len(ck_text)} chars")

    # Parse store entries using regex
    stores = []
    entry_pattern = re.compile(r'(\d{1,3})\.\s+(.+?)(?=\n)')
    date_pattern = re.compile(r'Ngày khai trương:\s*(\d{2}/\d{2}/\d{4})')

    # Split into blocks per store entry
    entries = re.split(r'\n(?=\d{1,3}\.\s)', ck_text)

    for entry in entries:
        entry = entry.strip()
        if not entry:
            continue

        header_match = entry_pattern.match(entry)
        if not header_match:
            continue

        stt = header_match.group(1)
        name_mail = header_match.group(2).strip()

        # Clean name
        name_mail = re.sub(r'https?://\S+', '', name_mail).strip()
        name_mail = name_mail.rstrip(' -')

        date_match = date_pattern.search(entry)
        if not date_match:
            continue

        opening_date = date_match.group(1)

        stores.append({
            "stt": int(stt),
            "name_mail": name_mail,
            "opening_date": opening_date,
        })

    stores.sort(key=lambda s: s["stt"])

    print(f"  ✅ Parsed {len(stores)} stores from mail")
    for s in stores[:5]:
        print(f"     #{s['stt']} | {s['name_mail'][:50]} | {s['opening_date']}")
    if len(stores) > 5:
        print(f"     ... +{len(stores) - 5} more")

    return stores


# ══════════════════════════════════════════════════════════════
#  DSST LOOKUP (download via Selenium since sheet needs auth)
# ══════════════════════════════════════════════════════════════

def download_dsst_via_browser(driver):
    """Download DSST sheet as xlsx through authenticated browser session.
    
    Returns path to downloaded file or None.
    """
    import tempfile
    import openpyxl

    print("  📥 Downloading DSST sheet...")

    # Set download path
    download_dir = os.path.join(REPO_ROOT, "output", "state", "nso")
    os.makedirs(download_dir, exist_ok=True)
    dsst_path = os.path.join(download_dir, "dsst_cache.xlsx")

    # Navigate to export URL — browser handles auth via Google session
    export_url = f"{DSST_SHEET_URL}&gid={DSST_GID}"

    try:
        # Navigate to Google domain first to ensure cookies are available
        driver.get("https://docs.google.com/spreadsheets/d/1byEE8KterdcRr10IydIjbPcJcQwhX2HtGBzd0VZ5N1k/edit")
        time.sleep(3)

        # Use async fetch API with Promise to download xlsx
        result = driver.execute_async_script("""
            var url = arguments[0];
            var callback = arguments[arguments.length - 1];
            fetch(url, {credentials: 'include'})
                .then(function(r) { return r.arrayBuffer(); })
                .then(function(buf) {
                    var arr = new Uint8Array(buf);
                    var chunks = [];
                    var chunkSize = 8192;
                    for (var i = 0; i < arr.length; i += chunkSize) {
                        var slice = arr.subarray(i, Math.min(i + chunkSize, arr.length));
                        chunks.push(String.fromCharCode.apply(null, slice));
                    }
                    callback(btoa(chunks.join('')));
                })
                .catch(function(e) { callback(null); });
        """, export_url)

        if result:
            import base64
            data = base64.b64decode(result)
            with open(dsst_path, 'wb') as f:
                f.write(data)
            print(f"  ✅ DSST downloaded ({len(data):,} bytes)")
            return dsst_path
        else:
            print("  ⚠ Fetch download failed, trying direct navigation...")
    except Exception as e:
        print(f"  ⚠ XHR approach failed: {e}")

    # Fallback: direct navigation (will trigger download)
    try:
        driver.get(export_url)
        time.sleep(8)

        # Check common download locations
        for dl_dir in [
            os.path.join(os.path.expanduser("~"), "Downloads"),
            os.path.join(os.path.expanduser("~"), "Desktop"),
        ]:
            if os.path.exists(dl_dir):
                files = [f for f in os.listdir(dl_dir) if f.endswith('.xlsx') and 'MDT' in f.upper()]
                if files:
                    src = os.path.join(dl_dir, sorted(files, key=lambda x: os.path.getmtime(os.path.join(dl_dir, x)))[-1])
                    import shutil
                    shutil.copy2(src, dsst_path)
                    print(f"  ✅ DSST found in downloads: {src}")
                    return dsst_path
    except Exception as e:
        print(f"  ⚠ Direct download failed: {e}")

    # Use cached version if exists
    if os.path.exists(dsst_path):
        age_hours = (time.time() - os.path.getmtime(dsst_path)) / 3600
        print(f"  ⚠ Using cached DSST ({age_hours:.0f}h old)")
        return dsst_path

    print("  ❌ Could not download DSST")
    return None


def load_dsst_lookup(xlsx_path):
    """Load DSST sheet and build lookup dict by code.

    Returns dict: code → {name_system, name_full, version}
    """
    import openpyxl

    if not xlsx_path or not os.path.exists(xlsx_path):
        return {}

    print("  📖 Loading DSST lookup...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Find the right sheet by gid or name
    ws = None
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Check if this is the right sheet by looking at headers
        header_row = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if any('version' in h.lower() or 'brand' in h.lower() or 'code' in h.lower() for h in header_row):
            break
    else:
        ws = wb.worksheets[0]

    # Read all rows
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 8:
            continue

        # Col B (idx 1) = name_system, Col C (idx 2) = code, Col H (idx 7) = version
        code = str(row[2] or '').strip()
        if not code or not re.match(r'^A\d+$', code):
            continue

        name_system = str(row[1] or '').strip() or None
        name_full = str(row[3] or '').strip() if len(row) > 3 else None
        version_raw = row[7] if len(row) > 7 else None

        version = None
        if version_raw is not None:
            try:
                version = int(float(str(version_raw)))
                if version not in (700, 1000, 1500, 2000):
                    version = None
            except (ValueError, TypeError):
                pass

        lookup[code] = {
            "name_system": name_system,
            "name_full": name_full,
            "version": version,
        }

    wb.close()
    print(f"  ✅ DSST lookup: {len(lookup)} stores")
    return lookup


# ══════════════════════════════════════════════════════════════
#  MERGE LOGIC
# ══════════════════════════════════════════════════════════════

def _normalize(s):
    """Normalize store name for fuzzy matching."""
    s = s.lower().strip()
    # Remove common noise
    for noise in ['- mới bổ sung', '- tbi', '- bth', '- q1', 'shophoue', 'shophouse']:
        s = s.replace(noise, '')
    return ' '.join(s.split())


def _name_match(name_mail, store):
    """Check if a mail store name matches an existing store entry."""
    nm = _normalize(name_mail)
    # Check against name_mail, name_full, name_system
    for key in ['name_mail', 'name_full']:
        sv = store.get(key)
        if sv:
            sv_n = _normalize(sv)
            # Substring match (either direction)
            if nm in sv_n or sv_n in nm:
                return True
            # First significant word match
            nm_words = [w for w in nm.split() if len(w) > 2]
            sv_words = [w for w in sv_n.split() if len(w) > 2]
            common = set(nm_words) & set(sv_words)
            if len(common) >= 2:
                return True
    return False


def merge_stores(current_stores, mail_stores, dsst_lookup):
    """Merge mail data into current stores list.

    Mail stores have: stt, name_mail, opening_date (no code!)
    Current stores have: code, name_system, name_full, name_mail, opening_date, version

    Match by name (fuzzy). New stores added without code (will be filled by DSST later).
    """
    added = []
    updated = []

    for ms in mail_stores:
        # Find matching existing store by name
        matched = None
        for cs in current_stores:
            if _name_match(ms["name_mail"], cs):
                matched = cs
                break

        if matched:
            # Check if opening_date changed
            if matched["opening_date"] != ms["opening_date"]:
                old_date = matched.get("original_date") or matched["opening_date"]
                matched["original_date"] = old_date
                matched["opening_date"] = ms["opening_date"]
                updated.append(matched.get("code") or ms["name_mail"][:20])
        else:
            # New store — no code yet
            new_store = {
                "code": None,
                "name_system": None,
                "name_full": ms["name_mail"],
                "name_mail": ms["name_mail"],
                "opening_date": ms["opening_date"],
                "version": None,
                "original_date": None,
            }
            current_stores.append(new_store)
            added.append(ms["name_mail"][:25])

    # Enrich with DSST data
    for store in current_stores:
        code = store.get("code")
        if code:
            dsst = dsst_lookup.get(code, {})
            if dsst.get("name_system") and not store.get("name_system"):
                store["name_system"] = dsst["name_system"]
            if dsst.get("version") and not store.get("version"):
                store["version"] = dsst["version"]

    return current_stores, added, updated


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="NSO Mail Scanner — Haraworks")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, don't write")
    parser.add_argument("--force", action="store_true", help="Skip day-of-week check")
    args = parser.parse_args()

    now = datetime.now()
    print(f"\n{'='*55}")
    print(f"  🏪 NSO Mail Scanner — {now.strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*55}")

    # Day-of-week check (Mon=0, Tue=1)
    if not args.force and now.weekday() not in (0, 1):
        print(f"\n  ⏭ Today is {now.strftime('%A')} — NSO scan runs Mon/Tue only.")
        print(f"  💡 Use --force to run anyway.")
        return

    driver = None
    try:
        # Setup browser
        print("\n  🌐 Starting browser...")
        driver = setup_driver()

        # Login
        if not ensure_logged_in(driver):
            print("  ❌ Cannot login to Haraworks")
            return

        # Find NSO mail
        if not find_nso_mail(driver):
            print("  ❌ Cannot find NSO mail")
            return

        # Parse table
        mail_stores = parse_nso_table(driver)
        if not mail_stores:
            print("  ❌ No stores parsed from mail")
            return

        # Download DSST
        dsst_path = download_dsst_via_browser(driver)
        dsst_lookup = load_dsst_lookup(dsst_path)

        # Load current stores
        current_stores = load_stores()
        print(f"\n  📊 Current stores: {len(current_stores)}")
        print(f"  📧 Mail stores: {len(mail_stores)}")
        print(f"  📖 DSST lookup: {len(dsst_lookup)}")

        # Merge
        merged, added, updated = merge_stores(current_stores, mail_stores, dsst_lookup)

        print(f"\n  {'─'*40}")
        print(f"  📊 Merge result:")
        print(f"     Total: {len(merged)}")
        print(f"     New:   {len(added)} {added[:5] if added else ''}")
        print(f"     Updated dates: {len(updated)} {updated[:5] if updated else ''}")

        if args.dry_run:
            print(f"\n  🏃 DRY RUN — not writing files")
            return

        if not added and not updated:
            print(f"\n  ✓ No changes needed")
        else:
            # Save
            save_stores(merged)
            print(f"\n  💾 Saved {len(merged)} stores to {JSON_STORES_PATH}")

            # Re-export + deploy
            import subprocess
            print(f"\n  📦 Re-exporting NSO data...")
            subprocess.run(
                [sys.executable, os.path.join(REPO_ROOT, "script", "dashboard", "export_data.py"),
                 "--domain", "nso"],
                cwd=REPO_ROOT, timeout=60
            )

            print(f"  🚀 Deploying...")
            subprocess.run(
                [sys.executable, os.path.join(REPO_ROOT, "script", "dashboard", "deploy.py"),
                 "--domain", "nso"],
                cwd=REPO_ROOT, timeout=120
            )

    except Exception as e:
        print(f"\n  ❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        print(f"\n  🎉 Done!")
        print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
