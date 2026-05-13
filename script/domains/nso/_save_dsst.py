# -*- coding: utf-8 -*-
"""
_save_dsst.py — Fetch DSST (Danh sách siêu thị) from Google Sheet → cache JSON
================================================================================
Fetches live data from the DSST Google Sheet every time it's called.
Saves to data/dsst_cache.json for use by NSO pipeline.

Sheet structure (sheet name: "DSST"):
  Col A (0): STT
  Col B (1): branch_name  (e.g. "KFM_HCM_Q07 - 31 Tân Mỹ")
  Col C (2): branch_id    (= code, e.g. "TMY")
  Col H (7): Version_ST_FMCG (= version, e.g. "2000")

Usage:
    python script/domains/nso/_save_dsst.py
"""
import json, re, sys, os, io
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(REPO_ROOT, "script"))

from lib.sources import DSST_SHEET_URL, DSST_GID

CACHE_PATH = os.path.join(REPO_ROOT, "data", "dsst_cache.json")


def fetch_dsst_from_sheet():
    """Download DSST Google Sheet and parse store data.
    
    Returns dict: code -> {name_system, name_full, branch_name, version}
    """
    import requests
    import openpyxl
    import time

    # Download xlsx with specific gid to get only DSST sheet
    url = f"{DSST_SHEET_URL}&gid={DSST_GID}"
    print(f"  📥 Downloading DSST sheet...")
    
    max_retries = 3
    last_err = None
    content = None
    
    for attempt in range(max_retries):
        try:
            r = requests.get(url, allow_redirects=True, timeout=30)
            r.raise_for_status()
            content = r.content
            if len(content) < 500:
                raise ValueError(f"Response too small ({len(content)} bytes)")
            break
        except Exception as e:
            last_err = e
            if attempt < max_retries - 1:
                wait = 3 * (attempt + 1)
                print(f"  ⚠ Attempt {attempt+1}/{max_retries} failed: {e}")
                print(f"    Retry in {wait}s...")
                time.sleep(wait)
    
    if content is None:
        raise last_err
    
    print(f"  ✅ Downloaded {len(content):,} bytes")

    # Parse xlsx
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    
    # Find DSST sheet
    ws = None
    for name in wb.sheetnames:
        if name.upper() == "DSST":
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
        print(f"  ⚠ 'DSST' sheet not found, using '{ws.title}'")

    # Fixed column mapping:
    # A(0)=STT, B(1)=branch_name, C(2)=branch_id(code), H(7)=Version_ST_FMCG
    COL_BRANCH = 1   # branch_name
    COL_CODE = 2      # branch_id = store code
    COL_VERSION = 7   # Version_ST_FMCG

    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        if not row or not any(row):
            continue
        
        code_val = row[COL_CODE] if COL_CODE < len(row) else None
        branch_val = row[COL_BRANCH] if COL_BRANCH < len(row) else None
        version_val = row[COL_VERSION] if COL_VERSION < len(row) else None

        if not code_val or not branch_val:
            continue

        code = str(code_val).strip()
        branch_name = str(branch_val).strip()
        
        if not code or code.lower() in ("branch_id", "code", "mã st"):
            continue

        # Parse "KFM_HCM_Q07 - 31 Tân Mỹ" → name_system + name_full
        parts = branch_name.split(" - ", 1)
        name_system = parts[0].strip()
        name_full = parts[1].strip() if len(parts) > 1 else branch_name

        # Version
        try:
            version = int(version_val) if version_val else None
        except (ValueError, TypeError):
            version = None

        lookup[code] = {
            "name_system": name_system,
            "name_full": name_full,
            "branch_name": branch_name,
            "version": version,
        }

    wb.close()
    return lookup


def save_dsst_cache(lookup):
    """Save DSST lookup to cache JSON."""
    os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
    
    cache_data = {
        "_meta": {
            "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "count": len(lookup),
            "source": "Google Sheet (DSST)",
        },
        "stores": lookup,
    }
    
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=2)
    
    a_codes = sorted([k for k in lookup if re.match(r'^A\d+$', k)])
    print(f"  💾 Saved {len(lookup)} stores ({len(a_codes)} A-codes) to dsst_cache.json")
    return CACHE_PATH


def refresh_dsst():
    """Fetch DSST from Google Sheet and save to cache. Returns lookup dict."""
    lookup = fetch_dsst_from_sheet()
    save_dsst_cache(lookup)
    return lookup


if __name__ == "__main__":
    print(f"\n  🏪 DSST Cache Refresh — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"  {'─'*40}")
    lookup = refresh_dsst()
    
    a_codes = sorted([k for k in lookup if re.match(r'^A\d+$', k)])
    print(f"\n  Sample A-codes ({len(a_codes)} total):")
    for c in a_codes[-5:]:
        info = lookup[c]
        print(f"    {c} → {info['name_full']} (v{info.get('version', '?')})")
