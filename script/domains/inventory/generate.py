# -*- coding: utf-8 -*-
"""
Generate Report Đối Soát Tồn Kho KFM vs ABA
=============================================
Reads reconciliation data, master data, and weight data.
Produces an HTML report with:
  1. Statistics table by category (Đông/Mát/TCNK)
  2. Trend charts for accuracy rate (Item/SKU/KG)
  3. Pie chart of KG proportion by category
"""

import openpyxl
import os
import sys
import io
import glob
import json
import re
import math
import requests
import base64
from datetime import datetime, timedelta
from collections import defaultdict

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np


# ─────────────────────────────────────────────
# Static Chart Image Generation (fallback for iOS file viewers)
# ─────────────────────────────────────────────
def generate_combo_chart_image(dates, kfm_data, aba_data, acc_data, y_label):
    """Generate a combo bar+line chart as base64 PNG for fallback display"""
    fig, ax1 = plt.subplots(figsize=(10, 4.2))
    fig.patch.set_facecolor('#1e293b')
    ax1.set_facecolor('#1e293b')

    x = np.arange(len(dates))
    width = 0.32

    # Bar charts
    ax1.bar(x - width/2, kfm_data, width, label='KFM',
            color='#60a5fa99', edgecolor='#60a5fa', linewidth=0.8)
    ax1.bar(x + width/2, aba_data, width, label='ABA',
            color='#34d39999', edgecolor='#34d399', linewidth=0.8)

    ax1.set_ylabel(y_label, color='#64748b', fontsize=10, fontweight='600')
    ax1.set_xticks(x)
    ax1.set_xticklabels(dates, color='#94a3b8', fontsize=9)
    ax1.tick_params(axis='y', colors='#94a3b8', labelsize=9)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:,.0f}'))
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax1.spines['left'].set_color('#ffffff10')
    ax1.spines['bottom'].set_color('#ffffff10')
    ax1.grid(axis='y', color='#ffffff0a', linewidth=0.5)

    # Line chart on secondary axis
    ax2 = ax1.twinx()
    ax2.plot(x, acc_data, color='#f472b6', linewidth=2.5, marker='o',
             markersize=7, markerfacecolor='#f472b6', markeredgecolor='white',
             markeredgewidth=1.5, zorder=5)

    # Data labels
    for i, v in enumerate(acc_data):
        ax2.annotate(f'{v:.2f}%', (i, v), textcoords="offset points",
                     xytext=(0, 14), ha='center', color='#f472b6',
                     fontsize=9, fontweight='bold')

    ax2.set_ylabel('Tỷ lệ chính xác (%)', color='#f472b6', fontsize=10, fontweight='600')
    ax2.set_ylim(0, 100)
    ax2.tick_params(axis='y', colors='#f472b6', labelsize=9)
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:.0f}%'))
    ax2.spines['top'].set_visible(False)
    ax2.spines['left'].set_visible(False)
    ax2.spines['right'].set_color('#f472b640')
    ax2.spines['bottom'].set_visible(False)

    # Legend
    from matplotlib.lines import Line2D
    legend_elements = [
        plt.Rectangle((0, 0), 1, 1, fc='#60a5fa99', ec='#60a5fa', label='KFM'),
        plt.Rectangle((0, 0), 1, 1, fc='#34d39999', ec='#34d399', label='ABA'),
        Line2D([0], [0], color='#f472b6', marker='o', markersize=6,
               markerfacecolor='#f472b6', markeredgecolor='white', label='Tỷ lệ chính xác (%)'),
    ]
    leg = ax1.legend(handles=legend_elements, loc='upper center',
                     bbox_to_anchor=(0.5, -0.12), ncol=3,
                     facecolor='#1e293b', edgecolor='none', labelcolor='#94a3b8',
                     fontsize=9)

    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight',
                facecolor='#1e293b', edgecolor='none', pad_inches=0.3)
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


def generate_pie_chart_image(labels, values, colors, unit):
    """Generate a doughnut chart as base64 PNG for fallback display"""
    fig, ax = plt.subplots(figsize=(4, 4))
    fig.patch.set_facecolor('#1e293b')
    ax.set_facecolor('#1e293b')

    total = sum(values)
    # Outer ring
    wedges, texts, autotexts = ax.pie(
        values, labels=None, colors=[c + 'cc' for c in colors],
        autopct=lambda pct: f'{pct:.1f}%',
        pctdistance=0.78, startangle=90,
        wedgeprops=dict(width=0.4, edgecolor='#1e293b', linewidth=2),
    )

    for t in autotexts:
        t.set_color('white')
        t.set_fontsize(11)
        t.set_fontweight('bold')

    # Legend below
    legend_texts = []
    for i, label in enumerate(labels):
        pct = (values[i] / total) * 100
        if isinstance(values[i], float):
            val_str = f'{values[i]:,.2f}'
        else:
            val_str = f'{values[i]:,}'
        legend_texts.append(f'{label}: {val_str} {unit} ({pct:.1f}%)')

    leg = ax.legend(legend_texts, loc='upper center', bbox_to_anchor=(0.5, -0.02),
                    ncol=1, facecolor='#1e293b', edgecolor='#ffffff14',
                    labelcolor='#94a3b8', fontsize=9,
                    handler_map={str: matplotlib.legend_handler.HandlerPatch()})
    for i, handle in enumerate(leg.legend_handles):
        handle.set_facecolor(colors[i])
        handle.set_edgecolor(colors[i])

    ax.set_aspect('equal')
    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight',
                facecolor='#1e293b', edgecolor='none', pad_inches=0.2)
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Navigate up to repo root: script/domains/inventory/ → transport_daily_report/
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(REPO_ROOT, "script"))
# Data paths (from shared lib)
from lib.sources import DOI_SOAT_DIR, MASTER_DATA_FILE, WEIGHT_DATA_FILE
OUTPUT_DIR = os.path.join(REPO_ROOT, 'output', 'artifacts', 'inventory')
TELEGRAM_CONFIG_FILE = os.path.join(REPO_ROOT, 'config', 'telegram.json')
SENT_MESSAGES_FILE = os.path.join(REPO_ROOT, 'output', 'state', 'inventory', 'sent_messages.json')

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.dirname(SENT_MESSAGES_FILE), exist_ok=True)


def acc_precise(discrepancy, total):
    """accuracy = max(0, 1 - discrepancy/total) * 100, truncated to 2 decimals"""
    if total == 0:
        return 100.0 if discrepancy == 0 else 0.0
    raw = max(0.0, (1.0 - discrepancy / total)) * 100.0
    return math.floor(raw * 100) / 100


# ─────────────────────────────────────────────
# 1. Load Master Data → barcode → category
# ─────────────────────────────────────────────
def load_master_data():
    """
    Returns dict: barcode (str) → category ('ĐÔNG' | 'MÁT' | 'TCNK')
    Rules:
      - Chi nhánh = 'KHO LƯU HÀNG' → use Phân Loại (col E): ĐÔNG or MÁT
      - Chi nhánh = 'KHO QUÁ CẢNH' → use Phân Loại (col E): ĐÔNG or MÁT
      - Chi nhánh contains 'TCNK' → TCNK
    """
    # Manual overrides for products not found in master data
    CATEGORY_OVERRIDES = {
        '8938527710565': 'MÁT',  # SEAPRODEX - TÔM SÚ THỊT SINH THÁI 71/90 200G
        '1102969': 'TCNK',       # TÁO GREEN DRAGON MỸ
    }

    wb = openpyxl.load_workbook(MASTER_DATA_FILE, data_only=True, read_only=True)
    ws = wb['Sheet1']
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        chi_nhanh = str(row[0]).strip() if row[0] else ''
        barcode = str(row[1]).strip() if row[1] else ''
        phan_loai = str(row[4]).strip().upper() if row[4] else ''

        if not barcode:
            continue

        # Determine category
        if 'TCNK' in chi_nhanh:
            category = 'TCNK'
        elif chi_nhanh in ('KHO LƯU HÀNG', 'KHO QUÁ CẢNH'):
            if phan_loai == 'ĐÔNG' or phan_loai == 'DONG':
                category = 'ĐÔNG'
            elif phan_loai == 'MÁT' or phan_loai == 'MAT':
                category = 'MÁT'
            else:
                category = None
        else:
            category = None

        if category:
            # Store original barcode
            if barcode not in mapping:
                mapping[barcode] = category
            # Also store zero-padded variants (12, 13 digits) for matching
            # doi_soat files may have leading zeros that master data lacks
            for pad_len in (12, 13, 14):
                padded = barcode.zfill(pad_len)
                if padded != barcode and padded not in mapping:
                    mapping[padded] = category
            # Also store stripped version (no leading zeros)
            stripped = barcode.lstrip('0')
            if stripped and stripped != barcode and stripped not in mapping:
                mapping[stripped] = category

    wb.close()

    # Apply manual overrides
    mapping.update(CATEGORY_OVERRIDES)

    return mapping


# ─────────────────────────────────────────────
# 2. Load Weight Data → barcode → kg_per_item
# ─────────────────────────────────────────────
def load_weight_data():
    """
    Returns dict: barcode (str) → weight in KG per item (float)
    Uses col 25 (Giá trị trọng lượng/thể tích) and col 26 (Đơn vị)
    """
    # Manual overrides for products not found in the weight source
    WEIGHT_OVERRIDES = {
        '4014500028835': 0.2,  # ZOTT - PHÔ MAI LÁT SANDWICH 200G → 200g = 0.2 KG
        '5701215042391': 0.125,  # EMBORG - PHÔ MAI CAMEMBERT 125G → 125g = 0.125 KG
    }

    wb = openpyxl.load_workbook(WEIGHT_DATA_FILE, data_only=True, read_only=True)
    ws = wb['THONG_TIN_CO_BAN']
    mapping = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        barcode_base = str(row[0]).strip() if row[0] else ''
        barcode = str(row[1]).strip() if row[1] else ''
        weight_val = row[25] if len(row) > 25 else None
        weight_unit = str(row[26]).strip().lower() if len(row) > 26 and row[26] else ''

        if weight_val is None or weight_val == '':
            kg = None
        else:
            try:
                wv = float(weight_val)
            except (ValueError, TypeError):
                kg = None
                continue

            if weight_unit in ('g', 'gr', 'gram'):
                kg = wv / 1000.0
            elif weight_unit in ('kg',):
                kg = wv
            elif weight_unit in ('ml',):
                kg = wv / 1000.0
            elif weight_unit in ('l', 'lít', 'lit'):
                kg = wv
            else:
                kg = wv / 1000.0  # default assume grams

        if kg is not None:
            if barcode and barcode not in mapping:
                mapping[barcode] = kg
            if barcode_base and barcode_base not in mapping:
                mapping[barcode_base] = kg

    wb.close()

    # Apply overrides
    mapping.update(WEIGHT_OVERRIDES)

    return mapping


# ─────────────────────────────────────────────
# 3. Load Reconciliation Data
# ─────────────────────────────────────────────
def load_doi_soat_files():
    """
    Returns list of dicts, each representing one day's data.
    Each dict: {
        'date': datetime,
        'date_str': 'dd/mm/yyyy',
        'items': [
            {'barcode': str, 'ten_hang': str, 'dvt': str,
             'ton_KFM': int, 'ton_ABA': int,
             'chenh_lech_thua': int, 'chenh_lech_thieu': int,
             'ghi_chu': str}
        ]
    }
    Columns:
      0: Ten_viet_tat, 1: barcode, 2: ten_hang, 3: dvt,
      4: ton_KDB (KFM), 5: ton_ABA, 6: NSO, 7: chenh_lech,
      8: chenh_lech_thua, 9: chenh_lech_thieu, 10: ghi_chu
    """
    files = glob.glob(os.path.join(DOI_SOAT_DIR, 'Đối soát tồn *.xlsx'))
    if not files:
        print("ERROR: No reconciliation files found!")
        return []

    all_days = []
    for fpath in sorted(files):
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=False)
        if 'Đối soát tồn' not in wb.sheetnames:
            print(f"WARNING: Sheet 'Đối soát tồn' not found in {os.path.basename(fpath)}")
            wb.close()
            continue

        ws = wb['Đối soát tồn']
        items = []
        date_val = None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            barcode = str(row[1]).strip() if row[1] else ''
            if not barcode:
                continue

            ten_hang = str(row[2]) if row[2] else ''
            dvt = str(row[3]) if row[3] else ''
            ton_kfm = int(row[4]) if row[4] is not None else 0
            ton_aba = int(row[5]) if row[5] is not None else 0
            chenh_lech_thua = int(row[8]) if row[8] is not None else 0
            chenh_lech_thieu = int(row[9]) if row[9] is not None else 0
            ghi_chu = str(row[10]).strip() if row[10] is not None else ''
            if ghi_chu == 'None':
                ghi_chu = ''
            ngay = row[11]  # ngay_ton_kho

            if ngay and date_val is None:
                if isinstance(ngay, datetime):
                    date_val = ngay
                elif isinstance(ngay, str):
                    # Try mm/dd/yyyy format
                    try:
                        date_val = datetime.strptime(ngay, '%m/%d/%Y')
                    except ValueError:
                        try:
                            date_val = datetime.strptime(ngay, '%d/%m/%Y')
                        except ValueError:
                            pass

            items.append({
                'barcode': barcode,
                'ten_hang': ten_hang,
                'dvt': dvt,
                'ton_KFM': ton_kfm,
                'ton_ABA': ton_aba,
                'chenh_lech_thua': chenh_lech_thua,
                'chenh_lech_thieu': chenh_lech_thieu,
                'ghi_chu': ghi_chu,
            })

        if date_val and items:
            all_days.append({
                'date': date_val,
                'date_str': date_val.strftime('%d/%m/%Y'),
                'items': items,
            })

        wb.close()

    # Sort by date
    all_days.sort(key=lambda x: x['date'])
    return all_days


# ─────────────────────────────────────────────
# 4. Compute Statistics
# ─────────────────────────────────────────────
def compute_stats(all_days, category_map, weight_map):
    """
    For each day, compute per-category stats:
      - sku_count, item_KFM, item_ABA, kg_KFM, kg_ABA
      - sku_lech, item_lech, kg_lech (absolute discrepancy)

    ═══════════════════════════════════════════════
    CÁCH TÍNH LỆCH & TỶ LỆ CHÍNH XÁC:
    ═══════════════════════════════════════════════
    1. Lệch (discrepancy) = abs(ton_KFM - ton_ABA) cho từng SKU
       → Luôn dùng GIÁ TRỊ TUYỆT ĐỐI, không để thừa/thiếu triệt tiêu nhau.
       → item_lech = tổng abs(KFM_i - ABA_i) cho tất cả SKU
       → kg_lech   = tổng abs(kg_KFM_i - kg_ABA_i)
       → sku_lech  = số SKU có abs(KFM - ABA) > 0

    2. Tỷ lệ chính xác (accuracy):
       → Item accuracy = (1 - item_lech / item_KFM_total) × 100
       → SKU accuracy  = (1 - sku_lech / sku_total) × 100
       → KG accuracy   = (1 - kg_lech / kg_KFM_total) × 100

    3. Làm tròn: TRUNCATE xuống 2 decimal (math.floor), KHÔNG round lên.
       Ví dụ: 99.9954% → 99.99% (không phải 100.00%)

    4. Trend chart: cột ABA = KFM - lệch_abs (để visual gap = lệch tuyệt đối)
    ═══════════════════════════════════════════════

    Returns:
      daily_stats, unmatched_category_details, unmatched_weight_details
    """
    unmatched_category = set()
    unmatched_weight = set()
    unmatched_category_details = {}  # barcode -> ten_hang
    unmatched_weight_details = {}

    daily_stats = []

    for day in all_days:
        cat_stats = {
            'ĐÔNG': {'sku': 0, 'item_KFM': 0, 'item_ABA': 0, 'kg_KFM': 0.0, 'kg_ABA': 0.0,
                     'sku_lech': 0, 'item_lech': 0, 'kg_lech': 0.0},
            'MÁT': {'sku': 0, 'item_KFM': 0, 'item_ABA': 0, 'kg_KFM': 0.0, 'kg_ABA': 0.0,
                    'sku_lech': 0, 'item_lech': 0, 'kg_lech': 0.0},
            'TCNK': {'sku': 0, 'item_KFM': 0, 'item_ABA': 0, 'kg_KFM': 0.0, 'kg_ABA': 0.0,
                     'sku_lech': 0, 'item_lech': 0, 'kg_lech': 0.0},
        }
        # Track lệch dương (ABA > KFM = thừa) and lệch âm (ABA < KFM = thiếu) — count by ITEMS
        day_lech_duong = 0  # total items where ABA > KFM
        day_lech_am = 0     # total items where ABA < KFM

        for item in day['items']:
            bc = item['barcode']
            cat = category_map.get(bc)
            if not cat:
                unmatched_category.add(bc)
                unmatched_category_details[bc] = item['ten_hang']
                continue

            kg_per = weight_map.get(bc)
            if kg_per is None:
                unmatched_weight.add(bc)
                unmatched_weight_details[bc] = item['ten_hang']
                kg_per = 0

            cs = cat_stats[cat]
            cs['sku'] += 1
            cs['item_KFM'] += item['ton_KFM']
            cs['item_ABA'] += item['ton_ABA']
            cs['kg_KFM'] += item['ton_KFM'] * kg_per
            cs['kg_ABA'] += item['ton_ABA'] * kg_per

            # Determine discrepancy using absolute difference
            item_diff = abs(item['ton_KFM'] - item['ton_ABA'])
            kg_diff = abs(item['ton_KFM'] * kg_per - item['ton_ABA'] * kg_per)
            has_discrepancy = (item_diff > 0)

            if has_discrepancy:
                cs['sku_lech'] += 1
                cs['item_lech'] += item_diff
                cs['kg_lech'] += kg_diff

            # Count lệch dương / lệch âm by item quantity
            if item['ton_ABA'] > item['ton_KFM']:
                day_lech_duong += (item['ton_ABA'] - item['ton_KFM'])
            elif item['ton_ABA'] < item['ton_KFM']:
                day_lech_am += (item['ton_KFM'] - item['ton_ABA'])

        # Compute per-day totals across all categories
        cats_list = ['ĐÔNG', 'MÁT', 'TCNK']
        t_sku = sum(cat_stats[c]['sku'] for c in cats_list)
        t_item_kfm = sum(cat_stats[c]['item_KFM'] for c in cats_list)
        t_item_aba = sum(cat_stats[c]['item_ABA'] for c in cats_list)
        t_item_lech = sum(cat_stats[c]['item_lech'] for c in cats_list)
        t_sku_lech = sum(cat_stats[c]['sku_lech'] for c in cats_list)
        t_kg_kfm = sum(cat_stats[c]['kg_KFM'] for c in cats_list)
        t_kg_aba = sum(cat_stats[c]['kg_ABA'] for c in cats_list)
        t_kg_lech = sum(cat_stats[c]['kg_lech'] for c in cats_list)

        daily_stats.append({
            'date_str': day['date_str'],
            'date': day['date'],
            'categories': cat_stats,
            'totals': {
                'sku': t_sku, 'item_KFM': t_item_kfm, 'item_ABA': t_item_aba,
                'kg_KFM': t_kg_kfm, 'kg_ABA': t_kg_aba,
                'sku_lech': t_sku_lech, 'item_lech': t_item_lech, 'kg_lech': t_kg_lech,
                'lech_duong': day_lech_duong, 'lech_am': day_lech_am,
                'tong_lech': day_lech_duong + day_lech_am,
            }
        })

    return daily_stats, unmatched_category_details, unmatched_weight_details


# ─────────────────────────────────────────────
# 5. Generate HTML Report
# ─────────────────────────────────────────────
def generate_html(daily_stats, target_date_str=None, all_days=None):
    """Generate a self-contained HTML report"""

    # If target_date_str specified, find that day for the summary table and pie
    if target_date_str:
        target_day = None
        for ds in daily_stats:
            if ds['date_str'] == target_date_str:
                target_day = ds
                break
        if not target_day:
            target_day = daily_stats[-1]
    else:
        target_day = daily_stats[-1]

    target_label = target_day['date_str']

    # ---- Build table data for Section 1 ----
    cats = ['ĐÔNG', 'MÁT', 'TCNK']
    cat_labels = {'ĐÔNG': 'Hàng Đông', 'MÁT': 'Hàng Mát', 'TCNK': 'TCNK'}
    cat_colors = {'ĐÔNG': '#60a5fa', 'MÁT': '#34d399', 'TCNK': '#fbbf24'}

    table_rows = []
    total_kfm_sku = 0
    total_aba_sku = 0
    total_kfm_item = 0
    total_aba_item = 0
    total_kfm_kg = 0.0
    total_aba_kg = 0.0

    for cat in cats:
        cs = target_day['categories'][cat]
        table_rows.append({
            'cat': cat_labels[cat],
            'color': cat_colors[cat],
            'sku': cs['sku'],
            'item_KFM': cs['item_KFM'],
            'item_ABA': cs['item_ABA'],
            'kg_KFM': cs['kg_KFM'],
            'kg_ABA': cs['kg_ABA'],
        })
        total_kfm_sku += cs['sku']
        total_aba_sku += cs['sku']  # same SKU count
        total_kfm_item += cs['item_KFM']
        total_aba_item += cs['item_ABA']
        total_kfm_kg += cs['kg_KFM']
        total_aba_kg += cs['kg_ABA']

    # ---- Build trend data for Section 2 ----
    trend_dates = []
    trend_item_kfm = []
    trend_item_aba = []
    trend_item_acc = []
    trend_sku_kfm = []
    trend_sku_aba = []
    trend_sku_acc = []
    trend_kg_kfm = []
    trend_kg_aba = []
    trend_kg_acc = []

    for ds in daily_stats:
        trend_dates.append(ds['date_str'])

        # Totals across categories
        t_item_kfm = sum(ds['categories'][c]['item_KFM'] for c in cats)
        t_item_aba = sum(ds['categories'][c]['item_ABA'] for c in cats)
        t_sku_total = sum(ds['categories'][c]['sku'] for c in cats)
        t_sku_lech = sum(ds['categories'][c]['sku_lech'] for c in cats)
        t_item_lech = sum(ds['categories'][c]['item_lech'] for c in cats)
        t_kg_kfm = sum(ds['categories'][c]['kg_KFM'] for c in cats)
        t_kg_aba = sum(ds['categories'][c]['kg_ABA'] for c in cats)
        t_kg_lech = sum(ds['categories'][c]['kg_lech'] for c in cats)

        # Accuracy (uses module-level acc_precise)
        item_acc = acc_precise(t_item_lech, t_item_kfm)
        sku_acc = acc_precise(t_sku_lech, t_sku_total)
        kg_acc = acc_precise(t_kg_lech, t_kg_kfm)

        trend_item_kfm.append(t_item_kfm)
        trend_item_aba.append(t_item_kfm - t_item_lech)
        trend_item_acc.append(item_acc)

        trend_sku_kfm.append(t_sku_total)
        trend_sku_aba.append(t_sku_total - t_sku_lech)
        trend_sku_acc.append(sku_acc)

        trend_kg_kfm.append(round(t_kg_kfm, 2))
        trend_kg_aba.append(round(t_kg_kfm - t_kg_lech, 2))
        trend_kg_acc.append(kg_acc)

    # ---- Build pie data for KG (Section 3) ----
    pie_labels = [cat_labels[c] for c in cats]
    pie_kg_values = [round(target_day['categories'][c]['kg_KFM'], 2) for c in cats]
    pie_colors = [cat_colors[c] for c in cats]

    # ---- Build pie data for Items (new Section) ----
    pie_item_values = [target_day['categories'][c]['item_KFM'] for c in cats]

    # ---- Build weekly summary data ----
    target_date = target_day['date']
    iso_year, iso_week, iso_weekday = target_date.isocalendar()
    # Monday of this week
    week_monday = target_date - timedelta(days=iso_weekday - 1)
    week_days = [week_monday + timedelta(days=i) for i in range(7)]
    thu_labels = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']

    # Map daily_stats by date for quick lookup
    stats_by_date = {ds['date'].strftime('%Y-%m-%d'): ds for ds in daily_stats}

    weekly_rows = []
    weekly_acc_values = []
    for i, wd in enumerate(week_days):
        key = wd.strftime('%Y-%m-%d')
        ds = stats_by_date.get(key)
        if ds:
            day_acc = acc_precise(ds['totals']['item_lech'], ds['totals']['item_KFM'])
            weekly_acc_values.append(day_acc)
            weekly_rows.append({
                'thu': thu_labels[i],
                'ngay': wd.strftime('%d/%m/%Y'),
                'acc': day_acc,
                'lech_duong': ds['totals']['lech_duong'],
                'lech_am': ds['totals']['lech_am'],
                'tong_lech': ds['totals']['tong_lech'],
                'ton_kdb': ds['totals']['item_KFM'],
                'has_data': True,
            })
        else:
            weekly_rows.append({
                'thu': thu_labels[i],
                'ngay': wd.strftime('%d/%m/%Y'),
                'acc': None, 'lech_duong': 0, 'lech_am': 0,
                'tong_lech': 0, 'ton_kdb': 0, 'has_data': False,
            })
    avg_acc = sum(weekly_acc_values) / len(weekly_acc_values) if weekly_acc_values else 0
    avg_acc = math.floor(avg_acc * 100) / 100

    # ---- Compute comparison data (vs hôm qua & vs LFL) ----
    def _fmt_change_acc(current, previous):
        """Format chênh lệch tỷ lệ chính xác (hiện chênh lệch trực tiếp dạng +0.01%)"""
        if previous is None:
            return None, None, None  # no data
        diff = round(current - previous, 2)
        return diff, current, previous

    def _fmt_change_qty(current, previous):
        """Format chênh lệch sản lượng (hiện % thay đổi)"""
        if previous is None or previous == 0:
            return None, None, None  # no data
        pct = round((current - previous) / previous * 100, 2)
        return pct, current, previous

    def _build_insight_html(label, diff, current, previous, is_acc=True):
        """Build HTML cho 1 insight badge"""
        if diff is None:
            return f'''<div class="insight-badge">
                <div class="insight-label">{label}</div>
                <div class="insight-value" style="color: var(--text-muted);">N/A</div>
            </div>'''
        if is_acc:
            # Tỷ lệ chính xác: hiện chênh lệch trực tiếp
            if diff > 0:
                arrow = '▲'
                color = 'var(--accent-green)'
                sign = '+'
            elif diff < 0:
                arrow = '▼'
                color = 'var(--accent-red)'
                sign = ''
            else:
                arrow = '—'
                color = 'var(--text-muted)'
                sign = ''
            return f'''<div class="insight-badge">
                <div class="insight-label">{label}</div>
                <div class="insight-value" style="color: {color};">{arrow} {sign}{diff:.2f}%</div>
                <div class="insight-detail">{previous:.2f}% → {current:.2f}%</div>
            </div>'''
        else:
            # Sản lượng: hiện % thay đổi
            if diff > 0:
                arrow = '▲'
                color = 'var(--accent-green)'
                sign = '+'
            elif diff < 0:
                arrow = '▼'
                color = 'var(--accent-red)'
                sign = ''
            else:
                arrow = '—'
                color = 'var(--text-muted)'
                sign = ''
            # Format giá trị
            if isinstance(current, float):
                cur_str = f'{current:,.2f}'
                prev_str = f'{previous:,.2f}'
            else:
                cur_str = f'{current:,}'
                prev_str = f'{previous:,}'
            return f'''<div class="insight-badge">
                <div class="insight-label">{label}</div>
                <div class="insight-value" style="color: {color};">{arrow} {sign}{diff:.2f}%</div>
                <div class="insight-detail">{prev_str} → {cur_str}</div>
            </div>'''

    # Find prev day and LFL day from daily_stats
    target_idx = None
    for i, ds in enumerate(daily_stats):
        if ds['date_str'] == target_day['date_str']:
            target_idx = i
            break

    prev_day = daily_stats[target_idx - 1] if target_idx and target_idx > 0 else None
    lfl_date = target_date - timedelta(days=7)
    lfl_day = stats_by_date.get(lfl_date.strftime('%Y-%m-%d'))
    lfl_label = f"v/s LFL ({lfl_date.strftime('%d/%m/%Y')})"

    # --- Trend charts: tỷ lệ chính xác ---
    # Item accuracy
    cur_item_acc = trend_item_acc[-1]
    prev_item_acc = trend_item_acc[-2] if len(trend_item_acc) >= 2 else None
    lfl_item_acc = None
    if lfl_day:
        t_item_kfm_lfl = sum(lfl_day['categories'][c]['item_KFM'] for c in cats)
        t_item_lech_lfl = sum(lfl_day['categories'][c]['item_lech'] for c in cats)
        lfl_item_acc = acc_precise(t_item_lech_lfl, t_item_kfm_lfl)

    avg_item_acc = round(sum(trend_item_acc) / len(trend_item_acc), 2) if trend_item_acc else None
    d1, c1, p1 = _fmt_change_acc(cur_item_acc, prev_item_acc)
    d2, c2, p2 = _fmt_change_acc(cur_item_acc, lfl_item_acc)
    d3, c3, p3 = _fmt_change_acc(cur_item_acc, avg_item_acc)
    insight_trend_item = f'''<div class="insight-bar">
        {_build_insight_html('v/s Hôm qua', d1, c1, p1, is_acc=True)}
        {_build_insight_html(lfl_label, d2, c2, p2, is_acc=True)}
        {_build_insight_html('v/s Trung bình', d3, c3, p3, is_acc=True)}
    </div>'''

    # SKU accuracy
    cur_sku_acc = trend_sku_acc[-1]
    prev_sku_acc = trend_sku_acc[-2] if len(trend_sku_acc) >= 2 else None
    lfl_sku_acc = None
    if lfl_day:
        t_sku_total_lfl = sum(lfl_day['categories'][c]['sku'] for c in cats)
        t_sku_lech_lfl = sum(lfl_day['categories'][c]['sku_lech'] for c in cats)
        lfl_sku_acc = acc_precise(t_sku_lech_lfl, t_sku_total_lfl)

    avg_sku_acc = round(sum(trend_sku_acc) / len(trend_sku_acc), 2) if trend_sku_acc else None
    d1, c1, p1 = _fmt_change_acc(cur_sku_acc, prev_sku_acc)
    d2, c2, p2 = _fmt_change_acc(cur_sku_acc, lfl_sku_acc)
    d3, c3, p3 = _fmt_change_acc(cur_sku_acc, avg_sku_acc)
    insight_trend_sku = f'''<div class="insight-bar">
        {_build_insight_html('v/s Hôm qua', d1, c1, p1, is_acc=True)}
        {_build_insight_html(lfl_label, d2, c2, p2, is_acc=True)}
        {_build_insight_html('v/s Trung bình', d3, c3, p3, is_acc=True)}
    </div>'''

    # KG accuracy
    cur_kg_acc = trend_kg_acc[-1]
    prev_kg_acc = trend_kg_acc[-2] if len(trend_kg_acc) >= 2 else None
    lfl_kg_acc = None
    if lfl_day:
        t_kg_kfm_lfl = sum(lfl_day['categories'][c]['kg_KFM'] for c in cats)
        t_kg_lech_lfl = sum(lfl_day['categories'][c]['kg_lech'] for c in cats)
        lfl_kg_acc = acc_precise(t_kg_lech_lfl, t_kg_kfm_lfl)

    avg_kg_acc = round(sum(trend_kg_acc) / len(trend_kg_acc), 2) if trend_kg_acc else None
    d1, c1, p1 = _fmt_change_acc(cur_kg_acc, prev_kg_acc)
    d2, c2, p2 = _fmt_change_acc(cur_kg_acc, lfl_kg_acc)
    d3, c3, p3 = _fmt_change_acc(cur_kg_acc, avg_kg_acc)
    insight_trend_kg = f'''<div class="insight-bar">
        {_build_insight_html('v/s Hôm qua', d1, c1, p1, is_acc=True)}
        {_build_insight_html(lfl_label, d2, c2, p2, is_acc=True)}
        {_build_insight_html('v/s Trung bình', d3, c3, p3, is_acc=True)}
    </div>'''

    # --- Pie charts: sản lượng ---
    # Pie Item: tổng Item KFM
    cur_total_item = total_kfm_item
    prev_total_item = sum(prev_day['categories'][c]['item_KFM'] for c in cats) if prev_day else None
    lfl_total_item = sum(lfl_day['categories'][c]['item_KFM'] for c in cats) if lfl_day else None

    avg_total_item = round(sum(sum(ds['categories'][c]['item_KFM'] for c in cats) for ds in daily_stats) / len(daily_stats)) if daily_stats else None
    d1, c1, p1 = _fmt_change_qty(cur_total_item, prev_total_item)
    d2, c2, p2 = _fmt_change_qty(cur_total_item, lfl_total_item)
    d3, c3, p3 = _fmt_change_qty(cur_total_item, avg_total_item)
    insight_pie_item = f'''<div class="insight-bar">
        {_build_insight_html('v/s Hôm qua', d1, c1, p1, is_acc=False)}
        {_build_insight_html(lfl_label, d2, c2, p2, is_acc=False)}
        {_build_insight_html('v/s Trung bình', d3, c3, p3, is_acc=False)}
    </div>'''

    # Pie KG: tổng KG KFM
    cur_total_kg = total_kfm_kg
    prev_total_kg = sum(prev_day['categories'][c]['kg_KFM'] for c in cats) if prev_day else None
    lfl_total_kg = sum(lfl_day['categories'][c]['kg_KFM'] for c in cats) if lfl_day else None

    avg_total_kg = round(sum(sum(ds['categories'][c]['kg_KFM'] for c in cats) for ds in daily_stats) / len(daily_stats), 2) if daily_stats else None
    d1, c1, p1 = _fmt_change_qty(cur_total_kg, prev_total_kg)
    d2, c2, p2 = _fmt_change_qty(cur_total_kg, lfl_total_kg)
    d3, c3, p3 = _fmt_change_qty(cur_total_kg, avg_total_kg)
    insight_pie_kg = f'''<div class="insight-bar">
        {_build_insight_html('v/s Hôm qua', d1, c1, p1, is_acc=False)}
        {_build_insight_html(lfl_label, d2, c2, p2, is_acc=False)}
        {_build_insight_html('v/s Trung bình', d3, c3, p3, is_acc=False)}
    </div>'''

    # ---- Generate static fallback chart images ----
    print("  Generating static chart images for mobile fallback...")
    fallback_chart_item = generate_combo_chart_image(
        trend_dates, trend_item_kfm, trend_item_aba, trend_item_acc, 'Số lượng Item')
    fallback_chart_sku = generate_combo_chart_image(
        trend_dates, trend_sku_kfm, trend_sku_aba, trend_sku_acc, 'Số SKU')
    fallback_chart_kg = generate_combo_chart_image(
        trend_dates, trend_kg_kfm, trend_kg_aba, trend_kg_acc, 'Trọng lượng (KG)')
    fallback_pie_item = generate_pie_chart_image(
        pie_labels, pie_item_values, pie_colors, 'Item')
    fallback_pie_kg = generate_pie_chart_image(
        pie_labels, pie_kg_values, pie_colors, 'KG')
    print("  ✓ Static chart images generated")

    # ---- HTML Template ----
    html = f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Đối Soát Tồn Kho KFM vs ABA - {target_label}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-primary: #0f172a;
            --bg-secondary: #1e293b;
            --bg-card: rgba(30, 41, 59, 0.7);
            --bg-glass: rgba(255, 255, 255, 0.03);
            --border-glass: rgba(255, 255, 255, 0.08);
            --text-primary: #f1f5f9;
            --text-secondary: #94a3b8;
            --text-muted: #64748b;
            --accent-blue: #60a5fa;
            --accent-green: #34d399;
            --accent-yellow: #fbbf24;
            --accent-purple: #a78bfa;
            --accent-pink: #f472b6;
            --accent-red: #f87171;
            --gradient-1: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            --gradient-2: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            --gradient-3: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        }}

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        html {{
            overflow-x: hidden;
            width: 100%;
        }}

        body {{
            font-family: 'Inter', -apple-system, sans-serif;
            background: var(--bg-primary);
            color: var(--text-primary);
            min-height: 100vh;
            line-height: 1.6;
            overflow-x: hidden;
            width: 100%;
            -webkit-text-size-adjust: 100%;
        }}

        .bg-pattern {{
            position: fixed;
            inset: 0;
            background:
                radial-gradient(ellipse at 20% 20%, rgba(96, 165, 250, 0.08) 0%, transparent 50%),
                radial-gradient(ellipse at 80% 80%, rgba(167, 139, 250, 0.06) 0%, transparent 50%),
                radial-gradient(ellipse at 50% 50%, rgba(52, 211, 153, 0.04) 0%, transparent 50%);
            pointer-events: none;
            z-index: 0;
        }}

        .container {{
            max-width: 1100px;
            margin: 0 auto;
            padding: 2rem 3rem;
            position: relative;
            z-index: 1;
            width: 100%;
        }}

        /* Header */
        .header {{
            text-align: center;
            margin-bottom: 3rem;
            padding: 2.5rem 2rem;
            background: #1e293b;
            border: 1px solid var(--border-glass);
            border-radius: 24px;
            position: relative;
            overflow: hidden;
        }}
        .header::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 3px;
            background: linear-gradient(90deg, var(--accent-blue), var(--accent-purple), var(--accent-pink));
        }}
        .header h1 {{
            font-size: 2rem;
            font-weight: 800;
            background: linear-gradient(135deg, #f1f5f9, #94a3b8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 0.5rem;
            letter-spacing: -0.02em;
        }}
        .header .subtitle {{
            color: var(--text-secondary);
            font-size: 1.05rem;
            font-weight: 400;
        }}
        .header .date-badge {{
            display: inline-flex;
            align-items: center;
            gap: 0.5rem;
            margin-top: 1rem;
            padding: 0.5rem 1.25rem;
            background: rgba(96, 165, 250, 0.1);
            border: 1px solid rgba(96, 165, 250, 0.2);
            border-radius: 100px;
            font-size: 0.9rem;
            font-weight: 600;
            color: var(--accent-blue);
        }}

        /* Cards */
        .card {{
            background: #1e293b;
            border: 1px solid var(--border-glass);
            border-radius: 20px;
            padding: 2rem;
            margin-bottom: 2rem;
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }}
        .card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
        }}
        .card-title {{
            font-size: 1.25rem;
            font-weight: 700;
            margin-bottom: 1.5rem;
            display: flex;
            align-items: center;
            gap: 0.75rem;
        }}
        .card-title .icon {{
            width: 36px;
            height: 36px;
            border-radius: 10px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.1rem;
        }}

        /* Summary Cards */
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 1rem;
            margin-bottom: 2rem;
        }}
        .summary-item {{
            padding: 1.5rem;
            border-radius: 16px;
            background: var(--bg-glass);
            border: 1px solid var(--border-glass);
            text-align: center;
        }}
        .summary-item .label {{
            font-size: 0.8rem;
            font-weight: 500;
            color: var(--text-secondary);
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 0.5rem;
        }}
        .summary-item .value {{
            font-size: 1.8rem;
            font-weight: 800;
            letter-spacing: -0.02em;
        }}
        .summary-item .sub {{
            font-size: 0.75rem;
            color: var(--text-muted);
            margin-top: 0.25rem;
        }}

        /* Table */
        .table-wrapper {{
            overflow-x: auto;
            border-radius: 12px;
            border: 1px solid var(--border-glass);
            max-width: 100%;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.9rem;
        }}
        thead th {{
            background: rgba(255, 255, 255, 0.04);
            padding: 0.85rem 1rem;
            text-align: center;
            font-weight: 600;
            color: var(--text-secondary);
            font-size: 0.8rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            border-bottom: 1px solid var(--border-glass);
            white-space: nowrap;
        }}
        thead th:first-child {{
            text-align: left;
        }}
        thead th.group-header {{
            border-bottom: none;
            padding-bottom: 0.4rem;
        }}
        thead th.sub-header {{
            font-size: 0.72rem;
            padding-top: 0.3rem;
            color: var(--text-muted);
            text-align: center !important;
        }}
        tbody td {{
            padding: 0.85rem 1rem;
            text-align: center;
            border-bottom: 1px solid rgba(255, 255, 255, 0.03);
        }}
        tbody td:first-child {{
            text-align: left;
            font-weight: 600;
        }}
        tbody tr:hover {{
            background: rgba(255, 255, 255, 0.02);
        }}
        tbody tr:last-child td {{
            border-bottom: none;
        }}
        .row-total {{
            background: rgba(255, 255, 255, 0.03);
            font-weight: 700 !important;
        }}
        .row-total td {{
            font-weight: 700;
            border-top: 2px solid var(--border-glass);
        }}
        .cat-dot {{
            display: inline-block;
            width: 10px;
            height: 10px;
            border-radius: 50%;
            margin-right: 8px;
        }}

        /* Weekly Table */
        .weekly-header {{
            display: flex;
            align-items: center;
            gap: 1rem;
            margin-bottom: 1.5rem;
        }}
        .weekly-header h3 {{
            font-size: 1.1rem;
            font-weight: 700;
            color: var(--accent-red);
            font-style: italic;
        }}
        .week-badge {{
            display: inline-flex;
            align-items: center;
            padding: 0.35rem 1rem;
            background: rgba(248, 113, 113, 0.12);
            border: 1px solid rgba(248, 113, 113, 0.25);
            border-radius: 8px;
            font-size: 0.85rem;
            font-weight: 700;
            color: var(--accent-red);
        }}
        .row-avg {{
            background: rgba(255, 255, 255, 0.03);
        }}
        .row-avg td {{
            font-weight: 600;
            color: var(--accent-red);
            font-style: italic;
            border-top: 2px solid var(--border-glass);
        }}
        .no-data td {{
            color: var(--text-muted);
        }}

        /* Insight Badges */
        .insight-bar {{
            display: flex;
            gap: 1rem;
            margin-top: 1.25rem;
            flex-wrap: wrap;
        }}
        .insight-badge {{
            flex: 1;
            min-width: 180px;
            padding: 0.85rem 1rem;
            border-radius: 12px;
            background: var(--bg-glass);
            border: 1px solid var(--border-glass);
        }}
        .insight-label {{
            font-size: 0.72rem;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.04em;
            margin-bottom: 0.35rem;
        }}
        .insight-value {{
            font-size: 1.15rem;
            font-weight: 800;
            letter-spacing: -0.01em;
        }}
        .insight-detail {{
            font-size: 0.72rem;
            color: var(--text-muted);
            margin-top: 0.2rem;
        }}

        /* Chart Fallback Images (for viewers without JS) */
        .chart-fallback {{
            width: 100%;
            height: auto;
            border-radius: 8px;
            display: block;
        }}
        .pie-chart-container .chart-fallback {{
            max-width: 320px;
            margin: 0 auto;
        }}

        /* Charts Grid */
        .charts-grid {{
            display: grid;
            grid-template-columns: 1fr;
            gap: 2rem;
        }}
        .chart-container {{
            position: relative;
            height: 380px;
            width: 100%;
            isolation: isolate;
        }}
        .pie-section {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 2rem;
            align-items: center;
        }}
        .pie-chart-container {{
            position: relative;
            height: 350px;
            display: flex;
            justify-content: center;
            isolation: isolate;
        }}
        .pie-legend {{
            display: flex;
            flex-direction: column;
            gap: 1.25rem;
        }}
        .pie-legend-item {{
            display: flex;
            align-items: center;
            gap: 1rem;
            padding: 1rem 1.25rem;
            border-radius: 12px;
            background: var(--bg-glass);
            border: 1px solid var(--border-glass);
        }}
        .pie-legend-dot {{
            width: 14px;
            height: 14px;
            border-radius: 50%;
            flex-shrink: 0;
        }}
        .pie-legend-label {{
            font-weight: 500;
            color: var(--text-secondary);
            font-size: 0.85rem;
        }}
        .pie-legend-value {{
            margin-left: auto;
            font-weight: 700;
            font-size: 1.05rem;
        }}
        .pie-legend-pct {{
            font-size: 0.8rem;
            color: var(--text-muted);
            margin-left: 0.5rem;
        }}

        /* Footer */
        .footer {{
            text-align: center;
            padding: 2rem;
            color: var(--text-muted);
            font-size: 0.8rem;
        }}

        /* Tablet */
        @media (max-width: 1024px) {{
            .pie-section {{
                grid-template-columns: 1fr;
                gap: 1.5rem;
            }}
            .pie-chart-container {{
                height: 300px;
            }}
        }}

        /* Mobile */
        @media (max-width: 768px) {{
            .container {{
                padding: 0.75rem;
            }}
            .header {{
                padding: 1.5rem 1rem;
                margin-bottom: 1.5rem;
                border-radius: 16px;
            }}
            .header h1 {{
                font-size: 1.25rem;
            }}
            .header .subtitle {{
                font-size: 0.9rem;
            }}
            .header .date-badge {{
                font-size: 0.8rem;
                padding: 0.4rem 1rem;
            }}
            .summary-grid {{
                grid-template-columns: repeat(2, 1fr);
                gap: 0.75rem;
                margin-bottom: 1.25rem;
            }}
            .summary-item {{
                padding: 1rem;
                border-radius: 12px;
            }}
            .summary-item .value {{
                font-size: 1.35rem;
            }}
            .summary-item .label {{
                font-size: 0.7rem;
            }}
            .card {{
                padding: 1.25rem 1rem;
                margin-bottom: 1.25rem;
                border-radius: 14px;
            }}
            .card-title {{
                font-size: 1rem;
                flex-wrap: wrap;
                gap: 0.5rem;
                margin-bottom: 1rem;
            }}
            .card-title .icon {{
                width: 30px;
                height: 30px;
                font-size: 0.95rem;
                flex-shrink: 0;
            }}
            .chart-container {{
                height: 280px;
            }}
            .pie-section {{
                grid-template-columns: 1fr;
                gap: 1.25rem;
            }}
            .pie-chart-container {{
                height: 260px;
            }}
            .pie-legend {{
                gap: 0.75rem;
            }}
            .pie-legend-item {{
                padding: 0.75rem 1rem;
                gap: 0.75rem;
            }}
            .pie-legend-label {{
                font-size: 0.8rem;
            }}
            .pie-legend-value {{
                font-size: 0.9rem;
            }}
            .pie-legend-pct {{
                font-size: 0.7rem;
            }}
            table {{
                font-size: 0.8rem;
            }}
            thead th {{
                padding: 0.6rem 0.5rem;
                font-size: 0.7rem;
            }}
            tbody td {{
                padding: 0.6rem 0.5rem;
            }}
            .weekly-header {{
                flex-wrap: wrap;
                gap: 0.75rem;
                margin-bottom: 1rem;
            }}
            .week-badge {{
                font-size: 0.75rem;
                padding: 0.25rem 0.75rem;
            }}
            .footer {{
                padding: 1.25rem;
                font-size: 0.7rem;
            }}
        }}

        /* Small mobile */
        @media (max-width: 400px) {{
            .summary-grid {{
                grid-template-columns: 1fr;
            }}
            .summary-item .value {{
                font-size: 1.2rem;
            }}
            .chart-container {{
                height: 240px;
            }}
            .pie-chart-container {{
                height: 220px;
            }}
            .card-title {{
                font-size: 0.9rem;
            }}
        }}
    </style>
</head>
<body>
    <div class="bg-pattern"></div>
    <div class="container">
        <!-- Section 1: Header + Summary + Stats -->
        <div id="section-thongke">
        <!-- Header -->
        <div class="header">
            <h1>📦 Đối Soát Tồn Kho KFM vs ABA</h1>
            <p class="subtitle">Báo cáo kiểm soát tồn kho hàng ngày</p>
            <div class="date-badge">
                📅 {target_label}
            </div>
        </div>

        <!-- Summary Cards -->
        <div class="summary-grid">
            <div class="summary-item">
                <div class="label">Tổng SKU</div>
                <div class="value" style="color: var(--accent-purple)">{total_kfm_sku:,}</div>
                <div class="sub">Mã hàng đang theo dõi</div>
            </div>
            <div class="summary-item">
                <div class="label">Tồn KFM (Item)</div>
                <div class="value" style="color: var(--accent-blue)">{total_kfm_item:,}</div>
                <div class="sub">Tổng số lượng tồn</div>
            </div>
            <div class="summary-item">
                <div class="label">Tồn ABA (Item)</div>
                <div class="value" style="color: var(--accent-green)">{total_aba_item:,}</div>
                <div class="sub">Tổng số lượng tồn</div>
            </div>
            <div class="summary-item">
                <div class="label">Tổng KG (KFM)</div>
                <div class="value" style="color: var(--accent-yellow)">{total_kfm_kg:,.2f}</div>
                <div class="sub">Trọng lượng quy đổi</div>
            </div>
            <div class="summary-item">
                <div class="label">Tỷ lệ chính xác (Item)</div>
                <div class="value" style="color: {'var(--accent-green)' if trend_item_acc[-1] >= 95 else 'var(--accent-red)'}">{trend_item_acc[-1]:.2f}%</div>
                <div class="sub">{'✅ Tốt' if trend_item_acc[-1] >= 95 else '⚠️ Cần kiểm tra'}</div>
            </div>
        </div>
        </div>

        <!-- Section 7: Weekly Summary Table -->
        <div class="card" id="section-weekly">
            <div class="weekly-header">
                <div class="card-title" style="margin-bottom: 0;">
                    <div class="icon" style="background: rgba(248, 113, 113, 0.15); color: var(--accent-red);">📋</div>
                    <span style="font-style: italic; color: var(--accent-red);">BÁO CÁO TỒN KHO</span>
                </div>
                <div class="week-badge">W{iso_week}-{iso_year}</div>
            </div>
            <div class="table-wrapper">
                <table>
                    <thead>
                        <tr>
                            <th style="text-align: center;">Thứ</th>
                            <th>Ngày</th>
                            <th>Tỷ lệ tồn kho chính xác</th>
                            <th>Lệch dương</th>
                            <th>Lệch âm</th>
                            <th>Tổng lệch</th>
                            <th>Tồn KDB</th>
                        </tr>
                    </thead>
                    <tbody>
"""

    for wr in weekly_rows:
        if wr['has_data']:
            html += f"""                        <tr>
                            <td style="text-align: center; font-weight: 500;">{wr['thu']}</td>
                            <td style="text-align: center;">{wr['ngay']}</td>
                            <td style="text-align: center; font-weight: 600; color: {'var(--accent-green)' if wr['acc'] >= 99.5 else 'var(--accent-red)'};">{wr['acc']:.2f}%</td>
                            <td style="text-align: center;">{wr['lech_duong']:,}</td>
                            <td style="text-align: center;">{wr['lech_am']:,}</td>
                            <td style="text-align: center;">{wr['tong_lech']:,}</td>
                            <td style="text-align: center;">{wr['ton_kdb']:,}</td>
                        </tr>
"""
        else:
            html += f"""                        <tr class="no-data">
                            <td style="text-align: center; font-weight: 500;">{wr['thu']}</td>
                            <td style="text-align: center;">{wr['ngay']}</td>
                            <td style="text-align: center;"></td>
                            <td style="text-align: center;">0</td>
                            <td style="text-align: center;">0</td>
                            <td style="text-align: center;">0</td>
                            <td style="text-align: center;">0</td>
                        </tr>
"""

    html += f"""                        <tr class="row-avg">
                            <td colspan="2" style="text-align: center; font-style: italic;">Trung bình tỷ lệ tồn kho chính xác</td>
                            <td style="text-align: center; font-weight: 700;">{avg_acc:.2f}%</td>
                            <td colspan="4"></td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>
        </div>

        <!-- Stats Table (part of section-thongke) -->
        <div class="card" id="section-stats">
            <div class="card-title">
                <div class="icon" style="background: rgba(96, 165, 250, 0.15); color: var(--accent-blue);">📊</div>
                Thống Kê Sản Lượng Theo Phân Nhóm — {target_label}
            </div>
            <div class="table-wrapper">
                <table>
                    <thead>
                        <tr>
                            <th rowspan="2" style="vertical-align: middle;">Phân nhóm</th>
                            <th rowspan="2" style="vertical-align: middle;">Số SKU</th>
                            <th colspan="2" class="group-header" style="color: var(--accent-blue);">Tồn kho (Item)</th>
                            <th colspan="2" class="group-header" style="color: var(--accent-yellow);">Trọng lượng (KG)</th>
                        </tr>
                        <tr>
                            <th class="sub-header">KFM</th>
                            <th class="sub-header">ABA</th>
                            <th class="sub-header">KFM</th>
                            <th class="sub-header">ABA</th>
                        </tr>
                    </thead>
                    <tbody>
"""
    for r in table_rows:
        html += f"""                        <tr>
                            <td><span class="cat-dot" style="background: {r['color']}"></span>{r['cat']}</td>
                            <td>{r['sku']:,}</td>
                            <td>{r['item_KFM']:,}</td>
                            <td>{r['item_ABA']:,}</td>
                            <td>{r['kg_KFM']:,.2f}</td>
                            <td>{r['kg_ABA']:,.2f}</td>
                        </tr>
"""

    html += f"""                        <tr class="row-total">
                            <td>Tổng cộng</td>
                            <td>{total_kfm_sku:,}</td>
                            <td>{total_kfm_item:,}</td>
                            <td>{total_aba_item:,}</td>
                            <td>{total_kfm_kg:,.2f}</td>
                            <td>{total_aba_kg:,.2f}</td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>
        </div>


        <!-- Section 2: Trend Item -->
        <div class="card" id="section-trend-item">
            <div class="card-title">
                <div class="icon" style="background: rgba(167, 139, 250, 0.15); color: var(--accent-purple);">📈</div>
                Trend Tỷ Lệ Chính Xác Theo Item — {target_label}
            </div>
            <div class="chart-container">
                <img class="chart-fallback" id="fallback-chartItem" src="data:image/png;base64,{fallback_chart_item}" alt="Trend Item Chart">
                <canvas id="chartItem" style="display:none;"></canvas>
            </div>
            {insight_trend_item}
        </div>

        <!-- Section 3: Trend SKU -->
        <div class="card" id="section-trend-sku">
            <div class="card-title">
                <div class="icon" style="background: rgba(167, 139, 250, 0.15); color: var(--accent-purple);">📈</div>
                Trend Tỷ Lệ Chính Xác Theo SKU — {target_label}
            </div>
            <div class="chart-container">
                <img class="chart-fallback" id="fallback-chartSku" src="data:image/png;base64,{fallback_chart_sku}" alt="Trend SKU Chart">
                <canvas id="chartSku" style="display:none;"></canvas>
            </div>
            {insight_trend_sku}
        </div>

        <!-- Section 4: Trend KG -->
        <div class="card" id="section-trend-kg">
            <div class="card-title">
                <div class="icon" style="background: rgba(167, 139, 250, 0.15); color: var(--accent-purple);">📈</div>
                Trend Tỷ Lệ Chính Xác Theo KG — {target_label}
            </div>
            <div class="chart-container">
                <img class="chart-fallback" id="fallback-chartKg" src="data:image/png;base64,{fallback_chart_kg}" alt="Trend KG Chart">
                <canvas id="chartKg" style="display:none;"></canvas>
            </div>
            {insight_trend_kg}
        </div>

        <!-- Section 5: Pie Chart - Items -->
        <div class="card" id="section-pie-item">
            <div class="card-title">
                <div class="icon" style="background: rgba(167, 139, 250, 0.15); color: var(--accent-purple);">🥧</div>
                Tỉ Trọng Theo Số Lượng (Item) — KFM — {target_label}
            </div>
            <div class="pie-section">
                <div class="pie-chart-container">
                    <img class="chart-fallback" id="fallback-chartPieItem" src="data:image/png;base64,{fallback_pie_item}" alt="Pie Item Chart">
                    <canvas id="chartPieItem" style="display:none;"></canvas>
                </div>
                <div class="pie-legend" id="pieLegendItem"></div>
            </div>
            {insight_pie_item}
        </div>

        <!-- Section 6: Pie Chart - KG -->
        <div class="card" id="section-pie-kg">
            <div class="card-title">
                <div class="icon" style="background: rgba(251, 191, 36, 0.15); color: var(--accent-yellow);">🥧</div>
                Tỉ Trọng Theo Trọng Lượng (KG) — KFM — {target_label}
            </div>
            <div class="pie-section">
                <div class="pie-chart-container">
                    <img class="chart-fallback" id="fallback-chartPieKg" src="data:image/png;base64,{fallback_pie_kg}" alt="Pie KG Chart">
                    <canvas id="chartPieKg" style="display:none;"></canvas>
                </div>
                <div class="pie-legend" id="pieLegendKg"></div>
            </div>
            {insight_pie_kg}
        </div>
"""

    # ── Section 8: Chi tiết mã bị lệch ──
    detail_rows_html = ''
    if all_days:
        # Find target day's items from all_days
        target_day_items = None
        for d in all_days:
            if d['date_str'] == target_label:
                target_day_items = d['items']
                break
        if target_day_items:
            # Filter items with discrepancy (ton_KFM != ton_ABA)
            discrepant = [it for it in target_day_items if it['ton_KFM'] != it['ton_ABA']]
            if discrepant:
                for idx, it in enumerate(discrepant, 1):
                    thua_val = it['chenh_lech_thua'] if it['chenh_lech_thua'] else ''
                    thieu_val = it['chenh_lech_thieu'] if it['chenh_lech_thieu'] else ''
                    ghi_chu_val = it.get('ghi_chu', '')
                    # Color for thua (green) and thieu (red)
                    thua_style = 'color: var(--accent-green); font-weight: 600;' if thua_val else ''
                    thieu_style = 'color: var(--accent-red); font-weight: 600;' if thieu_val else ''
                    detail_rows_html += f"""                        <tr>
                            <td style="text-align: center;">{idx}</td>
                            <td style="text-align: center;">{target_label}</td>
                            <td style="text-align: left; font-family: monospace; font-size: 0.82rem;">{it['barcode']}</td>
                            <td style="text-align: left;">{it['ten_hang']}</td>
                            <td style="text-align: center;">{it['dvt']}</td>
                            <td style="text-align: center;">{it['ton_KFM']:,}</td>
                            <td style="text-align: center;">{it['ton_ABA']:,}</td>
                            <td style="text-align: center; {thua_style}">{thua_val if thua_val else ''}</td>
                            <td style="text-align: center; {thieu_style}">{thieu_val if thieu_val else ''}</td>
                            <td style="text-align: left; font-size: 0.82rem; color: var(--text-secondary);">{ghi_chu_val}</td>
                        </tr>
"""

    if detail_rows_html:
        html += f"""
        <!-- Section 8: Chi tiết mã bị lệch -->
        <div class="card" id="section-detail">
            <div class="card-title">
                <div class="icon" style="background: rgba(248, 113, 113, 0.15); color: var(--accent-red);">🔍</div>
                Chi Tiết Mã Bị Lệch — {target_label}
            </div>
            <div class="table-wrapper">
                <table>
                    <thead>
                        <tr>
                            <th style="text-align: center; width: 40px;">STT</th>
                            <th style="text-align: center;">Ngày</th>
                            <th style="text-align: left;">Barcode</th>
                            <th style="text-align: left;">Tên hàng</th>
                            <th style="text-align: center;">ĐVT</th>
                            <th style="text-align: center;">Tồn KFM</th>
                            <th style="text-align: center;">Tồn ABA</th>
                            <th style="text-align: center;">Thừa</th>
                            <th style="text-align: center;">Thiếu</th>
                            <th style="text-align: left;">Ghi chú</th>
                        </tr>
                    </thead>
                    <tbody>
{detail_rows_html}
                    </tbody>
                </table>
            </div>
            <div style="margin-top: 1rem; padding: 0.75rem 1rem; border-radius: 10px; background: var(--bg-glass); border: 1px solid var(--border-glass); display: flex; align-items: center; gap: 0.75rem;">
                <span style="font-size: 1.1rem;">⚠️</span>
                <span style="font-size: 0.85rem; color: var(--text-secondary);">Tổng số mã lệch: <strong style="color: var(--accent-red);">{len([it for it in target_day_items if it['ton_KFM'] != it['ton_ABA']])}</strong> / {len(target_day_items)} SKU</span>
            </div>
        </div>
"""

    html += f"""
        <div class="footer">
            <p>Generated at {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} • KFM × ABA Inventory Reconciliation</p>
        </div>
    </div>

    <script>
        // Chart.js defaults
        Chart.defaults.color = '#94a3b8';
        Chart.defaults.borderColor = 'rgba(255,255,255,0.05)';
        Chart.defaults.font.family = "'Inter', sans-serif";
        Chart.register(ChartDataLabels);
        // Disable datalabels globally, enable per dataset
        Chart.defaults.plugins.datalabels = {{ display: false }};

        const dates = {json.dumps(trend_dates)};

        // Utility: Create combo chart (bars + line)
        function createComboChart(canvasId, kfmData, abaData, accData, yLabel) {{
            const ctx = document.getElementById(canvasId).getContext('2d');
            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: dates,
                    datasets: [
                        {{
                            label: 'KFM',
                            data: kfmData,
                            backgroundColor: 'rgba(96, 165, 250, 0.6)',
                            borderColor: 'rgba(96, 165, 250, 1)',
                            borderWidth: 1,
                            borderRadius: 6,
                            order: 2,
                            yAxisID: 'y',
                        }},
                        {{
                            label: 'ABA',
                            data: abaData,
                            backgroundColor: 'rgba(52, 211, 153, 0.6)',
                            borderColor: 'rgba(52, 211, 153, 1)',
                            borderWidth: 1,
                            borderRadius: 6,
                            order: 2,
                            yAxisID: 'y',
                        }},
                        {{
                            label: 'Tỷ lệ chính xác (%)',
                            data: accData,
                            type: 'line',
                            borderColor: '#f472b6',
                            backgroundColor: 'rgba(244, 114, 182, 0.1)',
                            borderWidth: 3,
                            pointBackgroundColor: '#f472b6',
                            pointBorderColor: '#fff',
                            pointBorderWidth: 2,
                            pointRadius: 5,
                            pointHoverRadius: 7,
                            tension: 0.35,
                            fill: false,
                            order: 1,
                            yAxisID: 'y1',
                            clip: false,
                            datalabels: {{
                                display: true,
                                align: 'top',
                                anchor: 'end',
                                offset: 6,
                                color: '#f472b6',
                                font: {{ size: 11, weight: 'bold' }},
                                formatter: function(value) {{ return value.toFixed(2) + '%'; }}
                            }}
                        }}
                    ]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    layout: {{
                        padding: {{ top: 40 }}
                    }},
                    interaction: {{
                        mode: 'index',
                        intersect: false,
                    }},
                    plugins: {{
                        legend: {{
                            position: 'bottom',
                            labels: {{
                                usePointStyle: true,
                                pointStyle: 'circle',
                                padding: 20,
                                font: {{ size: 12, weight: '500' }}
                            }}
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(15, 23, 42, 0.95)',
                            titleFont: {{ size: 13, weight: '600' }},
                            bodyFont: {{ size: 12 }},
                            padding: 14,
                            cornerRadius: 10,
                            borderColor: 'rgba(255,255,255,0.1)',
                            borderWidth: 1,
                            callbacks: {{
                                label: function(ctx) {{
                                    if (ctx.dataset.yAxisID === 'y1') {{
                                        return ctx.dataset.label + ': ' + ctx.parsed.y.toFixed(2) + '%';
                                    }}
                                    return ctx.dataset.label + ': ' + ctx.parsed.y.toLocaleString();
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            grid: {{ display: false }},
                            ticks: {{ font: {{ size: 11, weight: '500' }} }}
                        }},
                        y: {{
                            position: 'left',
                            title: {{
                                display: true,
                                text: yLabel,
                                font: {{ size: 12, weight: '600' }},
                                color: '#64748b'
                            }},
                            grid: {{ color: 'rgba(255,255,255,0.04)' }},
                            ticks: {{
                                font: {{ size: 11 }},
                                callback: function(v) {{ return v.toLocaleString(); }}
                            }}
                        }},
                        y1: {{
                            position: 'right',
                            min: 0,
                            max: 100,
                            title: {{
                                display: true,
                                text: 'Tỷ lệ chính xác (%)',
                                font: {{ size: 12, weight: '600' }},
                                color: '#f472b6'
                            }},
                            grid: {{ drawOnChartArea: false }},
                            ticks: {{
                                font: {{ size: 11 }},
                                callback: function(v) {{ return v + '%'; }}
                            }}
                        }}
                    }}
                }}
            }});
            // Show interactive canvas, hide static fallback
            document.getElementById(canvasId).style.display = 'block';
            var fb = document.getElementById('fallback-' + canvasId);
            if (fb) fb.style.display = 'none';
        }}

        // Create charts
        createComboChart('chartItem',
            {json.dumps(trend_item_kfm)},
            {json.dumps(trend_item_aba)},
            {json.dumps(trend_item_acc)},
            'Số lượng Item'
        );
        createComboChart('chartSku',
            {json.dumps(trend_sku_kfm)},
            {json.dumps(trend_sku_aba)},
            {json.dumps(trend_sku_acc)},
            'Số SKU'
        );
        createComboChart('chartKg',
            {json.dumps(trend_kg_kfm)},
            {json.dumps(trend_kg_aba)},
            {json.dumps(trend_kg_acc)},
            'Trọng lượng (KG)'
        );

        // Reusable pie chart creator
        const pieLabels = {json.dumps(pie_labels)};
        const pieColors = {json.dumps(pie_colors)};

        function createPieChart(canvasId, legendId, values, unit) {{
            const total = values.reduce((a, b) => a + b, 0);
            const ctx = document.getElementById(canvasId).getContext('2d');
            new Chart(ctx, {{
                type: 'doughnut',
                data: {{
                    labels: pieLabels,
                    datasets: [{{
                        data: values,
                        backgroundColor: pieColors.map(c => c + 'cc'),
                        borderColor: pieColors,
                        borderWidth: 2,
                        hoverOffset: 12,
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    cutout: '55%',
                    plugins: {{
                        legend: {{ display: false }},
                        datalabels: {{
                            display: true,
                            color: '#fff',
                            font: {{ size: 13, weight: 'bold' }},
                            formatter: function(value) {{
                                const pct = ((value / total) * 100).toFixed(1);
                                return pct + '%';
                            }},
                            anchor: 'center',
                            align: 'center',
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(15, 23, 42, 0.95)',
                            titleFont: {{ size: 13, weight: '600' }},
                            bodyFont: {{ size: 12 }},
                            padding: 14,
                            cornerRadius: 10,
                            callbacks: {{
                                label: function(ctx) {{
                                    const pct = ((ctx.parsed / total) * 100).toFixed(1);
                                    return ctx.label + ': ' + ctx.parsed.toLocaleString() + ' ' + unit + ' (' + pct + '%)';
                                }}
                            }}
                        }}
                    }}
                }}
            }});
            // Custom legend
            const legendContainer = document.getElementById(legendId);
            pieLabels.forEach((label, i) => {{
                const pct = ((values[i] / total) * 100).toFixed(1);
                const item = document.createElement('div');
                item.className = 'pie-legend-item';
                item.innerHTML = `
                    <div class="pie-legend-dot" style="background: ${{pieColors[i]}}"></div>
                    <div><div class="pie-legend-label">${{label}}</div></div>
                    <div class="pie-legend-value">${{unit === 'KG' ? values[i].toLocaleString(undefined, {{minimumFractionDigits: 2, maximumFractionDigits: 2}}) : values[i].toLocaleString()}} ${{unit}}<span class="pie-legend-pct">${{pct}}%</span></div>
                `;
                legendContainer.appendChild(item);
            }});
            // Show interactive canvas, hide static fallback
            document.getElementById(canvasId).style.display = 'block';
            var fb = document.getElementById('fallback-' + canvasId);
            if (fb) fb.style.display = 'none';
        }}

        // Create both pie charts
        createPieChart('chartPieItem', 'pieLegendItem', {json.dumps(pie_item_values)}, 'Item');
        createPieChart('chartPieKg', 'pieLegendKg', {json.dumps(pie_kg_values)}, 'KG');
    </script>
</body>
</html>"""

    return html


# ─────────────────────────────────────────────
# 6. Generate PNG Images (Selenium screenshot from HTML)
# ─────────────────────────────────────────────
def generate_images(daily_stats, target_date_str):
    """Generate 7 PNG images by screenshotting sections of the HTML report."""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time

    # Find target day for date_tag
    target_day = None
    for ds in daily_stats:
        if ds['date_str'] == target_date_str:
            target_day = ds
            break
    if not target_day:
        print(f"   ⚠️ No data for {target_date_str}")
        return []

    date_tag = target_day['date'].strftime('%d%m%Y')

    # HTML report path
    html_path = os.path.join(OUTPUT_DIR,
                             f"report_doi_soat_{target_day['date'].strftime('%d.%m.%Y')}.html")
    if not os.path.exists(html_path):
        print(f"   ⚠️ HTML report not found: {html_path}")
        return []

    html_url = 'file:///' + html_path.replace('\\', '/')

    # Section IDs -> output filenames
    sections = [
        ('section-trend-item', f'DOI_SOAT_{date_tag}_2_TREND_ITEM.png'),
        ('section-trend-sku',  f'DOI_SOAT_{date_tag}_3_TREND_SKU.png'),
        ('section-trend-kg',   f'DOI_SOAT_{date_tag}_4_TREND_KG.png'),
        ('section-pie-item',   f'DOI_SOAT_{date_tag}_5_PIE_ITEM.png'),
        ('section-pie-kg',     f'DOI_SOAT_{date_tag}_6_PIE_KG.png'),
        ('section-weekly',     f'DOI_SOAT_{date_tag}_7_WEEKLY.png'),
        ('section-detail',     f'DOI_SOAT_{date_tag}_8_CHITIET.png'),
    ]

    # Setup headless Chrome
    chrome_options = Options()
    chrome_options.add_argument('--headless=new')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--force-device-scale-factor=2')
    chrome_options.add_argument('--window-size=1400,8000')
    chrome_options.add_argument('--hide-scrollbars')

    driver = None
    output_files = []
    try:
        from PIL import Image as PILImage
        import io

        driver = webdriver.Chrome(options=chrome_options)
        driver.get(html_url)

        # Wait for Chart.js + datalabels to render
        time.sleep(3)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'section-weekly'))
        )
        time.sleep(3)  # Extra wait for chart animations + datalabels

        # ── Image 1: Merge header + summary-cards + stats-table ──
        merge_ids = ['section-thongke', 'section-stats']
        merge_parts = []
        for sid in merge_ids:
            try:
                el = driver.find_element(By.ID, sid)
                png = el.screenshot_as_png
                merge_parts.append(PILImage.open(io.BytesIO(png)))
            except Exception as e:
                print(f"   ⚠️ Failed to capture {sid}: {e}")

        if merge_parts:
            # Use the stats table (last part) width as reference
            target_w = merge_parts[-1].width
            resized = []
            for img in merge_parts:
                if img.width != target_w:
                    # Resize proportionally to match target width
                    ratio = target_w / img.width
                    new_h = int(img.height * ratio)
                    img = img.resize((target_w, new_h), PILImage.LANCZOS)
                resized.append(img)

            gap = 30  # px gap between sections
            total_h = sum(img.height for img in resized) + gap * (len(resized) - 1)
            # Dark background matching the report theme (#0f172a)
            merged = PILImage.new('RGBA', (target_w, total_h), (15, 23, 42, 255))
            y_offset = 0
            for i, img in enumerate(resized):
                x_offset = (target_w - img.width) // 2
                merged.paste(img, (x_offset, y_offset))
                y_offset += img.height + gap
            merge_path = os.path.join(OUTPUT_DIR, f'DOI_SOAT_{date_tag}_1_THONGKE.png')
            merged.save(merge_path)
            output_files.append(merge_path)
            print(f"   📷 DOI_SOAT_{date_tag}_1_THONGKE.png (merged)")

        # ── Images 2-7: Individual section screenshots ──
        for section_id, filename in sections:
            try:
                element = driver.find_element(By.ID, section_id)
                path = os.path.join(OUTPUT_DIR, filename)
                element.screenshot(path)
                output_files.append(path)
                print(f"   📷 {filename}")
            except Exception as e:
                print(f"   ⚠️ Failed to capture {section_id}: {e}")

    except Exception as e:
        print(f"   ❌ Selenium error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if driver:
            driver.quit()

    return output_files


# ─────────────────────────────────────────────
# 7. Telegram Integration (via shared lib)
# ─────────────────────────────────────────────
from lib.telegram import (
    load_telegram_config as _load_tg_config,
    send_telegram_photo as _send_tg_photo,
    send_telegram_document as _send_tg_doc,
    delete_messages_by_tag,
    load_sent_messages as _load_sent,
    save_sent_messages as _save_sent,
)

def load_telegram_config():
    bot_token, chat_id = _load_tg_config(TELEGRAM_CONFIG_FILE, domain="inventory")
    if not bot_token:
        return None
    return {'bot_token': bot_token, 'chat_id': chat_id}

def send_telegram_photo(photo_path, caption):
    config = load_telegram_config()
    if not config:
        return None
    return _send_tg_photo(photo_path, caption, config['bot_token'], config['chat_id'])

def send_telegram_document(doc_path, caption):
    config = load_telegram_config()
    if not config:
        return None
    return _send_tg_doc(doc_path, caption, config['bot_token'], config['chat_id'])

def delete_telegram_messages(date_tag):
    config = load_telegram_config()
    if not config:
        return
    delete_messages_by_tag(SENT_MESSAGES_FILE, date_tag, config['bot_token'], config['chat_id'])

def send_to_telegram(image_files, html_path, date_tag, date_str):
    """Send all images + HTML to Telegram group."""
    config = load_telegram_config()
    if not config:
        print("   ⚠️ Telegram config not found, skipping...")
        return

    delete_telegram_messages(date_tag)

    sent_msg_ids = []
    caption = f"📊 Đối soát tồn kho {date_str}"

    section_labels = [
        "📋 Bảng thống kê",
        "📈 Trend Item",
        "📈 Trend SKU",
        "📈 Trend KG",
        "🥧 Tỉ trọng Item",
        "🥧 Tỉ trọng KG",
        "📋 Weekly Summary",
        "🔍 Chi tiết mã lệch",
    ]

    print(f"\n📤 Sending to Telegram...")
    for i, fpath in enumerate(image_files):
        label = section_labels[i] if i < len(section_labels) else ""
        mid = send_telegram_photo(fpath, f"{caption}\n{label}")
        if mid:
            sent_msg_ids.append(mid)
            print(f"   ✅ Sent: {os.path.basename(fpath)}")

    if os.path.exists(html_path):
        mid = send_telegram_document(html_path,
                                     f"📋 Báo cáo HTML {date_str} — mở bằng trình duyệt để xem chi tiết")
        if mid:
            sent_msg_ids.append(mid)
            print(f"   ✅ Sent: {os.path.basename(html_path)}")

    if sent_msg_ids:
        sent_data = _load_sent(SENT_MESSAGES_FILE)
        sent_data[date_tag] = sent_msg_ids
        _save_sent(SENT_MESSAGES_FILE, sent_data)
        print(f"   💾 Đã lưu {len(sent_msg_ids)} message IDs")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    import argparse
    parser = argparse.ArgumentParser(description='Generate inventory reconciliation report')
    parser.add_argument('--date', type=str, help='Target date dd/mm/yyyy (default: latest)')
    parser.add_argument('--no-telegram', action='store_true', help='Skip Telegram sending')
    args = parser.parse_args()

    print("🔄 Loading master data...")
    category_map = load_master_data()
    print(f"   ✅ {len(category_map)} barcodes mapped to categories")

    print("🔄 Loading weight data...")
    weight_map = load_weight_data()
    print(f"   ✅ {len(weight_map)} barcodes with weight info")

    print("🔄 Loading reconciliation files...")
    all_days = load_doi_soat_files()
    print(f"   ✅ {len(all_days)} days loaded")

    if not all_days:
        print("❌ No data to process!")
        return

    for d in all_days:
        print(f"   📅 {d['date_str']} — {len(d['items'])} items")

    print("\n🔄 Computing statistics...")
    daily_stats, unmatched_cat, unmatched_wt = compute_stats(all_days, category_map, weight_map)

    # Report unmatched
    if unmatched_cat:
        print(f"\n⚠️  {len(unmatched_cat)} barcode(s) KHÔNG TÌM ĐƯỢC PHÂN LOẠI (Đông/Mát/TCNK):")
        for bc, name in sorted(unmatched_cat.items()):
            print(f"   ❌ {bc} — {name}")

    if unmatched_wt:
        print(f"\n⚠️  {len(unmatched_wt)} barcode(s) KHÔNG TÌM ĐƯỢC TRỌNG LƯỢNG:")
        for bc, name in sorted(unmatched_wt.items()):
            print(f"   ❌ {bc} — {name}")

    # Determine target date
    target_date = args.date if args.date else daily_stats[-1]['date_str']
    date_file = target_date.replace('/', '.')
    date_parts = target_date.split('/')
    date_tag = f"{date_parts[0]}{date_parts[1]}{date_parts[2]}"

    # Generate HTML
    print(f"\n🔄 Generating HTML report for {target_date}...")
    html = generate_html(daily_stats, target_date, all_days=all_days)
    out_path = os.path.join(OUTPUT_DIR, f'report_doi_soat_{date_file}.html')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"   ✅ Report saved to: {out_path}")

    # Generate images
    print(f"\n🔄 Generating images...")
    image_files = generate_images(daily_stats, target_date)
    print(f"   ✅ {len(image_files)} images generated")

    # Send to Telegram
    if not args.no_telegram:
        send_to_telegram(image_files, out_path, date_tag, target_date)

    print("\n🎉 Done!")


if __name__ == '__main__':
    main()
