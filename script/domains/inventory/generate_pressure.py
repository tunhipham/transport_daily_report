"""
Pressure Score Analysis — Xác định SKU tồn cao / xuất thấp để chuyển kho.

Usage:
    python script/domains/inventory/generate_pressure.py
    python script/domains/inventory/generate_pressure.py --date 14/05/2026

Output: output/artifacts/inventory/pressure_score_ddmmyyyy.xlsx
"""
import sys, io, os, glob, argparse, math
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# ─── Paths ───────────────────────────────────────────────────
XNT_DIR = r'G:\My Drive\DOCS\DAILY\ton_kfm'
MASTER_PATH = r'G:\My Drive\DOCS\DAILY\ton_aba\data\master_data\Master Data.xlsx'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..', '..'))
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'output', 'artifacts', 'inventory')

# ─── Constants ───────────────────────────────────────────────
LOOKBACK_DAYS = 14

FALLBACK_CAT = {
    '2.FROZEN FOODS': 'ĐÔNG', '2.ICE CREAM': 'ĐÔNG',
    '2.CHILLED FOODS': 'MÁT', '2.DAIRY': 'MÁT', '2.BAKERY': 'MÁT',
    '2.DELICA': 'MÁT', '2.FISH AND SEAFOOD': 'MÁT', '2.FRUITS': 'MÁT',
    '2.MEAT': 'MÁT', '2.VEGETABLES': 'MÁT', '2.READY TO COOK': 'MÁT',
}
SKIP_CAT1 = {'1.zHÀNG KHÔNG BÁN', '1.CÔNG CỤ DỤNG CỤ'}

# ─── Thresholds (data-driven from P75/P90 split ĐÔNG vs MÁT) ────
THRESHOLDS = {
    'ĐÔNG': {
        'transfer_cover': 60,    # > P75 (56d)
        'transfer_pressure': 5000,  # ~P90 (5,431)
        'review_cover': 36,      # > P50 (36d)
        'review_pressure': 2000, # ~P75 (2,090)
        'dead_days': 10,         # frozen hàng turnover chậm hơn
    },
    'MÁT': {
        'transfer_cover': 25,    # > P75 (23d)
        'transfer_pressure': 1000, # ~P90 (988)
        'review_cover': 14,      # > P50 (14d)
        'review_pressure': 400,  # ~P75 (476)
        'dead_days': 7,
    },
}

# ─── Excel styles ────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
DATA_FONT = Font(name='Segoe UI', size=10)
NEGATIVE_FONT = Font(name='Segoe UI', size=10, color='FF4444')

TRANSFER_FILL = PatternFill('solid', fgColor='FEE2E2')
REVIEW_FILL = PatternFill('solid', fgColor='FEF3C7')
KEEP_FILL = PatternFill('solid', fgColor='D1FAE5')
NEGATIVE_FILL = PatternFill('solid', fgColor='FCA5A5')

THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

ACTION_LABELS = {
    'transfer': '🔴 Transfer',
    'review': '🟡 Review',
    'keep': '🟢 Keep',
}
ACTION_FILLS = {
    'transfer': TRANSFER_FILL,
    'review': REVIEW_FILL,
    'keep': KEEP_FILL,
}


# ════════════════════════════════════════════════════════════
# 1. Load master data
# ════════════════════════════════════════════════════════════
def load_master():
    master = {}
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in row]
        barcode = str(vals[1]).strip() if vals[1] else ''
        chi_nhanh = (vals[0] or '').strip().upper()
        phan_loai = (vals[4] or '').strip().upper()
        if barcode and 'LƯU HÀNG' in chi_nhanh and 'TCNK' not in chi_nhanh:
            if barcode not in master:
                master[barcode] = phan_loai
    wb.close()
    return master


# ════════════════════════════════════════════════════════════
# 2. Scan & select XNT files for 14-day coverage
# ════════════════════════════════════════════════════════════
def _parse_date(v):
    if isinstance(v, str):
        return datetime.strptime(v, '%Y-%m-%d').date()
    if hasattr(v, 'date'):
        return v.date()
    return v


def scan_xnt_files(target_date=None):
    """Return list of file_info dicts sorted by end_date asc."""
    files = sorted(glob.glob(os.path.join(XNT_DIR, 'XNT_*.xlsx')))
    if not files:
        raise FileNotFoundError(f"No XNT files in {XNT_DIR}")

    infos = []
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb[wb.sheetnames[0]]
        r2 = list(ws.iter_rows(min_row=2, max_row=2))[0]
        start = _parse_date(r2[0].value)
        end = _parse_date(r2[1].value)
        infos.append({'path': f, 'start': start, 'end': end})
        wb.close()

    infos.sort(key=lambda x: x['end'])

    # If target_date given, filter
    if target_date:
        infos = [fi for fi in infos if fi['end'] <= target_date]

    return infos


def select_files_for_coverage(infos, lookback=LOOKBACK_DAYS):
    """Pick non-overlapping files to cover ~lookback days ending at latest."""
    if not infos:
        raise ValueError("No XNT files available")

    latest = infos[-1]
    target_start = latest['end'] - timedelta(days=lookback - 1)

    selected = [latest]
    need_before = latest['start'] - timedelta(days=1)

    for fi in reversed(infos[:-1]):
        if fi['end'] > need_before:
            continue  # overlaps with already-selected
        if fi['end'] < target_start:
            continue  # too old
        eff_start = max(fi['start'], target_start)
        selected.append({**fi, 'effective_start': eff_start})
        need_before = eff_start - timedelta(days=1)
        if need_before < target_start:
            break

    # Calculate actual days covered
    actual_days = 0
    for s in selected:
        es = s.get('effective_start', s['start'])
        actual_days += (s['end'] - es).days + 1

    return selected, actual_days


# ════════════════════════════════════════════════════════════
# 3. Extract & aggregate data
# ════════════════════════════════════════════════════════════
def extract_data(selected_files, master):
    sku_data = {}

    for idx, fi in enumerate(selected_files):
        wb = openpyxl.load_workbook(fi['path'], data_only=True)
        ws = wb[wb.sheetnames[0]]
        file_days = (fi['end'] - fi['start']).days + 1
        eff_start = fi.get('effective_start', fi['start'])
        eff_days = (fi['end'] - eff_start).days + 1
        scale = eff_days / file_days if file_days > 0 else 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            vals = [c.value for c in row]
            branch = vals[6]
            cat1 = vals[7] or ''

            if branch != 'KHO ABA LƯU HÀNG':
                continue
            if cat1 in SKIP_CAT1:
                continue

            barcode = str(vals[3]).strip() if vals[3] else ''
            if not barcode:
                continue

            name = vals[4] or ''
            dvt = vals[5] or ''
            cat2 = vals[8] or ''
            xuat_chuyen = abs(vals[33] or 0)
            ton_cuoi = vals[37] or 0

            if barcode not in sku_data:
                sku_data[barcode] = {
                    'name': name, 'dvt': dvt, 'cat2': cat2,
                    'outbound_total': 0, 'inventory': 0, 'last_out_date': None,
                }

            sku_data[barcode]['outbound_total'] += xuat_chuyen * scale

            # Inventory from latest file only (idx == 0)
            if idx == 0:
                sku_data[barcode]['inventory'] = ton_cuoi

            if xuat_chuyen > 0:
                prev = sku_data[barcode]['last_out_date']
                if prev is None or fi['end'] > prev:
                    sku_data[barcode]['last_out_date'] = fi['end']

        wb.close()

    return sku_data


# ════════════════════════════════════════════════════════════
# 4. Compute pressure scores & action
# ════════════════════════════════════════════════════════════
def classify_group(barcode, cat2, master):
    if barcode in master:
        g = master[barcode]
    elif cat2 in FALLBACK_CAT:
        g = FALLBACK_CAT[cat2]
    else:
        g = 'UNKNOWN'
    if g in ('DONG',): g = 'ĐÔNG'
    if g in ('MAT',): g = 'MÁT'
    return g


def decide_action(group, days_cover, pressure, days_since, inventory):
    """Return action key: 'transfer', 'review', or 'keep'."""
    if inventory < 0:
        return 'review'  # anomaly
    if inventory == 0:
        return 'keep'

    t = THRESHOLDS.get(group, THRESHOLDS['MÁT'])

    # Dead stock
    if days_since >= t['dead_days'] and inventory > 0:
        return 'transfer'
    # High cover + high pressure
    if days_cover > t['transfer_cover'] and pressure > t['transfer_pressure']:
        return 'transfer'
    # Either high cover or high pressure alone
    if days_cover > t['transfer_cover'] or pressure > t['transfer_pressure']:
        return 'transfer'
    # Review zone
    if days_cover > t['review_cover'] or pressure > t['review_pressure']:
        return 'review'

    return 'keep'


def compute_results(sku_data, master, report_date, actual_days):
    results = []
    for barcode, d in sku_data.items():
        group = classify_group(barcode, d['cat2'], master)
        inv = d['inventory']
        out_total = d['outbound_total']
        avg_daily = out_total / actual_days if actual_days > 0 else 0
        days_cover = inv / max(avg_daily, 0.01)
        pressure = (inv * days_cover) / (avg_daily + 1)

        days_since = (report_date - d['last_out_date']).days if d['last_out_date'] else 999
        action = decide_action(group, days_cover, pressure, days_since, inv)

        results.append({
            'barcode': barcode,
            'name': d['name'],
            'dvt': d['dvt'],
            'group': group,
            'inventory': inv,
            'out_14d': out_total,
            'avg_daily': round(avg_daily, 1),
            'days_cover': round(days_cover, 1),
            'pressure': round(pressure, 0),
            'days_since_last_out': days_since,
            'last_out_date': d['last_out_date'],
            'action': action,
        })

    results.sort(key=lambda x: x['pressure'], reverse=True)
    return results


# ════════════════════════════════════════════════════════════
# 5. Export Excel
# ════════════════════════════════════════════════════════════
COLUMNS = [
    ('#', 5),
    ('Mã hàng', 18),
    ('Tên hàng', 42),
    ('ĐVT', 8),
    ('Tồn', 10),
    ('Xuất 14d', 10),
    ('Xuất TB/ngày', 12),
    ('Days Cover', 12),
    ('Pressure Score', 15),
    ('Last Outbound', 16),
    ('Action', 14),
]


def _write_sheet(ws, rows, report_date):
    """Write data rows to a worksheet."""
    # Header
    for ci, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    for ri, r in enumerate(rows, 2):
        is_negative = r['inventory'] < 0
        action_key = r['action']
        row_fill = NEGATIVE_FILL if is_negative else ACTION_FILLS.get(action_key)

        # Last outbound text
        if r['days_since_last_out'] >= 999:
            lo_text = '⚠ Không xuất'
        elif r['days_since_last_out'] == 0:
            lo_text = 'Hôm nay'
        elif r['days_since_last_out'] == 1:
            lo_text = 'Hôm qua'
        else:
            lo_text = f"{r['days_since_last_out']} ngày trước"

        values = [
            ri - 1,
            r['barcode'],
            r['name'],
            r['dvt'],
            r['inventory'],
            round(r['out_14d'], 1),
            r['avg_daily'],
            r['days_cover'],
            r['pressure'],
            lo_text,
            ACTION_LABELS.get(action_key, action_key),
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = NEGATIVE_FONT if is_negative else DATA_FONT
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            # Alignment
            if ci in (1,):
                cell.alignment = Alignment(horizontal='center')
            elif ci in (5, 6, 7, 8, 9):
                cell.alignment = Alignment(horizontal='right')
                if ci == 9:
                    cell.number_format = '#,##0'
                elif ci in (5, 6):
                    cell.number_format = '#,##0.#'
            elif ci == 11:
                cell.alignment = Alignment(horizontal='center')


def _write_summary(ws, results, report_date, actual_days, selected_files):
    """Write summary sheet."""
    ws.sheet_properties.tabColor = '4F46E5'

    title_font = Font(name='Segoe UI', bold=True, size=14, color='1F2937')
    subtitle_font = Font(name='Segoe UI', size=11, color='6B7280')
    label_font = Font(name='Segoe UI', bold=True, size=11)
    value_font = Font(name='Segoe UI', size=11)
    section_font = Font(name='Segoe UI', bold=True, size=12, color='4F46E5')

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    row = 2
    ws.cell(row=row, column=2, value='📦 Pressure Score Analysis').font = title_font
    row += 1
    ws.cell(row=row, column=2, value=f'Report date: {report_date.strftime("%d/%m/%Y")}').font = subtitle_font
    row += 1
    ws.cell(row=row, column=2, value=f'Data coverage: {actual_days} days').font = subtitle_font
    row += 2

    # Per group stats
    for group in ['ĐÔNG', 'MÁT']:
        g_data = [r for r in results if r['group'] == group]
        if not g_data:
            continue

        ws.cell(row=row, column=2, value=f'── {group} ──').font = section_font
        row += 1

        total = len(g_data)
        transfer = sum(1 for r in g_data if r['action'] == 'transfer')
        review = sum(1 for r in g_data if r['action'] == 'review')
        keep = sum(1 for r in g_data if r['action'] == 'keep')
        negative = sum(1 for r in g_data if r['inventory'] < 0)
        transfer_qty = sum(r['inventory'] for r in g_data if r['action'] == 'transfer' and r['inventory'] > 0)

        stats = [
            ('Tổng SKU', total),
            ('🔴 Transfer', f'{transfer} SKU ({transfer_qty:,.0f} items)'),
            ('🟡 Review', f'{review} SKU'),
            ('🟢 Keep', f'{keep} SKU'),
            ('⚠ Tồn âm', f'{negative} SKU'),
        ]
        for label, val in stats:
            ws.cell(row=row, column=2, value=label).font = label_font
            ws.cell(row=row, column=3, value=val).font = value_font
            row += 1
        row += 1

    # Threshold reference
    row += 1
    ws.cell(row=row, column=2, value='Thresholds (data-driven)').font = section_font
    row += 1
    for col_name, ci in [('Metric', 2), ('ĐÔNG', 3), ('MÁT', 4)]:
        cell = ws.cell(row=row, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1
    t_rows = [
        ('Transfer: Days Cover >', THRESHOLDS['ĐÔNG']['transfer_cover'], THRESHOLDS['MÁT']['transfer_cover']),
        ('Transfer: Pressure >', THRESHOLDS['ĐÔNG']['transfer_pressure'], THRESHOLDS['MÁT']['transfer_pressure']),
        ('Review: Days Cover >', THRESHOLDS['ĐÔNG']['review_cover'], THRESHOLDS['MÁT']['review_cover']),
        ('Review: Pressure >', THRESHOLDS['ĐÔNG']['review_pressure'], THRESHOLDS['MÁT']['review_pressure']),
        ('Dead stock: no outbound ≥', f"{THRESHOLDS['ĐÔNG']['dead_days']}d", f"{THRESHOLDS['MÁT']['dead_days']}d"),
    ]
    for label, v_dong, v_mat in t_rows:
        ws.cell(row=row, column=2, value=label).font = DATA_FONT
        ws.cell(row=row, column=3, value=v_dong).font = DATA_FONT
        ws.cell(row=row, column=4, value=v_mat).font = DATA_FONT
        for ci in (2, 3, 4):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    # Files used
    row += 2
    ws.cell(row=row, column=2, value='Files used').font = section_font
    row += 1
    for fi in selected_files:
        bn = os.path.basename(fi['path'])
        eff = fi.get('effective_start', fi['start'])
        ws.cell(row=row, column=2, value=bn).font = DATA_FONT
        ws.cell(row=row, column=3, value=f"{eff} → {fi['end']}").font = subtitle_font
        row += 1


def export_excel(results, report_date, actual_days, selected_files):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    fname = f"pressure_score_{report_date.strftime('%d%m%Y')}.xlsx"
    fpath = os.path.join(OUTPUT_DIR, fname)

    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    _write_summary(ws_summary, results, report_date, actual_days, selected_files)

    # ĐÔNG sheet
    dong = [r for r in results if r['group'] == 'ĐÔNG']
    ws_dong = wb.create_sheet('ĐÔNG')
    ws_dong.sheet_properties.tabColor = '3B82F6'
    _write_sheet(ws_dong, dong, report_date)

    # MÁT sheet
    mat = [r for r in results if r['group'] == 'MÁT']
    ws_mat = wb.create_sheet('MÁT')
    ws_mat.sheet_properties.tabColor = '10B981'
    _write_sheet(ws_mat, mat, report_date)

    # Tồn âm sheet (if any)
    negative = [r for r in results if r['inventory'] < 0]
    if negative:
        ws_neg = wb.create_sheet('⚠ Tồn Âm')
        ws_neg.sheet_properties.tabColor = 'EF4444'
        _write_sheet(ws_neg, negative, report_date)

    wb.save(fpath)
    return fpath


# ════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description='Pressure Score Analysis')
    parser.add_argument('--date', help='Target date dd/mm/yyyy (default: latest file)')
    args = parser.parse_args()

    target_date = None
    if args.date:
        target_date = datetime.strptime(args.date, '%d/%m/%Y').date()

    print("=" * 60)
    print("📦 Pressure Score Analysis")
    print("=" * 60)

    # 1. Load master
    print("\n[1/5] Loading master data...")
    master = load_master()
    print(f"  → {len(master)} barcodes loaded")

    # 2. Scan files
    print("\n[2/5] Scanning XNT files...")
    infos = scan_xnt_files(target_date)
    print(f"  → {len(infos)} files found")

    # 3. Select files for coverage
    print("\n[3/5] Selecting files for {}-day coverage...".format(LOOKBACK_DAYS))
    selected, actual_days = select_files_for_coverage(infos)
    report_date = selected[0]['end']
    print(f"  → {len(selected)} files selected, {actual_days} days covered")
    for s in selected:
        eff = s.get('effective_start', s['start'])
        print(f"     {os.path.basename(s['path']):25s}  {eff} → {s['end']}")

    # 4. Extract & compute
    print("\n[4/5] Extracting data & computing scores...")
    sku_data = extract_data(selected, master)
    results = compute_results(sku_data, master, report_date, actual_days)

    total = len(results)
    dong_n = sum(1 for r in results if r['group'] == 'ĐÔNG')
    mat_n = sum(1 for r in results if r['group'] == 'MÁT')
    transfer_n = sum(1 for r in results if r['action'] == 'transfer')
    review_n = sum(1 for r in results if r['action'] == 'review')
    keep_n = sum(1 for r in results if r['action'] == 'keep')
    neg_n = sum(1 for r in results if r['inventory'] < 0)

    print(f"  → {total} SKUs (ĐÔNG={dong_n}, MÁT={mat_n})")
    print(f"  → 🔴 Transfer: {transfer_n}  🟡 Review: {review_n}  🟢 Keep: {keep_n}")
    if neg_n:
        print(f"  → ⚠ Tồn âm: {neg_n}")

    # 5. Export Excel
    print("\n[5/5] Exporting Excel...")
    fpath = export_excel(results, report_date, actual_days, selected)
    print(f"  → Saved: {fpath}")

    print("\n" + "=" * 60)
    print("✅ Done!")
    print("=" * 60)


if __name__ == '__main__':
    main()
