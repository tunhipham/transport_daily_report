import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Check KH HANG DONG files
for folder_name in ["KH HÀNG ĐÔNG", "KH HÀNG MÁT"]:
    local_dir = rf"G:\My Drive\DOCS\DAILY\{folder_name}"
    if not os.path.exists(local_dir):
        print(f"NOT FOUND: {local_dir}")
        continue
    files = sorted([f for f in os.listdir(local_dir) if f.endswith('.xlsx') and not f.startswith('~')])
    print(f"\n=== {folder_name} ({len(files)} files) ===")
    print(f"  Recent: {files[-3:]}")
    
    # Read one file to see structure
    fpath = os.path.join(local_dir, files[-1])
    print(f"\n  Reading: {files[-1]}")
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    # Headers
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    for i, cell in enumerate(hdr):
        print(f"    Col {i}: {cell.value}")
    
    # Sample rows
    print(f"\n  Sample rows:")
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        vals = [f"{i}:{str(c)[:25]}" for i, c in enumerate(row) if c]
        print(f"    {' | '.join(vals)}")
        count += 1
    
    wb.close()

# Also check trip data for vehicle (Số xe) vs trip mapping
print("\n\n=== Trip data: Vehicle vs Trip mapping ===")
d = r'G:\My Drive\DOCS\DAILY\DS chi tiet chuyen xe\T04.26'
files = sorted([f for f in os.listdir(d) if f.endswith('.xlsx') and not f.startswith('~')])
fpath = os.path.join(d, files[-1])
wb = openpyxl.load_workbook(fpath, read_only=True)
ws = wb['Sheet 1']

from collections import defaultdict
vehicle_trips = defaultdict(set)  # vehicle -> set of trips
trip_vehicles = defaultdict(set)  # trip -> set of vehicles

count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    noi_chuyen = str(row[8] or "").strip()
    if noi_chuyen != "QCABA":
        continue
    trip_id = str(row[0] or "").strip()
    vehicle = str(row[2] or "").strip()
    date = str(row[5] or "").strip()
    if trip_id and vehicle:
        vehicle_trips[(date, vehicle)].add(trip_id)
        trip_vehicles[trip_id].add(vehicle)
    count += 1

wb.close()

print(f"  Total QCABA rows: {count}")
print(f"  Unique (date,vehicle) pairs: {len(vehicle_trips)}")
print(f"  Unique trips: {len(trip_vehicles)}")

# Show some vehicles with multiple trips
multi = {k: v for k, v in vehicle_trips.items() if len(v) > 1}
print(f"\n  Vehicles with >1 trip on same day: {len(multi)}")
for (date, veh), trips in list(multi.items())[:5]:
    print(f"    {date} {veh}: {trips}")
