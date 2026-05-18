"""
update_daily_report.py - Read all data sources for Day D and append to DAILY REPORT.xlsx
Usage: python script/update_daily_report.py [--date DD/MM/YYYY]

Opens DAILY REPORT.xlsx only ONCE, writes all sheets, saves once.
After writing, prints a sanity-check summary comparing today vs previous days.
"""
import os, sys, glob, re
from datetime import datetime
from openpyxl import load_workbook
from collections import Counter

# Fix Windows console encoding for Vietnamese characters
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "data")
OUTPUT = os.path.join(BASE, "output")


# ──────────────────────────────────────────
#  Helper functions
# ──────────────────────────────────────────

def parse_time_hour(time_text):
    """Parse time string like '03:00' or '22:30' and return hour as int."""
    if not time_text:
        return -1
    m = re.match(r'(\d{1,2}):', str(time_text).strip())
    return int(m.group(1)) if m else -1


def safe_val(row, idx):
    """Safely get cell value by index, return '' if index out of range."""
    if idx < len(row):
        return str(row[idx].value or "").strip()
    return ""


def find_file(folder, date_for_file):
    """Find first file in folder matching the date pattern."""
    if not os.path.isdir(folder):
        return None
    for f in os.listdir(folder):
        if date_for_file in f and not f.startswith("~") and not f.startswith("desktop"):
            return os.path.join(folder, f)
    return None


# ──────────────────────────────────────────
#  Read DATA STHI + XE sources
# ──────────────────────────────────────────

def read_krc(date_str):
    """Read KRC data - sheet 'KRC', filter by date, C7+C8 non-empty."""
    rows = []
    krc_files = glob.glob(os.path.join(DATA, "*KRC.xlsx"))
    if not krc_files:
        print("  WARNING: KRC file not found!")
        return rows

    wb = load_workbook(krc_files[0], read_only=True, data_only=True)
    ws = wb["KRC"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        scv = str(row[0].value or "").strip()
        if scv == date_str:
            diem_den = str(row[6].value or "").strip()
            gio_den = str(row[7].value or "").strip()
            tuyen = str(row[10].value or "").strip()
            if diem_den and gio_den:
                rows.append({"scv": date_str, "kho": "KRC", "diem_den": diem_den, "tuyen": tuyen})
    wb.close()
    return rows


def read_kfm(date_str):
    """Read KFM - sheet containing 'DRY', classify Sang/Toi by time."""
    rows = []
    kfm_files = glob.glob(os.path.join(DATA, "*KFM.xlsx"))
    if not kfm_files:
        print("  WARNING: KFM file not found!")
        return rows, 0, 0

    wb = load_workbook(kfm_files[0], read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "DRY" in name:
            ws = wb[name]
            break
    if not ws:
        ws = wb.worksheets[1]
    print(f"  Using sheet: {ws.title}")

    sang_count = toi_count = 0
    for row in ws.iter_rows(min_row=3, values_only=False):
        scv = str(row[0].value or "").strip()
        if scv == date_str:
            diem_den = str(row[6].value or "").strip()
            gio_den = str(row[7].value or "").strip()
            tuyen = str(row[10].value or "").strip()
            if diem_den and gio_den:
                hour = parse_time_hour(gio_den)
                if 6 <= hour < 18:
                    kho = "KSL-Sáng"
                    sang_count += 1
                else:
                    kho = "KSL-Tối"
                    toi_count += 1
                rows.append({"scv": date_str, "kho": kho, "diem_den": diem_den, "tuyen": tuyen})
    wb.close()
    return rows, sang_count, toi_count


def read_kh_meat(date_str, date_for_file):
    """Read KH MEAT - Col C3=DiemDen, Col L(12)=Tuyen."""
    rows = []
    folder = os.path.join(DATA, "KH MEAT")
    f = find_file(folder, date_for_file)
    if not f:
        print(f"  WARNING: No KH MEAT file for {date_for_file}")
        return rows

    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2, values_only=False):
        diem_den = safe_val(row, 2)
        tuyen = safe_val(row, 11)
        if diem_den:
            rows.append({"scv": date_str, "kho": "THỊT CÁ", "diem_den": diem_den, "tuyen": tuyen})
    wb.close()
    print(f"  File: {os.path.basename(f)}")
    return rows


def read_kh_dong_mat(date_str, date_for_file):
    """Read KH ĐÔNG + MÁT files, Col C3=DiemDen, Col J(10)=Tuyen."""
    rows = []
    dong_folder = mat_folder = None
    data_dirs = [d for d in os.listdir(DATA) if os.path.isdir(os.path.join(DATA, d))]
    for d in data_dirs:
        dl = d.lower()
        if "meat" not in dl:
            if "đông" in dl or "dong" in dl:
                dong_folder = os.path.join(DATA, d)
            elif "mát" in dl or "mat" in dl:
                mat_folder = os.path.join(DATA, d)

    dong_count = mat_count = 0

    for label, folder in [("ĐÔNG", dong_folder), ("MÁT", mat_folder)]:
        if not folder:
            continue
        f = find_file(folder, date_for_file)
        if not f:
            print(f"  WARNING: No KH {label} file for {date_for_file}")
            continue
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            diem_den = safe_val(row, 2)
            tuyen = safe_val(row, 9)
            if diem_den:
                rows.append({"scv": date_str, "kho": "ĐÔNG MÁT", "diem_den": diem_den, "tuyen": tuyen})
                count += 1
        wb.close()
        print(f"  {label}: {os.path.basename(f)} ({count} rows)")
        if label == "ĐÔNG":
            dong_count = count
        else:
            mat_count = count

    return rows, dong_count, mat_count


# ──────────────────────────────────────────
#  Read PT sources
# ──────────────────────────────────────────

def read_transfer(date_str, filepath):
    """Read transfer file, filter by date (col A)."""
    rows = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2, values_only=False):
        ngay = str(row[0].value or "").strip()
        if ngay == date_str:
            kho = safe_val(row, 2)
            code = safe_val(row, 7)
            sl_raw = safe_val(row, 10)
            tl_raw = safe_val(row, 14)
            try:
                sl = float(sl_raw) if sl_raw else 0
                if sl == int(sl): sl = int(sl)
            except ValueError:
                sl = sl_raw
            try:
                tl = float(tl_raw) if tl_raw else 0
                if tl == int(tl): tl = int(tl)
            except ValueError:
                tl = tl_raw
            rows.append({"ngay": date_str, "kho": kho, "code": code, "sl": sl, "tl": tl})
    wb.close()
    return rows


def read_yeu_cau(date_str, filepath):
    """Read yeu_cau KSL file, sheet KF. All rows."""
    rows = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name == 'KF':
            ws = wb[name]
            break
    if not ws:
        ws = wb.worksheets[0]

    for row in ws.iter_rows(min_row=2, values_only=False):
        code = safe_val(row, 2)
        kho = safe_val(row, 23)
        sl_raw = safe_val(row, 17)
        try:
            sl = float(sl_raw) if sl_raw else 0
            if sl == int(sl): sl = int(sl)
        except ValueError:
            sl = sl_raw
        if code:
            rows.append({"ngay": date_str, "kho": kho, "code": code, "sl": sl, "tl": ""})
    wb.close()
    return rows


# ──────────────────────────────────────────
#  Sanity check: compare today vs recent days
# ──────────────────────────────────────────

def sanity_check(ws_sthi, ws_pt, date_str):
    """Compare today's aggregates with the previous day in the data.
    Print a comparison table so user can spot anomalies."""

    print("\n" + "=" * 60)
    print("  SANITY CHECK")
    print("=" * 60)

    # Collect STHI data per (date, kho)
    sthi_data = {}  # date -> kho -> {"sthi": set, "tuyen": set}
    for row in ws_sthi.iter_rows(min_row=2, values_only=False):
        d = str(row[0].value or "").strip()
        kho = str(row[1].value or "").strip()
        diem_den = str(row[2].value or "").strip()
        tuyen = str(row[3].value or "").strip()
        if not d or not kho:
            continue
        if d not in sthi_data:
            sthi_data[d] = {}
        if kho not in sthi_data[d]:
            sthi_data[d][kho] = {"sthi": set(), "tuyen": set()}
        if diem_den:
            sthi_data[d][kho]["sthi"].add(diem_den)
        if tuyen:
            sthi_data[d][kho]["tuyen"].add(tuyen)

    # Collect PT data per (date, kho) => count items
    pt_data = {}  # date -> kho -> {"items": count}
    for row in ws_pt.iter_rows(min_row=2, values_only=False):
        d = str(row[0].value or "").strip()
        kho = str(row[1].value or "").strip()
        if not d or not kho:
            continue
        if d not in pt_data:
            pt_data[d] = {}
        if kho not in pt_data[d]:
            pt_data[d][kho] = 0
        pt_data[d][kho] += 1

    # Get dates sorted, find today and previous day
    all_dates = sorted(set(list(sthi_data.keys()) + list(pt_data.keys())))
    if date_str not in all_dates:
        print(f"  Date {date_str} not found in data.")
        return

    idx = all_dates.index(date_str)
    prev_date = all_dates[idx - 1] if idx > 0 else None

    # Print comparison
    all_khos = sorted(set(
        list(sthi_data.get(date_str, {}).keys()) +
        list(sthi_data.get(prev_date, {}).keys()) if prev_date else
        list(sthi_data.get(date_str, {}).keys())
    ))

    print(f"\n  {'KHO':<15} {'SL STHI':>10} {'SL XE':>10} | {'SL STHI':>10} {'SL XE':>10}")
    print(f"  {'':15} {'(' + date_str + ')':>21} | {'(' + (prev_date or 'N/A') + ')':>21}")
    print(f"  {'-'*15} {'-'*10} {'-'*10} | {'-'*10} {'-'*10}")

    for kho in all_khos:
        today = sthi_data.get(date_str, {}).get(kho, {"sthi": set(), "tuyen": set()})
        prev = sthi_data.get(prev_date, {}).get(kho, {"sthi": set(), "tuyen": set()}) if prev_date else {"sthi": set(), "tuyen": set()}
        t_sthi = len(today["sthi"])
        t_xe = len(today["tuyen"])
        p_sthi = len(prev["sthi"])
        p_xe = len(prev["tuyen"])

        # Flag large differences
        flag = ""
        if prev_date and p_sthi > 0:
            diff = abs(t_sthi - p_sthi) / p_sthi
            if diff > 0.5:
                flag = " ⚠️"

        print(f"  {kho:<15} {t_sthi:>10} {t_xe:>10} | {p_sthi:>10} {p_xe:>10}{flag}")

    # PT summary
    print(f"\n  {'KHO':<25} {'PT items':>10} | {'PT items':>10}")
    print(f"  {'':25} {'(' + date_str + ')':>10} | {'(' + (prev_date or 'N/A') + ')':>10}")
    print(f"  {'-'*25} {'-'*10} | {'-'*10}")

    pt_khos = sorted(set(
        list(pt_data.get(date_str, {}).keys()) +
        (list(pt_data.get(prev_date, {}).keys()) if prev_date else [])
    ))
    for kho in pt_khos[:10]:  # Top 10 khos
        t = pt_data.get(date_str, {}).get(kho, 0)
        p = pt_data.get(prev_date, {}).get(kho, 0) if prev_date else 0
        flag = ""
        if prev_date and p > 0:
            diff = abs(t - p) / p
            if diff > 1.0:
                flag = " ⚠️"
        print(f"  {kho:<25} {t:>10} | {p:>10}{flag}")

    total_t = sum(pt_data.get(date_str, {}).values())
    total_p = sum(pt_data.get(prev_date, {}).values()) if prev_date else 0
    print(f"  {'TOTAL':<25} {total_t:>10} | {total_p:>10}")
    print()


# ──────────────────────────────────────────
#  Main: single file open/save
# ──────────────────────────────────────────

def main():
    # Parse date argument
    date_str = ""
    if "--date" in sys.argv:
        idx = sys.argv.index("--date")
        date_str = sys.argv[idx + 1]
    if not date_str:
        date_str = datetime.now().strftime("%d/%m/%Y")

    parts = date_str.split("/")
    date_for_file = f"{parts[0]}.{parts[1]}.{parts[2]}"

    print("=" * 45)
    print(f"  DAILY REPORT UPDATE - {date_str}")
    print("=" * 45)

    # ── Read all source data ──

    sthi_rows = []

    # 1. KRC
    print("\n[1/4] Reading KRC data...")
    krc_rows = read_krc(date_str)
    sthi_rows.extend(krc_rows)
    print(f"  Found {len(krc_rows)} rows for KRC")

    # 2. KFM
    print("\n[2/4] Reading KFM data (KSL-Sáng + KSL-Tối)...")
    kfm_rows, sang, toi = read_kfm(date_str)
    sthi_rows.extend(kfm_rows)
    print(f"  Found {sang} rows for KSL-Sáng")
    print(f"  Found {toi} rows for KSL-Tối")

    # 3. KH MEAT
    print("\n[3/4] Reading KH MEAT data...")
    meat_rows = read_kh_meat(date_str, date_for_file)
    sthi_rows.extend(meat_rows)
    print(f"  Found {len(meat_rows)} rows for THỊT CÁ")

    # 4. KH DONG + MAT
    print("\n[4/4] Reading KH ĐÔNG + KH MÁT data...")
    dm_rows, dong_c, mat_c = read_kh_dong_mat(date_str, date_for_file)
    sthi_rows.extend(dm_rows)
    print(f"  Total ĐÔNG MÁT: {dong_c + mat_c} rows")

    # 5. Transfer (PT)
    pt_rows = []
    print("\n" + "=" * 45)
    print("  Reading PT data...")
    print("=" * 45)

    transfer_files = glob.glob(os.path.join(DATA, "transfer_*.xlsx"))
    if transfer_files:
        transfer_file = sorted(transfer_files)[-1]
        print(f"\n  Transfer: {os.path.basename(transfer_file)}")
        t_rows = read_transfer(date_str, transfer_file)
        pt_rows.extend(t_rows)
        print(f"  Found {len(t_rows)} transfer rows")
    else:
        print("\n  WARNING: No transfer file found!")

    # 6. Yeu cau (PT)
    yc_files = glob.glob(os.path.join(DATA, "yeu_cau_chuyen_hang_thuong_*.xlsx"))
    if yc_files:
        yc_file = sorted(yc_files)[-1]
        print(f"  Yêu cầu: {os.path.basename(yc_file)}")
        yc_rows = read_yeu_cau(date_str, yc_file)
        pt_rows.extend(yc_rows)
        print(f"  Found {len(yc_rows)} yêu cầu rows")
    else:
        print("\n  WARNING: No yeu_cau file found!")

    # ── Summary ──
    print(f"\n  STHI+XE: {len(sthi_rows)} rows")
    print(f"  PT:      {len(pt_rows)} rows")

    if not sthi_rows and not pt_rows:
        print(f"\nNo data found for {date_str}. Nothing to update.")
        return

    # ── Open file ONCE, write all, save ONCE ──
    report_path = os.path.join(OUTPUT, "DAILY REPORT.xlsx")
    print(f"\nOpening DAILY REPORT.xlsx...")
    wb = load_workbook(report_path)

    # Write STHI+XE
    if sthi_rows:
        ws = wb["DATA STHI + XE"]
        last_row = ws.max_row
        while last_row > 1 and not ws.cell(last_row, 1).value:
            last_row -= 1
        start = last_row + 1
        for i, row_data in enumerate(sthi_rows):
            r = start + i
            ws.cell(r, 1, row_data["scv"])
            ws.cell(r, 2, row_data["kho"])
            ws.cell(r, 3, row_data["diem_den"])
            ws.cell(r, 4, row_data["tuyen"])
        print(f"  STHI+XE: appended {len(sthi_rows)} rows (rows {start} - {start + len(sthi_rows) - 1})")

    # Write PT
    if pt_rows:
        ws_pt = wb["PT"]
        last_row = ws_pt.max_row
        while last_row > 1 and not ws_pt.cell(last_row, 1).value:
            last_row -= 1
        start = last_row + 1
        for i, row_data in enumerate(pt_rows):
            r = start + i
            ws_pt.cell(r, 1, row_data["ngay"])
            ws_pt.cell(r, 2, row_data["kho"])
            ws_pt.cell(r, 4, row_data["code"])
            ws_pt.cell(r, 5, row_data["sl"])
            ws_pt.cell(r, 6, row_data["tl"])
        print(f"  PT:      appended {len(pt_rows)} rows (rows {start} - {start + len(pt_rows) - 1})")

    # Save once
    print(f"\nSaving...")
    wb.save(report_path)

    # ── Sanity check (read-only) ──
    try:
        sanity_check(wb["DATA STHI + XE"], wb["PT"], date_str)
    except Exception as e:
        print(f"  Sanity check error: {e}")

    wb.close()

    print("=" * 45)
    print("  DONE")
    print("=" * 45)


if __name__ == "__main__":
    main()
