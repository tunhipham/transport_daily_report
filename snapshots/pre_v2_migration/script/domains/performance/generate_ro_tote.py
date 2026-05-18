"""
generate_ro_tote.py — Báo cáo Rổ / Tote / Thùng Carton cho kho ĐÔNG MÁT

Đọc DS chi tiết chuyến xe, filter QCABA, xuất Excel 3 sheets:
  1. Hàng Mát — Theo ST   (Ngày, Mã ST, Tên ST, Mã Trip, Rổ, Thùng Carton)
  2. Hàng Đông — Theo ST  (Ngày, Mã ST, Tên ST, Mã Trip, Tote)
  3. Theo Trip             (Ngày, Mã Trip, Rổ, Tote, Thùng Carton)

Usage:
  python script/domains/performance/generate_ro_tote.py --months 3,4 --year 2026
"""
import os, sys, argparse, re
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
OUTPUT = os.path.join(BASE, "output")

# ── Container type classification ──
CONTAINER_RO = "Rổ ABA đông mát"
CONTAINER_THUNG = "Thùng Carton, Bịch nguyên"
CONTAINER_TOTE = "Tote ABA đông mát"


def parse_date(s):
    """Parse date string to datetime.date."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ["%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]:
        try:
            return datetime.strptime(s.split(" ")[0] if " " in s and fmt.startswith("%d") else s, fmt).date()
        except ValueError:
            continue
    return None


def load_trip_data(months, year):
    """Load all QCABA trip data from specified months."""
    all_rows = []

    for month in months:
        data_dir = rf'G:\My Drive\DOCS\DAILY\DS chi tiet chuyen xe\T{month:02d}.{year % 100}'
        if not os.path.exists(data_dir):
            print(f"  ⚠ Directory not found: {data_dir}")
            continue

        files = sorted([f for f in os.listdir(data_dir) if f.endswith('.xlsx') and not f.startswith('~')])
        print(f"  📁 T{month:02d}: {len(files)} files in {data_dir}")

        for fname in files:
            fpath = os.path.join(data_dir, fname)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True)
                ws = wb['Sheet 1']
                file_count = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    noi_chuyen = str(row[8] or "").strip()
                    if noi_chuyen != "QCABA":
                        continue

                    trip_id = str(row[0] or "").strip()
                    date_str = str(row[5] or "").strip()
                    store_code = str(row[9] or "").strip()
                    store_name = str(row[10] or "").strip()
                    container_type = str(row[18] or "").strip() if len(row) > 18 else ""
                    qty_raw = row[19] if len(row) > 19 else 0
                    qty_nhan_raw = row[20] if len(row) > 20 else 0

                    if not trip_id or not store_code or not container_type:
                        continue

                    try:
                        qty = int(qty_raw) if qty_raw else 0
                    except (ValueError, TypeError):
                        qty = 0

                    try:
                        qty_nhan = int(qty_nhan_raw) if qty_nhan_raw else 0
                    except (ValueError, TypeError):
                        qty_nhan = 0

                    date = parse_date(date_str)
                    if not date:
                        continue

                    all_rows.append({
                        "date": date,
                        "trip_id": trip_id,
                        "store_code": store_code,
                        "store_name": store_name,
                        "container_type": container_type,
                        "qty": qty,
                        "qty_nhan": qty_nhan,
                    })
                    file_count += 1

                wb.close()
                if file_count > 0:
                    print(f"    ✅ {fname}: {file_count} QCABA rows")
            except Exception as e:
                print(f"    ⚠ Error reading {fname}: {e}")

    print(f"\n  📊 Total: {len(all_rows)} QCABA rows loaded")
    return all_rows


def classify_mixed_deliveries(rows):
    """Handle trips where a (date, trip, store) has BOTH rổ AND tote.
    
    Rule: When a delivery has both container types, use SL nhận (qty_nhan)
    to determine classification:
    - SL nhận rổ/thùng ≠ 0, SL nhận tote = 0 → hàng mát → exclude tote
    - SL nhận rổ/thùng = 0, SL nhận tote ≠ 0 → hàng đông → exclude rổ/thùng
    
    SUM still uses SL giao (qty). SL nhận is only for classification.
    Returns rows with added 'exclude' flag.
    """
    # Group rows by (date, trip_id, store_code)
    groups = defaultdict(list)
    for i, r in enumerate(rows):
        key = (r["date"], r["trip_id"], r["store_code"])
        groups[key].append(i)
    
    excluded = set()  # indices of rows to exclude
    mixed_count = 0
    reclassified_mat = 0
    reclassified_dong = 0
    
    for key, indices in groups.items():
        # Check if group has both mát (rổ/thùng) and đông (tote)
        mat_indices = [i for i in indices if rows[i]["container_type"] in (CONTAINER_RO, CONTAINER_THUNG)]
        dong_indices = [i for i in indices if rows[i]["container_type"] == CONTAINER_TOTE]
        
        if not mat_indices or not dong_indices:
            continue  # Not mixed → no reclassification needed
        
        mixed_count += 1
        
        # Sum qty_nhan per category
        mat_nhan = sum(rows[i]["qty_nhan"] for i in mat_indices)
        dong_nhan = sum(rows[i]["qty_nhan"] for i in dong_indices)
        
        if mat_nhan != 0 and dong_nhan == 0:
            # Hàng mát → exclude tote rows
            for i in dong_indices:
                excluded.add(i)
            reclassified_mat += 1
        elif mat_nhan == 0 and dong_nhan != 0:
            # Hàng đông → exclude rổ/thùng rows
            for i in mat_indices:
                excluded.add(i)
            reclassified_dong += 1
        # If both ≠ 0 or both = 0 → keep all (no reclassification)
    
    print(f"  🔀 Mixed deliveries (both rổ+tote): {mixed_count}")
    if mixed_count > 0:
        print(f"     → Reclassified as MÁT (exclude tote): {reclassified_mat}")
        print(f"     → Reclassified as ĐÔNG (exclude rổ): {reclassified_dong}")
        print(f"     → Kept both (ambiguous): {mixed_count - reclassified_mat - reclassified_dong}")
        print(f"     → Total rows excluded: {len(excluded)}")
    
    # Mark excluded rows
    for i, r in enumerate(rows):
        r["excluded"] = (i in excluded)
    
    return rows


def aggregate_data(rows):
    """Aggregate rows into 3 sheet datasets."""

    # ── Sheet 1: Hàng Mát — Theo ST ──
    # Group by (date, store_code, trip_id) → SUM rổ, SUM thùng carton
    mat_agg = defaultdict(lambda: {"store_name": "", "ro": 0, "thung_carton": 0})
    for r in rows:
        if r.get("excluded"):
            continue
        ct = r["container_type"]
        if ct not in (CONTAINER_RO, CONTAINER_THUNG):
            continue
        key = (r["date"], r["store_code"], r["trip_id"])
        mat_agg[key]["store_name"] = r["store_name"]
        if ct == CONTAINER_RO:
            mat_agg[key]["ro"] += r["qty"]
        elif ct == CONTAINER_THUNG:
            mat_agg[key]["thung_carton"] += r["qty"]

    sheet1 = []
    for (date, store, trip), vals in sorted(mat_agg.items()):
        sheet1.append({
            "date": date,
            "store_code": store,
            "store_name": vals["store_name"],
            "trip_id": trip,
            "ro": vals["ro"],
            "thung_carton": vals["thung_carton"],
        })

    # ── Sheet 2: Hàng Đông — Theo ST ──
    # Group by (date, store_code, trip_id) → SUM tote
    dong_agg = defaultdict(lambda: {"store_name": "", "tote": 0})
    for r in rows:
        if r.get("excluded"):
            continue
        if r["container_type"] != CONTAINER_TOTE:
            continue
        key = (r["date"], r["store_code"], r["trip_id"])
        dong_agg[key]["store_name"] = r["store_name"]
        dong_agg[key]["tote"] += r["qty"]

    sheet2 = []
    for (date, store, trip), vals in sorted(dong_agg.items()):
        sheet2.append({
            "date": date,
            "store_code": store,
            "store_name": vals["store_name"],
            "trip_id": trip,
            "tote": vals["tote"],
        })

    # ── Sheet 3: Theo Trip ──
    # Group by (date, trip_id) → SUM rổ, tote, thùng carton (excluded rows skipped)
    trip_agg = defaultdict(lambda: {"ro": 0, "tote": 0, "thung_carton": 0})
    for r in rows:
        if r.get("excluded"):
            continue
        key = (r["date"], r["trip_id"])
        ct = r["container_type"]
        if ct == CONTAINER_RO:
            trip_agg[key]["ro"] += r["qty"]
        elif ct == CONTAINER_TOTE:
            trip_agg[key]["tote"] += r["qty"]
        elif ct == CONTAINER_THUNG:
            trip_agg[key]["thung_carton"] += r["qty"]

    sheet3 = []
    for (date, trip), vals in sorted(trip_agg.items()):
        sheet3.append({
            "date": date,
            "trip_id": trip,
            "ro": vals["ro"],
            "tote": vals["tote"],
            "thung_carton": vals["thung_carton"],
        })

    print(f"  📋 Sheet 1 (Hàng Mát — ST): {len(sheet1)} rows")
    print(f"  📋 Sheet 2 (Hàng Đông — ST): {len(sheet2)} rows")
    print(f"  📋 Sheet 3 (Theo Trip):       {len(sheet3)} rows")

    return sheet1, sheet2, sheet3


def write_excel(sheet1, sheet2, sheet3, output_path):
    """Write 3 sheets to Excel with formatting."""
    wb = openpyxl.Workbook()

    # ── Styles ──
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill_mat = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")   # Blue for Mát
    header_fill_dong = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")  # Red for Đông
    header_fill_trip = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # Green for Trip
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=11)
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_left = Alignment(horizontal="left", vertical="center")
    num_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    # Alternating row fill
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    even_fill_dong = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    even_fill_trip = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

    def auto_width(ws, col_count):
        for i in range(1, col_count + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=i, max_col=i, values_only=False):
                for cell in row:
                    val = str(cell.value or "")
                    max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 45)

    # ══════════════════════════════════════
    # Sheet 1: Hàng Mát — Theo ST
    # ══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Hàng Mát — Theo ST"

    headers1 = ["Ngày", "Mã ST", "Tên Siêu Thị", "Mã Trip", "Rổ", "Thùng Carton, Bịch nguyên"]
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill_mat
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(sheet1, 2):
        vals = [
            r["date"].strftime("%d/%m/%Y"),
            r["store_code"],
            r["store_name"],
            r["trip_id"],
            r["ro"],
            r["thung_carton"],
        ]
        is_even = row_idx % 2 == 0
        for col_idx, v in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (1, 2, 4):
                cell.alignment = data_align_center
            elif col_idx in (5, 6):
                cell.alignment = num_align
            else:
                cell.alignment = data_align_left
            if is_even:
                cell.fill = even_fill

    auto_width(ws1, len(headers1))
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{len(sheet1) + 1}"
    ws1.freeze_panes = "A2"

    # ══════════════════════════════════════
    # Sheet 2: Hàng Đông — Theo ST
    # ══════════════════════════════════════
    ws2 = wb.create_sheet("Hàng Đông — Theo ST")

    headers2 = ["Ngày", "Mã ST", "Tên Siêu Thị", "Mã Trip", "Tote"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill_dong
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(sheet2, 2):
        vals = [
            r["date"].strftime("%d/%m/%Y"),
            r["store_code"],
            r["store_name"],
            r["trip_id"],
            r["tote"],
        ]
        is_even = row_idx % 2 == 0
        for col_idx, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (1, 2, 4):
                cell.alignment = data_align_center
            elif col_idx == 5:
                cell.alignment = num_align
            else:
                cell.alignment = data_align_left
            if is_even:
                cell.fill = even_fill_dong

    auto_width(ws2, len(headers2))
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(sheet2) + 1}"
    ws2.freeze_panes = "A2"

    # ══════════════════════════════════════
    # Sheet 3: Theo Trip (Xe)
    # ══════════════════════════════════════
    ws3 = wb.create_sheet("Theo Trip")

    headers3 = ["Ngày", "Mã Trip", "Rổ", "Tote", "Thùng Carton, Bịch nguyên"]
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill_trip
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(sheet3, 2):
        vals = [
            r["date"].strftime("%d/%m/%Y"),
            r["trip_id"],
            r["ro"],
            r["tote"],
            r["thung_carton"],
        ]
        is_even = row_idx % 2 == 0
        for col_idx, v in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (1, 2):
                cell.alignment = data_align_center
            elif col_idx in (3, 4, 5):
                cell.alignment = num_align
            else:
                cell.alignment = data_align_left
            if is_even:
                cell.fill = even_fill_trip

    auto_width(ws3, len(headers3))
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers3))}{len(sheet3) + 1}"
    ws3.freeze_panes = "A2"

    # ── Save ──
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n  ✅ Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Báo cáo Rổ/Tote/Thùng Carton — ĐÔNG MÁT")
    parser.add_argument("--months", type=str, default="3,4", help="Comma-separated months, e.g. 3,4")
    parser.add_argument("--year", type=int, default=2026)
    args = parser.parse_args()

    months = [int(m.strip()) for m in args.months.split(",")]
    year = args.year
    month_str = "_".join([f"T{m:02d}" for m in months])

    print("=" * 60)
    print(f"  BÁO CÁO RỔ / TOTE — ĐÔNG MÁT")
    print(f"  Tháng: {', '.join([f'T{m:02d}' for m in months])} / {year}")
    print("=" * 60)

    # 1. Load data
    print("\n📥 Loading trip data...")
    rows = load_trip_data(months, year)
    if not rows:
        print("  ❌ No data found!")
        return

    # 2. Classify mixed deliveries
    print("\n🔀 Classifying mixed deliveries...")
    rows = classify_mixed_deliveries(rows)

    # 3. Aggregate
    print("\n📊 Aggregating...")
    sheet1, sheet2, sheet3 = aggregate_data(rows)

    # 4. Write Excel
    output_path = os.path.join(OUTPUT, f"RO_TOTE_DONGMAT_{month_str}_{year}.xlsx")
    print(f"\n📝 Writing Excel...")
    write_excel(sheet1, sheet2, sheet3, output_path)

    # 5. Summary
    print(f"\n{'=' * 60}")
    print(f"  📊 Summary:")
    print(f"    Hàng Mát (ST):  {len(sheet1)} rows")
    print(f"    Hàng Đông (ST): {len(sheet2)} rows")
    print(f"    Theo Trip:      {len(sheet3)} rows")
    print(f"  ✅ Output: {output_path}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
