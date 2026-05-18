# -*- coding: utf-8 -*-
"""
Generate KFM Tonnage Summary by Warehouse
==========================================
Reads reconciliation (đối soát) files and master data to produce
an Excel report of daily KFM inventory tonnage (tấn) by warehouse:
  - Kho lưu hàng - Phân loại MÁT
  - Kho lưu hàng - Phân loại ĐÔNG
  - Kho lưu hàng TCNK

Output: Excel file with daily breakdown + monthly averages + total
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import io
import glob
import re
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Navigate up to repo root
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(REPO_ROOT, "script"))
from lib.sources import DOI_SOAT_DIR, MASTER_DATA_FILE, WEIGHT_DATA_FILE

OUTPUT_DIR = os.path.join(REPO_ROOT, 'output', 'artifacts', 'inventory')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# Date range: 30/03/2026 → 09/05/2026
# ─────────────────────────────────────────────
DATE_START = datetime(2026, 3, 30)
DATE_END = datetime(2026, 5, 9)

# ─────────────────────────────────────────────
# Category overrides (from generate.py)
# ─────────────────────────────────────────────
CATEGORY_OVERRIDES = {
    '8938527710565': 'MÁT',
    '1102969': 'TCNK',
    '8935101607211': 'ĐÔNG',   # O'FOOD - SỦI CẢO NHÂN THỊT HEO RAU CỦ 450G
    '8935335402859': 'MÁT',    # THỌ PHÁT - BÁNH BAO KHOAI MÔN PHÔ MAI 280G
    '8938528582574': 'MÁT',    # NGỌC TÚ - BÁNH GIÒ THỊT HEO TRỨNG CÚT 160G
}


def load_master_data():
    """
    Returns dict: barcode (str) → { 'category': 'ĐÔNG'|'MÁT'|'TCNK', 'chi_nhanh': str, 'phan_loai': str }
    """
    wb = openpyxl.load_workbook(MASTER_DATA_FILE, data_only=True, read_only=True)
    ws = wb['Sheet1']
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        chi_nhanh = str(row[0]).strip() if row[0] else ''
        barcode = str(row[1]).strip() if row[1] else ''
        phan_loai = str(row[4]).strip().upper() if row[4] else ''

        if not barcode:
            continue

        # Determine category
        if 'TCNK' in chi_nhanh:
            category = 'TCNK'
        elif chi_nhanh in ('KHO LƯU HÀNG', 'KHO QUÁ CẢNH'):
            if phan_loai in ('ĐÔNG', 'DONG'):
                category = 'ĐÔNG'
            elif phan_loai in ('MÁT', 'MAT'):
                category = 'MÁT'
            else:
                category = None
        else:
            category = None

        if category:
            if barcode not in mapping:
                mapping[barcode] = category
            # Also store zero-padded variants for matching
            for pad_len in (12, 13, 14):
                padded = barcode.zfill(pad_len)
                if padded != barcode and padded not in mapping:
                    mapping[padded] = category
            stripped = barcode.lstrip('0')
            if stripped and stripped != barcode and stripped not in mapping:
                mapping[stripped] = category

    wb.close()

    # Apply overrides
    mapping.update(CATEGORY_OVERRIDES)
    return mapping


def load_weight_data():
    """Returns dict: barcode (str) → weight in KG per item (float)"""
    WEIGHT_OVERRIDES = {
        '4014500028835': 0.2,
        '5701215042391': 0.125,
    }
    wb = openpyxl.load_workbook(WEIGHT_DATA_FILE, data_only=True, read_only=True)
    ws = wb['THONG_TIN_CO_BAN']
    mapping = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        barcode_base = str(row[0]).strip() if row[0] else ''
        barcode = str(row[1]).strip() if row[1] else ''
        weight_val = row[25] if len(row) > 25 else None
        weight_unit = str(row[26]).strip().lower() if len(row) > 26 and row[26] else ''

        if weight_val is None or weight_val == '':
            kg = None
        else:
            try:
                wv = float(weight_val)
            except (ValueError, TypeError):
                continue

            if weight_unit in ('g', 'gr', 'gram'):
                kg = wv / 1000.0
            elif weight_unit in ('kg',):
                kg = wv
            elif weight_unit in ('ml',):
                kg = wv / 1000.0
            elif weight_unit in ('l', 'lít', 'lit'):
                kg = wv
            else:
                kg = wv / 1000.0  # default assume grams

        if kg is not None:
            if barcode and barcode not in mapping:
                mapping[barcode] = kg
            if barcode_base and barcode_base not in mapping:
                mapping[barcode_base] = kg

    wb.close()
    mapping.update(WEIGHT_OVERRIDES)
    return mapping


def load_doi_soat_for_date_range(date_start, date_end):
    """Load reconciliation files within the date range. Returns list of day dicts."""
    files = glob.glob(os.path.join(DOI_SOAT_DIR, 'Đối soát tồn *.xlsx'))
    if not files:
        print("ERROR: No reconciliation files found!")
        return []

    all_days = []
    for fpath in sorted(files):
        # Parse date from filename
        fname = os.path.basename(fpath)
        match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', fname)
        if not match:
            continue
        file_date = datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))

        if file_date < date_start or file_date > date_end:
            continue

        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=False)
        if 'Đối soát tồn' not in wb.sheetnames:
            wb.close()
            continue

        ws = wb['Đối soát tồn']
        items = []
        date_val = None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            barcode = str(row[1]).strip() if row[1] else ''
            if not barcode:
                continue

            ten_hang = str(row[2]) if row[2] else ''
            ton_kfm = int(row[4]) if row[4] is not None else 0
            ngay = row[11]

            if ngay and date_val is None:
                if isinstance(ngay, datetime):
                    date_val = ngay
                elif isinstance(ngay, str):
                    try:
                        date_val = datetime.strptime(ngay, '%m/%d/%Y')
                    except ValueError:
                        try:
                            date_val = datetime.strptime(ngay, '%d/%m/%Y')
                        except ValueError:
                            pass

            items.append({
                'barcode': barcode,
                'ten_hang': ten_hang,
                'ton_KFM': ton_kfm,
            })

        if not date_val:
            date_val = file_date

        if items:
            all_days.append({
                'date': date_val,
                'date_str': date_val.strftime('%d/%m/%Y'),
                'items': items,
            })

        wb.close()

    all_days.sort(key=lambda x: x['date'])
    return all_days


def compute_tonnage(all_days, category_map, weight_map):
    """
    Compute daily KFM tonnage (tấn = KG / 1000) by warehouse.
    Returns:
      - daily_data: list of { 'date': datetime, 'date_str': str,
                               'MÁT': float, 'ĐÔNG': float, 'TCNK': float, 'TOTAL': float }
      - unmatched: dict barcode → ten_hang (barcodes not in master)
    """
    unmatched = {}
    daily_data = []

    for day in all_days:
        tonnage = {'MÁT': 0.0, 'ĐÔNG': 0.0, 'TCNK': 0.0}

        for item in day['items']:
            bc = item['barcode']
            cat = category_map.get(bc)
            if not cat:
                unmatched[bc] = item['ten_hang']
                continue

            kg_per = weight_map.get(bc, 0)
            # ton_KFM is item count, multiply by kg_per to get KG
            kg_total = item['ton_KFM'] * kg_per
            tonnage[cat] += kg_total

        # Convert KG → tấn (tons)
        total = 0.0
        for cat in ('MÁT', 'ĐÔNG', 'TCNK'):
            tonnage[cat] = tonnage[cat] / 1000.0  # KG → tấn
            total += tonnage[cat]

        daily_data.append({
            'date': day['date'],
            'date_str': day['date_str'],
            'MÁT': tonnage['MÁT'],
            'ĐÔNG': tonnage['ĐÔNG'],
            'TCNK': tonnage['TCNK'],
            'TOTAL': total,
        })

    return daily_data, unmatched


def generate_excel(daily_data):
    """Generate the Excel report with daily tonnage by warehouse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sản lượng tấn KFM"

    # ── Styles ──
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='2F5496')
    sub_header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    mat_fill = PatternFill(start_color='34D399', end_color='34D399', fill_type='solid')
    dong_fill = PatternFill(start_color='60A5FA', end_color='60A5FA', fill_type='solid')
    tcnk_fill = PatternFill(start_color='FBBF24', end_color='FBBF24', fill_type='solid')
    total_fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
    avg_fill = PatternFill(start_color='F97316', end_color='F97316', fill_type='solid')
    data_font = Font(name='Calibri', size=11)
    bold_font = Font(name='Calibri', bold=True, size=11)
    avg_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # ── Title ──
    ws.merge_cells('A1:E1')
    ws['A1'] = 'SẢN LƯỢNG TẤN KFM THEO KHO (Tháng 3-4/2026 + Tháng 5/2026 đến 09/05)'
    ws['A1'].font = title_font
    ws['A1'].alignment = left

    # ── Headers ──
    headers = ['Ngày', 'Kho LH - MÁT', 'Kho LH - ĐÔNG', 'Kho LH TCNK', 'TỔNG']
    fills = [header_fill, mat_fill, dong_fill, tcnk_fill, total_fill]
    row_num = 3
    for col_idx, (h, f) in enumerate(zip(headers, fills), 1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.font = sub_header_font
        cell.fill = f
        cell.alignment = center
        cell.border = thin_border

    # ── Group by month ──
    months = defaultdict(list)
    for d in daily_data:
        month_key = d['date'].strftime('%Y-%m')
        months[month_key].append(d)

    row_num = 4
    month_start_rows = {}  # For averages
    month_labels = {
        '2026-03': 'Tháng 3/2026',
        '2026-04': 'Tháng 4/2026',
        '2026-05': 'Tháng 5/2026',
    }
    month_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    month_font = Font(name='Calibri', bold=True, size=11, color='2F5496')

    avg_rows = []  # to track average row positions

    for mk in sorted(months.keys()):
        days = months[mk]
        label = month_labels.get(mk, mk)

        # Month separator row
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = month_font
        cell.fill = month_fill
        cell.alignment = left
        for c in range(1, 6):
            ws.cell(row=row_num, column=c).border = thin_border
            ws.cell(row=row_num, column=c).fill = month_fill
        row_num += 1
        month_start_row = row_num

        for d in days:
            ws.cell(row=row_num, column=1, value=d['date_str']).font = data_font
            ws.cell(row=row_num, column=1).alignment = center
            ws.cell(row=row_num, column=2, value=round(d['MÁT'], 3)).font = data_font
            ws.cell(row=row_num, column=2).number_format = '#,##0.000'
            ws.cell(row=row_num, column=3, value=round(d['ĐÔNG'], 3)).font = data_font
            ws.cell(row=row_num, column=3).number_format = '#,##0.000'
            ws.cell(row=row_num, column=4, value=round(d['TCNK'], 3)).font = data_font
            ws.cell(row=row_num, column=4).number_format = '#,##0.000'
            ws.cell(row=row_num, column=5, value=round(d['TOTAL'], 3)).font = data_font
            ws.cell(row=row_num, column=5).number_format = '#,##0.000'
            for c in range(1, 6):
                ws.cell(row=row_num, column=c).border = thin_border
                ws.cell(row=row_num, column=c).alignment = center
            row_num += 1

        month_end_row = row_num - 1
        num_days = len(days)

        # Average row for this month
        ws.cell(row=row_num, column=1, value=f'TB {label}').font = avg_font
        ws.cell(row=row_num, column=1).fill = avg_fill
        ws.cell(row=row_num, column=1).alignment = center

        mat_avg = sum(d['MÁT'] for d in days) / num_days if num_days else 0
        dong_avg = sum(d['ĐÔNG'] for d in days) / num_days if num_days else 0
        tcnk_avg = sum(d['TCNK'] for d in days) / num_days if num_days else 0
        total_avg = sum(d['TOTAL'] for d in days) / num_days if num_days else 0

        ws.cell(row=row_num, column=2, value=round(mat_avg, 3)).font = avg_font
        ws.cell(row=row_num, column=2).number_format = '#,##0.000'
        ws.cell(row=row_num, column=3, value=round(dong_avg, 3)).font = avg_font
        ws.cell(row=row_num, column=3).number_format = '#,##0.000'
        ws.cell(row=row_num, column=4, value=round(tcnk_avg, 3)).font = avg_font
        ws.cell(row=row_num, column=4).number_format = '#,##0.000'
        ws.cell(row=row_num, column=5, value=round(total_avg, 3)).font = avg_font
        ws.cell(row=row_num, column=5).number_format = '#,##0.000'

        for c in range(1, 6):
            ws.cell(row=row_num, column=c).border = thin_border
            ws.cell(row=row_num, column=c).fill = avg_fill
            ws.cell(row=row_num, column=c).alignment = center

        avg_rows.append(row_num)
        row_num += 1

    # ── Grand Total Average ──
    row_num += 1
    all_days_count = len(daily_data)
    grand_mat = sum(d['MÁT'] for d in daily_data) / all_days_count if all_days_count else 0
    grand_dong = sum(d['ĐÔNG'] for d in daily_data) / all_days_count if all_days_count else 0
    grand_tcnk = sum(d['TCNK'] for d in daily_data) / all_days_count if all_days_count else 0
    grand_total = sum(d['TOTAL'] for d in daily_data) / all_days_count if all_days_count else 0

    grand_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    grand_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')

    ws.cell(row=row_num, column=1, value='TB TOÀN BỘ').font = grand_font
    ws.cell(row=row_num, column=1).fill = grand_fill
    ws.cell(row=row_num, column=1).alignment = center
    ws.cell(row=row_num, column=2, value=round(grand_mat, 3)).font = grand_font
    ws.cell(row=row_num, column=2).number_format = '#,##0.000'
    ws.cell(row=row_num, column=3, value=round(grand_dong, 3)).font = grand_font
    ws.cell(row=row_num, column=3).number_format = '#,##0.000'
    ws.cell(row=row_num, column=4, value=round(grand_tcnk, 3)).font = grand_font
    ws.cell(row=row_num, column=4).number_format = '#,##0.000'
    ws.cell(row=row_num, column=5, value=round(grand_total, 3)).font = grand_font
    ws.cell(row=row_num, column=5).number_format = '#,##0.000'

    for c in range(1, 6):
        ws.cell(row=row_num, column=c).border = thin_border
        ws.cell(row=row_num, column=c).fill = grand_fill
        ws.cell(row=row_num, column=c).alignment = center

    # ── Grand Total Sum ──
    row_num += 1
    total_sum_mat = sum(d['MÁT'] for d in daily_data)
    total_sum_dong = sum(d['ĐÔNG'] for d in daily_data)
    total_sum_tcnk = sum(d['TCNK'] for d in daily_data)
    total_sum_all = sum(d['TOTAL'] for d in daily_data)

    sum_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    ws.cell(row=row_num, column=1, value='TỔNG SẢN LƯỢNG').font = grand_font
    ws.cell(row=row_num, column=1).fill = sum_fill
    ws.cell(row=row_num, column=1).alignment = center
    ws.cell(row=row_num, column=2, value=round(total_sum_mat, 3)).font = grand_font
    ws.cell(row=row_num, column=2).number_format = '#,##0.000'
    ws.cell(row=row_num, column=3, value=round(total_sum_dong, 3)).font = grand_font
    ws.cell(row=row_num, column=3).number_format = '#,##0.000'
    ws.cell(row=row_num, column=4, value=round(total_sum_tcnk, 3)).font = grand_font
    ws.cell(row=row_num, column=4).number_format = '#,##0.000'
    ws.cell(row=row_num, column=5, value=round(total_sum_all, 3)).font = grand_font
    ws.cell(row=row_num, column=5).number_format = '#,##0.000'

    for c in range(1, 6):
        ws.cell(row=row_num, column=c).border = thin_border
        ws.cell(row=row_num, column=c).fill = sum_fill
        ws.cell(row=row_num, column=c).alignment = center

    # ── Column widths ──
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    return wb


def main():
    print("=" * 60)
    print("KFM Tonnage Summary Generator")
    print("=" * 60)
    print(f"Date range: {DATE_START.strftime('%d/%m/%Y')} → {DATE_END.strftime('%d/%m/%Y')}")

    print("\n[1/4] Loading master data...")
    category_map = load_master_data()
    print(f"  → {len(category_map)} barcode mappings loaded")

    print("\n[2/4] Loading weight data...")
    weight_map = load_weight_data()
    print(f"  → {len(weight_map)} weight mappings loaded")

    print("\n[3/4] Loading reconciliation files...")
    all_days = load_doi_soat_for_date_range(DATE_START, DATE_END)
    print(f"  → {len(all_days)} days loaded")

    if not all_days:
        print("ERROR: No data found for the date range!")
        return

    print("\n[4/4] Computing tonnage...")
    daily_data, unmatched = compute_tonnage(all_days, category_map, weight_map)

    # Print summary
    print("\n── Daily Tonnage Summary ──")
    print(f"{'Ngày':<14} {'MÁT':>10} {'ĐÔNG':>10} {'TCNK':>10} {'TOTAL':>10}")
    print("-" * 58)
    for d in daily_data:
        print(f"{d['date_str']:<14} {d['MÁT']:>10.3f} {d['ĐÔNG']:>10.3f} {d['TCNK']:>10.3f} {d['TOTAL']:>10.3f}")

    # Monthly averages
    from collections import defaultdict
    months = defaultdict(list)
    for d in daily_data:
        mk = d['date'].strftime('%Y-%m')
        months[mk].append(d)

    print("\n── Monthly Averages (tấn/ngày) ──")
    for mk in sorted(months.keys()):
        days = months[mk]
        n = len(days)
        mat_avg = sum(d['MÁT'] for d in days) / n
        dong_avg = sum(d['ĐÔNG'] for d in days) / n
        tcnk_avg = sum(d['TCNK'] for d in days) / n
        total_avg = sum(d['TOTAL'] for d in days) / n
        print(f"  {mk}: MÁT={mat_avg:.3f}  ĐÔNG={dong_avg:.3f}  TCNK={tcnk_avg:.3f}  TOTAL={total_avg:.3f} ({n} ngày)")

    # Generate Excel
    print("\n── Generating Excel ──")
    wb = generate_excel(daily_data)
    out_path = os.path.join(OUTPUT_DIR, 'san_luong_tan_KFM_T3_T5.xlsx')
    wb.save(out_path)
    print(f"  → Saved: {out_path}")

    # Report unmatched barcodes
    if unmatched:
        print(f"\n⚠ BARCODE CHƯA ĐƯỢC PHÂN LOẠI: {len(unmatched)} mã")
        print("-" * 70)
        for bc, name in sorted(unmatched.items()):
            print(f"  {bc:<20} {name}")
    else:
        print("\n✓ Tất cả barcode đều đã được phân loại!")

    print("\n✅ Done!")


if __name__ == '__main__':
    main()
