"""Generate Lich_Khai_Truong_NSO.xlsx archive from nso_stores.json."""
import json, sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
STORES_PATH = os.path.join(BASE, 'data', 'nso', 'nso_stores.json')
DSST_PATH = os.path.join(BASE, 'data', 'dsst_cache.json')
OUTPUT_DIR = os.path.join(BASE, 'output', 'artifacts', 'nso')
os.makedirs(OUTPUT_DIR, exist_ok=True)

stores = json.load(open(STORES_PATH, 'r', encoding='utf-8'))
dsst = json.load(open(DSST_PATH, 'r', encoding='utf-8')) if os.path.exists(DSST_PATH) else {}

# Sort by opening_date
from datetime import datetime
def parse_d(s):
    try: return datetime.strptime(s, '%d/%m/%Y')
    except: return datetime(2099, 1, 1)
stores.sort(key=lambda s: parse_d(s.get('opening_date', '')))

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Lịch Khai Trương NSO'

# Styles
hdr_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
data_font = Font(name='Calibri', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')

# Headers
headers = ['STT', 'Tên Siêu Thị', 'Code', 'Ngày Khai Trương', 'Version', 'Trạng Thái']
widths = [6, 45, 10, 18, 12, 25]
for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

# Data
from datetime import date
today = date.today()
for i, s in enumerate(stores, 1):
    code = s.get('code', '')
    d_entry = dsst.get(code, {})
    
    # Display name: DSST name_full if available, else store name_full
    name = d_entry.get('name_full') or s.get('name_full') or s.get('name_mail', '')
    if d_entry.get('name_system'):
        name = f"{d_entry['name_system']} - {name}"
    elif s.get('name_system'):
        name = f"{s['name_system']} - {name}"
    
    opening = s.get('opening_date', '')
    version = s.get('version', '')
    
    # Status
    try:
        d = datetime.strptime(opening, '%d/%m/%Y').date()
        delta = (today - d).days
        if delta > 3:
            status = 'Đã khai trương'
        elif 0 <= delta <= 3:
            status = 'Đang khai trương'
        else:
            status = 'Sắp khai trương'
    except:
        status = ''
    
    if s.get('original_date') and s.get('original_date') != opening:
        status = f'Dời lịch: {s["original_date"]} → {opening}'
    
    row = i + 1
    ws.cell(row=row, column=1, value=i).font = data_font
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=2, value=name).font = data_font
    ws.cell(row=row, column=2).alignment = left
    ws.cell(row=row, column=3, value=code or '—').font = data_font
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=4, value=opening).font = data_font
    ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=5, value=version or '').font = data_font
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=6, value=status).font = data_font
    ws.cell(row=row, column=6).alignment = left
    
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = thin_border

# Freeze header
ws.freeze_panes = 'A2'

# Save
out_path = os.path.join(OUTPUT_DIR, 'Lich_Khai_Truong_NSO.xlsx')
wb.save(out_path)
print(f"✅ {out_path} ({len(stores)} stores, {os.path.getsize(out_path):,} bytes)")
